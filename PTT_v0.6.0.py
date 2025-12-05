import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from pathlib import Path
import json

# Fichiers de données
import os
from pathlib import Path

# Dossier de données OneDrive
DATA_DIR = Path.home() / "OneDrive - STEF" / "O_BEL_transport_tubize - Documents Service Transport" / "[09] Planning transport [BETA]" / "_data"

# Créer le dossier s'il n'existe pas
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Chemins des fichiers
MISSIONS_FILE = DATA_DIR / "missions.json"
VOYAGES_FILE = DATA_DIR / "voyages.json"
CHAUFFEURS_FILE = DATA_DIR / "chauffeurs.json"
TARIFS_SST_FILE = DATA_DIR / "tarifs_sst.json"
REVENUS_FILE = DATA_DIR / "revenus_palettes.json"
SST_EMAILS_FILE = DATA_DIR / "sst_emails.json"
ANNOUNCEMENT_CONFIG_FILE = DATA_DIR / "announcement_config.json"
ANNOUNCEMENT_HISTORY_FILE = DATA_DIR / "announcement_history.json"

from datetime import date, datetime, timedelta
import uuid
import getpass
import os

# Imports optionnels pour l'export
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Import matplotlib pour les graphiques
try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
    from matplotlib.figure import Figure
    import matplotlib.dates as mdates
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

# Import Outlook (Windows uniquement)
try:
    import win32com.client
    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False

# Configuration
COMPANY_OD_FOLDER = "OneDrive - STEF"
SUBPATH = Path("O_BEL_transport_tubize - Documents Service Transport") / "[09] Planning transport [BETA]"

# =============================================================================
# SYSTÈME SAURON - Logging et surveillance des activités utilisateurs
# =============================================================================

class ActivityLogger:
    """
    Système de logging des activités utilisateurs.
    Enregistre toutes les actions dans des fichiers JSON par utilisateur.
    """
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self.logs_dir = None
        self.current_user = None
        self.session_id = None
        self.session_start = None
        self.user_log_file = None
        
    def initialize(self, root_dir, username):
        """Initialiser le logger avec le dossier racine et l'utilisateur"""
        self.logs_dir = root_dir / "_logs"
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        self.current_user = username.upper()
        self.session_id = str(uuid.uuid4())[:8]
        self.session_start = datetime.now()
        self.user_log_file = self.logs_dir / f"{self.current_user}.json"
        
        # Créer le fichier de log utilisateur si inexistant
        if not self.user_log_file.exists():
            self._save_user_logs({
                "user": self.current_user,
                "created": datetime.now().isoformat(),
                "sessions": [],
                "actions": []
            })
        
        # Enregistrer le début de session
        self.log_action("SESSION_START", {
            "session_id": self.session_id,
            "timestamp": self.session_start.isoformat()
        })
    
    def _load_user_logs(self):
        """Charger les logs de l'utilisateur courant"""
        try:
            if self.user_log_file and self.user_log_file.exists():
                with open(self.user_log_file, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            print(f"Erreur chargement logs: {e}")
        return {
            "user": self.current_user,
            "created": datetime.now().isoformat(),
            "sessions": [],
            "actions": []
        }
    
    def _save_user_logs(self, data):
        """Sauvegarder les logs de l'utilisateur courant"""
        try:
            if self.user_log_file:
                with open(self.user_log_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erreur sauvegarde logs: {e}")
    
    def log_action(self, action_type, details=None, before_state=None, after_state=None):
        """
        Enregistrer une action utilisateur.
        
        Args:
            action_type: Type d'action (ex: "MISSION_CREATE", "CHAUFFEUR_EDIT", etc.)
            details: Dictionnaire avec les détails de l'action
            before_state: État avant modification (pour les éditions)
            after_state: État après modification (pour les éditions)
        """
        if not self.logs_dir:
            return
        
        try:
            logs = self._load_user_logs()
            
            action_entry = {
                "id": str(uuid.uuid4())[:12],
                "session_id": self.session_id,
                "timestamp": datetime.now().isoformat(),
                "type": action_type,
                "details": details or {}
            }
            
            if before_state is not None:
                action_entry["before"] = before_state
            if after_state is not None:
                action_entry["after"] = after_state
            
            logs["actions"].append(action_entry)
            
            self._save_user_logs(logs)
            
        except Exception as e:
            print(f"Erreur log_action: {e}")
    
    def log_session_end(self):
        """Enregistrer la fin de session"""
        if not self.logs_dir or not self.session_start:
            return
        
        session_end = datetime.now()
        duration_seconds = (session_end - self.session_start).total_seconds()
        
        self.log_action("SESSION_END", {
            "session_id": self.session_id,
            "duration_seconds": int(duration_seconds),
            "duration_formatted": self._format_duration(duration_seconds)
        })
        
        # Mettre à jour les statistiques de session
        try:
            logs = self._load_user_logs()
            logs["sessions"].append({
                "session_id": self.session_id,
                "start": self.session_start.isoformat(),
                "end": session_end.isoformat(),
                "duration_seconds": int(duration_seconds)
            })
            self._save_user_logs(logs)
        except Exception as e:
            print(f"Erreur log_session_end: {e}")
    
    def _format_duration(self, seconds):
        """Formater une durée en heures:minutes:secondes"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    
    def get_all_users_logs(self):
        """Récupérer les logs de tous les utilisateurs"""
        if not self.logs_dir or not self.logs_dir.exists():
            return {}
        
        all_logs = {}
        try:
            for log_file in self.logs_dir.glob("*.json"):
                try:
                    with open(log_file, "r", encoding="utf-8") as f:
                        user_logs = json.load(f)
                        username = log_file.stem
                        all_logs[username] = user_logs
                except Exception as e:
                    print(f"Erreur lecture {log_file}: {e}")
        except Exception as e:
            print(f"Erreur get_all_users_logs: {e}")
        
        return all_logs
    
    def get_active_sessions(self):
        """
        Déterminer les utilisateurs potentiellement actifs
        (sessions démarrées mais pas terminées dans les dernières 24h)
        """
        all_logs = self.get_all_users_logs()
        active_users = []
        now = datetime.now()
        
        for username, logs in all_logs.items():
            actions = logs.get("actions", [])
            if not actions:
                continue
            
            # Trouver la dernière action
            last_action = actions[-1]
            try:
                last_time = datetime.fromisoformat(last_action["timestamp"])
                time_diff = (now - last_time).total_seconds()
                
                # Considérer actif si dernière action < 1 heure
                if time_diff < 3600:
                    # Vérifier si la session n'est pas terminée
                    session_id = last_action.get("session_id")
                    is_ended = any(
                        a.get("type") == "SESSION_END" and a.get("session_id") == session_id
                        for a in actions[-50:]  # Vérifier les 50 dernières actions
                    )
                    if not is_ended:
                        active_users.append({
                            "user": username,
                            "last_action": last_action,
                            "last_time": last_time
                        })
            except Exception:
                pass
        
        return active_users
    
    def get_user_stats(self, username):
        """Calculer les statistiques d'un utilisateur"""
        all_logs = self.get_all_users_logs()
        if username not in all_logs:
            return None
        
        logs = all_logs[username]
        sessions = logs.get("sessions", [])
        actions = logs.get("actions", [])
        
        # Calculer le temps total de connexion
        total_seconds = sum(s.get("duration_seconds", 0) for s in sessions)
        
        # Compter les actions par type
        action_counts = {}
        for action in actions:
            action_type = action.get("type", "UNKNOWN")
            action_counts[action_type] = action_counts.get(action_type, 0) + 1
        
        # Dernière connexion
        last_session = sessions[-1] if sessions else None
        last_login = last_session.get("start") if last_session else None
        
        return {
            "total_sessions": len(sessions),
            "total_time_seconds": total_seconds,
            "total_time_formatted": self._format_duration(total_seconds),
            "total_actions": len(actions),
            "action_counts": action_counts,
            "last_login": last_login
        }


# Instance globale du logger
activity_logger = ActivityLogger()
EU_COUNTRIES = [
    "Belgique","Allemagne","France","Luxembourg","Pays-Bas",
]

COUNTRY_COLORS = {
    "Belgique": "#FAFAFA",
    "France": "#E8F4FD",
    "Allemagne": "#FFF9E6",
    "Pays-Bas": "#FFF0E6",
    "Luxembourg": "#E8F8F0",
    "Espagne": "#FFEBEE",
    "Italie": "#F0FFF0",
}

def resolve_root_dir() -> Path:
    ptt_override = os.environ.get("PTT_ROOT_DIR")
    if ptt_override:
        return Path(ptt_override)
    
    odc = os.environ.get("OneDriveCommercial")
    if odc:
        return Path(odc) / SUBPATH
    
    home = Path.home()
    
    company_base = home / COMPANY_OD_FOLDER
    if company_base.exists():
        return company_base / SUBPATH
    
    od = os.environ.get("OneDrive")
    if od:
        od_base = Path(od)
        sibling_company = od_base.with_name(COMPANY_OD_FOLDER)
        if sibling_company.exists():
            return sibling_company / SUBPATH
        return (od_base / COMPANY_OD_FOLDER) / SUBPATH
    
    return (home / COMPANY_OD_FOLDER) / SUBPATH

ROOT_DIR = resolve_root_dir()

def get_desktop_path():
    home = Path.home()
    desktop = home / "Desktop"
    if not desktop.exists():
        desktop = home / "Bureau"
    if not desktop.exists():
        desktop = home
    return desktop

DEFAULT_ADMIN = "KEKENJ.J"

# ---------- Utilitaires fichiers ----------

def ensure_base_dirs():
    (ROOT_DIR / "_data").mkdir(parents=True, exist_ok=True)
    (ROOT_DIR / "_planning").mkdir(parents=True, exist_ok=True)
    (ROOT_DIR / "_export").mkdir(parents=True, exist_ok=True)

def load_json(filename, default=None):
    """Charger des données depuis JSON"""
    try:
        # Convertir en Path si c'est une string
        if isinstance(filename, str):
            from pathlib import Path
            path = Path(filename)
        else:
            path = filename
        
        if not path.exists():
            return default if default is not None else {}
        
        with open(str(path), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Erreur lors du chargement de {filename}: {e}")
        return default if default is not None else {}


def save_json(filename, data):
    """Sauvegarder des données en JSON"""
    try:
        # Convertir en Path si c'est une string
        if isinstance(filename, str):
            from pathlib import Path
            path = Path(filename)
        else:
            path = filename
        
        # Créer le fichier
        with open(str(path), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        # Vérifier que le fichier existe
        if path.exists():
            print(f"✓ Fichier sauvegardé: {path} ({path.stat().st_size} octets)")
        else:
            print(f"⚠ Problème: fichier {path} non créé")
            
    except Exception as e:
        print(f"✗ Erreur lors de la sauvegarde de {filename}: {e}")
        import traceback
        traceback.print_exc()


def init_default_data():
    ensure_base_dirs()
    data_dir = ROOT_DIR / "_data"

    sst_path = data_dir / "sst.json"
    if not sst_path.exists():
        save_json(sst_path, ["AMF", "FES", "DST", "BBS", "CLAYDRAY"])

    voyages_path = data_dir / "voyages.json"
    if not voyages_path.exists():
        save_json(voyages_path, [])

    chauffeurs_path = data_dir / "chauffeurs.json"
    if not chauffeurs_path.exists():
        save_json(chauffeurs_path, [])

    dispo_path = data_dir / "dispo_chauffeurs.json"
    if not dispo_path.exists():
        save_json(dispo_path, [])
    
    # Finance
    tarifs_sst_path = data_dir / "tarifs_sst.json"
    if not tarifs_sst_path.exists():
        save_json(tarifs_sst_path, {})
    
    revenus_palettes_path = data_dir / "revenus_palettes.json"
    if not revenus_palettes_path.exists():
        save_json(revenus_palettes_path, {})

    rights_path = data_dir / "users_rights.json"
    if not rights_path.exists():
        current_user = DEFAULT_ADMIN.upper()
        default_rights = {
            "roles": {
                "viewer": {
                    "view_planning": True,
                    "edit_planning": False,
                    "view_drivers": True,
                    "manage_drivers": False,
                    "edit_driver_planning": False,
                    "manage_rights": False,
                    "manage_voyages": False,
                    "generate_planning": False,
                    "edit_past_planning": False,
                    "edit_past_planning_advanced": False,
                    "view_finance": False,
                    "manage_finance": False,
                    "view_analyse": False,
                    "view_sauron": False,
                    "send_announcements": False,
                    "manage_announcements_config": False,
                },
                "planner": {
                    "view_planning": True,
                    "edit_planning": True,
                    "view_drivers": True,
                    "manage_drivers": False,
                    "edit_driver_planning": False,
                    "manage_rights": False,
                    "manage_voyages": True,
                    "generate_planning": False,
                    "edit_past_planning": False,
                    "edit_past_planning_advanced": False,
                    "view_finance": False,
                    "manage_finance": False,
                    "view_analyse": False,
                    "view_sauron": False,
                    "send_announcements": True,
                    "manage_announcements_config": False,
                },
                "planner_advanced": {
                    "view_planning": True,
                    "edit_planning": True,
                    "view_drivers": True,
                    "manage_drivers": False,
                    "edit_driver_planning": False,
                    "manage_rights": False,
                    "manage_voyages": True,
                    "generate_planning": False,
                    "edit_past_planning": True,
                    "edit_past_planning_advanced": True,
                    "view_finance": True,
                    "manage_finance": False,
                    "view_analyse": False,
                    "view_sauron": False,
                    "send_announcements": True,
                    "manage_announcements_config": True,
                },
                "driver_admin": {
                    "view_planning": True,
                    "edit_planning": False,
                    "view_drivers": True,
                    "manage_drivers": True,
                    "edit_driver_planning": True,
                    "manage_rights": False,
                    "manage_voyages": False,
                    "generate_planning": False,
                    "edit_past_planning": False,
                    "edit_past_planning_advanced": False,
                    "view_finance": False,
                    "manage_finance": False,
                    "view_analyse": False,
                    "view_sauron": False,
                    "send_announcements": False,
                    "manage_announcements_config": False,
                },
                "admin": {
                    "view_planning": True,
                    "edit_planning": True,
                    "view_drivers": True,
                    "manage_drivers": True,
                    "edit_driver_planning": True,
                    "manage_rights": True,
                    "manage_voyages": True,
                    "generate_planning": True,
                    "edit_past_planning": True,
                    "edit_past_planning_advanced": True,
                    "view_finance": True,
                    "manage_finance": True,
                    "view_analyse": True,
                    "view_sauron": True,
                    "send_announcements": True,
                    "manage_announcements_config": True,
                },
                "analyse": {
                    "view_planning": True,
                    "edit_planning": False,
                    "view_drivers": True,
                    "manage_drivers": False,
                    "edit_driver_planning": False,
                    "manage_rights": False,
                    "manage_voyages": False,
                    "generate_planning": False,
                    "edit_past_planning": False,
                    "edit_past_planning_advanced": False,
                    "view_finance": True,
                    "manage_finance": False,
                    "view_analyse": True,
                    "view_sauron": False,
                    "send_announcements": False,
                    "manage_announcements_config": False,
                },
                "finance": {
                    "view_planning": True,
                    "edit_planning": False,
                    "view_drivers": False,
                    "manage_drivers": False,
                    "edit_driver_planning": False,
                    "manage_rights": False,
                    "manage_voyages": False,
                    "generate_planning": False,
                    "edit_past_planning": False,
                    "edit_past_planning_advanced": False,
                    "view_finance": True,
                    "manage_finance": True,
                    "view_analyse": False,
                    "view_sauron": False,
                    "send_announcements": False,
                    "manage_announcements_config": False,
                },
            },
            "users": {
                current_user: ["admin"]
            },
        }
        save_json(rights_path, default_rights)

def get_week_folder(d: date) -> str:
    week = d.isocalendar()[1]
    return f"Semaine_{week:02d}"

def get_planning_day_dir(d: date) -> Path:
    year = f"{d.year:04d}"
    month = f"{d.month:02d}"
    week_folder = get_week_folder(d)
    return ROOT_DIR / "_planning" / year / month / week_folder / d.strftime("%Y-%m-%d")

def list_existing_dates():
    base = ROOT_DIR / "_planning"
    if not base.exists():
        return []
    dates = []
    for year_dir in base.iterdir():
        if not year_dir.is_dir():
            continue
        for month_dir in year_dir.iterdir():
            if not month_dir.is_dir():
                continue
            for week_dir in month_dir.iterdir():
                if not week_dir.is_dir():
                    continue
                for day_dir in week_dir.iterdir():
                    if day_dir.is_dir():
                        try:
                            d = datetime.strptime(day_dir.name, "%Y-%m-%d").date()
                            dates.append(format_date_display(d))
                        except ValueError:
                            pass
    dates.sort(key=lambda x: datetime.strptime(x, "%d/%m/%Y"))
    return dates

def generate_time_choices():
    times = []
    for minutes in range(0, 28 * 60 + 1, 15):
        h = minutes // 60
        m = minutes % 60
        times.append(f"{h:02d}:{m:02d}")
    return times

def format_date_display(d: date) -> str:
    return d.strftime("%d/%m/%Y")

def parse_date_input(s: str) -> date:
    s = s.strip()
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        pass
    raise ValueError(f"Format de date invalide: '{s}'. Formats acceptés: JJ/MM/AAAA ou AAAA-MM-JJ")

def format_date_internal(d: date) -> str:
    return d.strftime("%Y-%m-%d")

TIME_CHOICES = generate_time_choices()

# ---------- Paramètres locaux ----------
APP_NAME = "PTT"
DEFAULT_SETTINGS = {"auto_refresh_enabled": True, "auto_refresh_seconds": 10}

def _settings_path() -> Path:
    base = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA") or str(Path.home())
    p = Path(base) / APP_NAME / "settings.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p

def load_local_settings() -> dict:
    p = _settings_path()
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            return {**DEFAULT_SETTINGS, **data}
        except Exception:
            pass
    return DEFAULT_SETTINGS.copy()

def save_local_settings(data: dict) -> None:
    _settings_path().write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

# =============================================================================
# SYSTÈME DE CACHE LOCAL - Pré-téléchargement des plannings
# =============================================================================

import threading
import shutil
import time as time_module

class PlanningCache:
    """
    Système de cache local pour les plannings.
    Pré-télécharge les plannings des dates proches en arrière-plan
    pour éliminer la latence lors du changement de date.
    """

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True

        # Dossier de cache dans AppData
        base = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA") or str(Path.home())
        self.cache_dir = Path(base) / APP_NAME / "cache" / "planning"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        # Fichier de métadonnées du cache
        self.meta_file = self.cache_dir / "_cache_meta.json"
        self.cache_meta = self._load_meta()

        # Configuration du cache
        self.days_before = 2   # Jours avant aujourd'hui à mettre en cache
        self.days_after = 5    # Jours après aujourd'hui à mettre en cache
        self.refresh_interval = 30  # Intervalle de rafraîchissement en secondes

        # État du thread de fond
        self._stop_event = threading.Event()
        self._cache_thread = None
        self._lock = threading.Lock()

        # File d'attente pour les dates prioritaires
        self._priority_dates = []

        # Callback pour notifier l'UI
        self._on_cache_updated = None

    def _load_meta(self) -> dict:
        """Charger les métadonnées du cache"""
        try:
            if self.meta_file.exists():
                with open(self.meta_file, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            print(f"[Cache] Erreur chargement métadonnées: {e}")
        return {"dates": {}, "last_full_refresh": None}

    def _save_meta(self):
        """Sauvegarder les métadonnées du cache"""
        try:
            with open(self.meta_file, "w", encoding="utf-8") as f:
                json.dump(self.cache_meta, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[Cache] Erreur sauvegarde métadonnées: {e}")

    def _get_cache_path(self, d: date) -> Path:
        """Obtenir le chemin du cache pour une date"""
        return self.cache_dir / d.strftime("%Y-%m-%d")

    def _get_source_path(self, d: date) -> Path:
        """Obtenir le chemin source (OneDrive) pour une date"""
        return get_planning_day_dir(d)

    def _cache_date(self, d: date) -> bool:
        """
        Mettre en cache une date spécifique.
        Retourne True si le cache a été mis à jour.
        """
        source_dir = self._get_source_path(d)
        cache_path = self._get_cache_path(d)
        date_str = d.strftime("%Y-%m-%d")

        try:
            # Vérifier si le dossier source existe
            if not source_dir.exists():
                # Pas de planning pour cette date - supprimer le cache si existant
                if cache_path.exists():
                    shutil.rmtree(cache_path)
                    with self._lock:
                        if date_str in self.cache_meta["dates"]:
                            del self.cache_meta["dates"][date_str]
                            self._save_meta()
                return False

            # Calculer le timestamp le plus récent des fichiers source
            source_files = list(source_dir.glob("*.json"))
            if not source_files:
                return False

            latest_source_mtime = max(f.stat().st_mtime for f in source_files)

            # Vérifier si le cache est à jour
            with self._lock:
                cached_info = self.cache_meta["dates"].get(date_str, {})
                cached_mtime = cached_info.get("source_mtime", 0)

                if cache_path.exists() and cached_mtime >= latest_source_mtime:
                    # Cache déjà à jour
                    return False

            # Créer/mettre à jour le cache
            cache_path.mkdir(parents=True, exist_ok=True)

            # Copier tous les fichiers JSON
            for src_file in source_files:
                dst_file = cache_path / src_file.name
                shutil.copy2(src_file, dst_file)

            # Mettre à jour les métadonnées
            with self._lock:
                self.cache_meta["dates"][date_str] = {
                    "source_mtime": latest_source_mtime,
                    "cached_at": datetime.now().isoformat(),
                    "file_count": len(source_files)
                }
                self._save_meta()

            print(f"[Cache] Date {date_str} mise en cache ({len(source_files)} fichiers)")
            return True

        except Exception as e:
            print(f"[Cache] Erreur mise en cache {date_str}: {e}")
            return False

    def get_cached_planning(self, d: date) -> list:
        """
        Récupérer le planning depuis le cache.
        Retourne None si pas en cache ou si les fichiers source ont changé.
        Vérifie TOUJOURS les timestamps source pour détecter les modifications
        faites par d'autres utilisateurs.
        """
        cache_path = self._get_cache_path(d)
        date_str = d.strftime("%Y-%m-%d")
        source_dir = self._get_source_path(d)

        with self._lock:
            if date_str not in self.cache_meta["dates"]:
                return None
            if not cache_path.exists():
                return None
            cached_info = self.cache_meta["dates"].get(date_str, {})
            cached_mtime = cached_info.get("source_mtime", 0)

        # IMPORTANT: Vérifier si les fichiers source ont changé
        # (détecte les modifications faites par d'autres utilisateurs)
        try:
            if source_dir.exists():
                source_files = list(source_dir.glob("*.json"))
                if source_files:
                    latest_source_mtime = max(f.stat().st_mtime for f in source_files)
                    # Si les fichiers source ont été modifiés depuis la mise en cache
                    if latest_source_mtime > cached_mtime:
                        print(f"[Cache] Fichiers source modifiés pour {date_str}, invalidation du cache")
                        # Invalider le cache et forcer la mise à jour
                        with self._lock:
                            if date_str in self.cache_meta["dates"]:
                                del self.cache_meta["dates"][date_str]
                                self._save_meta()
                        self.prioritize_date(d)
                        return None  # Forcer rechargement depuis source

                    # Vérifier aussi si le nombre de fichiers a changé
                    cached_file_count = cached_info.get("file_count", 0)
                    if len(source_files) != cached_file_count:
                        print(f"[Cache] Nombre de fichiers changé pour {date_str}, invalidation du cache")
                        with self._lock:
                            if date_str in self.cache_meta["dates"]:
                                del self.cache_meta["dates"][date_str]
                                self._save_meta()
                        self.prioritize_date(d)
                        return None
        except Exception as e:
            print(f"[Cache] Erreur vérification source {date_str}: {e}")
            # En cas d'erreur, on utilise quand même le cache
            pass

        try:
            missions = []
            for file in cache_path.glob("*.json"):
                data = load_json(file, None)
                if data:
                    # Stocker le chemin original (pas le cache)
                    source_path = self._get_source_path(d) / file.name
                    data["_path"] = source_path.as_posix()
                    missions.append(data)
            return missions
        except Exception as e:
            print(f"[Cache] Erreur lecture cache {date_str}: {e}")
            return None

    def is_cached(self, d: date) -> bool:
        """Vérifier si une date est en cache"""
        date_str = d.strftime("%Y-%m-%d")
        with self._lock:
            return date_str in self.cache_meta["dates"]

    def prioritize_date(self, d: date):
        """Ajouter une date à la liste prioritaire pour mise en cache immédiate"""
        with self._lock:
            if d not in self._priority_dates:
                self._priority_dates.insert(0, d)

    def _get_dates_to_cache(self) -> list:
        """Obtenir la liste des dates à mettre en cache"""
        today = date.today()
        dates = []

        # D'abord les dates prioritaires
        with self._lock:
            priority = self._priority_dates.copy()
            self._priority_dates.clear()
        dates.extend(priority)

        # Puis les dates autour d'aujourd'hui
        for delta in range(-self.days_before, self.days_after + 1):
            d = today + timedelta(days=delta)
            if d not in dates:
                dates.append(d)

        return dates

    def _background_cache_loop(self):
        """Boucle de fond pour mettre en cache les plannings"""
        print("[Cache] Thread de cache démarré")

        while not self._stop_event.is_set():
            try:
                dates_to_cache = self._get_dates_to_cache()
                updated = False

                for d in dates_to_cache:
                    if self._stop_event.is_set():
                        break

                    if self._cache_date(d):
                        updated = True
                        # Petite pause entre chaque date pour ne pas surcharger
                        time_module.sleep(0.5)

                # Notifier l'UI si du nouveau contenu est disponible
                if updated and self._on_cache_updated:
                    try:
                        self._on_cache_updated()
                    except Exception:
                        pass

            except Exception as e:
                print(f"[Cache] Erreur boucle cache: {e}")

            # Attendre avant le prochain cycle
            self._stop_event.wait(self.refresh_interval)

        print("[Cache] Thread de cache arrêté")

    def start(self, on_cache_updated=None):
        """Démarrer le système de cache en arrière-plan"""
        if self._cache_thread is not None and self._cache_thread.is_alive():
            return

        self._on_cache_updated = on_cache_updated
        self._stop_event.clear()
        self._cache_thread = threading.Thread(target=self._background_cache_loop, daemon=True)
        self._cache_thread.start()
        print("[Cache] Système de cache démarré")

    def stop(self):
        """Arrêter le système de cache"""
        self._stop_event.set()
        if self._cache_thread is not None:
            self._cache_thread.join(timeout=2)
        print("[Cache] Système de cache arrêté")

    def force_refresh(self, d: date = None):
        """Forcer le rafraîchissement du cache pour une date (ou toutes)"""
        if d:
            date_str = d.strftime("%Y-%m-%d")
            cache_path = self._get_cache_path(d)
            with self._lock:
                if date_str in self.cache_meta["dates"]:
                    del self.cache_meta["dates"][date_str]
                    self._save_meta()
            # Supprimer aussi le dossier cache physique
            if cache_path.exists():
                try:
                    shutil.rmtree(cache_path)
                    print(f"[Cache] Dossier cache {date_str} supprimé")
                except Exception as e:
                    print(f"[Cache] Erreur suppression dossier {date_str}: {e}")
            self.prioritize_date(d)
        else:
            # Vider tout le cache
            with self._lock:
                self.cache_meta["dates"] = {}
                self._save_meta()
            # Supprimer tous les dossiers cache
            try:
                for item in self.cache_dir.iterdir():
                    if item.is_dir():
                        shutil.rmtree(item)
                print("[Cache] Tout le cache supprimé")
            except Exception as e:
                print(f"[Cache] Erreur suppression cache: {e}")

    def get_cache_status(self) -> dict:
        """Obtenir le statut du cache pour l'affichage"""
        with self._lock:
            return {
                "cached_dates": len(self.cache_meta["dates"]),
                "dates": list(self.cache_meta["dates"].keys()),
                "cache_dir": str(self.cache_dir),
                "running": self._cache_thread is not None and self._cache_thread.is_alive()
            }

    def clear_old_cache(self, max_age_days: int = 30):
        """Nettoyer les entrées de cache anciennes"""
        cutoff = datetime.now() - timedelta(days=max_age_days)

        with self._lock:
            dates_to_remove = []
            for date_str, info in self.cache_meta["dates"].items():
                try:
                    cached_at = datetime.fromisoformat(info.get("cached_at", ""))
                    if cached_at < cutoff:
                        dates_to_remove.append(date_str)
                except Exception:
                    pass

            for date_str in dates_to_remove:
                cache_path = self.cache_dir / date_str
                if cache_path.exists():
                    shutil.rmtree(cache_path)
                del self.cache_meta["dates"][date_str]

            if dates_to_remove:
                self._save_meta()
                print(f"[Cache] Nettoyé {len(dates_to_remove)} dates anciennes")


# Instance globale du cache
planning_cache = PlanningCache()


class _AutoRefresher:
    def __init__(self, root: tk.Misc, refresh_callable):
        self.root = root
        self.refresh_callable = refresh_callable
        self.settings = load_local_settings()
        self._after_id = None
        self.is_user_active = False
        self.last_activity = datetime.now()
        self.start()

    def start(self):
        self.cancel()
        if self.settings.get("auto_refresh_enabled", True):
            delay_ms = max(1, int(self.settings.get("auto_refresh_seconds", 10))) * 1000
            self._after_id = self.root.after(delay_ms, self._tick)

    def cancel(self):
        if self._after_id is not None:
            try:
                self.root.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def _tick(self):
        try:
            if not self.is_user_active:
                self.refresh_callable()
        finally:
            self.start()

    def force_now(self):
        self.refresh_callable()

    def set_user_active(self, active: bool):
        self.is_user_active = active
        if active:
            self.last_activity = datetime.now()

    def apply(self, enabled: bool, seconds: int):
        self.settings["auto_refresh_enabled"] = bool(enabled)
        self.settings["auto_refresh_seconds"] = max(1, int(seconds))
        save_local_settings(self.settings)
        self.start()

def install_setup_menu(root: tk.Tk, menubar: tk.Menu, refresh_callable):
    refresher = _AutoRefresher(root, refresh_callable)

    setup_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Setup", menu=setup_menu)

    var_enabled = tk.BooleanVar(value=refresher.settings.get("auto_refresh_enabled", True))

    def toggle_enabled():
        refresher.apply(var_enabled.get(), refresher.settings.get("auto_refresh_seconds", 10))

    setup_menu.add_checkbutton(
        label="Activer l'auto-rafraîchissement",
        variable=var_enabled,
        command=toggle_enabled
    )
    setup_menu.add_command(
        label="Forcer un rafraîchissement maintenant",
        command=refresher.force_now
    )
    setup_menu.add_separator()

    def open_params_window():
        win = tk.Toplevel(root)
        win.title("Paramétrage — Rafraîchissement")
        win.resizable(False, False)

        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Checkbutton(
            frm, text="Activer l'auto-rafraîchissement",
            variable=var_enabled, command=toggle_enabled
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,8))

        ttk.Label(frm, text="Intervalle (secondes) :").grid(row=1, column=0, sticky="w")
        sec_var = tk.StringVar(value=str(refresher.settings.get("auto_refresh_seconds", 10)))
        ttk.Spinbox(frm, from_=5, to=3600, textvariable=sec_var, width=8).grid(row=1, column=1, sticky="w")

        def save_and_close():
            try:
                seconds = int(float(sec_var.get()))
            except ValueError:
                messagebox.showerror("Valeur invalide", "Entre un entier ≥ 5.")
                return
            refresher.apply(var_enabled.get(), seconds)
            win.destroy()

        btns = ttk.Frame(frm); btns.grid(row=2, column=0, columnspan=2, pady=(12,0), sticky="e")
        ttk.Button(btns, text="Forcer maintenant", command=refresher.force_now).pack(side="left", padx=(0,6))
        ttk.Button(btns, text="Enregistrer", command=save_and_close).pack(side="left")
        ttk.Button(btns, text="Fermer", command=win.destroy).pack(side="left", padx=(6,0))

        win.transient(root); win.grab_set(); root.wait_window(win)

    setup_menu.add_command(label="Paramètres…", command=open_params_window)

    # Menu Cache
    setup_menu.add_separator()
    cache_menu = tk.Menu(setup_menu, tearoff=0)
    setup_menu.add_cascade(label="Cache local", menu=cache_menu)

    def show_cache_status():
        status = planning_cache.get_cache_status()
        dates_str = "\n".join(sorted(status["dates"])[-10:]) if status["dates"] else "(vide)"
        if len(status["dates"]) > 10:
            dates_str = f"... et {len(status['dates']) - 10} autres\n" + dates_str

        messagebox.showinfo(
            "Statut du cache",
            f"Dates en cache : {status['cached_dates']}\n"
            f"Thread actif : {'Oui' if status['running'] else 'Non'}\n"
            f"Dossier : {status['cache_dir']}\n\n"
            f"Dernières dates :\n{dates_str}"
        )

    def clear_cache():
        if messagebox.askyesno("Vider le cache", "Voulez-vous vraiment vider tout le cache local ?"):
            planning_cache.force_refresh()
            messagebox.showinfo("Cache vidé", "Le cache a été vidé. Il sera recréé automatiquement.")

    def force_cache_refresh():
        planning_cache.force_refresh()
        messagebox.showinfo("Rafraîchissement", "Le cache va être reconstruit en arrière-plan.")

    cache_menu.add_command(label="Voir le statut du cache", command=show_cache_status)
    cache_menu.add_command(label="Forcer la mise à jour du cache", command=force_cache_refresh)
    cache_menu.add_command(label="Vider le cache", command=clear_cache)

    return refresher

# ---------- Fonctions d'export ----------

def generate_export_filename(prefix="extraction_planning", extension="pdf"):
    from datetime import datetime
    import getpass
    now = datetime.now()
    user = getpass.getuser().upper()
    filename = f"{prefix}_{now.strftime('%d_%m_%Y_%H%M')}_{user}.{extension}"
    return filename

def export_treeview_to_excel(tree, filename, sheet_name="Planning", title="Planning Export"):
    if not EXCEL_AVAILABLE:
        return False, "Module openpyxl non disponible"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        columns = tree["columns"]
        headers = [tree.heading(col)["text"] for col in columns]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row_idx = 4
        for item in tree.get_children():
            values = tree.item(item)["values"]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            row_idx += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)
        return True, f"Fichier Excel créé: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export Excel: {str(e)}"

def open_exported_file(filepath):
    """Ouvre le fichier exporté avec l'application par défaut"""
    import os
    import platform
    import subprocess
    try:
        if platform.system() == 'Windows':
            os.startfile(str(filepath))
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', str(filepath)])
        else:  # Linux
            subprocess.run(['xdg-open', str(filepath)])
        return True
    except Exception:
        return False

def export_planning_excel_par_chauffeur(missions, voyages, current_date, filename):
    """Export Excel spécialisé pour la vue par Chauffeur - format professionnel"""
    if not EXCEL_AVAILABLE:
        return False, "Module openpyxl non disponible"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planning Chauffeurs"

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16, color="1F4E79")
        subtitle_font = Font(bold=True, size=12, color="666666")
        group_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
        group_font = Font(bold=True, size=11, color="1F4E79")
        data_font = Font(size=10)
        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # En-tête du document
        ws.merge_cells('A1:H1')
        ws['A1'] = "PLANNING PAR CHAUFFEUR"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Date: {format_date_display(current_date)}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align

        # Statistiques rapides
        total_missions = len([m for m in missions if not str(m.get("chauffeur_nom", "")).startswith("─")])
        total_palettes = sum(int(m.get("nb_pal", 0)) for m in missions if str(m.get("nb_pal", "")).isdigit())
        chauffeurs_uniques = len(set(m.get("chauffeur_nom", "") for m in missions if m.get("chauffeur_nom") and not str(m.get("chauffeur_nom", "")).startswith("─")))

        ws.merge_cells('A3:H3')
        ws['A3'] = f"Total: {total_missions} missions | {total_palettes} palettes | {chauffeurs_uniques} chauffeurs"
        ws['A3'].font = Font(size=10, italic=True, color="888888")
        ws['A3'].alignment = center_align

        # En-têtes de colonnes (ligne 5)
        headers = ["Chauffeur", "N°", "Heure", "Type", "Voyage", "Palettes", "SST", "Pays", "Infos"]
        col_widths = [20, 6, 10, 12, 12, 10, 15, 15, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Regrouper les missions par chauffeur
        v_by_code = {v.get("code"): v for v in voyages}
        missions_by_driver = {}
        for m in missions:
            driver = m.get("chauffeur_nom", "Non assigné")
            if driver and not str(driver).startswith("─"):
                if driver not in missions_by_driver:
                    missions_by_driver[driver] = []
                missions_by_driver[driver].append(m)

        row_idx = 6
        for driver_name in sorted(missions_by_driver.keys()):
            driver_missions = missions_by_driver[driver_name]

            # Ligne de groupe pour le chauffeur
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            driver_palettes = sum(int(m.get("nb_pal", 0)) for m in driver_missions if str(m.get("nb_pal", "")).isdigit())
            ws.cell(row=row_idx, column=1).value = f"  {driver_name} ({len(driver_missions)} missions, {driver_palettes} pal.)"
            ws.cell(row=row_idx, column=1).font = group_font
            ws.cell(row=row_idx, column=1).fill = group_fill
            ws.cell(row=row_idx, column=1).alignment = left_align
            row_idx += 1

            # Trier les missions par numéro
            sorted_missions = sorted(driver_missions, key=lambda x: int(x.get("numero", 0)) if str(x.get("numero", "")).isdigit() else 0)

            for m in sorted_missions:
                voyage_code = m.get("voyage", "")
                voyage = v_by_code.get(voyage_code, {})
                country = voyage.get("country", "Belgique")

                values = [
                    "",  # Chauffeur vide (déjà affiché dans le groupe)
                    m.get("numero", ""),
                    m.get("heure", ""),
                    m.get("type", ""),
                    voyage_code,
                    m.get("nb_pal", ""),
                    m.get("sst", ""),
                    country,
                    m.get("infos", "")
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = data_font
                    cell.border = border_thin
                    cell.alignment = center_align if col_idx in [2, 3, 4, 5, 6] else left_align

                # Colorer selon le type
                type_cell = ws.cell(row=row_idx, column=4)
                if m.get("type") == "LIVRAISON":
                    type_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif m.get("type") == "RAMASSE":
                    type_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

                row_idx += 1

            row_idx += 1  # Espace entre les groupes

        # Pied de page
        ws.cell(row=row_idx + 1, column=1).value = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws.cell(row=row_idx + 1, column=1).font = Font(size=8, italic=True, color="999999")

        wb.save(filename)
        open_exported_file(filename)
        return True, f"Export Excel créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export Excel: {str(e)}"

def export_planning_excel_par_heure(missions, voyages, current_date, filename):
    """Export Excel spécialisé pour la vue par Heure - format chronologique"""
    if not EXCEL_AVAILABLE:
        return False, "Module openpyxl non disponible"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planning Horaire"

        # Styles
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16, color="2E7D32")
        subtitle_font = Font(bold=True, size=12, color="666666")
        time_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        time_font = Font(bold=True, size=11, color="2E7D32")
        data_font = Font(size=10)
        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # En-tête
        ws.merge_cells('A1:I1')
        ws['A1'] = "PLANNING CHRONOLOGIQUE"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:I2')
        ws['A2'] = f"Date: {format_date_display(current_date)}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align

        # Statistiques
        total_missions = len([m for m in missions if m.get("heure")])
        livraisons = len([m for m in missions if m.get("type") == "LIVRAISON"])
        ramasses = len([m for m in missions if m.get("type") == "RAMASSE"])

        ws.merge_cells('A3:I3')
        ws['A3'] = f"Total: {total_missions} missions | {livraisons} livraisons | {ramasses} ramasses"
        ws['A3'].font = Font(size=10, italic=True, color="888888")
        ws['A3'].alignment = center_align

        # En-têtes
        headers = ["Heure", "Type", "Voyage", "Chauffeur", "N°", "Palettes", "SST", "Pays", "Infos"]
        col_widths = [10, 12, 12, 20, 6, 10, 15, 15, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Trier par heure
        v_by_code = {v.get("code"): v for v in voyages}

        def time_key(m):
            heure = m.get("heure", "99:99")
            try:
                parts = heure.replace("h", ":").replace("H", ":").split(":")
                return int(parts[0]) * 60 + int(parts[1]) if len(parts) >= 2 else 9999
            except:
                return 9999

        sorted_missions = sorted([m for m in missions if m.get("heure")], key=time_key)

        row_idx = 6
        last_hour = None

        for m in sorted_missions:
            heure = m.get("heure", "")
            current_hour = heure.split(":")[0] if ":" in heure else heure.split("h")[0] if "h" in heure else ""

            # Séparateur d'heure
            if current_hour != last_hour and current_hour:
                if last_hour is not None:
                    row_idx += 1
                ws.merge_cells(f'A{row_idx}:I{row_idx}')
                ws.cell(row=row_idx, column=1).value = f"  {current_hour}h00 - {current_hour}h59"
                ws.cell(row=row_idx, column=1).font = time_font
                ws.cell(row=row_idx, column=1).fill = time_fill
                ws.cell(row=row_idx, column=1).alignment = left_align
                last_hour = current_hour
                row_idx += 1

            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")

            values = [
                heure,
                m.get("type", ""),
                voyage_code,
                m.get("chauffeur_nom", ""),
                m.get("numero", ""),
                m.get("nb_pal", ""),
                m.get("sst", ""),
                country,
                m.get("infos", "")
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.border = border_thin
                cell.alignment = center_align if col_idx in [1, 2, 3, 5, 6] else left_align

            # Colorer selon le type
            type_cell = ws.cell(row=row_idx, column=2)
            if m.get("type") == "LIVRAISON":
                type_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                type_cell.font = Font(size=10, bold=True, color="2E7D32")
            elif m.get("type") == "RAMASSE":
                type_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                type_cell.font = Font(size=10, bold=True, color="E65100")

            row_idx += 1

        # Pied de page
        ws.cell(row=row_idx + 2, column=1).value = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws.cell(row=row_idx + 2, column=1).font = Font(size=8, italic=True, color="999999")

        wb.save(filename)
        open_exported_file(filename)
        return True, f"Export Excel créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export Excel: {str(e)}"

def export_planning_excel_par_voyage(missions, voyages, current_date, filename):
    """Export Excel spécialisé pour la vue par Voyage - format groupé par destination"""
    if not EXCEL_AVAILABLE:
        return False, "Module openpyxl non disponible"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planning Voyages"

        # Styles
        header_fill = PatternFill(start_color="5E35B1", end_color="5E35B1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16, color="5E35B1")
        subtitle_font = Font(bold=True, size=12, color="666666")
        voyage_fill = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")
        voyage_font = Font(bold=True, size=11, color="5E35B1")
        data_font = Font(size=10)
        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # En-tête
        ws.merge_cells('A1:I1')
        ws['A1'] = "PLANNING PAR VOYAGE"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:I2')
        ws['A2'] = f"Date: {format_date_display(current_date)}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align

        # Statistiques
        v_by_code = {v.get("code"): v for v in voyages}
        voyages_uniques = len(set(m.get("voyage", "") for m in missions if m.get("voyage")))
        total_palettes = sum(int(m.get("nb_pal", 0)) for m in missions if str(m.get("nb_pal", "")).isdigit())

        ws.merge_cells('A3:I3')
        ws['A3'] = f"Total: {voyages_uniques} voyages | {len(missions)} missions | {total_palettes} palettes"
        ws['A3'].font = Font(size=10, italic=True, color="888888")
        ws['A3'].alignment = center_align

        # En-têtes
        headers = ["Voyage", "Pays", "Type", "Heure", "Chauffeur", "N°", "Palettes", "SST", "Infos"]
        col_widths = [12, 15, 12, 10, 20, 6, 10, 15, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Regrouper par voyage
        missions_by_voyage = {}
        for m in missions:
            voyage_code = m.get("voyage", "Sans voyage")
            if voyage_code not in missions_by_voyage:
                missions_by_voyage[voyage_code] = []
            missions_by_voyage[voyage_code].append(m)

        row_idx = 6
        for voyage_code in sorted(missions_by_voyage.keys()):
            voyage_missions = missions_by_voyage[voyage_code]
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")

            # Ligne de groupe
            voyage_palettes = sum(int(m.get("nb_pal", 0)) for m in voyage_missions if str(m.get("nb_pal", "")).isdigit())
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            ws.cell(row=row_idx, column=1).value = f"  {voyage_code} - {country} ({len(voyage_missions)} missions, {voyage_palettes} pal.)"
            ws.cell(row=row_idx, column=1).font = voyage_font
            ws.cell(row=row_idx, column=1).fill = voyage_fill
            ws.cell(row=row_idx, column=1).alignment = left_align
            row_idx += 1

            # Trier par heure
            def time_key(m):
                heure = m.get("heure", "99:99")
                try:
                    parts = heure.replace("h", ":").replace("H", ":").split(":")
                    return int(parts[0]) * 60 + int(parts[1]) if len(parts) >= 2 else 9999
                except:
                    return 9999

            sorted_missions = sorted(voyage_missions, key=time_key)

            for m in sorted_missions:
                values = [
                    "",  # Voyage vide (déjà affiché)
                    "",  # Pays vide (déjà affiché)
                    m.get("type", ""),
                    m.get("heure", ""),
                    m.get("chauffeur_nom", ""),
                    m.get("numero", ""),
                    m.get("nb_pal", ""),
                    m.get("sst", ""),
                    m.get("infos", "")
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.font = data_font
                    cell.border = border_thin
                    cell.alignment = center_align if col_idx in [3, 4, 6, 7] else left_align

                # Colorer selon le type
                type_cell = ws.cell(row=row_idx, column=3)
                if m.get("type") == "LIVRAISON":
                    type_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif m.get("type") == "RAMASSE":
                    type_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

                row_idx += 1

            row_idx += 1  # Espace entre les groupes

        # Pied de page
        ws.cell(row=row_idx + 1, column=1).value = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws.cell(row=row_idx + 1, column=1).font = Font(size=8, italic=True, color="999999")

        wb.save(filename)
        open_exported_file(filename)
        return True, f"Export Excel créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export Excel: {str(e)}"

def export_planning_pdf_par_chauffeur(missions, voyages, current_date, filename):
    """Export PDF spécialisé pour la vue par Chauffeur - format professionnel"""
    if not PDF_AVAILABLE:
        return False, "Module reportlab non disponible"

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4),
                               leftMargin=1*cm, rightMargin=1*cm,
                               topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Styles personnalisés
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'],
                                     fontSize=18, textColor=colors.HexColor('#1F4E79'),
                                     alignment=TA_CENTER, spaceAfter=5)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'],
                                        fontSize=12, textColor=colors.HexColor('#666666'),
                                        alignment=TA_CENTER, spaceAfter=3)
        stats_style = ParagraphStyle('StatsStyle', parent=styles['Normal'],
                                     fontSize=9, textColor=colors.HexColor('#888888'),
                                     alignment=TA_CENTER, spaceAfter=15)
        group_style = ParagraphStyle('GroupStyle', parent=styles['Normal'],
                                     fontSize=10, textColor=colors.HexColor('#1F4E79'),
                                     fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=3)
        footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'],
                                      fontSize=8, textColor=colors.HexColor('#999999'),
                                      alignment=TA_CENTER)

        # En-tête
        elements.append(Paragraph("PLANNING PAR CHAUFFEUR", title_style))
        elements.append(Paragraph(f"Date: {format_date_display(current_date)}", subtitle_style))

        # Statistiques
        total_missions = len([m for m in missions if not str(m.get("chauffeur_nom", "")).startswith("─")])
        total_palettes = sum(int(m.get("nb_pal", 0)) for m in missions if str(m.get("nb_pal", "")).isdigit())
        chauffeurs_uniques = len(set(m.get("chauffeur_nom", "") for m in missions if m.get("chauffeur_nom") and not str(m.get("chauffeur_nom", "")).startswith("─")))
        elements.append(Paragraph(f"Total: {total_missions} missions | {total_palettes} palettes | {chauffeurs_uniques} chauffeurs", stats_style))

        # Regrouper par chauffeur
        v_by_code = {v.get("code"): v for v in voyages}
        missions_by_driver = {}
        for m in missions:
            driver = m.get("chauffeur_nom", "Non assigné")
            if driver and not str(driver).startswith("─"):
                if driver not in missions_by_driver:
                    missions_by_driver[driver] = []
                missions_by_driver[driver].append(m)

        # Couleurs
        header_color = colors.HexColor('#1F4E79')
        group_color = colors.HexColor('#D6E3F8')
        livraison_color = colors.HexColor('#E8F5E9')
        ramasse_color = colors.HexColor('#FFF3E0')

        for driver_name in sorted(missions_by_driver.keys()):
            driver_missions = missions_by_driver[driver_name]
            driver_palettes = sum(int(m.get("nb_pal", 0)) for m in driver_missions if str(m.get("nb_pal", "")).isdigit())

            elements.append(Paragraph(f"{driver_name} ({len(driver_missions)} missions, {driver_palettes} palettes)", group_style))

            # Tableau pour ce chauffeur
            table_data = [["N°", "Heure", "Type", "Voyage", "Palettes", "SST", "Pays", "Infos"]]

            sorted_missions = sorted(driver_missions, key=lambda x: int(x.get("numero", 0)) if str(x.get("numero", "")).isdigit() else 0)

            for m in sorted_missions:
                voyage_code = m.get("voyage", "")
                voyage = v_by_code.get(voyage_code, {})
                country = voyage.get("country", "Belgique")
                table_data.append([
                    str(m.get("numero", "")),
                    m.get("heure", ""),
                    m.get("type", ""),
                    voyage_code,
                    str(m.get("nb_pal", "")),
                    m.get("sst", ""),
                    country,
                    str(m.get("infos", ""))[:30]
                ])

            col_widths = [25, 50, 70, 60, 50, 80, 80, 150]
            t = Table(table_data, colWidths=col_widths)

            # Style du tableau
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (5, 1), (5, -1), 'LEFT'),
                ('ALIGN', (7, 1), (7, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]

            # Colorer les lignes selon le type
            for idx, m in enumerate(sorted_missions, start=1):
                if m.get("type") == "LIVRAISON":
                    table_style.append(('BACKGROUND', (2, idx), (2, idx), livraison_color))
                elif m.get("type") == "RAMASSE":
                    table_style.append(('BACKGROUND', (2, idx), (2, idx), ramasse_color))

            t.setStyle(TableStyle(table_style))
            elements.append(t)
            elements.append(Spacer(1, 10))

        # Pied de page
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style))

        doc.build(elements)
        open_exported_file(filename)
        return True, f"Export PDF créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export PDF: {str(e)}"

def export_planning_pdf_par_heure(missions, voyages, current_date, filename):
    """Export PDF spécialisé pour la vue par Heure - format chronologique"""
    if not PDF_AVAILABLE:
        return False, "Module reportlab non disponible"

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4),
                               leftMargin=1*cm, rightMargin=1*cm,
                               topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Styles
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'],
                                     fontSize=18, textColor=colors.HexColor('#2E7D32'),
                                     alignment=TA_CENTER, spaceAfter=5)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'],
                                        fontSize=12, textColor=colors.HexColor('#666666'),
                                        alignment=TA_CENTER, spaceAfter=3)
        stats_style = ParagraphStyle('StatsStyle', parent=styles['Normal'],
                                     fontSize=9, textColor=colors.HexColor('#888888'),
                                     alignment=TA_CENTER, spaceAfter=15)
        footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'],
                                      fontSize=8, textColor=colors.HexColor('#999999'),
                                      alignment=TA_CENTER)

        # En-tête
        elements.append(Paragraph("PLANNING CHRONOLOGIQUE", title_style))
        elements.append(Paragraph(f"Date: {format_date_display(current_date)}", subtitle_style))

        # Statistiques
        total_missions = len([m for m in missions if m.get("heure")])
        livraisons = len([m for m in missions if m.get("type") == "LIVRAISON"])
        ramasses = len([m for m in missions if m.get("type") == "RAMASSE"])
        total_palettes = sum(int(m.get("nb_pal", 0)) for m in missions if str(m.get("nb_pal", "")).isdigit())
        elements.append(Paragraph(f"Total: {total_missions} missions | {livraisons} livraisons | {ramasses} ramasses | {total_palettes} palettes", stats_style))

        # Couleurs
        header_color = colors.HexColor('#2E7D32')
        livraison_color = colors.HexColor('#C8E6C9')
        ramasse_color = colors.HexColor('#FFE0B2')

        # Trier par heure
        v_by_code = {v.get("code"): v for v in voyages}

        def time_key(m):
            heure = m.get("heure", "99:99")
            try:
                parts = heure.replace("h", ":").replace("H", ":").split(":")
                return int(parts[0]) * 60 + int(parts[1]) if len(parts) >= 2 else 9999
            except:
                return 9999

        sorted_missions = sorted([m for m in missions if m.get("heure")], key=time_key)

        # Tableau principal
        table_data = [["Heure", "Type", "Voyage", "Chauffeur", "N°", "Palettes", "SST", "Pays", "Infos"]]

        for m in sorted_missions:
            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")
            table_data.append([
                m.get("heure", ""),
                m.get("type", ""),
                voyage_code,
                m.get("chauffeur_nom", ""),
                str(m.get("numero", "")),
                str(m.get("nb_pal", "")),
                m.get("sst", ""),
                country,
                str(m.get("infos", ""))[:25]
            ])

        col_widths = [45, 70, 60, 100, 25, 45, 80, 80, 130]
        t = Table(table_data, colWidths=col_widths)

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),
            ('ALIGN', (6, 1), (6, -1), 'LEFT'),
            ('ALIGN', (8, 1), (8, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]

        # Colorer selon le type
        for idx, m in enumerate(sorted_missions, start=1):
            if m.get("type") == "LIVRAISON":
                table_style.append(('BACKGROUND', (1, idx), (1, idx), livraison_color))
            elif m.get("type") == "RAMASSE":
                table_style.append(('BACKGROUND', (1, idx), (1, idx), ramasse_color))

        t.setStyle(TableStyle(table_style))
        elements.append(t)

        # Pied de page
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style))

        doc.build(elements)
        open_exported_file(filename)
        return True, f"Export PDF créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export PDF: {str(e)}"

def export_planning_pdf_par_voyage(missions, voyages, current_date, filename):
    """Export PDF spécialisé pour la vue par Voyage - format groupé par destination"""
    if not PDF_AVAILABLE:
        return False, "Module reportlab non disponible"

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4),
                               leftMargin=1*cm, rightMargin=1*cm,
                               topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Styles
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'],
                                     fontSize=18, textColor=colors.HexColor('#5E35B1'),
                                     alignment=TA_CENTER, spaceAfter=5)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'],
                                        fontSize=12, textColor=colors.HexColor('#666666'),
                                        alignment=TA_CENTER, spaceAfter=3)
        stats_style = ParagraphStyle('StatsStyle', parent=styles['Normal'],
                                     fontSize=9, textColor=colors.HexColor('#888888'),
                                     alignment=TA_CENTER, spaceAfter=15)
        group_style = ParagraphStyle('GroupStyle', parent=styles['Normal'],
                                     fontSize=10, textColor=colors.HexColor('#5E35B1'),
                                     fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=3)
        footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'],
                                      fontSize=8, textColor=colors.HexColor('#999999'),
                                      alignment=TA_CENTER)

        # En-tête
        elements.append(Paragraph("PLANNING PAR VOYAGE", title_style))
        elements.append(Paragraph(f"Date: {format_date_display(current_date)}", subtitle_style))

        # Statistiques
        v_by_code = {v.get("code"): v for v in voyages}
        voyages_uniques = len(set(m.get("voyage", "") for m in missions if m.get("voyage")))
        total_palettes = sum(int(m.get("nb_pal", 0)) for m in missions if str(m.get("nb_pal", "")).isdigit())
        elements.append(Paragraph(f"Total: {voyages_uniques} voyages | {len(missions)} missions | {total_palettes} palettes", stats_style))

        # Couleurs
        header_color = colors.HexColor('#5E35B1')
        group_color = colors.HexColor('#EDE7F6')
        livraison_color = colors.HexColor('#E8F5E9')
        ramasse_color = colors.HexColor('#FFF3E0')

        # Regrouper par voyage
        missions_by_voyage = {}
        for m in missions:
            voyage_code = m.get("voyage", "Sans voyage")
            if voyage_code not in missions_by_voyage:
                missions_by_voyage[voyage_code] = []
            missions_by_voyage[voyage_code].append(m)

        for voyage_code in sorted(missions_by_voyage.keys()):
            voyage_missions = missions_by_voyage[voyage_code]
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")
            voyage_palettes = sum(int(m.get("nb_pal", 0)) for m in voyage_missions if str(m.get("nb_pal", "")).isdigit())

            elements.append(Paragraph(f"{voyage_code} - {country} ({len(voyage_missions)} missions, {voyage_palettes} palettes)", group_style))

            # Tableau
            table_data = [["Heure", "Type", "Chauffeur", "N°", "Palettes", "SST", "Infos"]]

            def time_key(m):
                heure = m.get("heure", "99:99")
                try:
                    parts = heure.replace("h", ":").replace("H", ":").split(":")
                    return int(parts[0]) * 60 + int(parts[1]) if len(parts) >= 2 else 9999
                except:
                    return 9999

            sorted_missions = sorted(voyage_missions, key=time_key)

            for m in sorted_missions:
                table_data.append([
                    m.get("heure", ""),
                    m.get("type", ""),
                    m.get("chauffeur_nom", ""),
                    str(m.get("numero", "")),
                    str(m.get("nb_pal", "")),
                    m.get("sst", ""),
                    str(m.get("infos", ""))[:35]
                ])

            col_widths = [50, 70, 120, 30, 50, 100, 180]
            t = Table(table_data, colWidths=col_widths)

            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('ALIGN', (5, 1), (5, -1), 'LEFT'),
                ('ALIGN', (6, 1), (6, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]

            # Colorer selon le type
            for idx, m in enumerate(sorted_missions, start=1):
                if m.get("type") == "LIVRAISON":
                    table_style.append(('BACKGROUND', (1, idx), (1, idx), livraison_color))
                elif m.get("type") == "RAMASSE":
                    table_style.append(('BACKGROUND', (1, idx), (1, idx), ramasse_color))

            t.setStyle(TableStyle(table_style))
            elements.append(t)
            elements.append(Spacer(1, 10))

        # Pied de page
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style))

        doc.build(elements)
        open_exported_file(filename)
        return True, f"Export PDF créé et ouvert: {filename}"

    except Exception as e:
        return False, f"Erreur lors de l'export PDF: {str(e)}"

def export_treeview_to_pdf(tree, filename, title="Planning Export", date_str=""):
    if not PDF_AVAILABLE:
        return False, "Module reportlab non disponible"
    
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4))
        elements = []
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        
        if date_str:
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER
            )
            date_para = Paragraph(f"Date: {date_str}", date_style)
            elements.append(date_para)
            elements.append(Spacer(1, 20))
        
        columns = tree["columns"]
        headers = [tree.heading(col)["text"] for col in columns]
        
        table_data = [headers]
        for item in tree.get_children():
            values = tree.item(item)["values"]
            table_data.append([str(v) for v in values])
        
        t = Table(table_data)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        t.setStyle(style)
        elements.append(t)
        
        doc.build(elements)
        return True, f"Fichier PDF créé: {filename}"
    
    except Exception as e:
        return False, f"Erreur lors de l'export PDF: {str(e)}"

# =============================================================================
# MODULE D'ANALYSE AVANCÉE INTÉGRÉ
# =============================================================================

# Constantes pour l'analyse
COLORS_ANALYSE = {
    'primary': '#2563eb', 'success': '#22c55e', 'danger': '#ef4444',
    'warning': '#f59e0b', 'info': '#06b6d4', 'chart_blue': '#3b82f6',
    'chart_green': '#22c55e', 'chart_red': '#ef4444', 'chart_orange': '#f97316',
    'chart_purple': '#a855f7', 'chart_teal': '#14b8a6',
}

COUNTRY_FLAGS = {
    "Belgique": "🇧🇪", "France": "🇫🇷", "Allemagne": "🇩🇪",
    "Pays-Bas": "🇳🇱", "Luxembourg": "🇱🇺",
}

def format_currency(value, symbol="€"):
    if value >= 0:
        return f"{value:,.2f} {symbol}".replace(",", " ").replace(".", ",")
    return f"-{abs(value):,.2f} {symbol}".replace(",", " ").replace(".", ",")

def format_number(value, decimals=0):
    if decimals == 0:
        return f"{int(value):,}".replace(",", " ")
    return f"{value:,.{decimals}f}".replace(",", " ").replace(".", ",")

def calculate_variation(current, previous):
    if previous == 0:
        return 100 if current > 0 else 0
    return ((current - previous) / abs(previous)) * 100


class AdvancedAnalyseModule:
    """Module d'analyse avancée avec dashboard, filtres, graphiques et exports."""
    
    def __init__(self, parent_app):
        self.app = parent_app
        self.root = parent_app.root
        self.notebook = parent_app.notebook
        self.data_dir = parent_app.data_dir
        self.missions = parent_app.missions
        self.voyages = parent_app.voyages
        self.chauffeurs = parent_app.chauffeurs
        self.tarifs_sst = parent_app.tarifs_sst
        self.revenus_palettes = parent_app.revenus_palettes
        
        self.current_data = None
        self.comparison_data = None
        self.analyse_figures = []
        self.filter_vars = {}
        
        self.build_analyse_tab()
    
    def build_analyse_tab(self):
        """Construire l'onglet Analyse avec interface avancée"""
        self.tab_analyse = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_analyse, text="📊 Analyse Avancée")
        
        if not MATPLOTLIB_AVAILABLE:
            self._show_matplotlib_error()
            return
        
        self.tab_analyse.columnconfigure(0, weight=1)
        self.tab_analyse.rowconfigure(0, weight=1)
        
        # PanedWindow principal
        self.main_paned = ttk.PanedWindow(self.tab_analyse, orient="horizontal")
        self.main_paned.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        
        # Panneau gauche - Filtres
        self.left_panel = self._build_left_panel()
        self.main_paned.add(self.left_panel, weight=1)
        
        # Panneau droit - Contenu
        self.right_panel = self._build_right_panel()
        self.main_paned.add(self.right_panel, weight=4)
    
    def _show_matplotlib_error(self):
        msg_frame = ttk.Frame(self.tab_analyse)
        msg_frame.pack(fill="both", expand=True, padx=20, pady=20)
        ttk.Label(msg_frame, text="⚠️ Module matplotlib non disponible",
                 font=("Segoe UI", 16, "bold")).pack(pady=20)
        ttk.Label(msg_frame, text="pip install matplotlib numpy",
                 font=("Segoe UI", 11)).pack(pady=10)
    
    def _build_left_panel(self):
        """Panneau de filtres"""
        frame = ttk.Frame(self.main_paned)
        frame.columnconfigure(0, weight=1)
        
        # Header
        header = ttk.Frame(frame)
        header.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        ttk.Label(header, text="🎛️ Filtres", font=("Segoe UI", 11, "bold")).pack(side="left")
        
        # Canvas scrollable
        filter_canvas = tk.Canvas(frame, width=260, highlightthickness=0)
        filter_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=filter_canvas.yview)
        self.filter_frame = ttk.Frame(filter_canvas)
        
        self.filter_frame.bind("<Configure>",
            lambda e: filter_canvas.configure(scrollregion=filter_canvas.bbox("all")))
        filter_canvas.create_window((0, 0), window=self.filter_frame, anchor="nw")
        filter_canvas.configure(yscrollcommand=filter_scrollbar.set)
        
        filter_canvas.grid(row=1, column=0, sticky="nsew", padx=5)
        filter_scrollbar.grid(row=1, column=1, sticky="ns")
        frame.rowconfigure(1, weight=1)
        
        self._build_period_filter()
        self._build_entity_filters()
        self._build_comparison_filter()
        self._build_action_buttons()
        
        return frame
    
    def _build_period_filter(self):
        """Section période"""
        period_frame = ttk.LabelFrame(self.filter_frame, text="📅 Période", padding=8)
        period_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(period_frame, text="Du :").pack(anchor="w")
        self.date_start_var = tk.StringVar(value=(date.today() - timedelta(days=30)).strftime("%d/%m/%Y"))
        ttk.Entry(period_frame, textvariable=self.date_start_var, width=15).pack(fill="x", pady=2)
        
        ttk.Label(period_frame, text="Au :").pack(anchor="w", pady=(5, 0))
        self.date_end_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(period_frame, textvariable=self.date_end_var, width=15).pack(fill="x", pady=2)
        
        quick_frame = ttk.Frame(period_frame)
        quick_frame.pack(fill="x", pady=5)
        
        periods = [("7j", 7), ("14j", 14), ("30j", 30), ("90j", 90), ("180j", 180), ("365j", 365)]
        for i, (label, days) in enumerate(periods):
            btn = ttk.Button(quick_frame, text=label, width=5, command=lambda d=days: self._set_period(d))
            btn.grid(row=i//3, column=i%3, padx=1, pady=1, sticky="ew")
        for i in range(3):
            quick_frame.columnconfigure(i, weight=1)
        
        extra_frame = ttk.Frame(period_frame)
        extra_frame.pack(fill="x", pady=2)
        ttk.Button(extra_frame, text="Ce mois", command=self._set_period_this_month).pack(side="left", expand=True, fill="x", padx=1)
        ttk.Button(extra_frame, text="Cette année", command=self._set_period_this_year).pack(side="left", expand=True, fill="x", padx=1)
    
    def _build_entity_filters(self):
        """Filtres par entité"""
        # Pays
        country_frame = ttk.LabelFrame(self.filter_frame, text="🌍 Pays", padding=8)
        country_frame.pack(fill="x", padx=5, pady=5)
        
        self.country_vars = {}
        for country in EU_COUNTRIES:
            var = tk.BooleanVar(value=True)
            self.country_vars[country] = var
            flag = COUNTRY_FLAGS.get(country, "")
            ttk.Checkbutton(country_frame, text=f"{flag} {country}", variable=var).pack(anchor="w")
        
        btn_frame = ttk.Frame(country_frame)
        btn_frame.pack(fill="x", pady=2)
        ttk.Button(btn_frame, text="Tous", width=8, command=lambda: self._select_all_countries(True)).pack(side="left", padx=1)
        ttk.Button(btn_frame, text="Aucun", width=8, command=lambda: self._select_all_countries(False)).pack(side="left", padx=1)
        
        # SST
        sst_frame = ttk.LabelFrame(self.filter_frame, text="🏢 SST", padding=8)
        sst_frame.pack(fill="x", padx=5, pady=5)
        
        self.sst_listbox = tk.Listbox(sst_frame, height=5, selectmode="multiple", exportselection=False)
        self.sst_listbox.pack(fill="x", pady=2)
        
        sst_btn_frame = ttk.Frame(sst_frame)
        sst_btn_frame.pack(fill="x")
        ttk.Button(sst_btn_frame, text="Tous", width=8, command=lambda: self._select_all_listbox(self.sst_listbox, True)).pack(side="left", padx=1)
        ttk.Button(sst_btn_frame, text="Aucun", width=8, command=lambda: self._select_all_listbox(self.sst_listbox, False)).pack(side="left", padx=1)
        
        # Type
        type_frame = ttk.LabelFrame(self.filter_frame, text="📦 Type", padding=8)
        type_frame.pack(fill="x", padx=5, pady=5)
        
        self.type_livraison_var = tk.BooleanVar(value=True)
        self.type_ramasse_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(type_frame, text="🚚 Livraisons", variable=self.type_livraison_var).pack(anchor="w")
        ttk.Checkbutton(type_frame, text="📥 Ramasses", variable=self.type_ramasse_var).pack(anchor="w")
        
        # Voyages
        voyage_frame = ttk.LabelFrame(self.filter_frame, text="🛣️ Voyages", padding=8)
        voyage_frame.pack(fill="x", padx=5, pady=5)
        
        self.voyage_listbox = tk.Listbox(voyage_frame, height=5, selectmode="multiple", exportselection=False)
        self.voyage_listbox.pack(fill="x", pady=2)
        
        voyage_btn_frame = ttk.Frame(voyage_frame)
        voyage_btn_frame.pack(fill="x")
        ttk.Button(voyage_btn_frame, text="Tous", width=8, command=lambda: self._select_all_listbox(self.voyage_listbox, True)).pack(side="left", padx=1)
        ttk.Button(voyage_btn_frame, text="Aucun", width=8, command=lambda: self._select_all_listbox(self.voyage_listbox, False)).pack(side="left", padx=1)
    
    def _build_comparison_filter(self):
        """Section comparaison"""
        comp_frame = ttk.LabelFrame(self.filter_frame, text="📊 Comparaison", padding=8)
        comp_frame.pack(fill="x", padx=5, pady=5)
        
        self.comparison_enabled_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(comp_frame, text="Activer", variable=self.comparison_enabled_var,
                       command=self._toggle_comparison).pack(anchor="w")
        
        self.comparison_type_var = tk.StringVar(value="previous_period")
        self.comp_options_frame = ttk.Frame(comp_frame)
        
        ttk.Radiobutton(self.comp_options_frame, text="Période précédente",
                       variable=self.comparison_type_var, value="previous_period").pack(anchor="w")
        ttk.Radiobutton(self.comp_options_frame, text="Même période N-1",
                       variable=self.comparison_type_var, value="previous_year").pack(anchor="w")
    
    def _build_action_buttons(self):
        """Boutons d'action"""
        action_frame = ttk.Frame(self.filter_frame)
        action_frame.pack(fill="x", padx=5, pady=10)
        
        self.analyse_btn = ttk.Button(action_frame, text="🔍 Analyser", command=self.run_analysis)
        self.analyse_btn.pack(fill="x", pady=2)
        
        ttk.Button(action_frame, text="🔄 Réinitialiser", command=self.reset_filters).pack(fill="x", pady=2)
        
        ttk.Separator(action_frame, orient="horizontal").pack(fill="x", pady=10)
        
        export_frame = ttk.LabelFrame(action_frame, text="📥 Export", padding=5)
        export_frame.pack(fill="x", pady=5)
        
        ttk.Button(export_frame, text="📊 Excel", command=self.export_to_excel).pack(fill="x", pady=1)
        ttk.Button(export_frame, text="📄 CSV", command=self.export_to_csv).pack(fill="x", pady=1)
        ttk.Button(export_frame, text="📑 PDF", command=self.export_to_pdf).pack(fill="x", pady=1)
    
    def _build_right_panel(self):
        """Panneau de contenu principal"""
        frame = ttk.Frame(self.main_paned)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        
        # Header
        header = ttk.Frame(frame)
        header.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        ttk.Label(header, text="📊 Tableau de Bord", font=("Segoe UI", 14, "bold")).pack(side="left")
        
        # Vue sélecteur
        view_frame = ttk.Frame(header)
        view_frame.pack(side="right")
        ttk.Label(view_frame, text="Vue :").pack(side="left", padx=5)
        self.view_var = tk.StringVar(value="Dashboard")
        views = ["Dashboard", "Graphiques", "Tableau", "Pivot Table", "Tendances"]
        self.view_combo = ttk.Combobox(view_frame, textvariable=self.view_var, values=views, width=15, state="readonly")
        self.view_combo.pack(side="left")
        self.view_combo.bind("<<ComboboxSelected>>", self._on_view_changed)
        
        # Notebook pour les vues
        self.content_notebook = ttk.Notebook(frame)
        self.content_notebook.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        self._build_dashboard_view()
        self._build_charts_view()
        self._build_table_view()
        self._build_pivot_view()
        self._build_trends_view()
        
        return frame
    
    def _build_dashboard_view(self):
        """Vue Dashboard"""
        self.dashboard_frame = ttk.Frame(self.content_notebook)
        self.content_notebook.add(self.dashboard_frame, text="📊 Dashboard")
        
        canvas = tk.Canvas(self.dashboard_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.dashboard_frame, orient="vertical", command=canvas.yview)
        self.dashboard_content = ttk.Frame(canvas)
        
        self.dashboard_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.dashboard_content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.dashboard_canvas = canvas
        
        ttk.Label(self.dashboard_content, text="📈 Cliquez sur 'Analyser' pour afficher le tableau de bord",
                 font=("Segoe UI", 12), foreground="gray").pack(pady=50)
    
    def _build_charts_view(self):
        """Vue Graphiques"""
        self.charts_frame = ttk.Frame(self.content_notebook)
        self.content_notebook.add(self.charts_frame, text="📈 Graphiques")
        
        canvas = tk.Canvas(self.charts_frame, highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(self.charts_frame, orient="vertical", command=canvas.yview)
        self.charts_content = ttk.Frame(canvas)
        
        self.charts_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.charts_content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set)
        
        scrollbar_y.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self.charts_canvas = canvas
        
        # Sélecteur type graphique
        chart_selector = ttk.Frame(self.charts_content)
        chart_selector.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(chart_selector, text="Type :").pack(side="left")
        self.chart_type_var = tk.StringVar(value="Tous")
        chart_types = ["Tous", "Évolution", "Par Voyage", "Par SST", "Par Chauffeur", "Par Pays"]
        self.chart_type_combo = ttk.Combobox(chart_selector, textvariable=self.chart_type_var,
                                             values=chart_types, width=15, state="readonly")
        self.chart_type_combo.pack(side="left", padx=5)
        self.chart_type_combo.bind("<<ComboboxSelected>>", lambda e: self._update_charts())
        
        self.charts_container = ttk.Frame(self.charts_content)
        self.charts_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        ttk.Label(self.charts_container, text="📈 Cliquez sur 'Analyser'",
                 font=("Segoe UI", 12), foreground="gray").pack(pady=50)
    
    def _build_table_view(self):
        """Vue Tableau"""
        self.table_frame = ttk.Frame(self.content_notebook)
        self.content_notebook.add(self.table_frame, text="📋 Tableau")
        
        # Toolbar
        toolbar = ttk.Frame(self.table_frame)
        toolbar.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(toolbar, text="Rechercher :").pack(side="left", padx=5)
        self.table_search_var = tk.StringVar()
        self.table_search_var.trace("w", lambda *args: self._filter_table())
        ttk.Entry(toolbar, textvariable=self.table_search_var, width=20).pack(side="left")
        
        # Treeview
        tree_frame = ttk.Frame(self.table_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        columns = ("date", "voyage", "pays", "type", "sst", "chauffeur", "palettes", "revenus", "couts", "marge")
        self.data_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        
        col_config = {
            "date": ("Date", 90), "voyage": ("Voyage", 80), "pays": ("Pays", 80),
            "type": ("Type", 80), "sst": ("SST", 70), "chauffeur": ("Chauffeur", 120),
            "palettes": ("Pal", 60), "revenus": ("Revenus", 90), "couts": ("Coûts", 90), "marge": ("Marge", 90),
        }
        
        for col, (heading, width) in col_config.items():
            self.data_tree.heading(col, text=heading, command=lambda c=col: self._sort_table(c))
            self.data_tree.column(col, width=width, anchor="center")
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.data_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.data_tree.xview)
        self.data_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.data_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        self.data_tree.tag_configure('evenrow', background='#f8f9fa')
        self.data_tree.tag_configure('oddrow', background='white')
        self.data_tree.tag_configure('positive', foreground='#16a34a')
        self.data_tree.tag_configure('negative', foreground='#dc2626')
        
        self.table_stats_frame = ttk.Frame(self.table_frame)
        self.table_stats_frame.pack(fill="x", padx=5, pady=5)
    
    def _build_pivot_view(self):
        """Vue Pivot Table"""
        self.pivot_frame = ttk.Frame(self.content_notebook)
        self.content_notebook.add(self.pivot_frame, text="🔄 Pivot")
        
        config_frame = ttk.LabelFrame(self.pivot_frame, text="Configuration", padding=10)
        config_frame.pack(fill="x", padx=10, pady=5)
        
        row1 = ttk.Frame(config_frame)
        row1.pack(fill="x", pady=2)
        
        ttk.Label(row1, text="Lignes :").pack(side="left")
        self.pivot_rows_var = tk.StringVar(value="SST")
        pivot_dims = ["Date", "Semaine", "Mois", "Voyage", "SST", "Chauffeur", "Pays", "Type"]
        ttk.Combobox(row1, textvariable=self.pivot_rows_var, values=pivot_dims, width=12, state="readonly").pack(side="left", padx=5)
        
        ttk.Label(row1, text="Colonnes :").pack(side="left", padx=(20, 5))
        self.pivot_cols_var = tk.StringVar(value="Mois")
        ttk.Combobox(row1, textvariable=self.pivot_cols_var, values=pivot_dims, width=12, state="readonly").pack(side="left", padx=5)
        
        ttk.Label(row1, text="Valeur :").pack(side="left", padx=(20, 5))
        self.pivot_value_var = tk.StringVar(value="Marge")
        ttk.Combobox(row1, textvariable=self.pivot_value_var, values=["Revenus", "Coûts", "Marge", "Palettes", "Missions"], width=12, state="readonly").pack(side="left", padx=5)
        
        ttk.Button(row1, text="🔄 Générer", command=self._generate_pivot).pack(side="left", padx=20)
        
        self.pivot_container = ttk.Frame(self.pivot_frame)
        self.pivot_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        ttk.Label(self.pivot_container, text="🔄 Configurez et cliquez sur 'Générer'",
                 font=("Segoe UI", 12), foreground="gray").pack(pady=50)
    
    def _build_trends_view(self):
        """Vue Tendances"""
        self.trends_frame = ttk.Frame(self.content_notebook)
        self.content_notebook.add(self.trends_frame, text="📈 Tendances")
        
        options_frame = ttk.Frame(self.trends_frame)
        options_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(options_frame, text="Métrique :").pack(side="left")
        self.trend_metric_var = tk.StringVar(value="Marge")
        ttk.Combobox(options_frame, textvariable=self.trend_metric_var,
                    values=["Revenus", "Coûts", "Marge", "Palettes", "Missions"], width=12, state="readonly").pack(side="left", padx=5)
        
        ttk.Label(options_frame, text="Moyenne mobile :").pack(side="left", padx=(20, 5))
        self.trend_ma_var = tk.StringVar(value="7 jours")
        ttk.Combobox(options_frame, textvariable=self.trend_ma_var,
                    values=["Aucune", "3 jours", "7 jours", "14 jours", "30 jours"], width=12, state="readonly").pack(side="left", padx=5)
        
        self.trend_forecast_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Prévision", variable=self.trend_forecast_var).pack(side="left", padx=20)
        
        ttk.Button(options_frame, text="📈 Actualiser", command=self._update_trends).pack(side="left", padx=10)
        
        self.trends_container = ttk.Frame(self.trends_frame)
        self.trends_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        ttk.Label(self.trends_container, text="📈 Cliquez sur 'Analyser' puis 'Actualiser'",
                 font=("Segoe UI", 12), foreground="gray").pack(pady=50)
    
    # === Méthodes de gestion des filtres ===
    
    def _set_period(self, days):
        end_date = date.today()
        start_date = end_date - timedelta(days=days)
        self.date_start_var.set(start_date.strftime("%d/%m/%Y"))
        self.date_end_var.set(end_date.strftime("%d/%m/%Y"))
    
    def _set_period_this_month(self):
        today = date.today()
        self.date_start_var.set(date(today.year, today.month, 1).strftime("%d/%m/%Y"))
        self.date_end_var.set(today.strftime("%d/%m/%Y"))
    
    def _set_period_this_year(self):
        today = date.today()
        self.date_start_var.set(date(today.year, 1, 1).strftime("%d/%m/%Y"))
        self.date_end_var.set(today.strftime("%d/%m/%Y"))
    
    def _select_all_countries(self, select):
        for var in self.country_vars.values():
            var.set(select)
    
    def _select_all_listbox(self, listbox, select):
        if select:
            listbox.select_set(0, tk.END)
        else:
            listbox.select_clear(0, tk.END)
    
    def _toggle_comparison(self):
        if self.comparison_enabled_var.get():
            self.comp_options_frame.pack(fill="x", pady=5)
        else:
            self.comp_options_frame.pack_forget()
    
    def _on_view_changed(self, event=None):
        view_map = {"Dashboard": 0, "Graphiques": 1, "Tableau": 2, "Pivot Table": 3, "Tendances": 4}
        self.content_notebook.select(view_map.get(self.view_var.get(), 0))
    
    def reset_filters(self):
        self._set_period(30)
        self._select_all_countries(True)
        self._select_all_listbox(self.sst_listbox, True)
        self._select_all_listbox(self.voyage_listbox, True)
        self.type_livraison_var.set(True)
        self.type_ramasse_var.set(True)
        self.comparison_enabled_var.set(False)
        self._toggle_comparison()
    
    def _populate_filter_lists(self):
        sst_set = set()
        for ch in self.chauffeurs:
            sst = ch.get("sst", "")
            if sst:
                sst_set.add(sst)
        
        self.sst_listbox.delete(0, tk.END)
        for sst in sorted(sst_set):
            self.sst_listbox.insert(tk.END, sst)
        self._select_all_listbox(self.sst_listbox, True)
        
        voyage_set = set()
        for v in self.voyages:
            code = v.get("code", "")
            if code:
                voyage_set.add(code)
        
        self.voyage_listbox.delete(0, tk.END)
        for voyage in sorted(voyage_set):
            self.voyage_listbox.insert(tk.END, voyage)
        self._select_all_listbox(self.voyage_listbox, True)
    
    # === Analyse principale ===
    
    def run_analysis(self):
        try:
            start_date = parse_date_input(self.date_start_var.get())
            end_date = parse_date_input(self.date_end_var.get())
            
            if start_date > end_date:
                messagebox.showerror("Erreur", "La date de début doit être avant la date de fin")
                return
            
            filters = self._get_current_filters()
            self.current_data = self._collect_data(start_date, end_date, filters)
            
            if self.comparison_enabled_var.get():
                comp_start, comp_end = self._get_comparison_period(start_date, end_date)
                self.comparison_data = self._collect_data(comp_start, comp_end, filters)
            else:
                self.comparison_data = None
            
            self._populate_filter_lists()
            self._update_dashboard()
            self._update_charts()
            self._update_table()
            
            nb_days = (end_date - start_date).days + 1
            nb_missions = len(self.current_data.get('missions_list', []))
            messagebox.showinfo("Analyse terminée", f"{nb_days} jours analysés\n{nb_missions} missions trouvées")
            
        except ValueError as e:
            messagebox.showerror("Erreur", str(e))
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur: {e}")
            import traceback
            traceback.print_exc()
    
    def _get_current_filters(self):
        filters = {
            'countries': [c for c, v in self.country_vars.items() if v.get()],
            'sst': [self.sst_listbox.get(i) for i in self.sst_listbox.curselection()],
            'voyages': [self.voyage_listbox.get(i) for i in self.voyage_listbox.curselection()],
            'types': [],
        }
        if self.type_livraison_var.get():
            filters['types'].append('LIVRAISON')
        if self.type_ramasse_var.get():
            filters['types'].append('RAMASSE')
        return filters
    
    def _get_comparison_period(self, start_date, end_date):
        period_days = (end_date - start_date).days
        comp_type = self.comparison_type_var.get()
        
        if comp_type == "previous_year":
            return start_date.replace(year=start_date.year - 1), end_date.replace(year=end_date.year - 1)
        else:
            comp_end = start_date - timedelta(days=1)
            return comp_end - timedelta(days=period_days), comp_end
    
    def _collect_data(self, start_date, end_date, filters):
        data = {
            'dates': [], 'revenus': [], 'couts': [], 'marges': [],
            'missions_count': [], 'palettes_liv': [], 'palettes_ram': [],
            'by_voyage': {}, 'by_sst': {}, 'by_driver': {}, 'by_country': {},
            'by_type': {'LIVRAISON': {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0},
                       'RAMASSE': {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}},
            'by_weekday': {i: {'revenus': 0, 'couts': 0, 'missions': 0} for i in range(7)},
            'by_week': {}, 'by_month': {},
            'missions_list': [],
            'start_date': start_date, 'end_date': end_date, 'filters': filters,
        }
        
        current = start_date
        while current <= end_date:
            day_data = self._collect_day_data(current, filters)
            
            data['dates'].append(current)
            data['revenus'].append(day_data['revenus'])
            data['couts'].append(day_data['couts'])
            data['marges'].append(day_data['revenus'] - day_data['couts'])
            data['missions_count'].append(day_data['missions'])
            data['palettes_liv'].append(day_data['pal_liv'])
            data['palettes_ram'].append(day_data['pal_ram'])
            
            weekday = current.weekday()
            data['by_weekday'][weekday]['revenus'] += day_data['revenus']
            data['by_weekday'][weekday]['couts'] += day_data['couts']
            data['by_weekday'][weekday]['missions'] += day_data['missions']
            
            week_key = f"{current.year}-W{current.isocalendar()[1]:02d}"
            if week_key not in data['by_week']:
                data['by_week'][week_key] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
            data['by_week'][week_key]['revenus'] += day_data['revenus']
            data['by_week'][week_key]['couts'] += day_data['couts']
            data['by_week'][week_key]['missions'] += day_data['missions']
            
            month_key = f"{current.year}-{current.month:02d}"
            if month_key not in data['by_month']:
                data['by_month'][month_key] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
            data['by_month'][month_key]['revenus'] += day_data['revenus']
            data['by_month'][month_key]['couts'] += day_data['couts']
            data['by_month'][month_key]['missions'] += day_data['missions']
            
            for key in ['by_voyage', 'by_sst', 'by_driver', 'by_country']:
                for entity, vals in day_data[key].items():
                    if entity not in data[key]:
                        data[key][entity] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
                    for metric in ['revenus', 'couts', 'missions', 'palettes']:
                        data[key][entity][metric] += vals.get(metric, 0)
            
            for mission in day_data['missions_list']:
                m_type = mission.get('type', 'LIVRAISON')
                if m_type in data['by_type']:
                    data['by_type'][m_type]['revenus'] += mission.get('revenus', 0)
                    data['by_type'][m_type]['missions'] += 1
                    data['by_type'][m_type]['palettes'] += mission.get('palettes', 0)
            
            data['missions_list'].extend(day_data['missions_list'])
            current += timedelta(days=1)
        
        return data
    
    def _collect_day_data(self, d, filters):
        day_data = {
            'revenus': 0, 'couts': 0, 'missions': 0, 'pal_liv': 0, 'pal_ram': 0,
            'by_voyage': {}, 'by_sst': {}, 'by_driver': {}, 'by_country': {},
            'missions_list': [],
        }
        
        date_str = format_date_internal(d)
        day_dir = get_planning_day_dir(d)
        
        if not day_dir or not day_dir.exists():
            return day_data
        
        missions = []
        for file in day_dir.glob("*.json"):
            mission_data = load_json(file, None)
            if mission_data:
                missions.append(mission_data)
        
        revenus_date = self.revenus_palettes.get(date_str, {})
        
        for mission in missions:
            m_type = mission.get("type", "LIVRAISON")
            if m_type not in filters['types']:
                continue
            
            voyage_code = mission.get("voyage", "")
            voyage = next((v for v in self.voyages if v.get("code") == voyage_code), None)
            country = self._normalize_country(voyage.get("country", "Belgique") if voyage else "Belgique")
            
            if country not in filters['countries']:
                continue
            
            if filters['voyages'] and voyage_code not in filters['voyages']:
                continue
            
            nb_pal = int(mission.get("nb_pal", 0) or 0)
            
            rev_config = revenus_date.get(country, revenus_date.get("Belgique", {}))
            if isinstance(rev_config, dict):
                rev_per_pal = rev_config.get("livraison" if m_type == "LIVRAISON" else "ramasse", 0)
            else:
                rev_per_pal = 0
            
            mission_rev = nb_pal * rev_per_pal
            
            chauffeur_nom = mission.get("chauffeur_nom", "")
            chauffeur = next((c for c in self.chauffeurs if f"{c.get('nom', '')} {c.get('prenom', '')}" == chauffeur_nom), None)
            sst = chauffeur.get("sst", "") if chauffeur else mission.get("sst", "")
            
            if filters['sst'] and sst not in filters['sst']:
                continue
            
            mission_cout = 0
            if sst and sst in self.tarifs_sst:
                tarifs_country = self.tarifs_sst[sst].get(country, {})
                tarif = tarifs_country.get(date_str, 0)
                if tarif == 0:
                    sorted_dates = sorted([dt for dt in tarifs_country.keys() if dt <= date_str], reverse=True)
                    if sorted_dates:
                        tarif = tarifs_country[sorted_dates[0]]
                mission_cout = tarif
            
            day_data['missions'] += 1
            day_data['revenus'] += mission_rev
            day_data['couts'] += mission_cout
            
            if m_type == "LIVRAISON":
                day_data['pal_liv'] += nb_pal
            else:
                day_data['pal_ram'] += nb_pal
            
            if voyage_code:
                if voyage_code not in day_data['by_voyage']:
                    day_data['by_voyage'][voyage_code] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
                day_data['by_voyage'][voyage_code]['revenus'] += mission_rev
                day_data['by_voyage'][voyage_code]['missions'] += 1
                day_data['by_voyage'][voyage_code]['palettes'] += nb_pal
            
            if sst:
                if sst not in day_data['by_sst']:
                    day_data['by_sst'][sst] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
                day_data['by_sst'][sst]['couts'] += mission_cout
                day_data['by_sst'][sst]['missions'] += 1
            
            if chauffeur_nom:
                if chauffeur_nom not in day_data['by_driver']:
                    day_data['by_driver'][chauffeur_nom] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
                day_data['by_driver'][chauffeur_nom]['revenus'] += mission_rev
                day_data['by_driver'][chauffeur_nom]['couts'] += mission_cout
                day_data['by_driver'][chauffeur_nom]['missions'] += 1
                day_data['by_driver'][chauffeur_nom]['palettes'] += nb_pal
            
            if country:
                if country not in day_data['by_country']:
                    day_data['by_country'][country] = {'revenus': 0, 'couts': 0, 'missions': 0, 'palettes': 0}
                day_data['by_country'][country]['revenus'] += mission_rev
                day_data['by_country'][country]['couts'] += mission_cout
                day_data['by_country'][country]['missions'] += 1
                day_data['by_country'][country]['palettes'] += nb_pal
            
            day_data['missions_list'].append({
                'date': d, 'voyage': voyage_code, 'pays': country, 'type': m_type,
                'sst': sst, 'chauffeur': chauffeur_nom, 'palettes': nb_pal,
                'revenus': mission_rev, 'couts': mission_cout, 'marge': mission_rev - mission_cout,
            })
        
        return day_data
    
    def _normalize_country(self, country):
        mapping = {
            "pays-bas": "Pays-Bas", "paysbas": "Pays-Bas", "netherlands": "Pays-Bas",
            "belgique": "Belgique", "belgium": "Belgique",
            "luxembourg": "Luxembourg", "france": "France", "allemagne": "Allemagne",
        }
        return mapping.get(country.lower(), country)
    
    # === Mise à jour des vues ===
    
    def _update_dashboard(self):
        if not self.current_data:
            return
        
        for widget in self.dashboard_content.winfo_children():
            widget.destroy()
        
        data = self.current_data
        comp_data = self.comparison_data
        
        total_revenus = sum(data['revenus'])
        total_couts = sum(data['couts'])
        total_marge = sum(data['marges'])
        total_missions = sum(data['missions_count'])
        total_palettes = sum(data['palettes_liv']) + sum(data['palettes_ram'])
        
        nb_days = (data['end_date'] - data['start_date']).days + 1
        days_with_data = len([r for r in data['revenus'] if r > 0])
        
        variations = {}
        if comp_data:
            variations = {
                'revenus': calculate_variation(total_revenus, sum(comp_data['revenus'])),
                'couts': calculate_variation(total_couts, sum(comp_data['couts'])),
                'marge': calculate_variation(total_marge, sum(comp_data['marges'])),
                'missions': calculate_variation(total_missions, sum(comp_data['missions_count'])),
            }
        
        # Header période
        kpi_frame = ttk.Frame(self.dashboard_content)
        kpi_frame.pack(fill="x", padx=10, pady=10)
        
        period_text = f"📅 {format_date_display(data['start_date'])} - {format_date_display(data['end_date'])} ({nb_days} jours)"
        ttk.Label(kpi_frame, text=period_text, font=("Segoe UI", 11)).pack(anchor="w", pady=(0, 10))
        
        # KPIs Grid
        kpis_grid = ttk.Frame(kpi_frame)
        kpis_grid.pack(fill="x")
        
        kpis = [
            ("💰 Revenus", format_currency(total_revenus), variations.get('revenus')),
            ("💸 Coûts", format_currency(total_couts), variations.get('couts')),
            ("📊 Marge", format_currency(total_marge), variations.get('marge')),
            ("📦 Missions", format_number(total_missions), variations.get('missions')),
            ("🎯 Palettes", format_number(total_palettes), None),
        ]
        
        for i, (title, value, variation) in enumerate(kpis):
            card = ttk.LabelFrame(kpis_grid, text=title, padding=10)
            card.grid(row=0, column=i, padx=5, pady=5, sticky="nsew")
            
            ttk.Label(card, text=value, font=("Segoe UI", 14, "bold")).pack()
            
            if variation is not None:
                arrow = "▲" if variation >= 0 else "▼"
                var_text = f"{arrow} {abs(variation):.1f}%"
                ttk.Label(card, text=var_text, font=("Segoe UI", 9)).pack()
        
        for i in range(5):
            kpis_grid.columnconfigure(i, weight=1)
        
        # Moyennes
        ttk.Separator(self.dashboard_content, orient="horizontal").pack(fill="x", padx=10, pady=10)
        
        avg_frame = ttk.LabelFrame(self.dashboard_content, text="📊 Moyennes", padding=10)
        avg_frame.pack(fill="x", padx=10, pady=5)
        
        avg_revenus = total_revenus / days_with_data if days_with_data > 0 else 0
        margin_pct = (total_marge / total_revenus * 100) if total_revenus > 0 else 0
        
        avgs = [
            f"Revenus/jour: {format_currency(avg_revenus)}",
            f"Missions/jour: {total_missions / days_with_data:.1f}" if days_with_data > 0 else "N/A",
            f"Taux de marge: {margin_pct:.1f}%",
        ]
        
        for text in avgs:
            ttk.Label(avg_frame, text=text, font=("Segoe UI", 10)).pack(side="left", padx=20)
        
        # Top performers
        self._create_top_performers_section()
        
        # Mini graphiques
        if MATPLOTLIB_AVAILABLE:
            self._create_mini_charts()
    
    def _create_top_performers_section(self):
        if not self.current_data:
            return
        
        top_frame = ttk.LabelFrame(self.dashboard_content, text="🏆 Top Performers", padding=10)
        top_frame.pack(fill="x", padx=10, pady=5)
        
        grid = ttk.Frame(top_frame)
        grid.pack(fill="x")
        
        data = self.current_data
        
        # Top Voyages
        col1 = ttk.Frame(grid)
        col1.grid(row=0, column=0, padx=10, sticky="nsew")
        ttk.Label(col1, text="🚛 Top Voyages", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        
        sorted_voyages = sorted(data['by_voyage'].items(), key=lambda x: x[1]['revenus'], reverse=True)[:5]
        for voyage, vals in sorted_voyages:
            ttk.Label(col1, text=f"  {voyage}: {format_currency(vals['revenus'])}", font=("Segoe UI", 9)).pack(anchor="w")
        
        # Top Chauffeurs
        col2 = ttk.Frame(grid)
        col2.grid(row=0, column=1, padx=10, sticky="nsew")
        ttk.Label(col2, text="👤 Top Chauffeurs", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        
        driver_margins = [(d, v['revenus'] - v['couts']) for d, v in data['by_driver'].items()]
        sorted_drivers = sorted(driver_margins, key=lambda x: x[1], reverse=True)[:5]
        for driver, marge in sorted_drivers:
            ttk.Label(col2, text=f"  {driver[:20]}: {format_currency(marge)}", font=("Segoe UI", 9)).pack(anchor="w")
        
        # Par Pays
        col3 = ttk.Frame(grid)
        col3.grid(row=0, column=2, padx=10, sticky="nsew")
        ttk.Label(col3, text="🌍 Par Pays", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        
        for country, vals in sorted(data['by_country'].items(), key=lambda x: x[1]['revenus'], reverse=True):
            marge = vals['revenus'] - vals['couts']
            flag = COUNTRY_FLAGS.get(country, "")
            ttk.Label(col3, text=f"  {flag} {country}: {format_currency(marge)}", font=("Segoe UI", 9)).pack(anchor="w")
        
        for i in range(3):
            grid.columnconfigure(i, weight=1)
    
    def _create_mini_charts(self):
        if not self.current_data:
            return
        
        charts_frame = ttk.LabelFrame(self.dashboard_content, text="📈 Aperçu", padding=10)
        charts_frame.pack(fill="x", padx=10, pady=5)
        
        data = self.current_data
        
        fig = Figure(figsize=(12, 3), dpi=100)
        fig.patch.set_facecolor('#f8fafc')
        
        # Evolution
        ax1 = fig.add_subplot(131)
        if data['dates']:
            ax1.plot(data['dates'], data['revenus'], 'g-', linewidth=1.5, label='Rev')
            ax1.plot(data['dates'], data['couts'], 'r-', linewidth=1.5, label='Coûts')
            ax1.fill_between(data['dates'], data['marges'], alpha=0.3, color='blue')
            ax1.set_title('Évolution', fontsize=9)
            ax1.tick_params(labelsize=7)
            ax1.legend(fontsize=7)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        
        # Par pays (pie)
        ax2 = fig.add_subplot(132)
        if data['by_country']:
            countries = list(data['by_country'].keys())
            revenus = [data['by_country'][c]['revenus'] for c in countries]
            if sum(revenus) > 0:
                ax2.pie(revenus, labels=countries, autopct='%1.0f%%', textprops={'fontsize': 7})
                ax2.set_title('Par Pays', fontsize=9)
        
        # Top voyages (bar)
        ax3 = fig.add_subplot(133)
        if data['by_voyage']:
            sorted_voyages = sorted(data['by_voyage'].items(), key=lambda x: x[1]['revenus'], reverse=True)[:5]
            codes = [v[0] for v in sorted_voyages]
            revs = [v[1]['revenus'] for v in sorted_voyages]
            ax3.barh(codes, revs, color=COLORS_ANALYSE['chart_blue'])
            ax3.set_title('Top 5 Voyages', fontsize=9)
            ax3.tick_params(labelsize=7)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=charts_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
    
    def _update_charts(self):
        if not MATPLOTLIB_AVAILABLE or not self.current_data:
            return
        
        for widget in self.charts_container.winfo_children():
            widget.destroy()
        self.analyse_figures = []
        
        data = self.current_data
        chart_type = self.chart_type_var.get()
        
        charts_to_create = []
        if chart_type in ["Tous", "Évolution"]:
            charts_to_create.append(("timeline", "📈 Évolution temporelle"))
        if chart_type in ["Tous", "Par Voyage"]:
            charts_to_create.append(("voyage", "🚛 Par Voyage (Top 15)"))
        if chart_type in ["Tous", "Par SST"]:
            charts_to_create.append(("sst", "🏢 Par SST"))
        if chart_type in ["Tous", "Par Chauffeur"]:
            charts_to_create.append(("driver", "👤 Par Chauffeur (Top 15)"))
        if chart_type in ["Tous", "Par Pays"]:
            charts_to_create.append(("country", "🌍 Par Pays"))
        
        for c_type, title in charts_to_create:
            chart_frame = ttk.LabelFrame(self.charts_container, text=title, padding=5)
            chart_frame.pack(fill="x", pady=5, padx=5)
            
            fig = Figure(figsize=(11, 4), dpi=100)
            fig.patch.set_facecolor('#f8fafc')
            
            if c_type == "timeline":
                self._create_timeline_chart(fig, data)
            elif c_type == "voyage":
                self._create_voyage_chart(fig, data)
            elif c_type == "sst":
                self._create_sst_chart(fig, data)
            elif c_type == "driver":
                self._create_driver_chart(fig, data)
            elif c_type == "country":
                self._create_country_chart(fig, data)
            
            canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            
            toolbar_frame = ttk.Frame(chart_frame)
            toolbar_frame.pack(fill="x")
            NavigationToolbar2Tk(canvas, toolbar_frame)
            
            self.analyse_figures.append((fig, canvas))
    
    def _create_timeline_chart(self, fig, data):
        ax = fig.add_subplot(111)
        if not data['dates']:
            ax.text(0.5, 0.5, "Aucune donnée", ha='center', va='center')
            return
        
        ax.plot(data['dates'], data['revenus'], color=COLORS_ANALYSE['success'], linewidth=2, label='Revenus', marker='o', markersize=3)
        ax.plot(data['dates'], data['couts'], color=COLORS_ANALYSE['danger'], linewidth=2, label='Coûts', marker='s', markersize=3)
        ax.fill_between(data['dates'], data['marges'], alpha=0.3, color=COLORS_ANALYSE['chart_blue'], label='Marge')
        ax.axhline(y=0, color='black', linestyle='-', linewidth=0.5)
        
        ax.set_xlabel('Date')
        ax.set_ylabel('Montant (€)')
        ax.legend(loc='upper left', fontsize=8)
        ax.grid(True, alpha=0.3)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        
        total_rev = sum(data['revenus'])
        total_cout = sum(data['couts'])
        total_marge = sum(data['marges'])
        ax.text(0.98, 0.98, f"Rev: {format_currency(total_rev)}\nCoûts: {format_currency(total_cout)}\nMarge: {format_currency(total_marge)}",
                transform=ax.transAxes, fontsize=9, verticalalignment='top', horizontalalignment='right',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        fig.autofmt_xdate()
        fig.tight_layout()
    
    def _create_voyage_chart(self, fig, data):
        ax = fig.add_subplot(111)
        if not data['by_voyage']:
            ax.text(0.5, 0.5, "Aucune donnée", ha='center', va='center')
            return
        
        sorted_voyages = sorted(data['by_voyage'].items(), key=lambda x: x[1]['revenus'], reverse=True)[:15]
        codes = [v[0] for v in sorted_voyages]
        revenus = [v[1]['revenus'] for v in sorted_voyages]
        missions = [v[1]['missions'] for v in sorted_voyages]
        
        colors = plt.cm.Blues([0.4 + 0.4 * i/len(codes) for i in range(len(codes))])
        bars = ax.bar(codes, revenus, color=colors)
        
        for bar, m_count in zip(bars, missions):
            ax.text(bar.get_x() + bar.get_width()/2., bar.get_height(), f'{m_count}m', ha='center', va='bottom', fontsize=8)
        
        ax.set_xlabel('Code Voyage')
        ax.set_ylabel('Revenus (€)')
        ax.tick_params(axis='x', rotation=45)
        ax.grid(True, alpha=0.3, axis='y')
        fig.tight_layout()
    
    def _create_sst_chart(self, fig, data):
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        
        if not data['by_sst']:
            ax1.text(0.5, 0.5, "Aucune donnée", ha='center', va='center')
            return
        
        ssts = list(data['by_sst'].keys())
        couts = [data['by_sst'][s]['couts'] for s in ssts]
        missions = [data['by_sst'][s]['missions'] for s in ssts]
        
        ax1.bar(ssts, couts, color=plt.cm.Reds([0.4 + 0.4 * i/len(ssts) for i in range(len(ssts))]))
        ax1.set_title('Coûts par SST', fontsize=10)
        ax1.set_ylabel('Coûts (€)')
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, alpha=0.3, axis='y')
        
        ax2.bar(ssts, missions, color=plt.cm.Blues([0.4 + 0.4 * i/len(ssts) for i in range(len(ssts))]))
        ax2.set_title('Missions par SST', fontsize=10)
        ax2.set_ylabel('Missions')
        ax2.tick_params(axis='x', rotation=45)
        ax2.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout()
    
    def _create_driver_chart(self, fig, data):
        ax = fig.add_subplot(111)
        if not data['by_driver']:
            ax.text(0.5, 0.5, "Aucune donnée", ha='center', va='center')
            return
        
        import numpy as np
        driver_data = [(d, v['revenus'], v['couts'], v['revenus'] - v['couts']) for d, v in data['by_driver'].items()]
        sorted_drivers = sorted(driver_data, key=lambda x: x[3], reverse=True)[:15]
        
        drivers = [d[0][:15] for d in sorted_drivers]
        revenus = [d[1] for d in sorted_drivers]
        couts = [d[2] for d in sorted_drivers]
        marges = [d[3] for d in sorted_drivers]
        
        x = np.arange(len(drivers))
        width = 0.35
        
        ax.bar(x - width/2, revenus, width, label='Revenus', color=COLORS_ANALYSE['success'], alpha=0.8)
        ax.bar(x + width/2, couts, width, label='Coûts', color=COLORS_ANALYSE['danger'], alpha=0.8)
        
        ax2 = ax.twinx()
        ax2.plot(x, marges, 'b-', marker='D', linewidth=2, label='Marge', markersize=6)
        ax2.axhline(y=0, color='blue', linestyle='--', alpha=0.5)
        ax2.set_ylabel('Marge (€)', color='blue')
        
        ax.set_xticks(x)
        ax.set_xticklabels(drivers, rotation=45, ha='right')
        ax.set_ylabel('Montant (€)')
        ax.legend(loc='upper left')
        ax2.legend(loc='upper right')
        ax.grid(True, alpha=0.3, axis='y')
        fig.tight_layout()
    
    def _create_country_chart(self, fig, data):
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        
        if not data['by_country']:
            ax1.text(0.5, 0.5, "Aucune donnée", ha='center', va='center')
            return
        
        import numpy as np
        countries = list(data['by_country'].keys())
        revenus = [data['by_country'][c]['revenus'] for c in countries]
        couts = [data['by_country'][c]['couts'] for c in countries]
        marges = [r - c for r, c in zip(revenus, couts)]
        
        if sum(revenus) > 0:
            ax1.pie(revenus, labels=countries, autopct='%1.1f%%', startangle=90)
            ax1.set_title('Répartition revenus')
        
        x = np.arange(len(countries))
        width = 0.35
        ax2.bar(x - width/2, revenus, width, label='Revenus', color=COLORS_ANALYSE['success'])
        ax2.bar(x + width/2, couts, width, label='Coûts', color=COLORS_ANALYSE['danger'])
        
        for i, marge in enumerate(marges):
            color = COLORS_ANALYSE['success'] if marge >= 0 else COLORS_ANALYSE['danger']
            ax2.text(i, max(revenus[i], couts[i]) + 50, f'{marge:+,.0f}€', ha='center', fontsize=8, color=color)
        
        ax2.set_xticks(x)
        ax2.set_xticklabels(countries)
        ax2.set_title('Revenus vs Coûts')
        ax2.legend()
        ax2.grid(True, alpha=0.3, axis='y')
        fig.tight_layout()
    
    def _update_table(self):
        if not self.current_data:
            return
        
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)
        
        for i, mission in enumerate(self.current_data['missions_list']):
            values = (
                format_date_display(mission['date']), mission['voyage'], mission['pays'],
                mission['type'], mission['sst'], mission['chauffeur'], mission['palettes'],
                format_currency(mission['revenus']), format_currency(mission['couts']),
                format_currency(mission['marge']),
            )
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.data_tree.insert("", "end", values=values, tags=(tag,))
        
        self._update_table_stats()
    
    def _update_table_stats(self):
        for widget in self.table_stats_frame.winfo_children():
            widget.destroy()
        
        if not self.current_data:
            return
        
        data = self.current_data
        total_rev = sum(m['revenus'] for m in data['missions_list'])
        total_cout = sum(m['couts'] for m in data['missions_list'])
        total_marge = sum(m['marge'] for m in data['missions_list'])
        total_pal = sum(m['palettes'] for m in data['missions_list'])
        nb_missions = len(data['missions_list'])
        
        stats = [f"📊 {nb_missions} missions", f"📦 {total_pal} pal", f"💰 {format_currency(total_rev)}",
                f"💸 {format_currency(total_cout)}", f"📈 {format_currency(total_marge)}"]
        
        for stat in stats:
            ttk.Label(self.table_stats_frame, text=stat, font=("Segoe UI", 9)).pack(side="left", padx=10)
    
    def _sort_table(self, col):
        items = [(self.data_tree.set(k, col), k) for k in self.data_tree.get_children()]
        items.sort(reverse=getattr(self, f'_sort_reverse_{col}', False))
        for index, (val, k) in enumerate(items):
            self.data_tree.move(k, '', index)
        setattr(self, f'_sort_reverse_{col}', not getattr(self, f'_sort_reverse_{col}', False))
    
    def _filter_table(self):
        search = self.table_search_var.get().lower()
        for item in self.data_tree.get_children():
            values = self.data_tree.item(item, 'values')
            if search in ' '.join(str(v).lower() for v in values):
                self.data_tree.reattach(item, '', 'end')
            else:
                self.data_tree.detach(item)
    
    def _generate_pivot(self):
        if not self.current_data:
            messagebox.showwarning("Attention", "Lancez d'abord une analyse")
            return
        
        for widget in self.pivot_container.winfo_children():
            widget.destroy()
        
        data = self.current_data
        rows_dim = self.pivot_rows_var.get()
        cols_dim = self.pivot_cols_var.get()
        value_metric = self.pivot_value_var.get()
        
        pivot_data = self._build_pivot_data(data, rows_dim, cols_dim, value_metric)
        
        if not pivot_data:
            ttk.Label(self.pivot_container, text="Pas de données", font=("Segoe UI", 11)).pack(pady=20)
            return
        
        rows = list(pivot_data.keys())
        cols = set()
        for row_data in pivot_data.values():
            cols.update(row_data.keys())
        cols = sorted(cols)
        
        columns = [rows_dim] + cols + ['Total']
        
        tree_frame = ttk.Frame(self.pivot_container)
        tree_frame.pack(fill="both", expand=True)
        
        pivot_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        for col in columns:
            pivot_tree.heading(col, text=col)
            pivot_tree.column(col, width=100, anchor="e" if col != rows_dim else "w")
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=pivot_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=pivot_tree.xview)
        pivot_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        pivot_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        for row_key in rows:
            row_values = [row_key]
            total = 0
            for col_key in cols:
                val = pivot_data[row_key].get(col_key, 0)
                row_values.append(format_number(val, 2) if val != 0 else "-")
                total += val
            row_values.append(format_number(total, 2))
            pivot_tree.insert("", "end", values=row_values)
        
        # Totaux
        total_row = ['TOTAL']
        grand_total = 0
        for col_key in cols:
            col_total = sum(pivot_data[r].get(col_key, 0) for r in rows)
            total_row.append(format_number(col_total, 2))
            grand_total += col_total
        total_row.append(format_number(grand_total, 2))
        pivot_tree.insert("", "end", values=total_row, tags=('total',))
        pivot_tree.tag_configure('total', background='#e2e8f0', font=('Segoe UI', 9, 'bold'))
    
    def _build_pivot_data(self, data, rows_dim, cols_dim, value_metric):
        pivot = {}
        metric_map = {'Revenus': 'revenus', 'Coûts': 'couts', 'Marge': 'marge', 'Palettes': 'palettes', 'Missions': 'missions'}
        metric_key = metric_map.get(value_metric, 'marge')
        
        for mission in data['missions_list']:
            row_key = self._get_dim_value(mission, rows_dim)
            col_key = self._get_dim_value(mission, cols_dim)
            
            value = 1 if metric_key == 'missions' else mission.get(metric_key, 0)
            
            if row_key not in pivot:
                pivot[row_key] = {}
            if col_key not in pivot[row_key]:
                pivot[row_key][col_key] = 0
            pivot[row_key][col_key] += value
        
        return pivot
    
    def _get_dim_value(self, mission, dim):
        if dim == "Date":
            return format_date_display(mission['date'])
        elif dim == "Semaine":
            return f"S{mission['date'].isocalendar()[1]:02d}"
        elif dim == "Mois":
            return mission['date'].strftime("%Y-%m")
        elif dim == "Voyage":
            return mission['voyage']
        elif dim == "SST":
            return mission['sst'] or "N/A"
        elif dim == "Chauffeur":
            return mission['chauffeur'] or "N/A"
        elif dim == "Pays":
            return mission['pays']
        elif dim == "Type":
            return mission['type']
        return "N/A"
    
    def _update_trends(self):
        if not MATPLOTLIB_AVAILABLE or not self.current_data:
            return
        
        for widget in self.trends_container.winfo_children():
            widget.destroy()
        
        data = self.current_data
        metric = self.trend_metric_var.get()
        ma_option = self.trend_ma_var.get()
        
        metric_map = {'Revenus': 'revenus', 'Coûts': 'couts', 'Marge': 'marges', 'Palettes': 'palettes_liv', 'Missions': 'missions_count'}
        metric_key = metric_map.get(metric, 'marges')
        
        dates = data['dates']
        values = data[metric_key]
        
        if metric_key == 'palettes_liv':
            values = [l + r for l, r in zip(data['palettes_liv'], data['palettes_ram'])]
        
        import numpy as np
        
        fig = Figure(figsize=(11, 6), dpi=100)
        fig.patch.set_facecolor('#f8fafc')
        ax = fig.add_subplot(111)
        
        ax.plot(dates, values, 'b-', linewidth=2, label=metric, marker='o', markersize=4)
        
        if ma_option != "Aucune" and len(values) >= 3:
            ma_days = int(ma_option.split()[0])
            if len(values) >= ma_days:
                ma_values = np.convolve(values, np.ones(ma_days)/ma_days, mode='valid')
                ax.plot(dates[ma_days-1:], ma_values, 'r--', linewidth=2, label=f'MM{ma_days}', alpha=0.8)
        
        if len(values) >= 2:
            x_numeric = np.arange(len(values))
            z = np.polyfit(x_numeric, values, 1)
            p = np.poly1d(z)
            ax.plot(dates, p(x_numeric), 'g--', linewidth=1.5, label=f'Tendance ({z[0]:+.1f}/j)', alpha=0.7)
        
        if self.trend_forecast_var.get() and len(values) >= 7:
            forecast_days = 7
            last_date = dates[-1]
            forecast_dates = [last_date + timedelta(days=i+1) for i in range(forecast_days)]
            
            x_numeric = np.arange(len(values))
            z = np.polyfit(x_numeric, values, 1)
            p = np.poly1d(z)
            
            forecast_x = np.arange(len(values), len(values) + forecast_days)
            forecast_values = p(forecast_x)
            
            ax.plot(forecast_dates, forecast_values, 'b:', linewidth=2, label='Prévision', alpha=0.6)
        
        ax.set_xlabel('Date')
        ax.set_ylabel(metric)
        ax.set_title(f'Tendance - {metric}')
        ax.legend(loc='upper left')
        ax.grid(True, alpha=0.3)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        fig.autofmt_xdate()
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.trends_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        
        toolbar_frame = ttk.Frame(self.trends_container)
        toolbar_frame.pack(fill="x")
        NavigationToolbar2Tk(canvas, toolbar_frame)
    
    # === Exports ===
    
    def export_to_excel(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Erreur", "pip install openpyxl")
            return
        if not self.current_data:
            messagebox.showwarning("Attention", "Lancez d'abord une analyse")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            data = self.current_data
            
            # Feuille résumé
            ws1 = wb.active
            ws1.title = "Résumé"
            ws1['A1'] = "RAPPORT D'ANALYSE"
            ws1['A1'].font = Font(bold=True, size=14)
            ws1['A3'] = f"Période: {format_date_display(data['start_date'])} - {format_date_display(data['end_date'])}"
            
            kpis = [("Revenus", sum(data['revenus'])), ("Coûts", sum(data['couts'])),
                   ("Marge", sum(data['marges'])), ("Missions", sum(data['missions_count']))]
            for i, (label, value) in enumerate(kpis, 5):
                ws1[f'A{i}'] = label
                ws1[f'B{i}'] = value
            
            # Feuille détail
            ws2 = wb.create_sheet("Détail")
            headers = ["Date", "Voyage", "Pays", "Type", "SST", "Chauffeur", "Palettes", "Revenus", "Coûts", "Marge"]
            for col, h in enumerate(headers, 1):
                ws2.cell(row=1, column=col, value=h).font = Font(bold=True)
            
            for row, m in enumerate(data['missions_list'], 2):
                ws2.cell(row=row, column=1, value=format_date_display(m['date']))
                ws2.cell(row=row, column=2, value=m['voyage'])
                ws2.cell(row=row, column=3, value=m['pays'])
                ws2.cell(row=row, column=4, value=m['type'])
                ws2.cell(row=row, column=5, value=m['sst'])
                ws2.cell(row=row, column=6, value=m['chauffeur'])
                ws2.cell(row=row, column=7, value=m['palettes'])
                ws2.cell(row=row, column=8, value=m['revenus'])
                ws2.cell(row=row, column=9, value=m['couts'])
                ws2.cell(row=row, column=10, value=m['marge'])
            
            export_dir = ROOT_DIR / "_export"
            export_dir.mkdir(parents=True, exist_ok=True)
            filename = export_dir / f"analyse_{data['start_date'].strftime('%Y%m%d')}_{data['end_date'].strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("Succès", f"Exporté: {filename}")
            import subprocess
            subprocess.Popen(f'explorer /select,"{filename}"')
        except Exception as e:
            messagebox.showerror("Erreur", str(e))
    
    def export_to_csv(self):
        if not self.current_data:
            messagebox.showwarning("Attention", "Lancez d'abord une analyse")
            return
        
        try:
            import csv
            data = self.current_data
            export_dir = ROOT_DIR / "_export"
            export_dir.mkdir(parents=True, exist_ok=True)
            filename = export_dir / f"analyse_{data['start_date'].strftime('%Y%m%d')}_{data['end_date'].strftime('%Y%m%d')}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(["Date", "Voyage", "Pays", "Type", "SST", "Chauffeur", "Palettes", "Revenus", "Coûts", "Marge"])
                for m in data['missions_list']:
                    writer.writerow([format_date_display(m['date']), m['voyage'], m['pays'], m['type'],
                                   m['sst'], m['chauffeur'], m['palettes'], m['revenus'], m['couts'], m['marge']])
            
            messagebox.showinfo("Succès", f"Exporté: {filename}")
            import subprocess
            subprocess.Popen(f'explorer /select,"{filename}"')
        except Exception as e:
            messagebox.showerror("Erreur", str(e))
    
    def export_to_pdf(self):
        if not PDF_AVAILABLE:
            messagebox.showerror("Erreur", "pip install reportlab")
            return
        if not self.current_data:
            messagebox.showwarning("Attention", "Lancez d'abord une analyse")
            return
        
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            
            data = self.current_data
            export_dir = ROOT_DIR / "_export"
            export_dir.mkdir(parents=True, exist_ok=True)
            filename = export_dir / f"rapport_{data['start_date'].strftime('%Y%m%d')}_{data['end_date'].strftime('%Y%m%d')}.pdf"
            
            doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            elements.append(Paragraph("RAPPORT D'ANALYSE - PLANNING TRANSPORT", styles['Heading1']))
            elements.append(Paragraph(f"Période: {format_date_display(data['start_date'])} - {format_date_display(data['end_date'])}", styles['Normal']))
            elements.append(Spacer(1, 20))
            
            kpi_data = [["Indicateur", "Valeur"],
                       ["Revenus", format_currency(sum(data['revenus']))],
                       ["Coûts", format_currency(sum(data['couts']))],
                       ["Marge", format_currency(sum(data['marges']))],
                       ["Missions", str(sum(data['missions_count']))]]
            
            t = Table(kpi_data, colWidths=[6*cm, 4*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            
            doc.build(elements)
            messagebox.showinfo("Succès", f"Exporté: {filename}")
            import subprocess
            subprocess.Popen(f'explorer /select,"{filename}"')
        except Exception as e:
            messagebox.showerror("Erreur", str(e))


class TransportPlannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Planning transport Tubize - BETA v0.6.0 - Analyse Avancée")
        self.root.minsize(1200, 700)
        self.root.geometry("1400x900")
        self.root.resizable(True, True)

        self.current_user = getpass.getuser().upper()
        # Informations sur la session et la dernière mise à jour
        self.last_refresh_dt = None  # datetime de la dernière mise à jour automatique/réussie
        self.status_var = tk.StringVar(value=f"Session : {self.current_user} | Dernière MAJ : -")

        init_default_data()
        self.data_dir = ROOT_DIR / "_data"
        
        # Initialiser le système de logging Sauron
        activity_logger.initialize(ROOT_DIR, self.current_user)
        
        # Enregistrer la fermeture de l'application
        self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)

        self.rights = self.load_rights()
        self.sst_list = load_json(self.data_dir / "sst.json", [])
        self.voyages = self.load_voyages_data()
        self.chauffeurs = load_json(self.data_dir / "chauffeurs.json", [])
        self.dispos = load_json(self.data_dir / "dispo_chauffeurs.json", [])
        self.tarifs_sst = load_json(self.data_dir / "tarifs_sst.json", {})
        self.revenus_palettes = load_json(self.data_dir / "revenus_palettes.json", {})

        self.current_date = date.today()
        self.missions = []

        self.country_trees = {}
        self.country_frames = {}
        
        self.sort_criteria = "heure"
        self.sort_reverse = False
        
        self.user_editing = False

        # État pour le système de rafraîchissement intelligent
        self.auto_refresh_enabled = True
        self.auto_refresh_interval = 60000  # Intervalle interne (ms) si utilisé
        self.is_editing = False
        self.refresh_timer_id = None
        self.file_timestamps = {}
        # Boucle d'auto-rafraîchissement simple (sans clignotement)
        self._auto_refresh_interval_ms = 15000  # 15 secondes
        self._auto_refresh_job = None


        self.build_gui()

        # Démarrer le système de cache en arrière-plan
        # Le cache pré-télécharge les plannings des dates proches
        planning_cache.start(on_cache_updated=self._on_cache_updated)
        # Nettoyer les anciennes entrées de cache
        planning_cache.clear_old_cache(max_age_days=30)

        self.load_planning_for_date(self.current_date)
        self._start_auto_refresh_loop()
        self.update_status_bar_initial()

        # Log du changement d'onglet
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _on_cache_updated(self):
        """Callback appelé quand le cache a été mis à jour en arrière-plan"""
        # On peut éventuellement rafraîchir l'UI si la date courante a été mise en cache
        # Pour l'instant, on ne fait rien car l'UI se met à jour automatiquement
        pass

    def _on_app_close(self):
        """Gérer la fermeture de l'application"""
        try:
            # Arrêter le système de cache
            planning_cache.stop()
        except Exception as e:
            print(f"Erreur arrêt cache: {e}")
        try:
            activity_logger.log_session_end()
        except Exception as e:
            print(f"Erreur log fermeture: {e}")
        finally:
            self.root.destroy()
    
    def _on_tab_changed(self, event):
        """Logger les changements d'onglet"""
        try:
            current_tab = self.notebook.tab(self.notebook.select(), "text")
            activity_logger.log_action("TAB_CHANGE", {"tab": current_tab})
        except Exception:
            pass


    def update_status_bar_initial(self):
        """Initialise la barre de statut au démarrage (session + heure de lancement)."""
        from datetime import datetime
        self.last_refresh_dt = datetime.now()
        try:
            self.status_var.set(
                f"Session : {self.current_user} | Dernière MAJ : {self.last_refresh_dt.strftime('%d/%m/%Y %H:%M:%S')}"
            )
        except Exception:
            pass

    def refresh_all(self):
        """
        Rafraîchissement global utilisé par le menu Setup.
        On recharge uniquement les fichiers qui ont changé et on met à jour
        les vues de manière intelligente (sans clignotement).
        """
        # Ne pas rafraîchir si l'utilisateur est en train d'éditer quelque chose
        if getattr(self, "user_editing", False) or getattr(self, "is_editing", False):
            return

        try:
            # Recharger les données JSON uniquement si des fichiers ont changé
            files_changed = self.reload_data_from_files()
            if not files_changed:
                return

            # Rafraîchir les différentes vues de façon incrémentale
            self.smart_refresh_all_views()

            # Mettre à jour le résumé en haut de l'écran
            if hasattr(self, "update_summary_stats"):
                try:
                    self.update_summary_stats()
                except Exception:
                    pass

            # Garder l'écran des droits utilisateurs à jour si besoin
            if hasattr(self, "refresh_users_view"):
                try:
                    self.refresh_users_view()
                except Exception:
                    pass

        except Exception as e:
            print(f"Erreur lors du refresh_all: {e}")
            import traceback
            traceback.print_exc()

    def load_rights(self):
        rights = load_json(self.data_dir / "users_rights.json", {})
        roles_def = rights.get("roles", {})
        users_def = rights.get("users", {})

        user_roles = users_def.get(self.current_user, ["viewer"])

        permissions = {
            "view_planning": False,
            "edit_planning": False,
            "view_drivers": False,
            "manage_drivers": False,
            "edit_driver_planning": False,
            "manage_rights": False,
            "manage_voyages": False,
            "generate_planning": False,
            "edit_past_planning": False,
            "edit_past_planning_advanced": False,
            "view_finance": False,
            "manage_finance": False,
            "view_analyse": False,
            "view_sauron": False,
            "send_announcements": False,
            "manage_announcements_config": False,
        }

        for role_name in user_roles:
            role = roles_def.get(role_name, {})
            for k, v in role.items():
                if v:
                    permissions[k] = True

        return {"roles_def": roles_def, "users_def": users_def, "permissions": permissions}

    def load_voyages_data(self):
        raw = load_json(self.data_dir / "voyages.json", [])
        voyages = []
        for v in raw:
            if isinstance(v, str):
                voyages.append({"code": v, "type": "LIVRAISON", "actif": True, "country": "Belgique", "duree": 60})
            elif isinstance(v, dict):
                voyages.append(
                    {
                        "code": v.get("code", ""),
                        "type": v.get("type", "LIVRAISON"),
                        "actif": v.get("actif", True),
                        "country": v.get("country", "Belgique"),
                        "duree": v.get("duree", 60),
                    }
                )
        return voyages

    def save_voyages_data(self):
        save_json(self.data_dir / "voyages.json", self.voyages)

    def build_gui(self):
        perms = self.rights["permissions"]

        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Quitter", command=self.root.destroy)
        menubar.add_cascade(label="Fichier", menu=file_menu)

        # Barre de statut en bas (session + dernière mise à jour)
        status_frame = ttk.Frame(self.root)
        status_frame.pack(side="bottom", fill="x")
        status_label = ttk.Label(status_frame, textvariable=self.status_var, anchor="e")
        status_label.pack(side="right", padx=5, pady=2)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        if perms["view_planning"]:
            self.build_planning_tab()
            self.build_suivi_missions_tab()  # Nouvel onglet Suivi missions
        if perms["view_drivers"]:
            self.build_chauffeurs_tab()
        if perms["manage_voyages"]:
            self.build_voyages_tab()
        if perms["view_finance"]:
            self.build_finance_tab()
        if perms["view_analyse"]:
            self.build_analyse_tab()
        if perms["generate_planning"]:
            self.build_admin_tab()
        if perms["manage_rights"]:
            self.build_rights_tab()
        if perms["view_sauron"]:
            self.build_sauron_tab()
        
        try:
            install_setup_menu(self.root, menubar, self.refresh_all)
        except Exception:
            pass

    def build_planning_tab(self):
        self.tab_planning = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_planning, text="Planning")

        top_frame = ttk.Frame(self.tab_planning)
        top_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(top_frame, text="Date :").pack(side="left")
        self.date_var = tk.StringVar(value=format_date_display(self.current_date))
        self.date_entry = ttk.Entry(top_frame, textvariable=self.date_var, width=12)
        self.date_entry.bind('<Return>', lambda e: self.on_load_date())
        self.date_entry.bind('<FocusOut>', lambda e: self.on_load_date())
        self.date_entry.pack(side="left", padx=(5, 15))

        ttk.Button(top_frame, text="◀◀ -2j", 
                  command=lambda: self.navigate_days(-2), width=8).pack(side="left", padx=2)
        ttk.Button(top_frame, text="◀ -1j", 
                  command=lambda: self.navigate_days(-1), width=8).pack(side="left", padx=2)
        ttk.Button(top_frame, text="📅 Aujourd'hui", 
                  command=self.set_today, width=12).pack(side="left", padx=5)
        ttk.Button(top_frame, text="+1j ▶", 
                  command=lambda: self.navigate_days(1), width=8).pack(side="left", padx=2)
        ttk.Button(top_frame, text="+2j ▶▶", 
                  command=lambda: self.navigate_days(2), width=8).pack(side="left", padx=(2, 15))
        
        ttk.Label(top_frame, text="Jours existants :").pack(side="left", padx=(0, 5))
        self.existing_dates_var = tk.StringVar()
        self.existing_dates_combo = ttk.Combobox(
            top_frame,
            textvariable=self.existing_dates_var,
            values=list_existing_dates(),
            width=12,
            state="readonly",
        )
        self.existing_dates_combo.pack(side="left")
        ttk.Button(top_frame, text="Ouvrir", command=self.on_open_existing_date).pack(side="left", padx=5)
        
        ttk.Separator(self.tab_planning, orient='horizontal').pack(fill='x', padx=5, pady=2)

        self.summary_frame = ttk.Frame(self.tab_planning, relief='solid', borderwidth=1)
        self.summary_frame.pack(fill='x', padx=5, pady=3)
        
        stats_container = ttk.Frame(self.summary_frame)
        stats_container.pack(fill='x', padx=5, pady=3)
        
        ttk.Label(stats_container, text="📊 Résumé:", font=('Arial', 9, 'bold')).pack(side='left', padx=(5, 10))
        
        ttk.Label(stats_container, text="📦 Liv:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_liv_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#2196F3')
        self.summary_liv_label.pack(side='left', padx=(0, 8))
        
        ttk.Label(stats_container, text="🚛 Ram:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_ram_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#4CAF50')
        self.summary_ram_label.pack(side='left', padx=(0, 8))
        
        ttk.Separator(stats_container, orient='vertical').pack(side='left', fill='y', padx=5)
        
        ttk.Label(stats_container, text="👥 Dispo:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_dispo_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#4CAF50')
        self.summary_dispo_label.pack(side='left', padx=(0, 8))
        
        ttk.Label(stats_container, text="🚗 Utilisés:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_used_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#FF9800')
        self.summary_used_label.pack(side='left', padx=(0, 8))
        
        ttk.Separator(stats_container, orient='vertical').pack(side='left', fill='y', padx=5)
        
        ttk.Label(stats_container, text="📦 Pal:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_pal_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#9C27B0')
        self.summary_pal_label.pack(side='left', padx=(0, 8))
        
        ttk.Label(stats_container, text="🌍 Pays:", font=('Arial', 8)).pack(side='left', padx=2)
        self.summary_pays_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#607D8B')
        self.summary_pays_label.pack(side='left')

        ttk.Separator(stats_container, orient='vertical').pack(side='left', fill='y', padx=5)

        # Label pour les voyages incomplets (rouge)
        self.summary_incomplete_label = ttk.Label(stats_container, text="", font=('Arial', 9, 'bold'), foreground='#D32F2F')
        self.summary_incomplete_label.pack(side='left', padx=5)

        controls_container = ttk.Frame(self.tab_planning)
        controls_container.pack(fill="x", padx=5, pady=2)
        
        sort_frame = ttk.LabelFrame(controls_container, text="🔽 Tri des plannings", padding=5)
        sort_frame.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        ttk.Label(sort_frame, text="Trier par :").pack(side="left", padx=(5, 2))
        
        style = ttk.Style()
        style.configure("Sort.TButton", padding=(10, 2))
        
        ttk.Button(sort_frame, text="⏰ Heure", command=lambda: self.sort_missions("heure"), 
                  style="Sort.TButton").pack(side="left", padx=2)
        ttk.Button(sort_frame, text="🚚 Voyage", command=lambda: self.sort_missions("voyage"),
                  style="Sort.TButton").pack(side="left", padx=2)
        ttk.Button(sort_frame, text="👤 Chauffeur", command=lambda: self.sort_missions("chauffeur"),
                  style="Sort.TButton").pack(side="left", padx=2)
        ttk.Button(sort_frame, text="#️⃣ N° tournée", command=lambda: self.sort_missions("numero"),
                  style="Sort.TButton").pack(side="left", padx=2)
        ttk.Button(sort_frame, text="🏢 SST", command=lambda: self.sort_missions("sst"),
                  style="Sort.TButton").pack(side="left", padx=2)
        ttk.Button(sort_frame, text="🌍 Pays", command=lambda: self.sort_missions("pays"),
                  style="Sort.TButton").pack(side="left", padx=2)
        
        self.sort_label = ttk.Label(sort_frame, text="(Tri actuel: Heure ↑)", foreground="blue")
        self.sort_label.pack(side="left", padx=10)

        view_frame = ttk.LabelFrame(controls_container, text="👁️ Vue planning", padding=5)
        view_frame.pack(side="left", fill="x", padx=(5, 0))
        
        ttk.Label(view_frame, text="Vue globale :").pack(side="left", padx=(5, 2))
        
        style.configure("View.TButton", padding=(10, 2))
        
        ttk.Button(view_frame, text="👥 Par Chauffeur", 
                  command=self.open_view_by_driver,
                  style="View.TButton").pack(side="left", padx=2)
        ttk.Button(view_frame, text="🕐 Par Heure", 
                  command=self.open_view_by_time,
                  style="View.TButton").pack(side="left", padx=2)
        ttk.Button(view_frame, text="📍 Par Voyage",
                  command=self.open_view_by_voyage,
                  style="View.TButton").pack(side="left", padx=2)

        ttk.Button(view_frame, text="📢 Annonces",
                  command=self.open_planning_announcements,
                  style="View.TButton").pack(side="left", padx=2)

        perms = self.rights["permissions"]
        btn_frame = ttk.Frame(self.tab_planning)
        btn_frame.pack(fill="x", padx=5, pady=5)
        if perms["edit_planning"]:
            ttk.Button(btn_frame, text="➕ Ajouter une ligne", command=self.on_add_mission).pack(side="left")
            ttk.Button(btn_frame, text="✏️ Modifier", command=self.on_edit_mission).pack(side="left", padx=5)
            ttk.Button(btn_frame, text="🗑️ Supprimer", command=self.on_delete_mission).pack(side="left", padx=5)
            
            ttk.Label(btn_frame, text="🔍 Rechercher :").pack(side="left", padx=(15, 5))
            self.search_var = tk.StringVar()
            self.search_var.trace('w', lambda *args: self.on_search())
            search_entry = ttk.Entry(btn_frame, textvariable=self.search_var, width=25)
            search_entry.pack(side="left")
            ttk.Button(btn_frame, text="❌", command=self.clear_search, width=3).pack(side="left", padx=2)

        ttk.Button(btn_frame, text="🔄 Rafraîchir", command=self.refresh_planning_view).pack(side="right")

        main_container = ttk.PanedWindow(self.tab_planning, orient=tk.HORIZONTAL)
        main_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        left_container = ttk.Frame(main_container)
        main_container.add(left_container, weight=3)
        
        self.left_container = left_container
        
        left_container.rowconfigure(0, weight=1)
        left_container.rowconfigure(1, weight=0)
        left_container.columnconfigure(0, weight=1)
        left_container.columnconfigure(1, weight=0)
        
        self.canvas = tk.Canvas(left_container, bg="white")
        self.scrollbar = ttk.Scrollbar(left_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")

        # Scroll avec la molette de souris
        def on_mousewheel_planning(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.canvas.bind("<MouseWheel>", on_mousewheel_planning)
        self.scrollable_frame.bind("<MouseWheel>", on_mousewheel_planning)

        def on_canvas_configure(event):
            try:
                canvas_width = event.width
                canvas_items = self.canvas.find_withtag("all")
                if canvas_items:
                    self.canvas.itemconfig(canvas_items[0], width=canvas_width)
            except (IndexError, tk.TclError):
                pass

        self.canvas.bind('<Configure>', on_canvas_configure)

        self.planning_container = ttk.Frame(self.scrollable_frame)
        self.planning_container.pack(fill="both", expand=True)

        # Mettre à jour le scroll region quand le sash est déplacé
        def on_sash_move(event=None):
            self.scrollable_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        self.planning_container.bind('<ButtonRelease-1>', on_sash_move)
        
        right_main_container = ttk.Frame(main_container)
        main_container.add(right_main_container, weight=1)
        
        right_main_container.rowconfigure(0, weight=1, minsize=150)
        right_main_container.rowconfigure(1, weight=1, minsize=150)
        right_main_container.columnconfigure(0, weight=1)
        
        self.available_container = ttk.LabelFrame(right_main_container, text="👥 Chauffeurs disponibles (Nb chauff dispo: 0)", padding=5)
        self.available_container.grid(row=0, column=0, sticky="nsew", pady=(0, 5))
        
        self.available_container.rowconfigure(0, weight=1)
        self.available_container.columnconfigure(0, weight=1)
        self.available_container.columnconfigure(1, weight=0)
        
        available_scrollbar = ttk.Scrollbar(self.available_container, orient="vertical")
        available_scrollbar.grid(row=0, column=1, sticky="ns")
        
        available_canvas = tk.Canvas(self.available_container, bg="white", 
                                   yscrollcommand=available_scrollbar.set)
        available_canvas.grid(row=0, column=0, sticky="nsew")
        available_scrollbar.config(command=available_canvas.yview)
        
        self.drivers_available_frame = ttk.Frame(available_canvas)
        available_canvas.create_window((0, 0), window=self.drivers_available_frame, anchor="nw")
        
        self.drivers_available_frame.bind(
            "<Configure>",
            lambda e: available_canvas.configure(scrollregion=available_canvas.bbox("all"))
        )
        
        self.used_container = ttk.LabelFrame(right_main_container, text="🚛 Chauffeurs utilisés (Nb chauff use: 0)", padding=5)
        self.used_container.grid(row=1, column=0, sticky="nsew")
        
        self.used_container.rowconfigure(0, weight=1)
        self.used_container.columnconfigure(0, weight=1)
        self.used_container.columnconfigure(1, weight=0)
        
        used_scrollbar = ttk.Scrollbar(self.used_container, orient="vertical")
        used_scrollbar.grid(row=0, column=1, sticky="ns")
        
        used_canvas = tk.Canvas(self.used_container, bg="#F5F5F5", 
                               yscrollcommand=used_scrollbar.set)
        used_canvas.grid(row=0, column=0, sticky="nsew")
        used_scrollbar.config(command=used_canvas.yview)
        
        self.drivers_used_frame = ttk.Frame(used_canvas)
        used_canvas.create_window((0, 0), window=self.drivers_used_frame, anchor="nw")
        
        self.drivers_used_frame.bind(
            "<Configure>",
            lambda e: used_canvas.configure(scrollregion=used_canvas.bbox("all"))
        )

        self.build_planning_form()
        self.hide_planning_form()

    # =============================================================================
    # ONGLET SUIVI MISSIONS - Suivi de réalisation des voyages
    # =============================================================================

    def build_suivi_missions_tab(self):
        """Construire l'onglet Suivi missions avec vue liste et Gantt"""
        self.tab_suivi = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_suivi, text="Suivi missions")

        # Variables pour le suivi
        self.suivi_current_date = date.today()
        self.suivi_missions = []
        self.suivi_missions_status = {}  # {mission_id: True/False} pour le statut effectué

        # Frame principal
        main_frame = ttk.Frame(self.tab_suivi)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Barre supérieure avec date et type de vue
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill="x", pady=5)

        ttk.Label(top_frame, text="Date :").pack(side="left")
        self.suivi_date_var = tk.StringVar(value=format_date_display(self.suivi_current_date))
        self.suivi_date_entry = ttk.Entry(top_frame, textvariable=self.suivi_date_var, width=12)
        self.suivi_date_entry.bind('<Return>', lambda e: self.suivi_on_load_date())
        self.suivi_date_entry.bind('<FocusOut>', lambda e: self.suivi_on_load_date())
        self.suivi_date_entry.pack(side="left", padx=(5, 15))

        ttk.Button(top_frame, text="◀ -1j",
                  command=lambda: self.suivi_navigate_days(-1), width=8).pack(side="left", padx=2)
        ttk.Button(top_frame, text="📅 Aujourd'hui",
                  command=self.suivi_set_today, width=12).pack(side="left", padx=5)
        ttk.Button(top_frame, text="+1j ▶",
                  command=lambda: self.suivi_navigate_days(1), width=8).pack(side="left", padx=(2, 15))

        ttk.Separator(top_frame, orient='vertical').pack(side="left", fill="y", padx=10)

        # Liste déroulante de type de vue
        ttk.Label(top_frame, text="Vue :").pack(side="left", padx=(0, 5))
        self.suivi_view_type = tk.StringVar(value="Liste")
        self.suivi_view_combo = ttk.Combobox(
            top_frame,
            textvariable=self.suivi_view_type,
            values=["Liste", "Gantt"],
            width=10,
            state="readonly"
        )
        self.suivi_view_combo.pack(side="left")
        self.suivi_view_combo.bind("<<ComboboxSelected>>", self.suivi_on_view_changed)

        ttk.Button(top_frame, text="🔄 Rafraîchir", command=self.suivi_refresh_view).pack(side="right")

        ttk.Separator(main_frame, orient='horizontal').pack(fill='x', pady=5)

        # Résumé des stats
        self.suivi_summary_frame = ttk.Frame(main_frame, relief='solid', borderwidth=1)
        self.suivi_summary_frame.pack(fill='x', pady=3)

        stats_container = ttk.Frame(self.suivi_summary_frame)
        stats_container.pack(fill='x', padx=5, pady=3)

        ttk.Label(stats_container, text="📊 Résumé:", font=('Arial', 9, 'bold')).pack(side='left', padx=(5, 10))

        ttk.Label(stats_container, text="📦 Total missions:", font=('Arial', 8)).pack(side='left', padx=2)
        self.suivi_total_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#2196F3')
        self.suivi_total_label.pack(side='left', padx=(0, 8))

        ttk.Label(stats_container, text="✅ Effectuées:", font=('Arial', 8)).pack(side='left', padx=2)
        self.suivi_done_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#4CAF50')
        self.suivi_done_label.pack(side='left', padx=(0, 8))

        ttk.Label(stats_container, text="⏳ En attente:", font=('Arial', 8)).pack(side='left', padx=2)
        self.suivi_pending_label = ttk.Label(stats_container, text="0", font=('Arial', 10, 'bold'), foreground='#FF9800')
        self.suivi_pending_label.pack(side='left', padx=(0, 8))

        # Container pour les vues (Liste et Gantt)
        self.suivi_view_container = ttk.Frame(main_frame)
        self.suivi_view_container.pack(fill="both", expand=True, pady=5)

        # Frame pour la vue Liste
        self.suivi_list_frame = ttk.Frame(self.suivi_view_container)

        # Frame pour la vue Gantt
        self.suivi_gantt_frame = ttk.Frame(self.suivi_view_container)

        # Construire la vue Liste
        self._build_suivi_list_view()

        # Construire la vue Gantt
        self._build_suivi_gantt_view()

        # Afficher la vue Liste par défaut
        self.suivi_list_frame.pack(fill="both", expand=True)

        # Formulaire d'édition (caché par défaut)
        self._build_suivi_edit_form()
        self.suivi_hide_form()

        # Message "Planning en construction" (caché par défaut)
        self.suivi_no_planning_frame = ttk.Frame(self.suivi_view_container)
        self.suivi_no_planning_label = ttk.Label(
            self.suivi_no_planning_frame,
            text="📋 Planning en construction",
            font=('Arial', 16, 'bold'),
            foreground='#666666'
        )
        self.suivi_no_planning_label.pack(expand=True, pady=50)

        # Charger les missions du jour
        self.suivi_load_missions()

    def _build_suivi_list_view(self):
        """Construire la vue liste du suivi missions"""
        # Canvas et scrollbar pour scroll vertical
        canvas = tk.Canvas(self.suivi_list_frame, bg="white")
        scrollbar = ttk.Scrollbar(self.suivi_list_frame, orient="vertical", command=canvas.yview)
        self.suivi_scrollable_frame = ttk.Frame(canvas)

        self.suivi_scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.suivi_scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Scroll avec molette
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<MouseWheel>", on_mousewheel)
        self.suivi_scrollable_frame.bind("<MouseWheel>", on_mousewheel)

        self.suivi_canvas = canvas

        # Container pour les sections par pays
        self.suivi_planning_container = ttk.Frame(self.suivi_scrollable_frame)
        self.suivi_planning_container.pack(fill="both", expand=True)

        # Dictionnaires pour stocker les frames et trees par pays
        self.suivi_country_frames = {}
        self.suivi_country_trees = {}

    def _build_suivi_gantt_view(self):
        """Construire la vue Gantt du suivi missions"""
        # Header avec légende et tri
        header_frame = ttk.Frame(self.suivi_gantt_frame)
        header_frame.pack(fill="x", pady=5)

        ttk.Label(header_frame, text="📊 Vue Gantt - Planning par chauffeur",
                 font=('Arial', 11, 'bold')).pack(side="left", padx=5)

        # Option de tri des chauffeurs
        sort_frame = ttk.Frame(header_frame)
        sort_frame.pack(side="left", padx=20)

        ttk.Label(sort_frame, text="Trier par:", font=('Arial', 8)).pack(side="left", padx=2)
        self.gantt_sort_var = tk.StringVar(value="Alphabétique")
        self.gantt_sort_combo = ttk.Combobox(
            sort_frame,
            textvariable=self.gantt_sort_var,
            values=["Alphabétique", "SST"],
            width=12,
            state="readonly"
        )
        self.gantt_sort_combo.pack(side="left", padx=2)
        self.gantt_sort_combo.bind("<<ComboboxSelected>>", lambda e: self.suivi_draw_gantt())

        # Légende
        legend_frame = ttk.Frame(header_frame)
        legend_frame.pack(side="right", padx=10)

        ttk.Label(legend_frame, text="Légende:", font=('Arial', 8)).pack(side="left", padx=5)

        # Créer une petite barre de couleur pour la légende
        liv_canvas = tk.Canvas(legend_frame, width=20, height=12, bg="#2196F3", highlightthickness=1)
        liv_canvas.pack(side="left", padx=2)
        ttk.Label(legend_frame, text="Livraison", font=('Arial', 8)).pack(side="left", padx=(0, 10))

        ram_canvas = tk.Canvas(legend_frame, width=20, height=12, bg="#4CAF50", highlightthickness=1)
        ram_canvas.pack(side="left", padx=2)
        ttk.Label(legend_frame, text="Ramasse", font=('Arial', 8)).pack(side="left", padx=(0, 10))

        done_canvas = tk.Canvas(legend_frame, width=20, height=12, bg="#9E9E9E", highlightthickness=1)
        done_canvas.pack(side="left", padx=2)
        ttk.Label(legend_frame, text="Effectué", font=('Arial', 8)).pack(side="left")

        # Container principal du Gantt avec deux parties: noms fixes + timeline scrollable
        self.suivi_gantt_container = ttk.Frame(self.suivi_gantt_frame)
        self.suivi_gantt_container.pack(fill="both", expand=True)

        # Frame pour la partie gauche (noms des chauffeurs - fixe)
        self.gantt_names_frame = ttk.Frame(self.suivi_gantt_container, width=200)
        self.gantt_names_frame.pack(side="left", fill="y")
        self.gantt_names_frame.pack_propagate(False)

        # Canvas pour les noms des chauffeurs (scroll vertical seulement)
        self.gantt_names_canvas = tk.Canvas(self.gantt_names_frame, bg="white",
                                            highlightthickness=0, width=200)
        self.gantt_names_canvas.pack(fill="both", expand=True)

        # Frame pour la partie droite (timeline - scrollable)
        self.gantt_timeline_frame = ttk.Frame(self.suivi_gantt_container)
        self.gantt_timeline_frame.pack(side="left", fill="both", expand=True)

        # Canvas pour la timeline (scroll horizontal et vertical)
        self.gantt_canvas = tk.Canvas(self.gantt_timeline_frame, bg="white", highlightthickness=0)
        self.gantt_h_scrollbar = ttk.Scrollbar(self.gantt_timeline_frame, orient="horizontal",
                                                command=self.gantt_canvas.xview)
        self.gantt_v_scrollbar = ttk.Scrollbar(self.gantt_timeline_frame, orient="vertical",
                                                command=self._gantt_yview_both)

        self.gantt_canvas.configure(xscrollcommand=self.gantt_h_scrollbar.set,
                                    yscrollcommand=self.gantt_v_scrollbar.set)

        self.gantt_v_scrollbar.pack(side="right", fill="y")
        self.gantt_h_scrollbar.pack(side="bottom", fill="x")
        self.gantt_canvas.pack(side="left", fill="both", expand=True)

        # Bind molette pour scroll
        def on_gantt_mousewheel(event):
            # Molette normale = scroll horizontal dans le Gantt
            self.gantt_canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

        def on_gantt_mousewheel_vertical(event):
            # Scroll vertical synchronisé
            self._gantt_yview_both("scroll", int(-1 * (event.delta / 120)), "units")

        self.gantt_canvas.bind("<MouseWheel>", on_gantt_mousewheel)
        self.gantt_names_canvas.bind("<MouseWheel>", on_gantt_mousewheel_vertical)
        # Shift + molette = scroll vertical
        self.gantt_canvas.bind("<Shift-MouseWheel>", on_gantt_mousewheel_vertical)

    def _gantt_yview_both(self, *args):
        """Synchroniser le scroll vertical entre les deux canvas du Gantt"""
        self.gantt_canvas.yview(*args)
        self.gantt_names_canvas.yview(*args)

    def _build_suivi_edit_form(self):
        """Construire le formulaire d'édition pour le suivi missions"""
        self.suivi_form_frame = ttk.LabelFrame(self.tab_suivi, text="Modifier le voyage")

        self.suivi_form_mission = None

        # Variables du formulaire
        self.suivi_form_type = tk.StringVar()
        self.suivi_form_heure = tk.StringVar()
        self.suivi_form_voyage = tk.StringVar()
        self.suivi_form_nb_pal = tk.StringVar()
        self.suivi_form_sst = tk.StringVar()
        self.suivi_form_chauffeur = tk.StringVar()
        self.suivi_form_ramasse = tk.StringVar()
        self.suivi_form_infos = tk.StringVar()
        self.suivi_form_numero = tk.StringVar()

        row = 0
        ttk.Label(self.suivi_form_frame, text="Type :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        rb_liv = ttk.Radiobutton(self.suivi_form_frame, text="Livraison",
                                 variable=self.suivi_form_type, value="LIVRAISON",
                                 command=self.suivi_on_form_type_changed)
        rb_ram = ttk.Radiobutton(self.suivi_form_frame, text="Ramasse",
                                 variable=self.suivi_form_type, value="RAMASSE",
                                 command=self.suivi_on_form_type_changed)
        rb_liv.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        rb_ram.grid(row=row, column=2, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Heure :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.suivi_form_heure_cb = ttk.Combobox(self.suivi_form_frame, textvariable=self.suivi_form_heure,
                                                values=TIME_CHOICES, width=8)
        self.suivi_form_heure_cb.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Voyage :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.suivi_form_voy_cb = ttk.Combobox(self.suivi_form_frame, textvariable=self.suivi_form_voyage,
                                              values=[v["code"] for v in self.voyages if v.get("actif", True)],
                                              width=15)
        self.suivi_form_voy_cb.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Ramasse :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.suivi_form_ram_entry = ttk.Entry(self.suivi_form_frame, textvariable=self.suivi_form_ramasse, width=25)
        self.suivi_form_ram_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Nb palettes :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(self.suivi_form_frame, textvariable=self.suivi_form_nb_pal, width=5).grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="SST :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.suivi_form_sst_cb = ttk.Combobox(self.suivi_form_frame, textvariable=self.suivi_form_sst,
                                              values=self.sst_list, width=15)
        self.suivi_form_sst_cb.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Chauffeur :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.suivi_form_ch_cb = ttk.Combobox(self.suivi_form_frame, textvariable=self.suivi_form_chauffeur,
                                             values=[c.get("nom_affichage", c.get("nom", "")) for c in self.chauffeurs],
                                             width=20)
        self.suivi_form_ch_cb.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Numéro tournée :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        ttk.Label(self.suivi_form_frame, textvariable=self.suivi_form_numero).grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.suivi_form_frame, text="Informations :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        ttk.Entry(self.suivi_form_frame, textvariable=self.suivi_form_infos, width=30).grid(
            row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2
        )
        row += 1

        btn_frame = ttk.Frame(self.suivi_form_frame)
        btn_frame.grid(row=row, column=0, columnspan=4, pady=5, sticky="e")
        ttk.Button(btn_frame, text="💾 Enregistrer", command=self.suivi_on_form_save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="❌ Annuler", command=self.suivi_hide_form).pack(side="left")

    def suivi_on_form_type_changed(self):
        """Mise à jour du formulaire lors du changement de type"""
        if self.suivi_form_type.get() == "LIVRAISON":
            self.suivi_form_ram_entry.configure(state="disabled")
            self.suivi_form_ramasse.set("")
        else:
            self.suivi_form_ram_entry.configure(state="normal")

    def suivi_show_form(self, mission):
        """Afficher le formulaire d'édition avec les données de la mission"""
        self.suivi_form_mission = mission

        self.suivi_form_type.set(mission.get("type", "LIVRAISON"))
        self.suivi_form_heure.set(mission.get("heure", TIME_CHOICES[0]))
        self.suivi_form_voyage.set(mission.get("voyage", ""))
        self.suivi_form_ramasse.set(mission.get("ramasse", ""))
        self.suivi_form_nb_pal.set(str(mission.get("nb_pal", 0)))
        self.suivi_form_sst.set(mission.get("sst", ""))
        self.suivi_form_chauffeur.set(mission.get("chauffeur_nom", ""))
        self.suivi_form_infos.set(mission.get("infos", ""))
        self.suivi_form_numero.set(str(mission.get("numero", "")))

        self.suivi_on_form_type_changed()
        self.suivi_form_frame.pack(fill="x", padx=5, pady=5, before=self.suivi_view_container)

    def suivi_hide_form(self):
        """Masquer le formulaire d'édition"""
        self.suivi_form_frame.pack_forget()
        self.suivi_form_mission = None

    def suivi_on_form_save(self):
        """Sauvegarder les modifications du formulaire"""
        if not self.suivi_form_mission:
            return

        mission = self.suivi_form_mission
        mid = mission["id"]

        # Récupérer les nouvelles valeurs
        type_ = self.suivi_form_type.get()
        heure = self.suivi_form_heure.get()
        voyage = self.suivi_form_voyage.get()
        ramasse = self.suivi_form_ramasse.get() if type_ == "RAMASSE" else ""

        try:
            nb_pal = int(self.suivi_form_nb_pal.get())
        except ValueError:
            nb_pal = 0

        sst = self.suivi_form_sst.get()
        chauffeur_nom = self.suivi_form_chauffeur.get()
        infos = self.suivi_form_infos.get()

        # Trouver l'ID du chauffeur
        chauffeur_id = mission.get("chauffeur_id", "")
        for ch in self.chauffeurs:
            if ch.get("nom_affichage") == chauffeur_nom or ch.get("nom") == chauffeur_nom:
                chauffeur_id = ch["id"]
                break

        # Mettre à jour la mission
        before_state = {k: v for k, v in mission.items() if k != "_path"}

        mission.update({
            "type": type_,
            "heure": heure,
            "voyage": voyage,
            "ramasse": ramasse,
            "nb_pal": nb_pal,
            "sst": sst,
            "chauffeur_nom": chauffeur_nom,
            "chauffeur_id": chauffeur_id,
            "infos": infos,
        })

        # Sauvegarder dans le fichier
        path = mission.get("_path")
        if path:
            save_json(path, {k: v for k, v in mission.items() if k != "_path"})
            planning_cache.force_refresh(self.suivi_current_date)

        # Logger l'action
        activity_logger.log_action("MISSION_EDIT", {
            "mission_id": mid,
            "voyage": voyage,
            "type": type_,
            "date": self.suivi_current_date.strftime("%Y-%m-%d"),
            "source": "suivi_missions",
        }, before_state=before_state, after_state={k: v for k, v in mission.items() if k != "_path"})

        self.suivi_hide_form()
        self.suivi_refresh_view()

        # Aussi rafraîchir la vue Planning si on est sur la même date
        if self.suivi_current_date == self.current_date:
            self.load_planning_for_date(self.current_date)

    def suivi_set_today(self):
        """Réinitialiser à la date d'aujourd'hui"""
        self.suivi_current_date = date.today()
        self.suivi_date_var.set(format_date_display(self.suivi_current_date))
        self.suivi_load_missions()

    def suivi_navigate_days(self, days):
        """Naviguer dans les dates"""
        try:
            current = datetime.strptime(self.suivi_date_var.get(), "%d/%m/%Y").date()
            new_date = current + timedelta(days=days)
            self.suivi_date_var.set(format_date_display(new_date))
            self.suivi_current_date = new_date
            self.suivi_load_missions()
        except ValueError:
            messagebox.showerror("Erreur", "Format de date invalide.")

    def suivi_on_load_date(self):
        """Charger le planning pour une date spécifique"""
        try:
            d = parse_date_input(self.suivi_date_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide.\nFormat attendu: JJ/MM/AAAA")
            return
        self.suivi_current_date = d
        self.suivi_load_missions()

    def suivi_on_view_changed(self, event=None):
        """Changer le type de vue (Liste/Gantt)"""
        view_type = self.suivi_view_type.get()

        # Masquer toutes les vues
        self.suivi_list_frame.pack_forget()
        self.suivi_gantt_frame.pack_forget()
        self.suivi_no_planning_frame.pack_forget()

        if not self.suivi_missions:
            self.suivi_no_planning_frame.pack(fill="both", expand=True)
            return

        if view_type == "Liste":
            self.suivi_list_frame.pack(fill="both", expand=True)
        else:
            self.suivi_gantt_frame.pack(fill="both", expand=True)
            self.suivi_draw_gantt()

    def suivi_load_missions(self):
        """Charger les missions pour la date courante du suivi"""
        d = self.suivi_current_date

        # Essayer le cache d'abord
        cached = planning_cache.get_cached_planning(d)
        if cached is not None:
            # Filtrer uniquement les missions valides (avec un id)
            self.suivi_missions = [m.copy() for m in cached if m and isinstance(m, dict) and "id" in m]
        else:
            # Charger depuis les fichiers
            day_dir = get_planning_day_dir(d)
            self.suivi_missions = []
            if day_dir and day_dir.exists():
                for file in day_dir.glob("*.json"):
                    # Ignorer les fichiers de métadonnées (commençant par _)
                    if file.name.startswith("_"):
                        continue
                    data = load_json(file, None)
                    if data and isinstance(data, dict) and "id" in data:  # Vérifier que c'est bien une mission
                        data["_path"] = file.as_posix()
                        self.suivi_missions.append(data)

        # Charger les statuts de suivi depuis le fichier de statut
        self.suivi_load_status()

        self.suivi_refresh_view()

    def suivi_load_status(self):
        """Charger les statuts de réalisation des missions"""
        d = self.suivi_current_date
        status_file = get_planning_day_dir(d)
        if status_file and status_file.exists():
            status_path = status_file / "_suivi_status.json"
            if status_path.exists():
                self.suivi_missions_status = load_json(status_path, {})
            else:
                self.suivi_missions_status = {}
        else:
            self.suivi_missions_status = {}

    def suivi_save_status(self):
        """Sauvegarder les statuts de réalisation des missions"""
        d = self.suivi_current_date
        day_dir = get_planning_day_dir(d)
        if day_dir and day_dir.exists():
            status_path = day_dir / "_suivi_status.json"
            save_json(status_path, self.suivi_missions_status)

    def suivi_refresh_view(self):
        """Rafraîchir la vue actuelle du suivi"""
        # Masquer le formulaire si visible
        self.suivi_hide_form()

        # Mettre à jour les stats
        total = len(self.suivi_missions)
        done = sum(1 for m in self.suivi_missions if self.suivi_missions_status.get(m["id"], False))
        pending = total - done

        self.suivi_total_label.config(text=str(total))
        self.suivi_done_label.config(text=str(done))
        self.suivi_pending_label.config(text=str(pending))

        # Afficher la vue appropriée
        self.suivi_on_view_changed()

        if self.suivi_missions:
            if self.suivi_view_type.get() == "Liste":
                self.suivi_refresh_list_view()
            else:
                self.suivi_draw_gantt()

    def suivi_refresh_list_view(self):
        """Rafraîchir la vue liste du suivi missions"""
        # Supprimer les anciennes sections
        for country, frame in self.suivi_country_frames.items():
            frame.destroy()
        self.suivi_country_frames.clear()
        self.suivi_country_trees.clear()

        # Organiser les missions par pays
        missions_by_country = {}
        v_by_code = {v.get("code"): v for v in self.voyages}

        for m in self.suivi_missions:
            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")

            if country not in missions_by_country:
                missions_by_country[country] = []
            missions_by_country[country].append(m)

        # Trier par heure
        for country in missions_by_country:
            missions_by_country[country] = sorted(
                missions_by_country[country],
                key=lambda m: self._time_key(m)
            )

        sorted_countries = sorted(missions_by_country.keys(), key=lambda x: (x != "Belgique", x))

        for country in sorted_countries:
            self.suivi_create_country_section(country, missions_by_country[country])

    def suivi_create_country_section(self, country, missions):
        """Créer une section par pays pour le suivi missions - vue adaptative"""
        bg_color = COUNTRY_COLORS.get(country, "#F5F5F5")

        flag_emoji = self.get_country_flag(country)
        country_frame = ttk.LabelFrame(
            self.suivi_planning_container,
            text=f"  {flag_emoji}  {country.upper()} ({len(missions)} missions)  ",
            padding=5
        )
        country_frame.pack(fill="both", expand=True, pady=5, padx=5)

        # Container pour le treeview avec scrollbars
        tree_container = ttk.Frame(country_frame)
        tree_container.pack(fill="both", expand=True)

        # Treeview avec colonnes incluant case à cocher
        columns = ("effectue", "heure", "type", "voyage", "nb_pal", "numero", "sst", "chauffeur", "ramasse", "infos")

        # Hauteur adaptative: min 5, max 15 lignes selon le contenu
        tree_height = max(5, min(15, len(missions)))
        tree = ttk.Treeview(tree_container, columns=columns, show="headings", height=tree_height)

        # Configuration des colonnes avec largeurs proportionnelles
        col_config = [
            ("effectue", "✅", 35, False),    # Fixe
            ("heure", "Heure", 55, False),    # Fixe
            ("type", "Type", 75, True),       # Extensible
            ("voyage", "Voyage", 85, True),   # Extensible
            ("nb_pal", "Pal", 40, False),     # Fixe
            ("numero", "N°", 35, False),      # Fixe
            ("sst", "SST", 60, True),         # Extensible
            ("chauffeur", "Chauffeur", 110, True),  # Extensible
            ("ramasse", "Ramasse", 100, True),     # Extensible
            ("infos", "Infos", 150, True)     # Extensible
        ]

        for col, txt, width, stretch in col_config:
            tree.heading(col, text=txt)
            tree.column(col, width=width, minwidth=width-10, stretch=stretch)

        # Scrollbars vertical et horizontal
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Layout avec grid pour un meilleur redimensionnement
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        # Tags pour le style
        tree.tag_configure('done', background='#E8F5E9', foreground='#2E7D32')
        tree.tag_configure('pending', background='white')
        tree.tag_configure('oddrow', background='#F5F5F5')
        tree.tag_configure('evenrow', background='white')

        # Remplir le treeview
        row_num = 0
        for m in missions:
            mid = m["id"]
            is_done = self.suivi_missions_status.get(mid, False)

            sst_display = "N/A" if m.get("sans_sst", False) else m.get("sst", "")
            chauffeur_display = "N/A" if m.get("sans_chauffeur", False) else m.get("chauffeur_nom", "")

            values = (
                "✔" if is_done else "☐",
                m.get("heure", ""),
                m.get("type", ""),
                m.get("voyage", ""),
                m.get("nb_pal", ""),
                m.get("numero", ""),
                sst_display,
                chauffeur_display,
                m.get("ramasse", ""),
                m.get("infos", ""),
            )

            tag = 'done' if is_done else ('oddrow' if row_num % 2 else 'evenrow')
            tree.insert("", "end", iid=mid, values=values, tags=(tag,))
            row_num += 1

        # Bindings
        tree.bind("<Double-1>", lambda e, t=tree: self.suivi_on_double_click(e, t))
        tree.bind("<Button-1>", lambda e, t=tree: self.suivi_on_click(e, t))

        self.suivi_country_frames[country] = country_frame
        self.suivi_country_trees[country] = tree

    def suivi_on_click(self, event, tree):
        """Gérer le clic simple pour basculer le statut effectué"""
        region = tree.identify_region(event.x, event.y)
        if region == "cell":
            column = tree.identify_column(event.x)
            if column == "#1":  # Colonne "effectue"
                item = tree.identify_row(event.y)
                if item:
                    self.suivi_toggle_status(item)

    def suivi_on_double_click(self, event, tree):
        """Gérer le double-clic pour éditer une mission"""
        region = tree.identify_region(event.x, event.y)
        if region == "cell":
            column = tree.identify_column(event.x)
            if column != "#1":  # Pas sur la colonne effectue
                item = tree.identify_row(event.y)
                if item:
                    mission = next((m for m in self.suivi_missions if m["id"] == item), None)
                    if mission:
                        self.suivi_show_form(mission)

    def suivi_toggle_status(self, mission_id):
        """Basculer le statut effectué d'une mission"""
        current = self.suivi_missions_status.get(mission_id, False)
        self.suivi_missions_status[mission_id] = not current
        self.suivi_save_status()

        # Logger l'action
        activity_logger.log_action("MISSION_STATUS_CHANGE", {
            "mission_id": mission_id,
            "date": self.suivi_current_date.strftime("%Y-%m-%d"),
            "new_status": "effectue" if not current else "non_effectue",
        })

        self.suivi_refresh_view()

    def suivi_draw_gantt(self):
        """Dessiner le diagramme de Gantt avec noms fixes à gauche"""
        canvas = self.gantt_canvas
        names_canvas = self.gantt_names_canvas
        canvas.delete("all")
        names_canvas.delete("all")

        if not self.suivi_missions:
            canvas.create_text(400, 200, text="Aucune mission pour cette date",
                             font=('Arial', 14), fill='gray')
            return

        # Configuration du Gantt
        DRIVER_HEIGHT = 40
        HOUR_WIDTH = 100
        TOP_MARGIN = 40    # Pour les heures
        NAMES_WIDTH = 200  # Largeur de la colonne des noms

        # Collecter les chauffeurs uniques avec des missions ce jour
        # Format: {chauffeur_nom: {"missions": [...], "sst": "..."}}
        drivers_with_missions = {}
        v_by_code = {v.get("code"): v for v in self.voyages}

        for m in self.suivi_missions:
            chauffeur = m.get("chauffeur_nom", "Sans chauffeur")
            if m.get("sans_chauffeur", False):
                chauffeur = "Sans chauffeur"
            if chauffeur not in drivers_with_missions:
                # Récupérer le SST du chauffeur (de sa première mission)
                sst = m.get("sst", "") if not m.get("sans_sst", False) else ""
                drivers_with_missions[chauffeur] = {"missions": [], "sst": sst}
            drivers_with_missions[chauffeur]["missions"].append(m)

        # Trier les chauffeurs selon l'option sélectionnée
        sort_mode = getattr(self, 'gantt_sort_var', None)
        sort_mode = sort_mode.get() if sort_mode else "Alphabétique"

        if sort_mode == "SST":
            # Trier par SST puis par nom
            drivers = sorted(drivers_with_missions.keys(),
                           key=lambda d: (drivers_with_missions[d]["sst"] or "ZZZ", d))
        else:
            # Tri alphabétique
            drivers = sorted(drivers_with_missions.keys())

        total_height = TOP_MARGIN + len(drivers) * DRIVER_HEIGHT + 20

        # === CANVAS DES NOMS (gauche, fixe) ===
        # En-tête "Chauffeur"
        names_canvas.create_text(NAMES_WIDTH/2, 20, text="Chauffeur",
                                font=('Arial', 9, 'bold'))
        names_canvas.create_line(0, TOP_MARGIN, NAMES_WIDTH, TOP_MARGIN, fill='#CCCCCC')

        # Dessiner les noms des chauffeurs
        for i, driver in enumerate(drivers):
            y = TOP_MARGIN + i * DRIVER_HEIGHT
            driver_info = drivers_with_missions[driver]
            sst = driver_info["sst"]

            # Fond alterné
            if i % 2 == 0:
                names_canvas.create_rectangle(0, y, NAMES_WIDTH, y + DRIVER_HEIGHT,
                                             fill='#FAFAFA', outline='')

            # Nom du chauffeur avec SST
            display_text = driver
            if sst:
                display_text = f"{driver} ({sst})"
            names_canvas.create_text(10, y + DRIVER_HEIGHT/2, text=display_text,
                                    anchor='w', font=('Arial', 9, 'bold'))

            # Ligne horizontale
            names_canvas.create_line(0, y + DRIVER_HEIGHT, NAMES_WIDTH, y + DRIVER_HEIGHT,
                                    fill='#E0E0E0')

        # === CANVAS TIMELINE (droite, scrollable) ===
        # Dessiner les en-têtes d'heures (0h à 28h pour couvrir les heures tardives)
        for hour in range(0, 28):
            x = hour * HOUR_WIDTH
            canvas.create_text(x + HOUR_WIDTH/2, 20, text=f"{hour:02d}:00",
                             font=('Arial', 9))
            canvas.create_line(x, TOP_MARGIN, x, TOP_MARGIN + len(drivers) * DRIVER_HEIGHT,
                             fill='#E0E0E0', dash=(2, 2))

        # Dessiner les lignes des chauffeurs sur la timeline
        for i, driver in enumerate(drivers):
            y = TOP_MARGIN + i * DRIVER_HEIGHT
            driver_info = drivers_with_missions[driver]

            # Fond alterné
            if i % 2 == 0:
                canvas.create_rectangle(0, y, 28 * HOUR_WIDTH, y + DRIVER_HEIGHT,
                                       fill='#FAFAFA', outline='')

            # Ligne horizontale
            canvas.create_line(0, y + DRIVER_HEIGHT, 28 * HOUR_WIDTH, y + DRIVER_HEIGHT,
                             fill='#E0E0E0')

            # Dessiner les barres des missions (sans LEFT_MARGIN maintenant)
            for m in driver_info["missions"]:
                self.suivi_draw_gantt_bar(canvas, m, i, v_by_code,
                                          HOUR_WIDTH, DRIVER_HEIGHT, 0, TOP_MARGIN)

        # Configurer les zones de scroll
        canvas.configure(scrollregion=(0, 0, 28 * HOUR_WIDTH, total_height))
        names_canvas.configure(scrollregion=(0, 0, NAMES_WIDTH, total_height))

    def suivi_draw_gantt_bar(self, canvas, mission, driver_idx, v_by_code,
                             hour_width, driver_height, left_margin, top_margin):
        """Dessiner une barre de mission dans le Gantt avec case à cocher"""
        try:
            heure_str = mission.get("heure", "08:00")
            h, m = map(int, heure_str.split(":"))
            start_hour = h + m / 60
        except:
            start_hour = 8

        # Récupérer la durée du voyage
        voyage_code = mission.get("voyage", "")
        voyage = v_by_code.get(voyage_code, {})
        duree_minutes = voyage.get("duree", 60)  # Par défaut 60 minutes
        duree_hours = duree_minutes / 60

        # Calculer les coordonnées
        x1 = left_margin + start_hour * hour_width
        x2 = x1 + duree_hours * hour_width
        y1 = top_margin + driver_idx * driver_height + 5
        y2 = y1 + driver_height - 10

        # Couleur selon le type et le statut
        is_done = self.suivi_missions_status.get(mission["id"], False)
        if is_done:
            fill_color = "#9E9E9E"  # Gris pour effectué
        elif mission.get("type") == "LIVRAISON":
            fill_color = "#2196F3"  # Bleu pour livraison
        else:
            fill_color = "#4CAF50"  # Vert pour ramasse

        # Dessiner la barre
        bar = canvas.create_rectangle(x1, y1, x2, y2, fill=fill_color, outline='#333333')

        # Case à cocher à gauche de la barre
        checkbox_size = 14
        cb_x1 = x1 + 3
        cb_y1 = (y1 + y2) / 2 - checkbox_size / 2
        cb_x2 = cb_x1 + checkbox_size
        cb_y2 = cb_y1 + checkbox_size

        # Dessiner la case à cocher
        checkbox = canvas.create_rectangle(cb_x1, cb_y1, cb_x2, cb_y2,
                                          fill='white', outline='#333333', width=1)

        # Si effectué, dessiner le coche
        if is_done:
            # Dessiner un coche vert
            canvas.create_line(cb_x1 + 2, cb_y1 + checkbox_size/2,
                             cb_x1 + checkbox_size/3, cb_y2 - 3,
                             fill='#2E7D32', width=2)
            canvas.create_line(cb_x1 + checkbox_size/3, cb_y2 - 3,
                             cb_x2 - 2, cb_y1 + 3,
                             fill='#2E7D32', width=2)

        # Texte dans la barre (décalé à droite de la checkbox)
        text = f"{voyage_code}"
        text_x = cb_x2 + 5
        bar_width = x2 - text_x
        if bar_width > 30:
            canvas.create_text(text_x + bar_width/2, (y1 + y2) / 2, text=text,
                             fill='white', font=('Arial', 8, 'bold'))

        # Tooltip au survol
        def show_tooltip(event):
            info = f"{mission.get('type')}: {voyage_code}\n"
            info += f"Heure: {heure_str}\n"
            info += f"Durée: {duree_minutes} min\n"
            info += f"SST: {mission.get('sst', 'N/A')}\n"
            info += f"Statut: {'✔ Effectué' if is_done else '☐ En attente'}"
            # Créer un tooltip simple
            tooltip = canvas.create_text(event.x + 10, event.y - 10,
                                        text=info, anchor='nw',
                                        fill='black', font=('Arial', 8),
                                        tags='tooltip')
            bg = canvas.create_rectangle(canvas.bbox(tooltip), fill='#FFFFCC', outline='black', tags='tooltip')
            canvas.tag_lower(bg, tooltip)

        def hide_tooltip(event):
            canvas.delete('tooltip')

        # Bindings pour la barre et la checkbox
        for item in [bar, checkbox]:
            canvas.tag_bind(item, '<Enter>', show_tooltip)
            canvas.tag_bind(item, '<Leave>', hide_tooltip)
            canvas.tag_bind(item, '<Button-1>', lambda e, mid=mission["id"]: self.suivi_toggle_status(mid))

    def open_view_by_driver(self):
        win = tk.Toplevel(self.root)
        win.title(f"Vue par Chauffeur - {format_date_display(self.current_date)}")
        win.geometry("1200x700")
        
        auto_refresh_active = {'active': True}
        
        header = ttk.Frame(win, padding=10)
        header.pack(fill="x")
        ttk.Label(header, text=f"👥 Planning Global par Chauffeur", 
                 font=('Arial', 14, 'bold')).pack(side="left")
        ttk.Label(header, text=f"Date: {self.current_date.strftime('%d/%m/%Y')}", 
                 font=('Arial', 11)).pack(side="right")
        
        toolbar = ttk.Frame(win, padding=5)
        toolbar.pack(fill="x", padx=10)
        
        search_var = tk.StringVar()
        
        ttk.Label(toolbar, text="🔍 Rechercher :").pack(side="left", padx=(0, 5))
        search_entry = ttk.Entry(toolbar, textvariable=search_var, width=30)
        search_entry.pack(side="left")
        
        clear_btn_frame = ttk.Frame(toolbar)
        clear_btn_frame.pack(side="left", padx=2)
        
        refresh_btn_frame = ttk.Frame(toolbar)
        refresh_btn_frame.pack(side="right", padx=5)
        
        tree_frame = ttk.Frame(win, padding=10)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("chauffeur", "numero", "heure", "type", "voyage", "nb_pal", 
                  "sst", "pays", "ramasse", "infos")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=25, selectmode="extended")
        
        col_config = [
            ("chauffeur", "Chauffeur", 120),
            ("numero", "N°", 40),
            ("heure", "Heure", 60),
            ("type", "Type", 80),
            ("voyage", "Voyage", 100),
            ("nb_pal", "Nb Pal", 60),
            ("sst", "SST", 70),
            ("pays", "Pays", 100),
            ("ramasse", "Ramasse", 100),
            ("infos", "Infos", 150)
        ]
        
        for col, header_text, width in col_config:
            tree.heading(col, text=header_text)
            tree.column(col, width=width)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        all_data = {'items': []}
        
        def do_search(*args):
            search_text = search_var.get().strip().upper()
            
            for item in tree.get_children():
                tree.delete(item)
            
            if not search_text:
                for item_id, values, tags in all_data['items']:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
                return
            
            for item_id, values, tags in all_data['items']:
                match = False
                for value in values:
                    if search_text in str(value).upper():
                        match = True
                        break
                if match:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
        
        search_var.trace('w', do_search)
        
        def refresh_tree():
            if auto_refresh_active['active']:
                try:
                    self.load_planning_for_date(self.current_date, preserve_ui=True)
                    for item in tree.get_children():
                        tree.delete(item)
                    self._fill_consolidated_view(tree, sort_by="driver")
                    all_data['items'] = []
                    for item_id in tree.get_children():
                        values = tree.item(item_id)["values"]
                        tags = tree.item(item_id)["tags"]
                        all_data['items'].append((item_id, values, tags))
                    do_search()
                    # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
                except Exception:
                    pass
        
        self._fill_consolidated_view(tree, sort_by="driver")
        
        all_data['items'] = []
        for item_id in tree.get_children():
            values = tree.item(item_id)["values"]
            tags = tree.item(item_id)["tags"]
            all_data['items'].append((item_id, values, tags))
        
        def clear_search():
            search_var.set("")
        
        ttk.Button(clear_btn_frame, text="❌", command=clear_search, width=3).pack()
        
        def manual_refresh():
            self.load_planning_for_date(self.current_date, preserve_ui=True)
            for item in tree.get_children():
                tree.delete(item)
            self._fill_consolidated_view(tree, sort_by="driver")
            all_data['items'] = []
            for item_id in tree.get_children():
                values = tree.item(item_id)["values"]
                tags = tree.item(item_id)["tags"]
                all_data['items'].append((item_id, values, tags))
            do_search()
        
        ttk.Button(refresh_btn_frame, text="🔄 Rafraîchir", command=manual_refresh).pack()
        
        # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
        
        def on_close():
            auto_refresh_active['active'] = False
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)
        
        export_frame = ttk.Frame(win)
        export_frame.pack(pady=5)
        
        def export_excel_par_chauffeur():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_chauffeur', extension='xlsx')
            success, message = export_planning_excel_par_chauffeur(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)

        def export_pdf_par_chauffeur():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_chauffeur', extension='pdf')
            success, message = export_planning_pdf_par_chauffeur(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)
        
        if EXCEL_AVAILABLE:
            ttk.Button(export_frame, text='📊 Exporter Excel', command=export_excel_par_chauffeur).pack(side='left', padx=5)
        if PDF_AVAILABLE:
            ttk.Button(export_frame, text='📄 Exporter PDF', command=export_pdf_par_chauffeur).pack(side='left', padx=5)
        
        ttk.Button(export_frame, text="Fermer", command=on_close).pack(side="left", padx=10)
    
    def open_view_by_time(self):
        win = tk.Toplevel(self.root)
        win.title(f"Vue par Heure - {format_date_display(self.current_date)}")
        win.geometry("1200x700")
        
        auto_refresh_active = {'active': True}
        
        header = ttk.Frame(win, padding=10)
        header.pack(fill="x")
        ttk.Label(header, text=f"🕐 Planning Global par Heure", 
                 font=('Arial', 14, 'bold')).pack(side="left")
        ttk.Label(header, text=f"Date: {self.current_date.strftime('%d/%m/%Y')}", 
                 font=('Arial', 11)).pack(side="right")
        
        toolbar = ttk.Frame(win, padding=5)
        toolbar.pack(fill="x", padx=10)
        
        search_var = tk.StringVar()
        
        ttk.Label(toolbar, text="🔍 Rechercher :").pack(side="left", padx=(0, 5))
        search_entry = ttk.Entry(toolbar, textvariable=search_var, width=30)
        search_entry.pack(side="left")
        
        clear_btn_frame = ttk.Frame(toolbar)
        clear_btn_frame.pack(side="left", padx=2)
        
        refresh_btn_frame = ttk.Frame(toolbar)
        refresh_btn_frame.pack(side="right", padx=5)
        
        tree_frame = ttk.Frame(win, padding=10)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("heure", "type", "voyage", "chauffeur", "numero", 
                  "nb_pal", "sst", "pays", "ramasse", "infos")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=25, selectmode="extended")
        
        col_config = [
            ("heure", "Heure", 60),
            ("type", "Type", 80),
            ("voyage", "Voyage", 100),
            ("chauffeur", "Chauffeur", 120),
            ("numero", "N°", 40),
            ("nb_pal", "Nb Pal", 60),
            ("sst", "SST", 70),
            ("pays", "Pays", 100),
            ("ramasse", "Ramasse", 100),
            ("infos", "Infos", 150)
        ]
        
        for col, header_text, width in col_config:
            tree.heading(col, text=header_text)
            tree.column(col, width=width)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        all_data = {'items': []}
        
        def do_search(*args):
            search_text = search_var.get().strip().upper()
            
            for item in tree.get_children():
                tree.delete(item)
            
            if not search_text:
                for item_id, values, tags in all_data['items']:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
                return
            
            for item_id, values, tags in all_data['items']:
                match = False
                for value in values:
                    if search_text in str(value).upper():
                        match = True
                        break
                if match:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
        
        search_var.trace('w', do_search)
        
        def refresh_tree():
            if auto_refresh_active['active']:
                try:
                    self.load_planning_for_date(self.current_date, preserve_ui=True)
                    for item in tree.get_children():
                        tree.delete(item)
                    self._fill_consolidated_view(tree, sort_by="time")
                    all_data['items'] = []
                    for item_id in tree.get_children():
                        values = tree.item(item_id)["values"]
                        tags = tree.item(item_id)["tags"]
                        all_data['items'].append((item_id, values, tags))
                    do_search()
                    # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
                except Exception:
                    pass
        
        self._fill_consolidated_view(tree, sort_by="time")
        
        all_data['items'] = []
        for item_id in tree.get_children():
            values = tree.item(item_id)["values"]
            tags = tree.item(item_id)["tags"]
            all_data['items'].append((item_id, values, tags))
        
        def clear_search():
            search_var.set("")
        
        ttk.Button(clear_btn_frame, text="❌", command=clear_search, width=3).pack()
        
        def manual_refresh():
            self.load_planning_for_date(self.current_date, preserve_ui=True)
            for item in tree.get_children():
                tree.delete(item)
            self._fill_consolidated_view(tree, sort_by="time")
            all_data['items'] = []
            for item_id in tree.get_children():
                values = tree.item(item_id)["values"]
                tags = tree.item(item_id)["tags"]
                all_data['items'].append((item_id, values, tags))
            do_search()
        
        ttk.Button(refresh_btn_frame, text="🔄 Rafraîchir", command=manual_refresh).pack()
        
        # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
        
        def on_close():
            auto_refresh_active['active'] = False
            win.destroy()
        
        win.protocol("WM_DELETE_WINDOW", on_close)
        
        control_frame = ttk.Frame(win, padding=5)
        control_frame.pack(fill="x")
        
        export_frame = ttk.Frame(win)
        export_frame.pack(pady=5)
        
        def export_excel_par_heure():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_heure', extension='xlsx')
            success, message = export_planning_excel_par_heure(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)

        def export_pdf_par_heure():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_heure', extension='pdf')
            success, message = export_planning_pdf_par_heure(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)
        
        if EXCEL_AVAILABLE:
            ttk.Button(export_frame, text='📊 Exporter Excel', command=export_excel_par_heure).pack(side='left', padx=5)
        if PDF_AVAILABLE:
            ttk.Button(export_frame, text='📄 Exporter PDF', command=export_pdf_par_heure).pack(side='left', padx=5)
        
        ttk.Button(export_frame, text="Fermer", command=on_close).pack(side="left", padx=10)
    
    def open_view_by_voyage(self):
        win = tk.Toplevel(self.root)
        win.title(f"Vue par Voyage - {format_date_display(self.current_date)}")
        win.geometry("1200x700")
        
        auto_refresh_active = {'active': True}
        
        header = ttk.Frame(win, padding=10)
        header.pack(fill="x")
        ttk.Label(header, text=f"📍 Planning Global par Voyage", 
                 font=('Arial', 14, 'bold')).pack(side="left")
        ttk.Label(header, text=f"Date: {self.current_date.strftime('%d/%m/%Y')}", 
                 font=('Arial', 11)).pack(side="right")
        
        toolbar = ttk.Frame(win, padding=5)
        toolbar.pack(fill="x", padx=10)
        
        search_var = tk.StringVar()
        
        all_data = {'items': []}
        
        def do_search(*args):
            search_text = search_var.get().strip().upper()
            
            for item in tree.get_children():
                tree.delete(item)
            
            if not search_text:
                for item_id, values, tags in all_data['items']:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
                return
            
            for item_id, values, tags in all_data['items']:
                match = False
                for value in values:
                    if search_text in str(value).upper():
                        match = True
                        break
                if match:
                    tree.insert("", "end", iid=item_id, values=values, tags=tags)
        
        ttk.Label(toolbar, text="🔍 Rechercher :").pack(side="left", padx=(0, 5))
        search_entry = ttk.Entry(toolbar, textvariable=search_var, width=30)
        search_entry.pack(side="left")
        
        def clear_search():
            search_var.set("")
            do_search()
        
        ttk.Button(toolbar, text="❌", command=clear_search, width=3).pack(side="left", padx=2)
        
        def manual_refresh():
            self.load_planning_for_date(self.current_date, preserve_ui=True)
            for item in tree.get_children():
                tree.delete(item)
            self._fill_consolidated_view(tree, sort_by="voyage")
            all_data['items'] = []
            for item_id in tree.get_children():
                values = tree.item(item_id)["values"]
                tags = tree.item(item_id)["tags"]
                all_data['items'].append((item_id, values, tags))
            do_search()
        
        ttk.Button(toolbar, text="🔄 Rafraîchir", command=manual_refresh).pack(side="right", padx=5)
        
        search_var.trace('w', do_search)
        
        tree_frame = ttk.Frame(win, padding=10)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("voyage", "pays", "type", "heure", "chauffeur", 
                  "numero", "nb_pal", "sst", "ramasse", "infos")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=25, selectmode="extended")
        
        col_config = [
            ("voyage", "Voyage", 100),
            ("pays", "Pays", 100),
            ("type", "Type", 80),
            ("heure", "Heure", 60),
            ("chauffeur", "Chauffeur", 120),
            ("numero", "N°", 40),
            ("nb_pal", "Nb Pal", 60),
            ("sst", "SST", 70),
            ("ramasse", "Ramasse", 100),
            ("infos", "Infos", 150)
        ]
        
        for col, header_text, width in col_config:
            tree.heading(col, text=header_text)
            tree.column(col, width=width)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        def refresh_tree():
            if auto_refresh_active['active']:
                try:
                    self.load_planning_for_date(self.current_date, preserve_ui=True)
                    for item in tree.get_children():
                        tree.delete(item)
                    self._fill_consolidated_view(tree, sort_by="voyage")
                    all_data['items'] = []
                    for item_id in tree.get_children():
                        values = tree.item(item_id)["values"]
                        tags = tree.item(item_id)["tags"]
                        all_data['items'].append((item_id, values, tags))
                    do_search()
                    # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
                except Exception:
                    pass
        
        self._fill_consolidated_view(tree, sort_by="voyage")
        
        all_data['items'] = []
        for item_id in tree.get_children():
            values = tree.item(item_id)["values"]
            tags = tree.item(item_id)["tags"]
            all_data['items'].append((item_id, values, tags))
        
        # win.after(5000, refresh_tree)  # auto-refresh désactivé (rafraîchissement manuel uniquement)
        
        def on_close():
            auto_refresh_active['active'] = False
            win.destroy()
        
        win.protocol("WM_DELETE_WINDOW", on_close)
        
        summary_frame = ttk.LabelFrame(win, text="📊 Résumé par voyage", padding=10)
        summary_frame.pack(fill="x", padx=10, pady=5)
        
        summary_text = tk.Text(summary_frame, height=5, width=100)
        summary_text.pack()
        
        voyage_stats = {}
        v_by_code = {v.get("code"): v for v in self.voyages}
        
        for m in self.missions:
            voyage_code = m.get("voyage", "")
            if voyage_code not in voyage_stats:
                voyage_stats[voyage_code] = {"count": 0, "palettes": 0}
            voyage_stats[voyage_code]["count"] += 1
            voyage_stats[voyage_code]["palettes"] += int(m.get("nb_pal", 0))
        
        summary = "Résumé des voyages:\n"
        for code in sorted(voyage_stats.keys()):
            voyage = v_by_code.get(code, {})
            country = voyage.get("country", "Belgique")
            stats = voyage_stats[code]
            summary += f"  • {code} ({country}): {stats['count']} missions, {stats['palettes']} palettes\n"
        
        summary_text.insert("1.0", summary)
        summary_text.config(state="disabled")
        
        export_frame = ttk.Frame(win)
        export_frame.pack(pady=5)
        
        def export_excel_par_voyage():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_voyage', extension='xlsx')
            success, message = export_planning_excel_par_voyage(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)

        def export_pdf_par_voyage():
            desktop = get_desktop_path()
            filename = desktop / generate_export_filename(prefix='planning_par_voyage', extension='pdf')
            success, message = export_planning_pdf_par_voyage(self.missions, self.voyages, self.current_date, filename)
            if success:
                messagebox.showinfo('Export réussi', f"Fichier exporté sur le bureau et ouvert:\n{filename.name}")
            else:
                messagebox.showerror('Erreur export', message)
        
        if EXCEL_AVAILABLE:
            ttk.Button(export_frame, text='📊 Exporter Excel', command=export_excel_par_voyage).pack(side='left', padx=5)
        if PDF_AVAILABLE:
            ttk.Button(export_frame, text='📄 Exporter PDF', command=export_pdf_par_voyage).pack(side='left', padx=5)
        
        ttk.Button(export_frame, text="Fermer", command=on_close).pack(side="left", padx=10)

    def open_planning_announcements(self):
        """Ouvrir la fenêtre des annonces SST pour J+1"""
        # Vérifier la permission
        if not self.rights["permissions"].get("send_announcements", False):
            messagebox.showwarning("Permission refusée",
                "Vous n'avez pas la permission d'envoyer des annonces.\n"
                "Contactez un administrateur pour obtenir cette permission.")
            return

        # Vérifier si on est sur J+1
        today = date.today()
        tomorrow = today + timedelta(days=1)

        if self.current_date != tomorrow:
            messagebox.showwarning("Date non autorisée",
                f"Les annonces ne peuvent être envoyées que pour J+1.\n\n"
                f"Date actuelle du planning: {self.current_date.strftime('%d/%m/%Y')}\n"
                f"Date autorisée (J+1): {tomorrow.strftime('%d/%m/%Y')}\n\n"
                f"Veuillez sélectionner la date de demain pour envoyer des annonces.")
            return

        # Charger les configurations
        sst_emails = load_json(SST_EMAILS_FILE, default={})
        announcement_config = load_json(ANNOUNCEMENT_CONFIG_FILE, default={
            "reply_to": "",
            "subject": "Annonce missions transport - {date}",
            "cc_addresses": [],
            "body_template": """Bonjour,

Pour le {date}, les missions suivantes sont planifiées pour votre société :

{chauffeurs_list}

Lieu de présentation des chauffeurs : {lieu_presentation}

Merci de confirmer :
- Que le(s) chauffeur(s) annoncé(s) est/sont bien les bons
- Que vous acceptez la/les mission(s)

Merci de répondre avant 20h00.

Cordialement,
{expediteur}""",
            "lieu_presentation": "Tubize",
            "variables_help": [
                "{date} - Date de la mission (format: JJ/MM/AAAA)",
                "{sst_name} - Nom du sous-traitant",
                "{chauffeurs_list} - Liste des chauffeurs avec leur première heure",
                "{lieu_presentation} - Lieu de présentation",
                "{expediteur} - Nom de l'expéditeur",
                "{nb_chauffeurs} - Nombre de chauffeurs",
                "{nb_missions} - Nombre de missions"
            ]
        })
        history = load_json(ANNOUNCEMENT_HISTORY_FILE, default=[])

        # Récupérer les SST du planning J+1
        sst_missions = {}
        for m in self.missions:
            sst = m.get("sst", "").strip()
            if sst:
                if sst not in sst_missions:
                    sst_missions[sst] = []
                sst_missions[sst].append(m)

        if not sst_missions:
            messagebox.showinfo("Aucun SST",
                f"Aucun sous-traitant trouvé dans le planning du {self.current_date.strftime('%d/%m/%Y')}.")
            return

        # Créer la fenêtre
        win = tk.Toplevel(self.root)
        win.title(f"📢 Annonces SST - {self.current_date.strftime('%d/%m/%Y')}")
        win.geometry("1000x700")
        win.transient(self.root)

        # Header
        header = ttk.Frame(win, padding=10)
        header.pack(fill="x")
        ttk.Label(header, text="📢 Envoi d'annonces aux sous-traitants",
                 font=('Arial', 14, 'bold')).pack(side="left")
        ttk.Label(header, text=f"Date mission: {self.current_date.strftime('%d/%m/%Y')} (J+1)",
                 font=('Arial', 11), foreground="green").pack(side="right")

        # Notebook pour les onglets
        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=10, pady=5)

        # === ONGLET 1: Liste des SST ===
        tab_sst = ttk.Frame(notebook, padding=10)
        notebook.add(tab_sst, text="📋 SST à contacter")

        # Frame pour la liste des SST
        list_frame = ttk.Frame(tab_sst)
        list_frame.pack(fill="both", expand=True)

        columns = ("sst", "nb_missions", "chauffeurs", "premiere_heure", "email", "status")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=15)

        tree.heading("sst", text="SST")
        tree.heading("nb_missions", text="Missions")
        tree.heading("chauffeurs", text="Chauffeurs")
        tree.heading("premiere_heure", text="1ère heure")
        tree.heading("email", text="Email configuré")
        tree.heading("status", text="Statut")

        tree.column("sst", width=100)
        tree.column("nb_missions", width=70)
        tree.column("chauffeurs", width=200)
        tree.column("premiere_heure", width=80)
        tree.column("email", width=200)
        tree.column("status", width=150)

        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def check_already_sent(sst_name, target_date):
            """Vérifier si un mail a déjà été envoyé pour ce SST et cette date"""
            for h in history:
                if h.get("sst") == sst_name and h.get("mission_date") == target_date:
                    return h
            return None

        def refresh_sst_list():
            for item in tree.get_children():
                tree.delete(item)

            target_date_str = self.current_date.strftime("%Y-%m-%d")

            for sst, missions in sorted(sst_missions.items()):
                # Trouver les chauffeurs uniques
                chauffeurs = list(set(m.get("chauffeur_nom", "N/A") for m in missions if m.get("chauffeur_nom")))
                chauffeurs_str = ", ".join(chauffeurs) if chauffeurs else "Non assigné"

                # Trouver la première heure
                heures = [m.get("heure", "99:99") for m in missions]
                premiere_heure = min(heures) if heures else "N/A"

                # Email configuré
                email_config = sst_emails.get(sst, {})
                emails = email_config.get("emails", [])
                email_str = emails[0] if emails else "❌ Non configuré"

                # Vérifier si déjà envoyé
                already_sent = check_already_sent(sst, target_date_str)
                if already_sent:
                    status = f"✅ Envoyé le {already_sent.get('sent_at', '')[:16]}"
                    tag = "sent"
                else:
                    status = "⏳ En attente"
                    tag = "pending"

                tree.insert("", "end", values=(
                    sst,
                    len(missions),
                    chauffeurs_str[:50] + ("..." if len(chauffeurs_str) > 50 else ""),
                    premiere_heure,
                    email_str[:40] + ("..." if len(email_str) > 40 else ""),
                    status
                ), tags=(tag,))

            tree.tag_configure("sent", background="#d4edda")
            tree.tag_configure("pending", background="#fff3cd")

        refresh_sst_list()

        # Boutons d'action SST
        btn_frame = ttk.Frame(tab_sst, padding=5)
        btn_frame.pack(fill="x")

        def send_announcement_to_selected():
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("Attention", "Veuillez sélectionner un SST.")
                return

            item = tree.item(selected[0])
            sst_name = item['values'][0]

            # Vérifier si déjà envoyé
            target_date_str = self.current_date.strftime("%Y-%m-%d")
            already_sent = check_already_sent(sst_name, target_date_str)

            if already_sent:
                msg = (f"Un mail a déjà été envoyé pour {sst_name} "
                      f"le {already_sent.get('sent_at', 'N/A')[:16]}\n"
                      f"par {already_sent.get('sent_by', 'N/A')}.\n\n"
                      "Voulez-vous renvoyer un mail ?")
                if not messagebox.askyesno("Mail déjà envoyé", msg):
                    return

            # Vérifier si email configuré
            email_config = sst_emails.get(sst_name, {})
            emails = email_config.get("emails", [])

            if not emails:
                messagebox.showerror("Email non configuré",
                    f"Aucune adresse email configurée pour {sst_name}.\n"
                    "Veuillez configurer l'email dans l'onglet 'Config Emails SST'.")
                return

            # Générer et envoyer le mail
            send_announcement_email(sst_name)

        def send_announcement_email(sst_name):
            """Générer et ouvrir le mail dans Outlook"""
            if not OUTLOOK_AVAILABLE:
                messagebox.showerror("Outlook non disponible",
                    "Le module Outlook (win32com) n'est pas installé.\n"
                    "Installez-le avec: pip install pywin32")
                return

            missions = sst_missions.get(sst_name, [])
            email_config = sst_emails.get(sst_name, {})
            emails = email_config.get("emails", [])

            # Préparer les données
            # Construire la liste des chauffeurs avec leur première heure
            # Regrouper les missions par chauffeur et trouver la première heure
            chauffeurs_dict = {}
            for m in missions:
                chauffeur_nom = m.get("chauffeur_nom", "").strip()
                if chauffeur_nom:
                    heure = m.get("heure", "99:99")
                    if chauffeur_nom not in chauffeurs_dict:
                        chauffeurs_dict[chauffeur_nom] = heure
                    else:
                        # Garder la première heure (la plus tôt)
                        if heure < chauffeurs_dict[chauffeur_nom]:
                            chauffeurs_dict[chauffeur_nom] = heure

            # Générer la liste formatée
            chauffeurs_lines = []
            for i, (nom, heure) in enumerate(sorted(chauffeurs_dict.items(), key=lambda x: x[1]), 1):
                chauffeurs_lines.append(f"Chauffeur {i} : {nom} – Attendu pour {heure}")
            chauffeurs_list = "\n".join(chauffeurs_lines) if chauffeurs_lines else "Aucun chauffeur assigné"

            chauffeurs = list(chauffeurs_dict.keys())

            # Remplacer les variables dans le template
            body = announcement_config.get("body_template", "")
            subject = announcement_config.get("subject", "Annonce missions - {date}")

            replacements = {
                "{date}": self.current_date.strftime("%d/%m/%Y"),
                "{sst_name}": sst_name,
                "{chauffeurs_list}": chauffeurs_list,
                "{lieu_presentation}": announcement_config.get("lieu_presentation", "Tubize"),
                "{expediteur}": self.current_user,
                "{nb_chauffeurs}": str(len(chauffeurs)),
                "{nb_missions}": str(len(missions))
            }

            for key, value in replacements.items():
                body = body.replace(key, value)
                subject = subject.replace(key, value)

            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)

                mail.To = "; ".join(emails)
                mail.Subject = subject
                mail.Body = body

                # Ajouter CC si configuré
                cc_addresses = announcement_config.get("cc_addresses", [])
                if cc_addresses:
                    mail.CC = "; ".join(cc_addresses)

                # Configurer Reply-To si spécifié
                reply_to = announcement_config.get("reply_to", "")
                if reply_to:
                    mail.ReplyRecipients.Add(reply_to)

                mail.Display()  # Afficher le mail (ne pas envoyer automatiquement)

                # Enregistrer dans l'historique
                history_entry = {
                    "id": str(uuid.uuid4()),
                    "sst": sst_name,
                    "mission_date": self.current_date.strftime("%Y-%m-%d"),
                    "sent_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "sent_by": self.current_user,
                    "recipients": emails,
                    "subject": subject,
                    "nb_missions": len(missions),
                    "chauffeurs": chauffeurs
                }
                history.append(history_entry)
                save_json(ANNOUNCEMENT_HISTORY_FILE, history)

                # Log de l'activité
                if hasattr(self, 'activity_logger'):
                    self.activity_logger.log_action(
                        action="send_sst_announcement",
                        entity_type="announcement",
                        entity_id=history_entry["id"],
                        details={"sst": sst_name, "mission_date": self.current_date.strftime("%Y-%m-%d")}
                    )

                refresh_sst_list()
                messagebox.showinfo("Mail créé",
                    f"Le mail pour {sst_name} a été créé dans Outlook.\n"
                    "Vérifiez le contenu et envoyez-le manuellement.")

            except Exception as e:
                messagebox.showerror("Erreur Outlook", f"Erreur lors de la création du mail:\n{e}")

        def send_all_pending():
            """Envoyer les annonces à tous les SST en attente"""
            target_date_str = self.current_date.strftime("%Y-%m-%d")
            pending_sst = []

            for sst in sst_missions.keys():
                email_config = sst_emails.get(sst, {})
                emails = email_config.get("emails", [])
                if emails and not check_already_sent(sst, target_date_str):
                    pending_sst.append(sst)

            if not pending_sst:
                messagebox.showinfo("Aucun envoi", "Tous les SST ont déjà reçu leur annonce ou n'ont pas d'email configuré.")
                return

            msg = f"Créer les mails pour {len(pending_sst)} SST ?\n\n" + "\n".join(f"  - {s}" for s in pending_sst)
            if not messagebox.askyesno("Confirmer envoi groupé", msg):
                return

            for sst in pending_sst:
                send_announcement_email(sst)

        ttk.Button(btn_frame, text="📧 Envoyer au SST sélectionné",
                  command=send_announcement_to_selected).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="📬 Envoyer à tous les SST en attente",
                  command=send_all_pending).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="🔄 Rafraîchir",
                  command=refresh_sst_list).pack(side="left", padx=5)

        # === ONGLET 2: Configuration emails SST ===
        tab_emails = ttk.Frame(notebook, padding=10)
        notebook.add(tab_emails, text="📧 Config Emails SST")

        email_list_frame = ttk.LabelFrame(tab_emails, text="Emails par SST", padding=10)
        email_list_frame.pack(fill="both", expand=True)

        email_columns = ("sst", "emails")
        email_tree = ttk.Treeview(email_list_frame, columns=email_columns, show="headings", height=12)
        email_tree.heading("sst", text="SST")
        email_tree.heading("emails", text="Adresses email")
        email_tree.column("sst", width=150)
        email_tree.column("emails", width=400)

        email_vsb = ttk.Scrollbar(email_list_frame, orient="vertical", command=email_tree.yview)
        email_tree.configure(yscrollcommand=email_vsb.set)
        email_tree.pack(side="left", fill="both", expand=True)
        email_vsb.pack(side="right", fill="y")

        def refresh_email_list():
            for item in email_tree.get_children():
                email_tree.delete(item)

            # Récupérer tous les SST connus
            all_sst = set(sst_missions.keys())
            sst_list = load_json(self.data_dir / "sst.json", [])
            all_sst.update(sst_list)

            for sst in sorted(all_sst):
                config = sst_emails.get(sst, {})
                emails = config.get("emails", [])
                email_tree.insert("", "end", values=(sst, ", ".join(emails) if emails else "Non configuré"))

        refresh_email_list()

        # Form pour éditer les emails
        edit_frame = ttk.LabelFrame(tab_emails, text="Modifier l'email d'un SST", padding=10)
        edit_frame.pack(fill="x", pady=10)

        ttk.Label(edit_frame, text="SST sélectionné:").grid(row=0, column=0, sticky="w", padx=5)
        selected_sst_var = tk.StringVar()
        selected_sst_label = ttk.Label(edit_frame, textvariable=selected_sst_var, font=('Arial', 10, 'bold'))
        selected_sst_label.grid(row=0, column=1, sticky="w", padx=5)

        ttk.Label(edit_frame, text="Emails (séparés par ;):").grid(row=1, column=0, sticky="w", padx=5)
        email_entry_var = tk.StringVar()
        email_entry = ttk.Entry(edit_frame, textvariable=email_entry_var, width=60)
        email_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        def on_email_tree_select(event):
            selected = email_tree.selection()
            if selected:
                item = email_tree.item(selected[0])
                sst = item['values'][0]
                selected_sst_var.set(sst)
                config = sst_emails.get(sst, {})
                emails = config.get("emails", [])
                email_entry_var.set("; ".join(emails))

        email_tree.bind("<<TreeviewSelect>>", on_email_tree_select)

        def save_sst_email():
            sst = selected_sst_var.get()
            if not sst:
                messagebox.showwarning("Attention", "Veuillez sélectionner un SST.")
                return

            emails_str = email_entry_var.get().strip()
            emails = [e.strip() for e in emails_str.split(";") if e.strip()]

            sst_emails[sst] = {"emails": emails}
            save_json(SST_EMAILS_FILE, sst_emails)
            refresh_email_list()
            refresh_sst_list()
            messagebox.showinfo("Sauvegardé", f"Email(s) sauvegardé(s) pour {sst}.")

        ttk.Button(edit_frame, text="💾 Sauvegarder", command=save_sst_email).grid(row=2, column=1, sticky="w", padx=5, pady=5)

        # === ONGLET 3: Configuration du modèle ===
        if self.rights["permissions"].get("manage_announcements_config", False):
            tab_config = ttk.Frame(notebook, padding=10)
            notebook.add(tab_config, text="⚙️ Admin Config")

            # Variables du template
            vars_frame = ttk.LabelFrame(tab_config, text="Variables disponibles dans le mail", padding=10)
            vars_frame.pack(fill="x", pady=5)

            vars_text = "\n".join(announcement_config.get("variables_help", []))
            ttk.Label(vars_frame, text=vars_text, font=('Courier', 9), justify="left").pack(anchor="w")

            # Configuration
            config_frame = ttk.LabelFrame(tab_config, text="Paramètres du mail", padding=10)
            config_frame.pack(fill="both", expand=True, pady=5)

            row = 0
            ttk.Label(config_frame, text="Objet du mail:").grid(row=row, column=0, sticky="w", padx=5, pady=2)
            subject_var = tk.StringVar(value=announcement_config.get("subject", ""))
            ttk.Entry(config_frame, textvariable=subject_var, width=60).grid(row=row, column=1, sticky="w", padx=5, pady=2)

            row += 1
            ttk.Label(config_frame, text="Adresse Reply-To:").grid(row=row, column=0, sticky="w", padx=5, pady=2)
            reply_to_var = tk.StringVar(value=announcement_config.get("reply_to", ""))
            ttk.Entry(config_frame, textvariable=reply_to_var, width=60).grid(row=row, column=1, sticky="w", padx=5, pady=2)

            row += 1
            ttk.Label(config_frame, text="CC fixe (séparés par ;):").grid(row=row, column=0, sticky="w", padx=5, pady=2)
            cc_var = tk.StringVar(value="; ".join(announcement_config.get("cc_addresses", [])))
            ttk.Entry(config_frame, textvariable=cc_var, width=60).grid(row=row, column=1, sticky="w", padx=5, pady=2)

            row += 1
            ttk.Label(config_frame, text="Lieu de présentation:").grid(row=row, column=0, sticky="w", padx=5, pady=2)
            lieu_var = tk.StringVar(value=announcement_config.get("lieu_presentation", "Tubize"))
            ttk.Entry(config_frame, textvariable=lieu_var, width=40).grid(row=row, column=1, sticky="w", padx=5, pady=2)

            row += 1
            ttk.Label(config_frame, text="Modèle du corps du mail:").grid(row=row, column=0, sticky="nw", padx=5, pady=2)
            body_text = tk.Text(config_frame, height=12, width=70)
            body_text.grid(row=row, column=1, sticky="w", padx=5, pady=2)
            body_text.insert("1.0", announcement_config.get("body_template", ""))

            def save_config():
                announcement_config["subject"] = subject_var.get()
                announcement_config["reply_to"] = reply_to_var.get()
                announcement_config["cc_addresses"] = [e.strip() for e in cc_var.get().split(";") if e.strip()]
                announcement_config["lieu_presentation"] = lieu_var.get()
                announcement_config["body_template"] = body_text.get("1.0", "end-1c")

                save_json(ANNOUNCEMENT_CONFIG_FILE, announcement_config)
                messagebox.showinfo("Sauvegardé", "Configuration sauvegardée.")

            ttk.Button(config_frame, text="💾 Sauvegarder la configuration", command=save_config).grid(row=row+1, column=1, sticky="w", padx=5, pady=10)

        # === ONGLET 4: Historique ===
        tab_history = ttk.Frame(notebook, padding=10)
        notebook.add(tab_history, text="📜 Historique")

        history_frame = ttk.Frame(tab_history)
        history_frame.pack(fill="both", expand=True)

        history_columns = ("sent_at", "sent_by", "sst", "mission_date", "recipients", "nb_missions")
        history_tree = ttk.Treeview(history_frame, columns=history_columns, show="headings", height=20)

        history_tree.heading("sent_at", text="Envoyé le")
        history_tree.heading("sent_by", text="Par")
        history_tree.heading("sst", text="SST")
        history_tree.heading("mission_date", text="Date mission")
        history_tree.heading("recipients", text="Destinataires")
        history_tree.heading("nb_missions", text="Nb missions")

        history_tree.column("sent_at", width=140)
        history_tree.column("sent_by", width=100)
        history_tree.column("sst", width=80)
        history_tree.column("mission_date", width=100)
        history_tree.column("recipients", width=250)
        history_tree.column("nb_missions", width=80)

        history_vsb = ttk.Scrollbar(history_frame, orient="vertical", command=history_tree.yview)
        history_tree.configure(yscrollcommand=history_vsb.set)
        history_tree.pack(side="left", fill="both", expand=True)
        history_vsb.pack(side="right", fill="y")

        def refresh_history():
            for item in history_tree.get_children():
                history_tree.delete(item)

            for h in sorted(history, key=lambda x: x.get("sent_at", ""), reverse=True):
                history_tree.insert("", "end", values=(
                    h.get("sent_at", "")[:16],
                    h.get("sent_by", ""),
                    h.get("sst", ""),
                    h.get("mission_date", ""),
                    "; ".join(h.get("recipients", []))[:50],
                    h.get("nb_missions", 0)
                ))

        refresh_history()

        # Bouton fermer
        close_frame = ttk.Frame(win, padding=10)
        close_frame.pack(fill="x")
        ttk.Button(close_frame, text="Fermer", command=win.destroy).pack(side="right", padx=5)

    def _fill_consolidated_view(self, tree, sort_by="time"):
        for item in tree.get_children():
            tree.delete(item)
        
        v_by_code = {v.get("code"): v for v in self.voyages}
        missions_with_info = []
        
        for m in self.missions:
            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")
            
            mission_info = m.copy()
            mission_info["pays"] = country
            missions_with_info.append(mission_info)
        
        if sort_by == "driver":
            missions_sorted = sorted(missions_with_info, 
                                   key=lambda x: (x.get("chauffeur_nom", ""), 
                                                 int(x.get("numero", 0)) if str(x.get("numero", "")).isdigit() else 0))
        elif sort_by == "time":
            missions_sorted = sorted(missions_with_info, key=self._time_key)
        elif sort_by == "voyage":
            missions_sorted = sorted(missions_with_info, 
                                   key=lambda x: (x.get("voyage", ""), self._time_key(x)))
        else:
            missions_sorted = missions_with_info
        
        row_num = 0
        current_group = None
        
        for m in missions_sorted:
            tag = 'evenrow' if row_num % 2 == 0 else 'oddrow'
            
            if sort_by == "driver":
                if current_group != m.get("chauffeur_nom", ""):
                    current_group = m.get("chauffeur_nom", "")
                    if row_num > 0:
                        tree.insert("", "end", values=["─"*20]*10, tags=('separator',))
                        row_num += 1
                
                values = (
                    m.get("chauffeur_nom", ""),
                    m.get("numero", ""),
                    m.get("heure", ""),
                    m.get("type", ""),
                    m.get("voyage", ""),
                    m.get("nb_pal", ""),
                    m.get("sst", ""),
                    m.get("pays", ""),
                    m.get("ramasse", ""),
                    m.get("infos", "")
                )
                
            elif sort_by == "time":
                values = (
                    m.get("heure", ""),
                    m.get("type", ""),
                    m.get("voyage", ""),
                    m.get("chauffeur_nom", ""),
                    m.get("numero", ""),
                    m.get("nb_pal", ""),
                    m.get("sst", ""),
                    m.get("pays", ""),
                    m.get("ramasse", ""),
                    m.get("infos", "")
                )
                
            elif sort_by == "voyage":
                if current_group != m.get("voyage", ""):
                    current_group = m.get("voyage", "")
                    if row_num > 0:
                        tree.insert("", "end", values=["─"*20]*10, tags=('separator',))
                        row_num += 1
                
                values = (
                    m.get("voyage", ""),
                    m.get("pays", ""),
                    m.get("type", ""),
                    m.get("heure", ""),
                    m.get("chauffeur_nom", ""),
                    m.get("numero", ""),
                    m.get("nb_pal", ""),
                    m.get("sst", ""),
                    m.get("ramasse", ""),
                    m.get("infos", "")
                )
            
            tree.insert("", "end", values=values, tags=(tag,))
            row_num += 1
        
        tree.tag_configure('oddrow', background='white')
        tree.tag_configure('evenrow', background='#F0F0F0')
        tree.tag_configure('separator', background='#CCCCCC', font=('Arial', 1))

    def update_summary_stats(self):
        if not hasattr(self, 'summary_frame'):
            return
        
        nb_liv = sum(1 for m in self.missions if m.get("type") == "LIVRAISON")
        nb_ram = sum(1 for m in self.missions if m.get("type") == "RAMASSE")
        
        active_drivers = [ch for ch in self.chauffeurs if ch.get("actif", True)]
        driver_missions = {}
        for ch in active_drivers:
            driver_missions[ch["id"]] = []
        
        date_str = self.current_date.strftime("%Y-%m-%d")
        dispo_map = {}
        for entry in self.dispos:
            if entry.get("date") == date_str:
                dispo_map[entry["id_chauffeur"]] = bool(entry.get("disponible", True))
        
        for mission in self.missions:
            ch_id = mission.get("chauffeur_id")
            voyage_code = mission.get("voyage", "")
            if ch_id and voyage_code:
                if ch_id in driver_missions:
                    driver_missions[ch_id].append(voyage_code)
            elif not ch_id:
                ch_nom = mission.get("chauffeur_nom", "")
                for ch in active_drivers:
                    if ch.get("nom_affichage") == ch_nom and voyage_code:
                        driver_missions[ch["id"]].append(voyage_code)
                        break
        
        nb_dispo = 0
        nb_used = 0
        for ch in active_drivers:
            ch_id = ch["id"]
            is_available = dispo_map.get(ch_id, True)
            has_missions = len(driver_missions[ch_id]) > 0
            if is_available and not has_missions:
                nb_dispo += 1
            elif is_available and has_missions:
                nb_used += 1
        
        # Calculer les palettes par type (livraison et ramasse)
        pal_liv = sum(int(m.get("nb_pal", 0)) for m in self.missions if m.get("type") == "LIVRAISON")
        pal_ram = sum(int(m.get("nb_pal", 0)) for m in self.missions if m.get("type") == "RAMASSE")
        total_pal = pal_liv + pal_ram
        
        v_by_code = {v.get("code"): v for v in self.voyages}
        countries = set()
        for m in self.missions:
            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")
            countries.add(country)
        nb_pays = len(countries)
        
        self.summary_liv_label.config(text=str(nb_liv))
        self.summary_ram_label.config(text=str(nb_ram))
        self.summary_dispo_label.config(text=str(nb_dispo))
        self.summary_used_label.config(text=str(nb_used))
        
        # Afficher les palettes au format "XXX Liv | XXX Ram (TOTAL)"
        pal_text = f"{pal_liv} Liv | {pal_ram} Ram ({total_pal})"
        self.summary_pal_label.config(text=pal_text)
        self.summary_pays_label.config(text=str(nb_pays))

        # Compter les voyages incomplets (sans SST ou sans chauffeur)
        nb_incomplete = sum(1 for m in self.missions
                          if m.get("sans_sst", False) or m.get("sans_chauffeur", False))

        if nb_incomplete > 0:
            if nb_incomplete == 1:
                incomplete_text = f"⚠️ Attention: {nb_incomplete} voyage pas complet"
            else:
                incomplete_text = f"⚠️ Attention: {nb_incomplete} voyages pas complets"
            self.summary_incomplete_label.config(text=incomplete_text)
        else:
            self.summary_incomplete_label.config(text="")

    def sort_missions(self, criteria):
        if self.sort_criteria == criteria:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_criteria = criteria
            self.sort_reverse = False
            
        arrow = "↓" if self.sort_reverse else "↑"
        criteria_names = {
            "heure": "Heure", "voyage": "Voyage", "chauffeur": "Chauffeur",
            "numero": "N° tournée", "sst": "SST", "pays": "Pays"
        }
        self.sort_label.config(text=f"(Tri actuel: {criteria_names.get(criteria, criteria)} {arrow})")
        
        self.refresh_planning_view()

    def create_country_planning_section(self, country):
        if country in self.country_frames:
            return

        bg_color = COUNTRY_COLORS.get(country, "#F5F5F5")
        
        style = ttk.Style()
        style_name = f"{country.replace(' ', '_')}.TLabelframe"
        style.configure(style_name, background=bg_color, relief="solid", borderwidth=2)
        style.configure(f"{style_name}.Label", background=bg_color, font=('Arial', 11, 'bold'))
        
        flag_emoji = self.get_country_flag(country)
        country_frame = ttk.LabelFrame(self.planning_container,
                                       text=f"  {flag_emoji}  PLANNING {country.upper()}  ",
                                       style=style_name, padding=15, height=450)
        country_frame.pack(fill="x", expand=False, pady=5)
        country_frame.pack_propagate(False)

        # Poignée de redimensionnement en bas de la section (pack en premier avec side=bottom)
        resize_handle = tk.Frame(country_frame, bg="#999999", height=12, cursor="sb_v_double_arrow", relief="raised", bd=1)
        resize_handle.pack(fill="x", side="bottom", padx=5, pady=2)

        resize_label = tk.Label(resize_handle, text="≡ ≡ ≡", bg="#999999", fg="white", font=('Arial', 8), cursor="sb_v_double_arrow")
        resize_label.pack(expand=True)

        def start_resize(event):
            resize_handle._drag_start_y = event.y_root
            resize_handle._start_height = country_frame.winfo_height()

        def do_resize(event):
            delta = event.y_root - resize_handle._drag_start_y
            new_height = max(200, resize_handle._start_height + delta)
            country_frame.configure(height=new_height)
            self.scrollable_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

        resize_handle.bind("<Button-1>", start_resize)
        resize_handle.bind("<B1-Motion>", do_resize)
        resize_label.bind("<Button-1>", start_resize)
        resize_label.bind("<B1-Motion>", do_resize)

        inner_frame = tk.Frame(country_frame, bg=bg_color, highlightbackground="#DDD",
                               highlightthickness=1, relief="flat")
        inner_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        inner_frame.rowconfigure(0, weight=1)
        inner_frame.columnconfigure(0, weight=1)
        
        middle_paned = ttk.PanedWindow(inner_frame, orient=tk.HORIZONTAL)
        middle_paned.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        left_frame = ttk.Frame(middle_paned)
        middle_paned.add(left_frame, weight=1)
        
        right_frame = ttk.Frame(middle_paned)
        middle_paned.add(right_frame, weight=1)
        
        left_frame.rowconfigure(0, weight=0)
        left_frame.rowconfigure(1, weight=1)
        left_frame.columnconfigure(0, weight=1)
        left_frame.columnconfigure(1, weight=0)
        
        right_frame.rowconfigure(0, weight=0)
        right_frame.rowconfigure(1, weight=1)
        right_frame.columnconfigure(0, weight=1)
        right_frame.columnconfigure(1, weight=0)

        header_frame_l = tk.Frame(left_frame, bg="#2196F3", height=35)
        header_frame_l.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        self.country_liv_label = tk.Label(header_frame_l, text=f"📦 LIVRAISONS - {country} (Nb livraison: 0)", 
                bg="#2196F3", fg="white", font=('Arial', 10, 'bold'))
        self.country_liv_label.pack(pady=8)
        
        vsb_liv = ttk.Scrollbar(left_frame, orient="vertical")
        vsb_liv.grid(row=1, column=1, sticky="ns")
        
        columns = ("heure", "voyage", "nb_pal", "numero", "sst", "chauffeur", "infos")
        tree_liv = ttk.Treeview(left_frame, columns=columns, show="headings", height=20,
                                yscrollcommand=vsb_liv.set, selectmode="extended")
        tree_liv.grid(row=1, column=0, sticky="nsew")
        vsb_liv.config(command=tree_liv.yview)
        
        col_config = [
            ("heure", "Heure", 60),
            ("voyage", "Voyage", 90),
            ("nb_pal", "Nb Pal", 60),
            ("numero", "N°", 40),
            ("sst", "SST", 70),
            ("chauffeur", "Nom", 100),
            ("infos", "Infos", 120)
        ]
        for col, txt, width in col_config:
            tree_liv.heading(col, text=txt)
            tree_liv.column(col, width=width, minwidth=40, stretch=True)
        
        tree_liv.tag_configure('oddrow', background='white')
        tree_liv.tag_configure('evenrow', background='#F0F0F0')
        tree_liv.tag_configure('incomplete', background='#FFCDD2', foreground='#B71C1C')  # Rouge clair pour voyages incomplets

        header_frame_r = tk.Frame(right_frame, bg="#4CAF50", height=35)
        header_frame_r.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 5))
        self.country_ram_label = tk.Label(header_frame_r, text=f"🚛 RAMASSES - {country} (Nb ramasse: 0)", 
                bg="#4CAF50", fg="white", font=('Arial', 10, 'bold'))
        self.country_ram_label.pack(pady=8)
        
        vsb_ram = ttk.Scrollbar(right_frame, orient="vertical")
        vsb_ram.grid(row=1, column=1, sticky="ns")
        
        columns_r = ("heure", "voyage", "nb_pal", "numero", "sst", "chauffeur", "ramasse", "infos")
        tree_ram = ttk.Treeview(right_frame, columns=columns_r, show="headings", height=20,
                                yscrollcommand=vsb_ram.set, selectmode="extended")
        tree_ram.grid(row=1, column=0, sticky="nsew")
        vsb_ram.config(command=tree_ram.yview)
        
        col_config_r = [
            ("heure", "Heure", 60),
            ("voyage", "Voyage", 90),
            ("nb_pal", "Nb Pal", 50),
            ("numero", "N°", 40),
            ("sst", "SST", 70),
            ("chauffeur", "Nom", 90),
            ("ramasse", "Ramasse", 90),
            ("infos", "Infos", 100)
        ]
        for col, txt, width in col_config_r:
            tree_ram.heading(col, text=txt)
            tree_ram.column(col, width=width, minwidth=40, stretch=True)
        
        tree_ram.tag_configure('oddrow', background='white')
        tree_ram.tag_configure('evenrow', background='#F0F0F0')
        tree_ram.tag_configure('incomplete', background='#FFCDD2', foreground='#B71C1C')  # Rouge clair pour voyages incomplets

        self.country_frames[country] = country_frame
        self.country_trees[country] = {"livraison": tree_liv, "ramasse": tree_ram}
        
        # Empêcher la sélection simultanée dans plusieurs pays
        tree_liv.bind("<<TreeviewSelect>>", lambda e, c=country, t="livraison": self.on_country_tree_select(c, t))
        tree_ram.bind("<<TreeviewSelect>>", lambda e, c=country, t="ramasse": self.on_country_tree_select(c, t))
        if not hasattr(self, 'country_headers'):
            self.country_headers = {}
        self.country_headers[country] = {
            "livraison_label": self.country_liv_label,
            "ramasse_label": self.country_ram_label
        }
        
    def get_country_flag(self, country):
        flags = {
            "Belgique": "🇧🇪",
            "France": "🇫🇷",
            "Allemagne": "🇩🇪",
            "Pays-Bas": "🇳🇱",
            "Luxembourg": "🇱🇺",
            "Espagne": "🇪🇸",
            "Italie": "🇮🇹",
            "Portugal": "🇵🇹",
            "Pologne": "🇵🇱",
            "Autriche": "🇦🇹",
            "Suisse": "🇨🇭",
            "Royaume-Uni": "🇬🇧",
        }
        return flags.get(country, "🌍")

    def setup_combobox_autocomplete(self, combobox, all_values_var_name, on_select_callback=None):
        def filter_and_update():
            typed = combobox.get()
            all_values = getattr(self, all_values_var_name, [])
            
            if not typed:
                combobox['values'] = all_values
                return all_values
            
            typed_upper = typed.upper()
            filtered = [v for v in all_values if str(v).upper().startswith(typed_upper)]
            combobox['values'] = filtered
            
            return filtered
        
        def on_keypress(event):
            key = event.keysym
            
            if key in ('Return', 'KP_Enter'):
                values = combobox['values']
                if values and len(values) > 0:
                    combobox.set(values[0])
                    if on_select_callback:
                        combobox.after(10, on_select_callback)
                return "break"
            
            if key == 'Tab':
                values = combobox['values']
                if values and len(values) > 0 and combobox.get():
                    combobox.set(values[0])
                    if on_select_callback:
                        combobox.after(10, on_select_callback)
        
        def on_keyrelease(event):
            key = event.keysym
            
            if key in ('Up', 'Down', 'Left', 'Right', 'Shift', 'Control', 'Alt', 
                      'Return', 'KP_Enter', 'Tab', 'Escape'):
                return
            
            filtered = filter_and_update()
            
            if len(filtered) == 1:
                typed = combobox.get()
                match = filtered[0]
                if typed.upper() != match.upper():
                    cursor_pos = combobox.index(tk.INSERT)
                    combobox.set(match)
                    combobox.icursor(cursor_pos)
                    combobox.selection_range(cursor_pos, tk.END)
                    if on_select_callback:
                        combobox.after(150, on_select_callback)
        
        def on_focusin(event):
            all_values = getattr(self, all_values_var_name, [])
            combobox['values'] = all_values
        
        def on_click(event):
            all_values = getattr(self, all_values_var_name, [])
            combobox['values'] = all_values
        
        def on_select(event):
            if on_select_callback:
                combobox.after(10, on_select_callback)
        
        combobox.bind('<KeyPress>', on_keypress)
        combobox.bind('<KeyRelease>', on_keyrelease)
        combobox.bind('<FocusIn>', on_focusin)
        combobox.bind('<Button-1>', on_click)
        combobox.bind('<<ComboboxSelected>>', on_select)

    def build_planning_form(self):
        self.plan_form_frame = ttk.LabelFrame(self.left_container, text="Saisie / édition d'une ligne de planning")
        self.plan_form_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

        self.form_mode = None
        self.form_existing = None

        self.form_type = tk.StringVar()
        self.form_heure = tk.StringVar()
        self.form_voyage = tk.StringVar()
        self.form_nb_pal = tk.StringVar()
        self.form_sst = tk.StringVar()
        self.form_chauffeur = tk.StringVar()
        self.form_ramasse = tk.StringVar()
        self.form_infos = tk.StringVar()
        self.form_numero = tk.StringVar()
        self.form_sans_sst = tk.BooleanVar(value=False)
        self.form_sans_chauffeur = tk.BooleanVar(value=False)

        row = 0
        ttk.Label(self.plan_form_frame, text="Type :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        rb_liv = ttk.Radiobutton(self.plan_form_frame, text="Livraison", variable=self.form_type, value="LIVRAISON",
                                 command=self.on_form_type_changed)
        rb_ram = ttk.Radiobutton(self.plan_form_frame, text="Ramasse", variable=self.form_type, value="RAMASSE",
                                 command=self.on_form_type_changed)
        rb_liv.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        rb_ram.grid(row=row, column=2, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.plan_form_frame, text="Heure :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.form_heure_cb = ttk.Combobox(self.plan_form_frame, textvariable=self.form_heure,
                                          values=TIME_CHOICES, width=8)
        self.form_heure_cb.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.plan_form_frame, text="Voyage :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.all_voyage_values = [v["code"] for v in self.voyages if v.get("actif", True)]
        self.form_voy_cb = ttk.Combobox(self.plan_form_frame, textvariable=self.form_voyage,
                                        values=self.all_voyage_values, width=15)
        self.form_voy_cb.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        self.setup_combobox_autocomplete(self.form_voy_cb, 'all_voyage_values', self.on_form_voyage_changed)
        self.form_country_label = ttk.Label(self.plan_form_frame, text="", foreground="blue")
        self.form_country_label.grid(row=row, column=3, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.plan_form_frame, text="Ramasse (client / lieu) :").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        self.form_ram_entry = ttk.Entry(self.plan_form_frame, textvariable=self.form_ramasse, width=25)
        self.form_ram_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        row += 1

        ttk.Label(self.plan_form_frame, text="Nb palettes :").grid(row=row, column=0, sticky="w")
        ttk.Entry(self.plan_form_frame, textvariable=self.form_nb_pal, width=5).grid(row=row, column=1, sticky="w")
        row += 1

        ttk.Label(self.plan_form_frame, text="SST :").grid(row=row, column=0, sticky="w")
        self.all_sst_values = self.sst_list.copy()
        self.form_sst_cb = ttk.Combobox(self.plan_form_frame, textvariable=self.form_sst,
                                        values=self.sst_list, width=15)
        self.form_sst_cb.grid(row=row, column=1, sticky="w")
        self.setup_combobox_autocomplete(self.form_sst_cb, 'all_sst_values', lambda: self.on_form_sst_changed())
        self.form_sans_sst_cb = ttk.Checkbutton(self.plan_form_frame, text="Sans SST",
                                                 variable=self.form_sans_sst,
                                                 command=self.on_form_sans_sst_changed)
        self.form_sans_sst_cb.grid(row=row, column=2, sticky="w", padx=5)
        row += 1

        ttk.Label(self.plan_form_frame, text="Chauffeur :").grid(row=row, column=0, sticky="w")
        self.all_chauffeur_values = [c.get("nom_affichage", c.get("nom", "")) for c in self.chauffeurs]
        self.form_ch_cb = ttk.Combobox(self.plan_form_frame, textvariable=self.form_chauffeur,
                                       values=self.all_chauffeur_values, width=20)
        self.form_ch_cb.grid(row=row, column=1, sticky="w")
        self.setup_combobox_autocomplete(self.form_ch_cb, 'all_chauffeur_values', self.on_form_chauffeur_changed)
        self.form_sans_chauffeur_cb = ttk.Checkbutton(self.plan_form_frame, text="Sans Chauffeur",
                                                       variable=self.form_sans_chauffeur,
                                                       command=self.on_form_sans_chauffeur_changed)
        self.form_sans_chauffeur_cb.grid(row=row, column=2, sticky="w", padx=5)
        row += 1

        ttk.Label(self.plan_form_frame, text="Numéro tournée :").grid(row=row, column=0, sticky="w")
        self.form_num_label = ttk.Label(self.plan_form_frame, textvariable=self.form_numero)
        self.form_num_label.grid(row=row, column=1, sticky="w")
        row += 1

        ttk.Label(self.plan_form_frame, text="Informations :").grid(row=row, column=0, sticky="w")
        ttk.Entry(self.plan_form_frame, textvariable=self.form_infos, width=30).grid(
            row=row, column=1, columnspan=2, sticky="w"
        )
        row += 1

        btn_frame = ttk.Frame(self.plan_form_frame)
        btn_frame.grid(row=row, column=0, columnspan=4, pady=5, sticky="e")
        ttk.Button(btn_frame, text="💾 Enregistrer", command=self.on_form_save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="❌ Annuler", command=self.on_form_cancel).pack(side="left")
        
        for widget in [self.form_heure_cb, self.form_voy_cb, self.form_ram_entry,
                      self.form_sst_cb, self.form_ch_cb]:
            widget.bind("<FocusIn>", lambda e: self.set_user_editing(True))
            widget.bind("<FocusOut>", lambda e: self.set_user_editing(False))

    def set_user_editing(self, editing):
        self.user_editing = editing
        self.is_editing = editing  # Pour le système de refresh automatique

    def on_form_voyage_changed(self, event=None):
        voyage_code = self.form_voyage.get()
        if voyage_code:
            voyage = next((v for v in self.voyages if v.get("code") == voyage_code), None)
            if voyage:
                country = voyage.get("country", "Belgique")
                if country != "Belgique":
                    flag = self.get_country_flag(country)
                    self.form_country_label.config(text=f"→ {flag} {country}")
                else:
                    self.form_country_label.config(text="")
            else:
                self.form_country_label.config(text="")
        else:
            self.form_country_label.config(text="")

    def hide_planning_form(self):
        self.plan_form_frame.grid_forget()
        self.form_mode = None
        self.form_existing = None
        self.set_user_editing(False)

    def show_planning_form(self, mode="add", existing=None):
        self.form_mode = mode
        self.form_existing = existing
        self.set_user_editing(True)

        if mode == "add":
            self.form_type.set("LIVRAISON")
            self.form_heure.set(TIME_CHOICES[0])
            self.form_voyage.set("")
            self.form_ramasse.set("")
            self.form_nb_pal.set("0")
            self.form_sst.set(self.sst_list[0] if self.sst_list else "")
            self.form_chauffeur.set("")
            self.form_infos.set("")
            self.form_numero.set("")
            self.form_sans_sst.set(False)
            self.form_sans_chauffeur.set(False)
            # Réactiver les combobox
            self.form_sst_cb.configure(state="normal")
            self.form_ch_cb.configure(state="normal")
        else:
            self.form_type.set(existing.get("type", "LIVRAISON"))
            self.form_heure.set(existing.get("heure", TIME_CHOICES[0]))
            self.form_voyage.set(existing.get("voyage", ""))
            self.form_ramasse.set(existing.get("ramasse", ""))
            self.form_nb_pal.set(str(existing.get("nb_pal", 0)))

            # Gérer les flags sans_sst et sans_chauffeur
            sans_sst = existing.get("sans_sst", False)
            sans_chauffeur = existing.get("sans_chauffeur", False)
            self.form_sans_sst.set(sans_sst)
            self.form_sans_chauffeur.set(sans_chauffeur)

            if sans_sst:
                self.form_sst.set("")
                self.form_sst_cb.configure(state="disabled")
            else:
                self.form_sst.set(existing.get("sst", self.sst_list[0] if self.sst_list else ""))
                self.form_sst_cb.configure(state="normal")

            if sans_chauffeur:
                self.form_chauffeur.set("")
                self.form_ch_cb.configure(state="disabled")
            else:
                self.form_chauffeur.set(existing.get("chauffeur_nom", ""))
                self.form_ch_cb.configure(state="normal")

            self.form_infos.set(existing.get("infos", ""))
            self.form_numero.set(str(existing.get("numero", "")))

        self.on_form_type_changed()
        self.on_form_voyage_changed()
        if not self.form_sans_sst.get():
            self.on_form_sst_changed(update_numero=(mode == "add"))
        if mode == "add" and not self.form_sans_chauffeur.get():
            self.update_form_numero()

        self.plan_form_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

    def on_form_type_changed(self):
        type_ = self.form_type.get()
        if type_ == "LIVRAISON":
            codes = [v["code"] for v in self.voyages
                     if v.get("actif", True) and v.get("type", "LIVRAISON") == "LIVRAISON"]
            self.all_voyage_values = codes
            self.form_voy_cb["values"] = codes
            if self.form_voyage.get() not in codes:
                self.form_voyage.set(codes[0] if codes else "")
            self.form_ram_entry.configure(state="disabled")
            self.form_ramasse.set("")
        else:
            codes = [v["code"] for v in self.voyages
                     if v.get("actif", True) and v.get("type", "LIVRAISON") == "RAMASSE"]
            self.all_voyage_values = codes
            self.form_voy_cb["values"] = codes
            if self.form_voyage.get() not in codes:
                self.form_voyage.set(codes[0] if codes else "")
            self.form_ram_entry.configure(state="normal")
        self.on_form_voyage_changed()

    def on_form_sst_changed(self, event=None, update_numero=True):
        sst = self.form_sst.get()
        dispo = self.get_chauffeurs_disponibles_for_date(self.current_date, sst)
        names = [c.get("nom_affichage", c.get("nom", "")) for c in dispo]
        self.form_ch_cb["values"] = names
        self.all_chauffeur_values = names
        if self.form_chauffeur.get() not in names:
            if names:
                self.form_chauffeur.set(names[0])
            else:
                self.form_chauffeur.set("")
        if update_numero:
            self.update_form_numero()

    def on_form_chauffeur_changed(self, event=None):
        self.update_form_numero()

    def on_form_sans_sst_changed(self):
        """Handler quand la case 'Sans SST' est cochée/décochée"""
        if self.form_sans_sst.get():
            self.form_sst.set("")
            self.form_sst_cb.configure(state="disabled")
            # Si Sans SST est coché, on coche aussi Sans Chauffeur automatiquement
            self.form_sans_chauffeur.set(True)
            self.on_form_sans_chauffeur_changed()
        else:
            self.form_sst_cb.configure(state="normal")
            if self.sst_list:
                self.form_sst.set(self.sst_list[0])
            self.on_form_sst_changed()

    def on_form_sans_chauffeur_changed(self):
        """Handler quand la case 'Sans Chauffeur' est cochée/décochée"""
        if self.form_sans_chauffeur.get():
            self.form_chauffeur.set("")
            self.form_ch_cb.configure(state="disabled")
            self.form_numero.set("")
        else:
            self.form_ch_cb.configure(state="normal")
            self.on_form_sst_changed(update_numero=True)

    def update_form_numero(self):
        if self.form_mode != "add":
            return
        nom = self.form_chauffeur.get().strip()
        if not nom:
            self.form_numero.set("")
            return
        max_num = 0
        for m in self.missions:
            if m.get("chauffeur_nom", "").strip() == nom:
                try:
                    n = int(m.get("numero", 0))
                except Exception:
                    n = 0
                if n > max_num:
                    max_num = n
        self.form_numero.set(str(max_num + 1 if max_num >= 0 else 1))
    
    def check_delivery_duplicate(self, voyage_code, new_heure):
        try:
            h, m = map(int, new_heure.split(":"))
            new_time_minutes = h * 60 + m
            
            for mission in self.missions:
                if self.form_mode == "edit" and self.form_existing:
                    if mission.get("id") == self.form_existing.get("id"):
                        continue
                
                if mission.get("type") == "LIVRAISON" and mission.get("voyage") == voyage_code:
                    mission_heure = mission.get("heure", "00:00")
                    mh, mm = map(int, mission_heure.split(":"))
                    mission_time_minutes = mh * 60 + mm
                    
                    diff_minutes = abs(new_time_minutes - mission_time_minutes)
                    
                    if diff_minutes < 120:
                        hours = diff_minutes // 60
                        minutes = diff_minutes % 60
                        
                        chauffeur = mission.get("chauffeur_nom", "Non assigné")
                        
                        if diff_minutes == 0:
                            return f"⚠️ Une livraison '{voyage_code}' existe déjà à {mission_heure} (Chauffeur: {chauffeur}).\n\nÉcart: MÊME HEURE"
                        elif hours == 0:
                            return f"⚠️ Une livraison '{voyage_code}' existe déjà à {mission_heure} (Chauffeur: {chauffeur}).\n\nÉcart: {minutes} minute(s) seulement"
                        else:
                            return f"⚠️ Une livraison '{voyage_code}' existe déjà à {mission_heure} (Chauffeur: {chauffeur}).\n\nÉcart: {hours}h{minutes:02d} (moins de 2h)"
            
            return None
        except Exception as e:
            return None

    def on_form_save(self):
        type_ = self.form_type.get()
        voy = self.form_voyage.get().strip()
        if not voy:
            messagebox.showerror("Erreur", "Code voyage obligatoire.")
            return

        # Vérifier que le voyage existe dans la base de données
        voyage_exists = any(v.get("code") == voy for v in self.voyages)
        if not voyage_exists:
            messagebox.showerror("Erreur", f"Le voyage '{voy}' n'existe pas dans la base de données.\n\nVeuillez d'abord l'ajouter dans l'onglet Référentiels > Voyages.")
            return

        try:
            nb_pal = int(self.form_nb_pal.get() or 0)
        except Exception:
            messagebox.showerror("Erreur", "Nb palettes invalide.")
            return

        ram = self.form_ramasse.get().strip() if type_ == "RAMASSE" else ""
        
        if type_ == "LIVRAISON":
            new_heure = self.form_heure.get()
            warning_msg = self.check_delivery_duplicate(voy, new_heure)
            if warning_msg:
                response = messagebox.askokcancel(
                    "⚠️ Attention - Doublon détecté",
                    warning_msg + "\n\nVoulez-vous quand même enregistrer cette mission ?"
                )
                if not response:
                    return

        if self.form_mode == "edit" and self.form_existing:
            mid = self.form_existing["id"]
            numero = self.form_existing.get("numero", 1)
        else:
            mid = str(uuid.uuid4())
            try:
                numero = int(self.form_numero.get() or 1)
            except Exception:
                numero = 1

        # Récupérer l'état des cases "Sans SST" et "Sans Chauffeur"
        sans_sst = self.form_sans_sst.get()
        sans_chauffeur = self.form_sans_chauffeur.get()

        # Vérifier que le SST existe dans la base de données (sauf si case cochée)
        sst = self.form_sst.get()
        if not sans_sst:
            if sst and sst not in self.sst_list:
                messagebox.showerror("Erreur", f"Le sous-traitant '{sst}' n'existe pas dans la base de données.\n\nVeuillez d'abord l'ajouter dans l'onglet Référentiels > Chauffeurs.")
                return
        else:
            sst = ""  # Vider le SST si la case est cochée

        # Vérifier que le chauffeur existe dans la base de données (sauf si case cochée)
        chauffeur_nom = self.form_chauffeur.get()
        chauffeur_id = None
        if not sans_chauffeur:
            if chauffeur_nom:
                for ch in self.chauffeurs:
                    if ch.get("nom_affichage") == chauffeur_nom:
                        chauffeur_id = ch["id"]
                        break
                if chauffeur_id is None:
                    messagebox.showerror("Erreur", f"Le chauffeur '{chauffeur_nom}' n'existe pas dans la base de données.\n\nVeuillez d'abord l'ajouter dans l'onglet Référentiels > Chauffeurs.")
                    return
        else:
            chauffeur_nom = ""  # Vider le chauffeur si la case est cochée
            chauffeur_id = None

        mission = self.form_existing.copy() if (self.form_mode == "edit" and self.form_existing) else {}
        mission.update(
            {
                "id": mid,
                "date": self.current_date.strftime("%Y-%m-%d"),
                "type": type_,
                "heure": self.form_heure.get(),
                "voyage": voy,
                "nb_pal": nb_pal,
                "numero": numero,
                "sst": sst,
                "chauffeur_nom": chauffeur_nom,
                "chauffeur_id": chauffeur_id,
                "ramasse": ram,
                "infos": self.form_infos.get(),
                "sans_sst": sans_sst,
                "sans_chauffeur": sans_chauffeur,
            }
        )

        day_dir = self.ensure_day_dir(self.current_date)
        if day_dir is None:
            return

        if self.form_mode == "edit" and self.form_existing and "_path" in self.form_existing:
            path = Path(self.form_existing["_path"])
        else:
            path = day_dir / f"{mid}.json"

        save_json(path, {k: v for k, v in mission.items() if k != "_path"})
        mission["_path"] = path.as_posix()

        # Invalider le cache pour cette date (le fichier a été modifié)
        planning_cache.force_refresh(self.current_date)

        # Logger l'action Sauron
        if self.form_mode == "edit":
            before_state = {k: v for k, v in self.form_existing.items() if k != "_path"} if self.form_existing else {}
            activity_logger.log_action("MISSION_EDIT", {
                "mission_id": mid,
                "voyage": voy,
                "type": type_,
                "date": self.current_date.strftime("%Y-%m-%d"),
            }, before_state=before_state, after_state={k: v for k, v in mission.items() if k != "_path"})
        else:
            activity_logger.log_action("MISSION_CREATE", {
                "mission_id": mid,
                "voyage": voy,
                "type": type_,
                "date": self.current_date.strftime("%Y-%m-%d"),
                "chauffeur": chauffeur_nom,
                "nb_pal": nb_pal,
            })

        found = False
        for i, m in enumerate(self.missions):
            if m["id"] == mid:
                self.missions[i] = mission
                found = True
                break
        if not found:
            self.missions.append(mission)

        self.refresh_planning_view()
        self.hide_planning_form()

    def on_form_cancel(self):
        self.hide_planning_form()

    def set_today(self):
        self.current_date = date.today()
        self.date_var.set(format_date_display(self.current_date))
        self.load_planning_for_date(self.current_date)

    def navigate_days(self, days):
        from datetime import timedelta
        try:
            current = datetime.strptime(self.date_var.get(), "%d/%m/%Y").date()
            new_date = current + timedelta(days=days)
            self.date_var.set(format_date_display(new_date))
            self.current_date = new_date
            self.load_planning_for_date(new_date)
        except ValueError:
            messagebox.showerror("Erreur", "Format de date invalide.")
    
    def on_load_date(self):
        try:
            d = parse_date_input(self.date_var.get())
        except ValueError as e:
            messagebox.showerror("Erreur", f"Date invalide.\nFormat attendu: JJ/MM/AAAA\nExemple: 25/12/2024")
            return
        self.current_date = d
        self.load_planning_for_date(d)

    def on_open_existing_date(self):
        value = self.existing_dates_var.get()
        if not value:
            return
        try:
            d = parse_date_input(value)
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide.")
            return
        self.date_var.set(format_date_display(d))
        self.current_date = d
        self.load_planning_for_date(d)

    def ensure_day_dir(self, d: date, allow_creation=False):
        day_dir = get_planning_day_dir(d)
        if not day_dir.exists():
            return None
        return day_dir

    def load_planning_for_date(self, d: date, preserve_ui=False, force_source=False):
        """
        Charger le planning pour une date donnée.

        Args:
            d: Date à charger
            preserve_ui: Préserver l'état de l'UI lors du rafraîchissement
            force_source: Forcer le chargement depuis la source (ignorer le cache)
        """
        # Prioritiser les dates adjacentes pour le pré-téléchargement
        planning_cache.prioritize_date(d + timedelta(days=1))
        planning_cache.prioritize_date(d - timedelta(days=1))

        # 1. Essayer d'abord le cache local (rapide)
        if not force_source:
            cached_missions = planning_cache.get_cached_planning(d)
            if cached_missions is not None:
                # Filtrer uniquement les missions valides (avec un id)
                self.missions = [m for m in cached_missions if m and isinstance(m, dict) and "id" in m]
                print(f"[Cache] Planning {d} chargé depuis le cache ({len(self.missions)} missions)")
                # Compléter les chauffeur_id manquants
                for data in self.missions:
                    if "chauffeur_nom" in data and "chauffeur_id" not in data:
                        chauffeur_nom = data["chauffeur_nom"]
                        for ch in self.chauffeurs:
                            if ch.get("nom_affichage") == chauffeur_nom:
                                data["chauffeur_id"] = ch["id"]
                                break
                self.refresh_planning_view(preserve_ui=preserve_ui)
                if hasattr(self, "existing_dates_combo"):
                    self.existing_dates_combo["values"] = list_existing_dates()
                self._update_views_after_planning_load()
                return

        # 2. Charger depuis la source (OneDrive)
        day_dir = self.ensure_day_dir(d)
        if day_dir is None:
            # Vérifier si c'est vraiment inexistant ou juste pas en cache
            if not planning_cache.is_cached(d):
                messagebox.showinfo(
                    "Planning inexistant",
                    f"Le planning pour le {format_date_display(d)} n'existe pas encore.\n\n"
                    "Contactez votre responsable pour générer ce planning."
                )
            self.missions = []
            self.refresh_planning_view(preserve_ui=preserve_ui)
        else:
            # Charger les missions depuis les fichiers JSON
            self.missions = []
            for file in day_dir.glob("*.json"):
                # Ignorer les fichiers de métadonnées (commençant par _)
                if file.name.startswith("_"):
                    continue
                data = load_json(file, None)
                if not data or not isinstance(data, dict) or "id" not in data:
                    continue
                data["_path"] = file.as_posix()

                if "chauffeur_nom" in data and "chauffeur_id" not in data:
                    chauffeur_nom = data["chauffeur_nom"]
                    for ch in self.chauffeurs:
                        if ch.get("nom_affichage") == chauffeur_nom:
                            data["chauffeur_id"] = ch["id"]
                            break

                self.missions.append(data)
            self.refresh_planning_view(preserve_ui=preserve_ui)
            if hasattr(self, "existing_dates_combo"):
                self.existing_dates_combo["values"] = list_existing_dates()

        self._update_views_after_planning_load()

    def _update_views_after_planning_load(self):
        """Mettre à jour les vues liées aux chauffeurs après rechargement du planning"""
        try:
            if hasattr(self, 'drivers_used_frame'):
                self.refresh_drivers_used_view()
            if hasattr(self, 'drivers_available_frame'):
                self.refresh_drivers_availability_view()
            if hasattr(self, 'calendar_container'):
                self.refresh_calendar()
        except Exception as e:
            print(f"Erreur lors de la mise à jour des vues chauffeurs après rechargement du planning: {e}")

    def refresh_planning_view(self, preserve_ui=False):
        if preserve_ui:
            selected_items = {}
            for country, trees in self.country_trees.items():
                for tree_type, tree in trees.items():
                    sel = tree.selection()
                    if sel:
                        selected_items[f"{country}_{tree_type}"] = sel[0]
        
        for country, frame in self.country_frames.items():
            frame.destroy()
        self.country_frames.clear()
        self.country_trees.clear()

        missions_by_country = {}
        v_by_code = {v.get("code"): v for v in self.voyages}
        
        for m in self.missions:
            voyage_code = m.get("voyage", "")
            voyage = v_by_code.get(voyage_code, {})
            country = voyage.get("country", "Belgique")
            
            if country not in missions_by_country:
                missions_by_country[country] = []
            missions_by_country[country].append(m)

        sort_key_functions = {
            "heure": lambda m: self._time_key(m),
            "voyage": lambda m: m.get("voyage", ""),
            "chauffeur": lambda m: m.get("chauffeur_nom", ""),
            "numero": lambda m: int(m.get("numero", 0)) if str(m.get("numero", "")).isdigit() else 0,
            "sst": lambda m: m.get("sst", ""),
            "pays": lambda m: (v_by_code.get(m.get("voyage", ""), {}).get("country", "Belgique"), self._time_key(m))
        }
        
        sort_func = sort_key_functions.get(self.sort_criteria, sort_key_functions["heure"])
        
        for country in missions_by_country:
            missions_by_country[country] = sorted(missions_by_country[country], 
                                                 key=sort_func, 
                                                 reverse=self.sort_reverse)

        sorted_countries = sorted(missions_by_country.keys(), key=lambda x: (x != "Belgique", x))

        for country in sorted_countries:
            self.create_country_planning_section(country)
            
            nb_liv_country = sum(1 for m in missions_by_country[country] if m.get("type") == "LIVRAISON")
            nb_ram_country = sum(1 for m in missions_by_country[country] if m.get("type") == "RAMASSE")
            
            if hasattr(self, 'country_headers') and country in self.country_headers:
                self.country_headers[country]["livraison_label"].config(
                    text=f"📦 LIVRAISONS - {country} ({nb_liv_country})"
                )
                self.country_headers[country]["ramasse_label"].config(
                    text=f"🚛 RAMASSES - {country} ({nb_ram_country})"
                )
            
            tree_liv = self.country_trees[country]["livraison"]
            tree_ram = self.country_trees[country]["ramasse"]
            
            row_num = 0
            for m in missions_by_country[country]:
                # Afficher "N/A" si la mission est marquée sans SST ou sans chauffeur
                sst_display = "N/A" if m.get("sans_sst", False) else m.get("sst", "")
                chauffeur_display = "N/A" if m.get("sans_chauffeur", False) else m.get("chauffeur_nom", "")

                values_common = (
                    m.get("heure", ""),
                    m.get("voyage", ""),
                    m.get("nb_pal", ""),
                    m.get("numero", ""),
                    sst_display,
                    chauffeur_display,
                    m.get("infos", ""),
                )

                # Déterminer le tag: incomplete si voyage incomplet, sinon alternance pair/impair
                is_incomplete = m.get("sans_sst", False) or m.get("sans_chauffeur", False)
                if is_incomplete:
                    tag = 'incomplete'
                else:
                    tag = 'evenrow' if row_num % 2 == 0 else 'oddrow'

                if m.get("type") == "LIVRAISON":
                    tree_liv.insert("", "end", iid=m["id"], values=values_common, tags=(tag,))
                else:
                    values_ram = values_common[:-1] + (m.get("ramasse", ""), m.get("infos", ""))
                    tree_ram.insert("", "end", iid=m["id"], values=values_ram, tags=(tag,))
                row_num += 1
        
        if preserve_ui and selected_items:
            for key, item_id in selected_items.items():
                parts = key.split('_')
                country = '_'.join(parts[:-1])
                tree_type = parts[-1]
                if country in self.country_trees and tree_type in self.country_trees[country]:
                    tree = self.country_trees[country][tree_type]
                    if item_id in tree.get_children(""):
                        tree.selection_set(item_id)
                        tree.see(item_id)
        
        self.refresh_drivers_availability_view()
        self.refresh_drivers_used_view()
        self.update_summary_stats()
    
    def _time_key(self, m):
        try:
            h, mi = map(int, m.get("heure", "00:00").split(":"))
            return h * 60 + mi
        except Exception:
            return 0
    
    def on_search(self):
        search_text = self.search_var.get().strip().upper()
        
        if not search_text:
            self.refresh_planning_view()
            return
        
        for country, trees in self.country_trees.items():
            for tree_type, tree in trees.items():
                all_items = []
                for item_id in tree.get_children():
                    values = tree.item(item_id)["values"]
                    all_items.append((item_id, values))
                
                for item_id in tree.get_children():
                    tree.delete(item_id)
                
                row_num = 0
                for item_id, values in all_items:
                    match = False
                    for value in values:
                        if search_text in str(value).upper():
                            match = True
                            break
                    
                    if match:
                        tag = 'evenrow' if row_num % 2 == 0 else 'oddrow'
                        tree.insert("", "end", iid=item_id, values=values, tags=(tag,))
                        row_num += 1
    
    def clear_search(self):
        self.search_var.set("")
        self.refresh_planning_view()
    
    def refresh_drivers_availability_view(self):
        if not hasattr(self, 'drivers_available_frame'):
            return
        
        for widget in self.drivers_available_frame.winfo_children():
            widget.destroy()
        
        active_drivers = [ch for ch in self.chauffeurs if ch.get("actif", True)]
        
        if not active_drivers:
            tk.Label(self.drivers_available_frame, text="Aucun chauffeur actif", 
                    font=("Arial", 10), fg="gray").pack(pady=10)
            if hasattr(self, 'available_container'):
                self.available_container.config(text="👥 Chauffeurs disponibles (Nb chauff dispo: 0)")
            return
        
        driver_missions = {}
        for ch in active_drivers:
            driver_missions[ch["id"]] = []
        
        for mission in self.missions:
            ch_id = mission.get("chauffeur_id")
            voyage_code = mission.get("voyage", "")
            if ch_id and voyage_code:
                if ch_id in driver_missions:
                    driver_missions[ch_id].append(voyage_code)
            elif not ch_id:
                ch_nom = mission.get("chauffeur_nom", "")
                for ch in active_drivers:
                    if ch.get("nom_affichage") == ch_nom and voyage_code:
                        driver_missions[ch["id"]].append(voyage_code)
                        break
        
        date_str = self.current_date.strftime("%Y-%m-%d")
        dispo_map = {}
        for entry in self.dispos:
            if entry.get("date") == date_str:
                dispo_map[entry["id_chauffeur"]] = bool(entry.get("disponible", True))
        
        available_drivers = []
        for ch in active_drivers:
            ch_id = ch["id"]
            is_available = dispo_map.get(ch_id, True)
            has_missions = len(driver_missions[ch_id]) > 0
            if is_available and not has_missions:
                available_drivers.append(ch)
        
        if not available_drivers:
            tk.Label(self.drivers_available_frame, text="Tous les chauffeurs sont utilisés", 
                    font=("Arial", 10, "italic"), fg="gray").pack(pady=10)
            if hasattr(self, 'available_container'):
                self.available_container.config(text="👥 Chauffeurs disponibles (Nb chauff dispo: 0)")
            return
        
        drivers_by_sst = {}
        for ch in available_drivers:
            sst = ch.get("sst", "Sans SST")
            if sst not in drivers_by_sst:
                drivers_by_sst[sst] = []
            drivers_by_sst[sst].append(ch)
        
        sorted_ssts = sorted(drivers_by_sst.keys())
        
        for sst in sorted_ssts:
            sst_frame = ttk.Frame(self.drivers_available_frame)
            sst_frame.pack(fill="x", pady=(10, 2))
            
            sst_label = tk.Label(sst_frame, text=f"📋 {sst}", 
                               font=("Arial", 11, "bold"), 
                               fg="#2E86DE", anchor="w")
            sst_label.pack(fill="x")
            
            sep = ttk.Separator(self.drivers_available_frame, orient="horizontal")
            sep.pack(fill="x", pady=(0, 5))
            
            sorted_drivers = sorted(drivers_by_sst[sst], key=lambda x: x.get("nom", ""))
            
            for ch in sorted_drivers:
                driver_frame = ttk.Frame(self.drivers_available_frame)
                driver_frame.pack(fill="x", padx=(10, 5), pady=2)
                
                name_text = f"{ch.get('nom', '')} {ch.get('prenom', '')}"
                name_label = tk.Label(driver_frame, text=name_text, 
                                    font=("Arial", 10, "bold"),
                                    bg="#90EE90", fg="black",
                                    padx=5, pady=2, anchor="w")
                name_label.pack(side="left", fill="x", expand=True)
        
        if hasattr(self, 'available_container'):
            self.available_container.config(text=f"👥 Chauffeurs disponibles ({len(available_drivers)})")
        
        self.drivers_available_frame.update_idletasks()
    
    def refresh_drivers_used_view(self):
        if not hasattr(self, 'drivers_used_frame'):
            return
        
        for widget in self.drivers_used_frame.winfo_children():
            widget.destroy()
        
        active_drivers = [ch for ch in self.chauffeurs if ch.get("actif", True)]
        
        if not active_drivers:
            return
        
        driver_missions = {}
        for ch in active_drivers:
            driver_missions[ch["id"]] = []
        
        for mission in self.missions:
            ch_id = mission.get("chauffeur_id")
            voyage_code = mission.get("voyage", "")
            if ch_id and voyage_code:
                if ch_id in driver_missions:
                    driver_missions[ch_id].append(voyage_code)
            elif not ch_id:
                ch_nom = mission.get("chauffeur_nom", "")
                for ch in active_drivers:
                    if ch.get("nom_affichage") == ch_nom and voyage_code:
                        driver_missions[ch["id"]].append(voyage_code)
                        break
        
        date_str = self.current_date.strftime("%Y-%m-%d")
        dispo_map = {}
        for entry in self.dispos:
            if entry.get("date") == date_str:
                dispo_map[entry["id_chauffeur"]] = bool(entry.get("disponible", True))
        
        def get_gray_for_missions(nb_missions):
            if nb_missions == 1:
                return "#D3D3D3"
            elif nb_missions == 2:
                return "#B0B0B0"
            elif nb_missions == 3:
                return "#909090"
            elif nb_missions == 4:
                return "#707070"
            else:
                return "#505050"
        
        used_drivers = []
        for ch in active_drivers:
            ch_id = ch["id"]
            is_available = dispo_map.get(ch_id, True)
            has_missions = len(driver_missions[ch_id]) > 0
            if is_available and has_missions:
                used_drivers.append(ch)
        
        if not used_drivers:
            tk.Label(self.drivers_used_frame, text="Aucun chauffeur utilisé", 
                    font=("Arial", 10, "italic"), fg="gray", bg="#F5F5F5").pack(pady=10)
            if hasattr(self, 'used_container'):
                self.used_container.config(text="🚛 Chauffeurs utilisés (Nb chauff use: 0)")
            return
        
        drivers_by_sst = {}
        for ch in used_drivers:
            sst = ch.get("sst", "Sans SST")
            if sst not in drivers_by_sst:
                drivers_by_sst[sst] = []
            drivers_by_sst[sst].append(ch)
        
        sorted_ssts = sorted(drivers_by_sst.keys())
        
        for sst in sorted_ssts:
            sst_frame = ttk.Frame(self.drivers_used_frame)
            sst_frame.pack(fill="x", pady=(10, 2))
            
            sst_label = tk.Label(sst_frame, text=f"📋 {sst}", 
                               font=("Arial", 11, "bold"), 
                               fg="#666666", anchor="w", bg="#F5F5F5")
            sst_label.pack(fill="x")
            
            sep = ttk.Separator(self.drivers_used_frame, orient="horizontal")
            sep.pack(fill="x", pady=(0, 5))
            
            sorted_drivers = sorted(drivers_by_sst[sst], key=lambda x: x.get("nom", ""))
            
            for ch in sorted_drivers:
                ch_id = ch["id"]
                nb_missions = len(driver_missions[ch_id])
                voyages = driver_missions[ch_id]
                
                driver_frame = ttk.Frame(self.drivers_used_frame)
                driver_frame.pack(fill="x", padx=(10, 5), pady=2)
                
                color = get_gray_for_missions(nb_missions)
                fg_color = "white" if nb_missions >= 4 else "black"
                
                name_text = f"{ch.get('nom', '')} {ch.get('prenom', '')}"
                name_label = tk.Label(driver_frame, text=name_text, 
                                    font=("Arial", 10, "bold"),
                                    bg=color, fg=fg_color,
                                    padx=5, pady=2, anchor="w")
                name_label.pack(side="left", fill="x", expand=False)
                
                count_text = f"({nb_missions})"
                count_label = tk.Label(driver_frame, text=count_text,
                                     font=("Arial", 10),
                                     fg="#555", anchor="w", bg="#F5F5F5")
                count_label.pack(side="left", padx=(5, 5))
                
                if voyages:
                    voyages_text = f"[{', '.join(voyages)}]"
                    voyages_label = tk.Label(driver_frame, text=voyages_text,
                                           font=("Arial", 9),
                                           fg="#777", anchor="w", bg="#F5F5F5",
                                           wraplength=150)
                    voyages_label.pack(side="left", fill="x", expand=True)
        
        if hasattr(self, 'used_container'):
            self.used_container.config(text=f"🚛 Chauffeurs utilisés ({len(used_drivers)})")
        
        self.drivers_used_frame.update_idletasks()

    def on_country_tree_select(self, selected_country, selected_type):
        """Désélectionner les autres pays quand on sélectionne dans un pays"""
        if not hasattr(self, 'country_trees'):
            return
        
        # Désélectionner tous les autres pays et réinitialiser leur style
        for country, trees in self.country_trees.items():
            if country != selected_country:
                # Désélectionner livraison et ramasse de ce pays
                for tree_type, tree in trees.items():
                    tree.selection_remove(tree.selection())
                
                # Réinitialiser le style du frame (bordure normale)
                if hasattr(self, 'country_frames') and country in self.country_frames:
                    frame = self.country_frames[country]
                    # ttk.LabelFrame ne supporte pas highlightthickness → on reste sur relief + borderwidth
                    frame.config(relief="groove", borderwidth=2)
        
        # Mettre en évidence le pays sélectionné (bordure verte épaisse)
        if hasattr(self, 'country_frames') and selected_country in self.country_frames:
            frame = self.country_frames[selected_country]
            # On simule une mise en évidence sans utiliser highlightthickness
            frame.config(relief="solid", borderwidth=3)
    
    def get_selected_mission(self):
        for country, trees in self.country_trees.items():
            for tree_type, tree in trees.items():
                sel = tree.selection()
                if sel:
                    return sel[0]
        return None

    def get_chauffeurs_disponibles_for_date(self, d: date, sst: str):
        dispo_map = {}
        for entry in self.dispos:
            if entry.get("date") == d.strftime("%Y-%m-%d"):
                dispo_map[entry["id_chauffeur"]] = bool(entry.get("disponible", True))

        res = []
        for ch in self.chauffeurs:
            if not ch.get("actif", True):
                continue
            if ch.get("sst") != sst:
                continue
            available = dispo_map.get(ch["id"], True)
            if available:
                res.append(ch)
        return res

    def can_edit_date(self, d: date) -> bool:
        perms = self.rights["permissions"]
        
        if not perms.get("edit_planning", False):
            return False
        
        if perms.get("edit_past_planning_advanced", False):
            return True
        
        today = date.today()
        delta_days = (today - d).days
        
        return delta_days < 2
    
    def check_edit_permission(self, action="modifier") -> bool:
        perms = self.rights["permissions"]
        
        if not perms.get("edit_planning", False):
            messagebox.showwarning(
                "Action non autorisée",
                f"Vous n'avez pas le droit de {action} le planning."
            )
            return False
        
        if not self.can_edit_date(self.current_date):
            today = date.today()
            delta = (today - self.current_date).days
            if delta >= 2:
                messagebox.showwarning(
                    "Modification interdite",
                    f"Vous ne pouvez pas {action} cette date (J-{delta}).\n\n"
                    f"Vous avez besoin du droit 'edit_past_planning_advanced' "
                    f"pour modifier J-2 et antérieur."
                )
            else:
                messagebox.showwarning(
                    "Action non autorisée",
                    f"Vous n'avez pas les droits pour {action} ce planning."
                )
            return False
        
        return True
        
    def on_add_mission(self):
        if not self.check_edit_permission("ajouter une mission"):
            return
        self.show_planning_form(mode="add")

    def on_edit_mission(self):
        if not self.check_edit_permission("modifier une mission"):
            return
        
        mid = self.get_selected_mission()
        if not mid:
            messagebox.showinfo("Info", "Sélectionnez une ligne à modifier.")
            return
        
        mission = None
        for m in self.missions:
            if m["id"] == mid:
                mission = m
                break
        if not mission:
            messagebox.showerror("Erreur", "Mission introuvable.")
            return
        self.show_planning_form(mode="edit", existing=mission)

    def on_delete_mission(self):
        if not self.check_edit_permission("supprimer une mission"):
            return
        
        mid = self.get_selected_mission()
        if not mid:
            messagebox.showinfo("Info", "Sélectionnez une ligne à supprimer.")
            return
        
        mission = None
        for m in self.missions:
            if m["id"] == mid:
                mission = m
                break
        if not mission:
            return
        if not messagebox.askyesno("Confirmation", "Supprimer cette ligne de planning ?"):
            return
        
        # Logger la suppression Sauron
        activity_logger.log_action("MISSION_DELETE", {
            "mission_id": mid,
            "voyage": mission.get("voyage", ""),
            "type": mission.get("type", ""),
            "date": mission.get("date", ""),
            "chauffeur": mission.get("chauffeur_nom", ""),
        }, before_state={k: v for k, v in mission.items() if k != "_path"})
        
        path = mission.get("_path")
        if path and os.path.exists(path):
            os.remove(path)
        self.missions = [m for m in self.missions if m["id"] != mid]

        # Invalider le cache pour cette date (le fichier a été supprimé)
        planning_cache.force_refresh(self.current_date)

        self.refresh_planning_view()

    def build_chauffeurs_tab(self):
        self.tab_ch = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_ch, text="Chauffeurs")

        main_frame = ttk.Frame(self.tab_ch)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # === PANNEAU GAUCHE : Liste des chauffeurs ===
        left_frame = ttk.LabelFrame(main_frame, text="Liste des chauffeurs")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Barre de recherche et filtres
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(search_frame, text="🔍").pack(side="left")
        self.ch_search_var = tk.StringVar()
        self.ch_search_var.trace("w", lambda *args: self.filter_chauffeurs_view())
        search_entry = ttk.Entry(search_frame, textvariable=self.ch_search_var, width=15)
        search_entry.pack(side="left", padx=(5, 10))

        ttk.Label(search_frame, text="Type:").pack(side="left")
        self.ch_filter_type_var = tk.StringVar(value="Tous")
        filter_type_cb = ttk.Combobox(search_frame, textvariable=self.ch_filter_type_var,
                                      values=["Tous", "interne", "externe"], width=10, state="readonly")
        filter_type_cb.pack(side="left", padx=5)
        filter_type_cb.bind("<<ComboboxSelected>>", lambda e: self.filter_chauffeurs_view())

        ttk.Label(search_frame, text="Statut:").pack(side="left", padx=(10, 0))
        self.ch_filter_actif_var = tk.StringVar(value="Actifs")
        filter_actif_cb = ttk.Combobox(search_frame, textvariable=self.ch_filter_actif_var,
                                       values=["Tous", "Actifs", "Inactifs"], width=10, state="readonly")
        filter_actif_cb.pack(side="left", padx=5)
        filter_actif_cb.bind("<<ComboboxSelected>>", lambda e: self.filter_chauffeurs_view())

        ttk.Label(search_frame, text="SST:").pack(side="left", padx=(10, 0))
        self.ch_filter_sst_var = tk.StringVar(value="Tous")
        sst_values = ["Tous"] + (self.sst_list if self.sst_list else [])
        filter_sst_cb = ttk.Combobox(search_frame, textvariable=self.ch_filter_sst_var,
                                     values=sst_values, width=12, state="readonly")
        filter_sst_cb.pack(side="left", padx=5)
        filter_sst_cb.bind("<<ComboboxSelected>>", lambda e: self.filter_chauffeurs_view())

        # Compteur de chauffeurs
        self.ch_count_label = ttk.Label(search_frame, text="", foreground="gray")
        self.ch_count_label.pack(side="right", padx=5)

        # Instruction sélection multiple
        ttk.Label(left_frame, text="💡 Ctrl+clic pour sélection multiple",
                  foreground="gray", font=("Segoe UI", 8)).pack(anchor="w", padx=5)

        # Treeview avec scrollbar
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        cols = ("id", "nom", "prenom", "sst", "type", "actif")
        self.tree_ch = ttk.Treeview(tree_frame, columns=cols, show="headings", height=18, selectmode="extended")

        # En-têtes cliquables pour tri
        self.tree_ch.heading("id", text="ID", command=lambda: self.sort_chauffeurs_view("id"))
        self.tree_ch.heading("nom", text="Nom ▼", command=lambda: self.sort_chauffeurs_view("nom"))
        self.tree_ch.heading("prenom", text="Prénom", command=lambda: self.sort_chauffeurs_view("prenom"))
        self.tree_ch.heading("sst", text="SST", command=lambda: self.sort_chauffeurs_view("sst"))
        self.tree_ch.heading("type", text="Type", command=lambda: self.sort_chauffeurs_view("type"))
        self.tree_ch.heading("actif", text="Actif", command=lambda: self.sort_chauffeurs_view("actif"))

        self.tree_ch.column("id", width=100, minwidth=80)
        self.tree_ch.column("nom", width=120, minwidth=80)
        self.tree_ch.column("prenom", width=100, minwidth=80)
        self.tree_ch.column("sst", width=100, minwidth=80)
        self.tree_ch.column("type", width=80, minwidth=60, anchor="center")
        self.tree_ch.column("actif", width=60, minwidth=50, anchor="center")

        # Scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_ch.yview)
        self.tree_ch.configure(yscrollcommand=vsb.set)

        self.tree_ch.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Variable de tri
        self.ch_sort_column = "nom"
        self.ch_sort_reverse = False

        self.refresh_chauffeurs_view()

        # Boutons d'action
        perms = self.rights["permissions"]
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", padx=5, pady=5)
        if perms["manage_drivers"]:
            ttk.Button(btn_frame, text="➕ Ajouter", command=self.on_add_chauffeur, width=12).pack(side="left", padx=2)
            ttk.Button(btn_frame, text="✏️ Modifier", command=self.on_edit_chauffeur, width=12).pack(side="left", padx=2)
            ttk.Button(btn_frame, text="🏢 Ajouter SST", command=self.on_add_sst, width=14).pack(side="left", padx=2)
        if perms["edit_driver_planning"]:
            ttk.Button(btn_frame, text="🔄 Activer/Désactiver", command=self.on_toggle_chauffeur, width=18).pack(
                side="left", padx=2
            )

        # === PANNEAU DROIT : Planning / disponibilités ===
        right = ttk.LabelFrame(main_frame, text="📅 Planning / disponibilités du chauffeur")
        right.pack(side="left", fill="both", expand=True)

        select_frame = ttk.Frame(right)
        select_frame.pack(fill="x", pady=5)
        ttk.Label(select_frame, text="ID chauffeur :").pack(side="left")
        self.dispo_ch_id_var = tk.StringVar()
        ttk.Entry(select_frame, textvariable=self.dispo_ch_id_var, width=15, 
                 state="readonly").pack(side="left", padx=5)
        
        nav_frame = ttk.Frame(right)
        nav_frame.pack(fill="x", pady=5)
        ttk.Button(nav_frame, text="◀ Mois précédent", command=self.calendar_prev_month).pack(side="left")
        self.calendar_month_var = tk.StringVar()
        ttk.Label(nav_frame, textvariable=self.calendar_month_var, font=("Arial", 10, "bold")).pack(side="left", padx=20)
        ttk.Button(nav_frame, text="Mois suivant ▶", command=self.calendar_next_month).pack(side="left")
        ttk.Button(nav_frame, text="Aujourd'hui", command=self.calendar_goto_today).pack(side="left", padx=(20, 0))
        
        self.calendar_container = ttk.Frame(right, relief="sunken", borderwidth=1)
        self.calendar_container.pack(fill="both", expand=True, pady=5)
        
        self.multi_selection_label = ttk.Label(right, text="", font=("Arial", 9, "bold"), 
                                               foreground="green", wraplength=500)
        self.multi_selection_label.pack(fill="x", pady=(2, 5))
        
        if perms["edit_driver_planning"]:
            btn_row1 = ttk.Frame(right)
            btn_row1.pack(fill="x", pady=2)
            ttk.Button(btn_row1, text="✓ Marquer DISPONIBLE", command=lambda: self.calendar_set_availability(True), 
                      width=20).pack(side="left", padx=2)
            ttk.Button(btn_row1, text="✗ Marquer INDISPONIBLE", command=lambda: self.calendar_set_availability(False),
                      width=20).pack(side="left", padx=2)
            ttk.Button(btn_row1, text="🗑️ Supprimer sélection", command=self.calendar_clear_selection).pack(side="left", padx=2)
        
        if perms["edit_driver_planning"]:
            recur_frame = ttk.LabelFrame(right, text="Récurrence", padding=5)
            recur_frame.pack(fill="x", pady=5)
            
            recur_opts_frame = ttk.Frame(recur_frame)
            recur_opts_frame.pack(fill="x")
            
            ttk.Label(recur_opts_frame, text="Appliquer à :").pack(side="left")
            self.recur_type_var = tk.StringVar(value="none")
            ttk.Radiobutton(recur_opts_frame, text="Aucun", variable=self.recur_type_var, 
                           value="none").pack(side="left", padx=5)
            ttk.Radiobutton(recur_opts_frame, text="Tous les lundis", variable=self.recur_type_var, 
                           value="monday").pack(side="left", padx=5)
            ttk.Radiobutton(recur_opts_frame, text="Tous les samedis", variable=self.recur_type_var, 
                           value="saturday").pack(side="left", padx=5)
            ttk.Radiobutton(recur_opts_frame, text="Tous les dimanches", variable=self.recur_type_var, 
                           value="sunday").pack(side="left", padx=5)
            
            recur_period_frame = ttk.Frame(recur_frame)
            recur_period_frame.pack(fill="x", pady=(5, 0))
            ttk.Label(recur_period_frame, text="Sur les prochains :").pack(side="left")
            self.recur_months_var = tk.StringVar(value="3")
            ttk.Spinbox(recur_period_frame, from_=1, to=12, textvariable=self.recur_months_var, 
                       width=5).pack(side="left", padx=5)
            ttk.Label(recur_period_frame, text="mois").pack(side="left")
            
            ttk.Button(recur_frame, text="Appliquer la récurrence", 
                      command=self.calendar_apply_recurrence).pack(pady=5)
        
        legend_frame = ttk.LabelFrame(right, text="Légende", padding=5)
        legend_frame.pack(fill="x", pady=5)
        
        legend_items = [
            ("Disponible", "#90EE90"),
            ("Indisponible", "#FFB6C1"),
            ("Sélectionné", "#87CEEB"),
            ("Aujourd'hui", "#FFD700")
        ]
        for text, color in legend_items:
            item = ttk.Frame(legend_frame)
            item.pack(side="left", padx=5)
            color_box = tk.Label(item, width=2, bg=color, relief="solid", borderwidth=1)
            color_box.pack(side="left", padx=(0, 3))
            ttk.Label(item, text=text).pack(side="left")
        
        self.calendar_current_year = date.today().year
        self.calendar_current_month = date.today().month
        self.calendar_selected_dates = set()
        self.calendar_buttons = {}

        self.tree_ch.bind("<<TreeviewSelect>>", self.on_select_chauffeur)
        
        self.tree_ch.bind("<Button-1>", lambda e: self.set_user_editing(True))
        self.tree_ch.bind("<ButtonRelease-1>", lambda e: self.root.after(1000, lambda: self.set_user_editing(False)))

    def refresh_chauffeurs_view(self):
        if not hasattr(self, "tree_ch"):
            return

        selected = None
        if hasattr(self, "tree_ch"):
            sel = self.tree_ch.selection()
            if sel:
                selected = sel[0]

        for row in self.tree_ch.get_children():
            self.tree_ch.delete(row)

        # Appliquer les filtres
        filtered = self.get_filtered_chauffeurs()

        # Appliquer le tri
        sort_col = getattr(self, 'ch_sort_column', 'nom')
        reverse = getattr(self, 'ch_sort_reverse', False)

        def sort_key(ch):
            val = ch.get(sort_col, "")
            if sort_col == "actif":
                return 0 if ch.get("actif", True) else 1
            return str(val).lower()

        filtered.sort(key=sort_key, reverse=reverse)

        for ch in filtered:
            self.tree_ch.insert(
                "",
                "end",
                iid=ch["id"],
                values=(
                    ch.get("id", ""),
                    ch.get("nom", ""),
                    ch.get("prenom", ""),
                    ch.get("sst", ""),
                    ch.get("type", ""),
                    "Oui" if ch.get("actif", True) else "Non",
                ),
            )

        if selected and selected in self.tree_ch.get_children(""):
            self.tree_ch.selection_set(selected)
            self.tree_ch.see(selected)

        # Mettre à jour le compteur
        if hasattr(self, 'ch_count_label'):
            total = len(self.chauffeurs)
            shown = len(filtered)
            if total == shown:
                self.ch_count_label.config(text=f"{total} chauffeur(s)")
            else:
                self.ch_count_label.config(text=f"{shown}/{total} chauffeur(s)")

    def get_filtered_chauffeurs(self):
        """Retourne la liste des chauffeurs filtrés selon les critères de recherche."""
        if not hasattr(self, 'ch_search_var'):
            return list(self.chauffeurs)

        search_text = self.ch_search_var.get().strip().lower()
        filter_type = getattr(self, 'ch_filter_type_var', tk.StringVar()).get()
        filter_actif = getattr(self, 'ch_filter_actif_var', tk.StringVar()).get()
        filter_sst = getattr(self, 'ch_filter_sst_var', tk.StringVar()).get()

        filtered = []
        for ch in self.chauffeurs:
            # Filtre par recherche texte (nom, prénom, ID)
            if search_text:
                nom = ch.get("nom", "").lower()
                prenom = ch.get("prenom", "").lower()
                cid = ch.get("id", "").lower()
                if search_text not in nom and search_text not in prenom and search_text not in cid:
                    continue

            # Filtre par type
            if filter_type != "Tous" and ch.get("type", "") != filter_type:
                continue

            # Filtre par statut actif
            if filter_actif == "Actifs" and not ch.get("actif", True):
                continue
            if filter_actif == "Inactifs" and ch.get("actif", True):
                continue

            # Filtre par SST
            if filter_sst != "Tous" and ch.get("sst", "") != filter_sst:
                continue

            filtered.append(ch)

        return filtered

    def filter_chauffeurs_view(self):
        """Rafraîchit la vue avec les filtres appliqués."""
        self.refresh_chauffeurs_view()

    def sort_chauffeurs_view(self, column):
        """Trie la liste des chauffeurs par la colonne spécifiée."""
        if not hasattr(self, 'tree_ch'):
            return

        # Toggle l'ordre si on clique sur la même colonne
        if getattr(self, 'ch_sort_column', None) == column:
            self.ch_sort_reverse = not getattr(self, 'ch_sort_reverse', False)
        else:
            self.ch_sort_column = column
            self.ch_sort_reverse = False

        # Mettre à jour les en-têtes
        cols = {"id": "ID", "nom": "Nom", "prenom": "Prénom", "sst": "SST", "type": "Type", "actif": "Actif"}
        for col, text in cols.items():
            if col == column:
                arrow = "▼" if self.ch_sort_reverse else "▲"
                self.tree_ch.heading(col, text=f"{text} {arrow}")
            else:
                self.tree_ch.heading(col, text=text)

        self.refresh_chauffeurs_view()

    def on_add_chauffeur(self):
        self.set_user_editing(True)
        self.open_chauffeur_dialog()

    def on_edit_chauffeur(self):
        sel = self.tree_ch.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez un chauffeur.")
            return
        cid = sel[0]
        ch = next((c for c in self.chauffeurs if c["id"] == cid), None)
        if not ch:
            return
        self.set_user_editing(True)
        self.open_chauffeur_dialog(existing=ch)

    def on_toggle_chauffeur(self):
        sel = self.tree_ch.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez un chauffeur.")
            return
        cid = sel[0]
        for ch in self.chauffeurs:
            if ch["id"] == cid:
                old_status = ch.get("actif", True)
                ch["actif"] = not old_status
                # Logger l'action Sauron
                activity_logger.log_action("CHAUFFEUR_TOGGLE", {
                    "chauffeur_id": cid,
                    "nom": ch.get("nom", ""),
                    "old_status": "actif" if old_status else "inactif",
                    "new_status": "actif" if ch["actif"] else "inactif",
                })
                break
        save_json(self.data_dir / "chauffeurs.json", self.chauffeurs)
        self.refresh_chauffeurs_view()

    def on_add_sst(self):
        self.set_user_editing(True)
        name = simpledialog.askstring("Nouveau SST", "Nom du sous-traitant / SST :")
        if not name:
            self.set_user_editing(False)
            return
        name = name.strip()
        if not name:
            self.set_user_editing(False)
            return
        if name in self.sst_list:
            messagebox.showinfo("Info", "Ce SST existe déjà.")
            self.set_user_editing(False)
            return
        self.sst_list.append(name)
        save_json(self.data_dir / "sst.json", self.sst_list)
        messagebox.showinfo("Info", f"SST '{name}' ajouté.")
        self.set_user_editing(False)

    def open_chauffeur_dialog(self, existing=None):
        win = tk.Toplevel(self.root)
        win.transient(self.root)
        win.grab_set()
        win.title("Nouveau chauffeur" if not existing else f"Modifier : {existing.get('nom', '')} {existing.get('prenom', '')}")
        win.resizable(False, False)

        # Variables
        id_var = tk.StringVar(value=existing.get("id") if existing else "")
        nom_var = tk.StringVar(value=existing.get("nom") if existing else "")
        prenom_var = tk.StringVar(value=existing.get("prenom") if existing else "")
        sst_var = tk.StringVar(
            value=existing.get("sst") if existing else (self.sst_list[0] if self.sst_list else "")
        )
        type_var = tk.StringVar(value=existing.get("type") if existing else "externe")
        tel_var = tk.StringVar(value=existing.get("telephone") if existing else "")
        actif_var = tk.BooleanVar(value=existing.get("actif", True) if existing else True)

        # Frame principal avec padding
        main_frame = ttk.Frame(win, padding=15)
        main_frame.pack(fill="both", expand=True)

        # === Section: Informations personnelles ===
        info_frame = ttk.LabelFrame(main_frame, text="Informations personnelles", padding=10)
        info_frame.pack(fill="x", pady=(0, 10))

        row = 0
        ttk.Label(info_frame, text="Nom *").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        nom_entry = ttk.Entry(info_frame, textvariable=nom_var, width=25)
        nom_entry.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        nom_entry.focus_set()
        row += 1

        ttk.Label(info_frame, text="Prénom").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(info_frame, textvariable=prenom_var, width=25).grid(row=row, column=1, sticky="w", padx=5, pady=5)
        row += 1

        ttk.Label(info_frame, text="Téléphone").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        ttk.Entry(info_frame, textvariable=tel_var, width=25).grid(row=row, column=1, sticky="w", padx=5, pady=5)

        # === Section: Paramètres professionnels ===
        pro_frame = ttk.LabelFrame(main_frame, text="Paramètres professionnels", padding=10)
        pro_frame.pack(fill="x", pady=(0, 10))

        row = 0
        ttk.Label(pro_frame, text="Type").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        type_frame = ttk.Frame(pro_frame)
        type_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        ttk.Radiobutton(type_frame, text="Interne", variable=type_var, value="interne").pack(side="left")
        ttk.Radiobutton(type_frame, text="Externe", variable=type_var, value="externe").pack(side="left", padx=(15, 0))
        row += 1

        ttk.Label(pro_frame, text="Sous-traitant (SST)").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        sst_cb = ttk.Combobox(pro_frame, textvariable=sst_var, values=self.sst_list, width=22)
        sst_cb.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        row += 1

        ttk.Label(pro_frame, text="Statut").grid(row=row, column=0, sticky="w", padx=5, pady=5)
        actif_frame = ttk.Frame(pro_frame)
        actif_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        ttk.Checkbutton(actif_frame, text="Actif", variable=actif_var).pack(side="left")

        # === Section: Paramètres techniques (collapsée par défaut pour les utilisateurs) ===
        tech_frame = ttk.LabelFrame(main_frame, text="Paramètres techniques", padding=10)
        tech_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(tech_frame, text="ID (auto si vide)").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        id_entry = ttk.Entry(tech_frame, textvariable=id_var, width=18)
        id_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        if existing:
            id_entry.configure(state="readonly")

        # === Boutons d'action ===
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        # Note champs obligatoires
        ttk.Label(btn_frame, text="* Champ obligatoire", foreground="gray", font=("Segoe UI", 8)).pack(side="left")

        def on_ok():
            nom = nom_var.get().strip()
            if not nom:
                messagebox.showerror("Erreur", "Le nom est obligatoire.")
                nom_entry.focus_set()
                return
            cid = id_var.get().strip()
            if not cid:
                cid = "CH_" + uuid.uuid4().hex[:8].upper()
            ch = existing.copy() if existing else {}
            ch.update(
                {
                    "id": cid,
                    "nom": nom,
                    "prenom": prenom_var.get().strip(),
                    "sst": sst_var.get(),
                    "type": type_var.get(),
                    "telephone": tel_var.get().strip(),
                    "actif": bool(actif_var.get()),
                    "nom_affichage": prenom_var.get().strip() or nom,
                }
            )
            found = False
            for i, c in enumerate(self.chauffeurs):
                if c["id"] == cid:
                    # Logger la modification
                    activity_logger.log_action("CHAUFFEUR_EDIT", {
                        "chauffeur_id": cid,
                        "nom": nom,
                    }, before_state=c.copy(), after_state=ch)
                    self.chauffeurs[i] = ch
                    found = True
                    break
            if not found:
                # Logger la création
                activity_logger.log_action("CHAUFFEUR_CREATE", {
                    "chauffeur_id": cid,
                    "nom": nom,
                    "prenom": prenom_var.get().strip(),
                    "sst": sst_var.get(),
                    "type": type_var.get(),
                })
                self.chauffeurs.append(ch)
            save_json(self.data_dir / "chauffeurs.json", self.chauffeurs)
            self.refresh_chauffeurs_view()
            self.set_user_editing(False)
            win.destroy()

        def on_cancel():
            self.set_user_editing(False)
            win.destroy()

        ttk.Button(btn_frame, text="↩ Annuler", command=on_cancel, width=12).pack(side="right", padx=(5, 0))
        ttk.Button(btn_frame, text="💾 Enregistrer", command=on_ok, width=14).pack(side="right")

        # Raccourcis clavier
        win.bind("<Return>", lambda e: on_ok())
        win.bind("<Escape>", lambda e: on_cancel())

    def on_select_chauffeur(self, event):
        sel = self.tree_ch.selection()
        if not sel:
            return
        if len(sel) == 1:
            cid = sel[0]
            self.dispo_ch_id_var.set(cid)
        else:
            self.dispo_ch_id_var.set(", ".join(sel))
        self.refresh_calendar()
    
    def refresh_calendar(self):
        cid_str = self.dispo_ch_id_var.get().strip()
        
        for widget in self.calendar_container.winfo_children():
            widget.destroy()
        self.calendar_buttons.clear()
        
        if not cid_str:
            return
        
        cids = [c.strip() for c in cid_str.split(",")]
        
        cid = cids[0] if len(cids) == 1 else None
        
        months_fr = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                     "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        self.calendar_month_var.set(f"{months_fr[self.calendar_current_month]} {self.calendar_current_year}")
        
        days_header = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
        for i, day_name in enumerate(days_header):
            label = tk.Label(self.calendar_container, text=day_name, font=("Arial", 9, "bold"),
                           bg="#E0E0E0", relief="ridge", borderwidth=1)
            label.grid(row=0, column=i, sticky="nsew", padx=1, pady=1)
        
        first_day = date(self.calendar_current_year, self.calendar_current_month, 1)
        start_weekday = first_day.weekday()
        
        if self.calendar_current_month == 12:
            next_month = date(self.calendar_current_year + 1, 1, 1)
        else:
            next_month = date(self.calendar_current_year, self.calendar_current_month + 1, 1)
        days_in_month = (next_month - first_day).days
        
        dispos_dict = {}
        if cid:
            for dispo in self.dispos:
                if dispo["id_chauffeur"] == cid:
                    try:
                        dispo_date = date.fromisoformat(dispo["date"])
                        if (dispo_date.year == self.calendar_current_year and 
                            dispo_date.month == self.calendar_current_month):
                            dispos_dict[dispo_date.day] = dispo["disponible"]
                    except Exception:
                        pass
        
        current_day = 1
        for week in range(6):
            if current_day > days_in_month:
                break
            
            for weekday in range(7):
                row = week + 1
                col = weekday
                
                if week == 0 and weekday < start_weekday:
                    empty = tk.Label(self.calendar_container, text="", bg="white")
                    empty.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
                    continue
                
                if current_day > days_in_month:
                    empty = tk.Label(self.calendar_container, text="", bg="white")
                    empty.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
                    continue
                
                day_date = date(self.calendar_current_year, self.calendar_current_month, current_day)
                
                bg_color = "white"
                is_today = day_date == date.today()
                is_available = dispos_dict.get(current_day)
                is_selected = day_date in self.calendar_selected_dates
                
                if is_selected:
                    bg_color = "#87CEEB"
                elif is_today:
                    bg_color = "#FFD700"
                elif is_available is True:
                    bg_color = "#90EE90"
                elif is_available is False:
                    bg_color = "#FFB6C1"
                
                btn = tk.Button(self.calendar_container, text=str(current_day),
                              bg=bg_color, relief="raised", borderwidth=1,
                              font=("Arial", 9))
                btn.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
                
                btn.bind("<Button-1>", lambda e, d=day_date: self.calendar_toggle_date(d))
                
                self.calendar_buttons[day_date] = btn
                current_day += 1
        
        for i in range(7):
            self.calendar_container.grid_columnconfigure(i, weight=1, minsize=40)
        for i in range(7):
            self.calendar_container.grid_rowconfigure(i, weight=1, minsize=30)
        
        if hasattr(self, 'multi_selection_label'):
            if len(cids) > 1:
                chauffeurs_names = []
                for ch_id in cids:
                    ch = next((c for c in self.chauffeurs if c["id"] == ch_id), None)
                    if ch:
                        nom = f"{ch.get('nom', '')} {ch.get('prenom', '')}"
                        chauffeurs_names.append(nom)
                
                text = f"✓ {len(cids)} chauffeurs sélectionnés : {', '.join(chauffeurs_names[:5])}"
                if len(chauffeurs_names) > 5:
                    text += f" ... et {len(chauffeurs_names)-5} autre(s)"
                
                self.multi_selection_label.config(text=text, foreground="green")
            else:
                self.multi_selection_label.config(text="")
    
    def calendar_toggle_date(self, day_date):
        if day_date in self.calendar_selected_dates:
            self.calendar_selected_dates.remove(day_date)
        else:
            self.calendar_selected_dates.add(day_date)
        self.refresh_calendar()
    
    def calendar_prev_month(self):
        if self.calendar_current_month == 1:
            self.calendar_current_month = 12
            self.calendar_current_year -= 1
        else:
            self.calendar_current_month -= 1
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
    
    def calendar_next_month(self):
        if self.calendar_current_month == 12:
            self.calendar_current_month = 1
            self.calendar_current_year += 1
        else:
            self.calendar_current_month += 1
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
    
    def calendar_goto_today(self):
        self.calendar_current_month = date.today().month
        self.calendar_current_year = date.today().year
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
    
    def calendar_set_availability(self, available):
        cid_str = self.dispo_ch_id_var.get().strip()
        if not cid_str:
            messagebox.showerror("Erreur", "Sélectionnez un ou plusieurs chauffeurs.")
            return
        
        if not self.calendar_selected_dates:
            messagebox.showinfo("Info", "Aucune date sélectionnée.")
            return
        
        cids = [c.strip() for c in cid_str.split(",")]
        nb_chauffeurs = len(cids)
        nb_dates = len(self.calendar_selected_dates)
        
        if nb_chauffeurs > 1:
            chauffeurs_names = []
            for cid in cids:
                ch = next((c for c in self.chauffeurs if c["id"] == cid), None)
                if ch:
                    nom = f"{ch.get('nom', '')} {ch.get('prenom', '')}"
                    chauffeurs_names.append(nom)
            
            dates_list = sorted(list(self.calendar_selected_dates))
            dates_str = f"{format_date_display(dates_list[0])} au {format_date_display(dates_list[-1])}" if len(dates_list) > 1 else format_date_display(dates_list[0])
            
            confirmation_msg = (
                f"Vous allez modifier {nb_chauffeurs} chauffeur(s) :\n\n"
                f"{chr(10).join(['• ' + name for name in chauffeurs_names[:10]])}\n"
                f"{'... et ' + str(len(chauffeurs_names)-10) + ' autre(s)' + chr(10) if len(chauffeurs_names) > 10 else ''}\n"
                f"Pour {nb_dates} jour(s) : {dates_str}\n"
                f"Statut : {'DISPONIBLE' if available else 'INDISPONIBLE'}\n\n"
                f"Confirmez-vous ?"
            )
            
            if not messagebox.askyesno("Confirmation", confirmation_msg):
                return
        
        for cid in cids:
            self.dispos = [
                d for d in self.dispos
                if not (d["id_chauffeur"] == cid and 
                       date.fromisoformat(d["date"]) in self.calendar_selected_dates)
            ]
            
            for day_date in self.calendar_selected_dates:
                self.dispos.append({
                    "id_chauffeur": cid,
                    "date": day_date.strftime("%Y-%m-%d"),
                    "disponible": available
                })
        
        save_json(self.data_dir / "dispo_chauffeurs.json", self.dispos)
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
        
        if nb_chauffeurs > 1:
            messagebox.showinfo("Succès", 
                f"✓ {nb_chauffeurs} chauffeur(s) modifié(s)\n"
                f"✓ {nb_dates} jour(s) par chauffeur\n"
                f"✓ Total : {nb_chauffeurs * nb_dates} entrées\n"
                f"✓ Statut : {'DISPONIBLE' if available else 'INDISPONIBLE'}")
        else:
            messagebox.showinfo("Info", f"{nb_dates} jour(s) marqué(s) comme {'disponible' if available else 'indisponible'}(s).")
    
    def calendar_clear_selection(self):
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
    
    def calendar_apply_recurrence(self):
        cid_str = self.dispo_ch_id_var.get().strip()
        if not cid_str:
            messagebox.showerror("Erreur", "Sélectionnez un ou plusieurs chauffeurs.")
            return
        
        if not self.calendar_selected_dates:
            messagebox.showinfo("Info", "Sélectionnez au moins une date de référence.")
            return
        
        recur_type = self.recur_type_var.get()
        if recur_type == "none":
            messagebox.showinfo("Info", "Veuillez sélectionner un type de récurrence.")
            return
        
        try:
            nb_months = int(self.recur_months_var.get())
        except Exception:
            nb_months = 3
        
        weekday_target = None
        if recur_type == "monday":
            weekday_target = 0
        elif recur_type == "saturday":
            weekday_target = 5
        elif recur_type == "sunday":
            weekday_target = 6
        
        if weekday_target is None:
            return
        
        available = messagebox.askyesno("Récurrence", 
                                       f"Marquer tous les jours comme DISPONIBLES ?\n"
                                       f"(Non = marquer comme indisponibles)")
        
        start_date = min(self.calendar_selected_dates)
        end_date = start_date + timedelta(days=30 * nb_months)
        
        current = start_date
        recurrence_dates = []
        while current <= end_date:
            if current.weekday() == weekday_target:
                recurrence_dates.append(current)
            current += timedelta(days=1)
        
        cids = [c.strip() for c in cid_str.split(",")]
        
        for cid in cids:
            self.dispos = [
                d for d in self.dispos
                if not (d["id_chauffeur"] == cid and 
                       date.fromisoformat(d["date"]) in recurrence_dates)
            ]
            
            for day_date in recurrence_dates:
                self.dispos.append({
                    "id_chauffeur": cid,
                    "date": day_date.strftime("%Y-%m-%d"),
                    "disponible": available
                })
        
        save_json(self.data_dir / "dispo_chauffeurs.json", self.dispos)
        self.calendar_selected_dates.clear()
        self.refresh_calendar()
        
        nb_chauffeurs = len(cids)
        if nb_chauffeurs > 1:
            messagebox.showinfo("Succès", 
                f"✓ {nb_chauffeurs} chauffeur(s) modifié(s)\n"
                f"✓ {len(recurrence_dates)} jour(s) par chauffeur\n"
                f"✓ Total : {nb_chauffeurs * len(recurrence_dates)} entrées")
        else:
            messagebox.showinfo("Info", f"Récurrence appliquée : {len(recurrence_dates)} jour(s) modifié(s).")

    def build_admin_tab(self):
        self.tab_admin = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_admin, text="⚙️ Admin")
        
        main = ttk.Frame(self.tab_admin, padding=20)
        main.pack(fill="both", expand=True)
        
        title = ttk.Label(main, text="⚙️ Administration - Génération de plannings", 
                         font=("Arial", 14, "bold"))
        title.pack(pady=(0, 20))
        
        desc = ttk.Label(main, text="Cet onglet permet de créer des plannings pour des jours/semaines/mois.\n"
                                   "Seuls les administrateurs ont accès à cette fonctionnalité.",
                        font=("Arial", 10))
        desc.pack(pady=(0, 20))
        
        day_frame = ttk.LabelFrame(main, text="📅 Générer un jour spécifique", padding=15)
        day_frame.pack(fill="x", pady=(0, 15))
        
        day_row = ttk.Frame(day_frame)
        day_row.pack(fill="x")
        
        ttk.Label(day_row, text="Date (JJ/MM/AAAA) :").pack(side="left", padx=(0, 5))
        self.admin_day_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(day_row, textvariable=self.admin_day_var, width=12).pack(side="left", padx=5)
        ttk.Button(day_row, text="Générer ce jour", command=self.admin_generate_day).pack(side="left", padx=5)
        
        week_frame = ttk.LabelFrame(main, text="📆 Générer une semaine", padding=15)
        week_frame.pack(fill="x", pady=(0, 15))
        
        week_desc = ttk.Label(week_frame, text="Génère tous les jours d'une semaine (du lundi au dimanche)", 
                             font=("Arial", 9), foreground="gray")
        week_desc.pack(pady=(0, 10))
        
        days_frame = ttk.LabelFrame(week_frame, text="Jours à générer", padding=5)
        days_frame.pack(fill="x", pady=(0, 10))
        
        self.week_days_vars = {}
        days_names = [("Lundi", "monday"), ("Mardi", "tuesday"), ("Mercredi", "wednesday"),
                      ("Jeudi", "thursday"), ("Vendredi", "friday"), ("Samedi", "saturday"), ("Dimanche", "sunday")]
        
        days_grid = ttk.Frame(days_frame)
        days_grid.pack()
        
        for idx, (day_name, day_key) in enumerate(days_names):
            var = tk.BooleanVar(value=True)
            self.week_days_vars[day_key] = var
            cb = ttk.Checkbutton(days_grid, text=day_name, variable=var)
            cb.grid(row=0, column=idx, padx=5, sticky="w")
        
        quick_buttons = ttk.Frame(days_frame)
        quick_buttons.pack(pady=(5, 0))
        
        def select_all_week():
            for var in self.week_days_vars.values():
                var.set(True)
        
        def deselect_all_week():
            for var in self.week_days_vars.values():
                var.set(False)
        
        def select_workdays_week():
            days_keys = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
            for idx, key in enumerate(days_keys):
                self.week_days_vars[key].set(idx < 6)
        
        ttk.Button(quick_buttons, text="Tout sélectionner", command=select_all_week).pack(side="left", padx=2)
        ttk.Button(quick_buttons, text="Tout désélectionner", command=deselect_all_week).pack(side="left", padx=2)
        ttk.Button(quick_buttons, text="Lun-Sam", command=select_workdays_week).pack(side="left", padx=2)
        
        week_row = ttk.Frame(week_frame)
        week_row.pack(fill="x")
        
        ttk.Label(week_row, text="Semaine de l'année :").pack(side="left", padx=(0, 5))
        self.admin_week_var = tk.StringVar(value=str(date.today().isocalendar()[1]))
        ttk.Spinbox(week_row, from_=1, to=53, textvariable=self.admin_week_var, width=5).pack(side="left", padx=5)
        
        ttk.Label(week_row, text="Année :").pack(side="left", padx=(10, 5))
        self.admin_week_year_var = tk.StringVar(value=str(date.today().year))
        ttk.Spinbox(week_row, from_=2020, to=2030, textvariable=self.admin_week_year_var, width=8).pack(side="left", padx=5)
        
        ttk.Button(week_row, text="Générer la semaine", command=self.admin_generate_week).pack(side="left", padx=5)
        
        month_frame = ttk.LabelFrame(main, text="📋 Générer un mois complet", padding=15)
        month_frame.pack(fill="x", pady=(0, 15))
        
        month_desc = ttk.Label(month_frame, text="Génère tous les jours d'un mois", 
                              font=("Arial", 9), foreground="gray")
        month_desc.pack(pady=(0, 10))
        
        days_frame_month = ttk.LabelFrame(month_frame, text="Jours à générer", padding=5)
        days_frame_month.pack(fill="x", pady=(0, 10))
        
        self.month_days_vars = {}
        days_names_month = [("Lundi", "monday"), ("Mardi", "tuesday"), ("Mercredi", "wednesday"),
                            ("Jeudi", "thursday"), ("Vendredi", "friday"), ("Samedi", "saturday"), ("Dimanche", "sunday")]
        
        days_grid_month = ttk.Frame(days_frame_month)
        days_grid_month.pack()
        
        for idx, (day_name, day_key) in enumerate(days_names_month):
            var = tk.BooleanVar(value=True)
            self.month_days_vars[day_key] = var
            cb = ttk.Checkbutton(days_grid_month, text=day_name, variable=var)
            cb.grid(row=0, column=idx, padx=5, sticky="w")
        
        quick_buttons_month = ttk.Frame(days_frame_month)
        quick_buttons_month.pack(pady=(5, 0))
        
        def select_all_month():
            for var in self.month_days_vars.values():
                var.set(True)
        
        def deselect_all_month():
            for var in self.month_days_vars.values():
                var.set(False)
        
        def select_workdays_month():
            days_keys = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
            for idx, key in enumerate(days_keys):
                self.month_days_vars[key].set(idx < 6)
        
        ttk.Button(quick_buttons_month, text="Tout sélectionner", command=select_all_month).pack(side="left", padx=2)
        ttk.Button(quick_buttons_month, text="Tout désélectionner", command=deselect_all_month).pack(side="left", padx=2)
        ttk.Button(quick_buttons_month, text="Lun-Sam", command=select_workdays_month).pack(side="left", padx=2)
        
        month_row = ttk.Frame(month_frame)
        month_row.pack(fill="x")
        
        ttk.Label(month_row, text="Mois :").pack(side="left", padx=(0, 5))
        months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                 "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        self.admin_month_var = tk.StringVar(value=months[date.today().month - 1])
        ttk.Combobox(month_row, textvariable=self.admin_month_var, values=months, 
                    width=12, state="readonly").pack(side="left", padx=5)
        
        ttk.Label(month_row, text="Année :").pack(side="left", padx=(10, 5))
        self.admin_month_year_var = tk.StringVar(value=str(date.today().year))
        ttk.Spinbox(month_row, from_=2020, to=2030, textvariable=self.admin_month_year_var, width=8).pack(side="left", padx=5)
        
        ttk.Button(month_row, text="Générer le mois", command=self.admin_generate_month).pack(side="left", padx=5)
        
        info_frame = ttk.LabelFrame(main, text="ℹ️ Informations", padding=15)
        info_frame.pack(fill="x", pady=(15, 0))
        
        info_text = (
            "• La génération crée les dossiers nécessaires pour les plannings.\n"
            "• Si un jour existe déjà, il ne sera pas recréé.\n"
            "• Les plannings générés sont vides au départ.\n"
            "• Les utilisateurs avec le droit 'view_planning' peuvent les consulter.\n"
            "• Les utilisateurs avec le droit 'edit_planning' peuvent les modifier."
        )
        ttk.Label(info_frame, text=info_text, font=("Arial", 9), justify="left").pack()
    
    def admin_generate_day(self):
        try:
            date_str = self.admin_day_var.get()
            d = datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError:
            messagebox.showerror("Erreur", "Format de date invalide. Utilisez JJ/MM/AAAA (ex: 25/12/2024)")
            return
        
        day_dir = get_planning_day_dir(d)
        if day_dir.exists():
            messagebox.showinfo("Info", f"Le planning pour le {format_date_display(d)} existe déjà.")
            return
        
        if messagebox.askyesno("Confirmer", f"Créer le planning pour le {format_date_display(d)} ?"):
            day_dir.mkdir(parents=True, exist_ok=True)
            messagebox.showinfo("Succès", f"Planning créé pour le {format_date_display(d)}")
    
    def admin_generate_week(self):
        try:
            week_num = int(self.admin_week_var.get())
            year = int(self.admin_week_year_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Numéro de semaine ou année invalide.")
            return
        
        if not 1 <= week_num <= 53:
            messagebox.showerror("Erreur", "Le numéro de semaine doit être entre 1 et 53.")
            return
        
        selected_days = [key for key, var in self.week_days_vars.items() if var.get()]
        if not selected_days:
            messagebox.showerror("Erreur", "Veuillez sélectionner au moins un jour.")
            return
        
        jan_4 = date(year, 1, 4)
        week_1_monday = jan_4 - timedelta(days=jan_4.weekday())
        target_monday = week_1_monday + timedelta(weeks=week_num - 1)
        
        day_mapping = {"monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
                       "friday": 4, "saturday": 5, "sunday": 6}
        
        days_to_create = []
        for day_key in selected_days:
            day_offset = day_mapping[day_key]
            day = target_monday + timedelta(days=day_offset)
            days_to_create.append(day)
        
        days_to_create.sort()
        
        days_names_fr = {0: "Lundi", 1: "Mardi", 2: "Mercredi", 3: "Jeudi",
                         4: "Vendredi", 5: "Samedi", 6: "Dimanche"}
        selected_days_str = ", ".join([days_names_fr[day_mapping[k]] for k in selected_days if k in day_mapping])
        
        if not messagebox.askyesno("Confirmer",
                                   f"Créer les plannings pour la semaine {week_num} de {year}\n"
                                   f"Jours sélectionnés : {selected_days_str}\n"
                                   f"({len(days_to_create)} jour(s)) ?"):
            return
        
        created = 0
        skipped = 0
        for d in days_to_create:
            day_dir = get_planning_day_dir(d)
            if day_dir.exists():
                skipped += 1
            else:
                day_dir.mkdir(parents=True, exist_ok=True)
                created += 1
        
        messagebox.showinfo("Succès",
                           f"Semaine {week_num} de {year} générée.\n"
                           f"• {created} jour(s) créé(s)\n"
                           f"• {skipped} jour(s) existait(ent) déjà")
    
    def admin_generate_month(self):
        try:
            months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                     "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            month_name = self.admin_month_var.get()
            month_num = months.index(month_name) + 1
            year = int(self.admin_month_year_var.get())
        except (ValueError, IndexError):
            messagebox.showerror("Erreur", "Mois ou année invalide.")
            return
        
        selected_days = {key for key, var in self.month_days_vars.items() if var.get()}
        if not selected_days:
            messagebox.showerror("Erreur", "Veuillez sélectionner au moins un jour.")
            return
        
        day_mapping = {0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday",
                       4: "friday", 5: "saturday", 6: "sunday"}
        
        first_day = date(year, month_num, 1)
        if month_num == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month_num + 1, 1) - timedelta(days=1)
        
        days_to_create = []
        current = first_day
        while current <= last_day:
            weekday = current.weekday()
            if day_mapping.get(weekday) in selected_days:
                days_to_create.append(current)
            current += timedelta(days=1)
        
        days_names_fr = {"monday": "Lundi", "tuesday": "Mardi", "wednesday": "Mercredi",
                         "thursday": "Jeudi", "friday": "Vendredi", "saturday": "Samedi", "sunday": "Dimanche"}
        selected_days_str = ", ".join([days_names_fr[k] for k in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"] if k in selected_days])
        
        if not messagebox.askyesno("Confirmer",
                                   f"Créer les plannings pour {month_name} {year}\n"
                                   f"Jours sélectionnés : {selected_days_str}\n"
                                   f"({len(days_to_create)} jour(s)) ?"):
            return
        
        created = 0
        skipped = 0
        for d in days_to_create:
            day_dir = get_planning_day_dir(d)
            if day_dir.exists():
                skipped += 1
            else:
                day_dir.mkdir(parents=True, exist_ok=True)
                created += 1
        
        messagebox.showinfo("Succès",
                           f"{month_name} {year} généré.\n"
                           f"• {created} jour(s) créé(s)\n"
                           f"• {skipped} jour(s) existait(ent) déjà")
    
    def build_voyages_tab(self):
        self.tab_voy = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_voy, text="Tournées / Voyages")

        main = ttk.Frame(self.tab_voy)
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # === PANNEAU GAUCHE : Liste des voyages ===
        left_frame = ttk.LabelFrame(main, text="Liste des codes voyages")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # Barre de recherche et filtres
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(search_frame, text="🔍").pack(side="left")
        self.voy_search_var = tk.StringVar()
        self.voy_search_var.trace("w", lambda *args: self.filter_voyages_view())
        search_entry = ttk.Entry(search_frame, textvariable=self.voy_search_var, width=15)
        search_entry.pack(side="left", padx=(5, 10))

        ttk.Label(search_frame, text="Type:").pack(side="left")
        self.voy_filter_type_var = tk.StringVar(value="Tous")
        filter_type_cb = ttk.Combobox(search_frame, textvariable=self.voy_filter_type_var,
                                      values=["Tous", "LIVRAISON", "RAMASSE"], width=12, state="readonly")
        filter_type_cb.pack(side="left", padx=5)
        filter_type_cb.bind("<<ComboboxSelected>>", lambda e: self.filter_voyages_view())

        ttk.Label(search_frame, text="Statut:").pack(side="left", padx=(10, 0))
        self.voy_filter_actif_var = tk.StringVar(value="Tous")
        filter_actif_cb = ttk.Combobox(search_frame, textvariable=self.voy_filter_actif_var,
                                       values=["Tous", "Actifs", "Inactifs"], width=10, state="readonly")
        filter_actif_cb.pack(side="left", padx=5)
        filter_actif_cb.bind("<<ComboboxSelected>>", lambda e: self.filter_voyages_view())

        # Compteur de voyages
        self.voy_count_label = ttk.Label(search_frame, text="", foreground="gray")
        self.voy_count_label.pack(side="right", padx=5)

        # Treeview avec scrollbars
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        cols = ("code", "type", "actif", "country", "duree")
        self.tree_voy = ttk.Treeview(tree_frame, columns=cols, show="headings", height=18, selectmode="browse")

        # En-têtes cliquables pour tri
        self.tree_voy.heading("code", text="Code ▼", command=lambda: self.sort_voyages_view("code"))
        self.tree_voy.heading("type", text="Type", command=lambda: self.sort_voyages_view("type"))
        self.tree_voy.heading("actif", text="Actif", command=lambda: self.sort_voyages_view("actif"))
        self.tree_voy.heading("country", text="Pays", command=lambda: self.sort_voyages_view("country"))
        self.tree_voy.heading("duree", text="Durée", command=lambda: self.sort_voyages_view("duree"))

        self.tree_voy.column("code", width=120, minwidth=80)
        self.tree_voy.column("type", width=100, minwidth=80)
        self.tree_voy.column("actif", width=60, minwidth=50, anchor="center")
        self.tree_voy.column("country", width=100, minwidth=80)
        self.tree_voy.column("duree", width=70, minwidth=50, anchor="center")

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_voy.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_voy.xview)
        self.tree_voy.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree_voy.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        self.tree_voy.bind("<<TreeviewSelect>>", self.on_voy_select)
        self.tree_voy.bind("<Double-1>", self.on_voy_select)

        # Variable de tri
        self.voy_sort_column = "code"
        self.voy_sort_reverse = False

        # === PANNEAU DROIT : Formulaire ===
        right_frame = ttk.Frame(main, width=320)
        right_frame.pack(side="right", fill="y", padx=(10, 0))
        right_frame.pack_propagate(False)

        # Titre dynamique du formulaire
        self.voy_form_title_var = tk.StringVar(value="Nouveau voyage")
        title_label = ttk.Label(right_frame, textvariable=self.voy_form_title_var, font=("Segoe UI", 12, "bold"))
        title_label.pack(pady=(0, 15))

        self.voy_form_mode = None
        self.voy_form_existing = None

        # === Section: Informations de base ===
        info_frame = ttk.LabelFrame(right_frame, text="Informations de base")
        info_frame.pack(fill="x", padx=5, pady=5)

        self.voy_code_var = tk.StringVar()
        self.voy_type_var = tk.StringVar(value="LIVRAISON")
        self.voy_actif_var = tk.BooleanVar(value=True)

        row = 0
        ttk.Label(info_frame, text="Code voyage *").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        self.voy_code_entry = ttk.Entry(info_frame, textvariable=self.voy_code_var, width=22)
        self.voy_code_entry.grid(row=row, column=1, sticky="w", padx=5, pady=3)
        row += 1

        ttk.Label(info_frame, text="Type").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        self.voy_type_cb = ttk.Combobox(info_frame, textvariable=self.voy_type_var,
                                        values=["LIVRAISON", "RAMASSE"], width=19, state="readonly")
        self.voy_type_cb.grid(row=row, column=1, sticky="w", padx=5, pady=3)
        row += 1

        ttk.Label(info_frame, text="Statut").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        actif_frame = ttk.Frame(info_frame)
        actif_frame.grid(row=row, column=1, sticky="w", padx=5, pady=3)
        ttk.Checkbutton(actif_frame, text="Actif", variable=self.voy_actif_var).pack(side="left")

        # === Section: Localisation ===
        loc_frame = ttk.LabelFrame(right_frame, text="Localisation")
        loc_frame.pack(fill="x", padx=5, pady=5)

        self.voy_foreign_var = tk.BooleanVar(value=False)
        self.voy_country_var = tk.StringVar(value="Belgique")

        row = 0
        ttk.Label(loc_frame, text="Destination").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        dest_frame = ttk.Frame(loc_frame)
        dest_frame.grid(row=row, column=1, sticky="w", padx=5, pady=3)
        ttk.Radiobutton(dest_frame, text="Belgique", variable=self.voy_foreign_var,
                        value=False, command=self.on_voy_foreign_toggle).pack(side="left")
        ttk.Radiobutton(dest_frame, text="Étranger", variable=self.voy_foreign_var,
                        value=True, command=self.on_voy_foreign_toggle).pack(side="left", padx=(10, 0))
        row += 1

        ttk.Label(loc_frame, text="Pays").grid(row=row, column=0, sticky="w", padx=5, pady=3)
        self.voy_country_cb = ttk.Combobox(loc_frame, textvariable=self.voy_country_var,
                                           values=[c for c in EU_COUNTRIES if c != "Belgique"],
                                           state="disabled", width=19)
        self.voy_country_cb.grid(row=row, column=1, sticky="w", padx=5, pady=3)

        # === Section: Paramètres ===
        param_frame = ttk.LabelFrame(right_frame, text="Paramètres")
        param_frame.pack(fill="x", padx=5, pady=5)

        self.voy_duree_var = tk.StringVar(value="60")

        ttk.Label(param_frame, text="Durée estimée").grid(row=0, column=0, sticky="w", padx=5, pady=3)
        duree_inner = ttk.Frame(param_frame)
        duree_inner.grid(row=0, column=1, sticky="w", padx=5, pady=3)
        self.voy_duree_entry = ttk.Entry(duree_inner, textvariable=self.voy_duree_var, width=8)
        self.voy_duree_entry.pack(side="left")
        ttk.Label(duree_inner, text="min", foreground="gray").pack(side="left", padx=5)

        # === Boutons d'action ===
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(fill="x", padx=5, pady=15)

        # Boutons principaux
        main_btn_frame = ttk.Frame(btn_frame)
        main_btn_frame.pack(fill="x")

        self.voy_save_btn = ttk.Button(main_btn_frame, text="💾 Enregistrer", command=self.on_voy_save, width=15)
        self.voy_save_btn.pack(side="left", padx=2)

        ttk.Button(main_btn_frame, text="🔄 Nouveau", command=self.on_voy_new, width=12).pack(side="left", padx=2)
        ttk.Button(main_btn_frame, text="↩ Annuler", command=self.on_voy_cancel, width=10).pack(side="left", padx=2)

        # Bouton suppression (séparé pour éviter les clics accidentels)
        del_frame = ttk.Frame(btn_frame)
        del_frame.pack(fill="x", pady=(10, 0))

        self.voy_delete_btn = ttk.Button(del_frame, text="🗑 Supprimer", command=self.on_voy_delete, width=12)
        self.voy_delete_btn.pack(side="right")
        self.voy_delete_btn.configure(state="disabled")

        # Note en bas
        note_label = ttk.Label(right_frame, text="* Champ obligatoire", foreground="gray", font=("Segoe UI", 8))
        note_label.pack(side="bottom", pady=5)

        self.refresh_voyages_view()
        self.on_voy_new()

        for widget in [self.voy_code_var, self.voy_type_var, self.voy_country_cb]:
            if hasattr(widget, 'trace'):
                widget.trace('w', lambda *args: self.set_user_editing(True))

    def refresh_voyages_view(self):
        if not hasattr(self, "tree_voy"):
            return
        sel_before = None
        cur_sel = self.tree_voy.selection()
        if cur_sel:
            sel_before = cur_sel[0]
        for row in self.tree_voy.get_children():
            self.tree_voy.delete(row)

        # Appliquer les filtres
        filtered = self.get_filtered_voyages()

        # Appliquer le tri
        sort_col = getattr(self, 'voy_sort_column', 'code')
        reverse = getattr(self, 'voy_sort_reverse', False)

        def sort_key(v):
            val = v.get(sort_col, "")
            if sort_col == "actif":
                return 0 if v.get("actif", True) else 1
            if sort_col == "duree":
                return v.get("duree", 60)
            return str(val).lower()

        filtered.sort(key=sort_key, reverse=reverse)

        for v in filtered:
            code = v.get("code", "")
            duree = v.get("duree", 60)
            self.tree_voy.insert(
                "",
                "end",
                iid=code,
                values=(
                    code,
                    v.get("type", "LIVRAISON"),
                    "Oui" if v.get("actif", True) else "Non",
                    v.get("country", "Belgique"),
                    f"{duree} min",
                ),
            )
        if sel_before and sel_before in self.tree_voy.get_children(""):
            self.tree_voy.selection_set(sel_before)

        # Mettre à jour le compteur
        if hasattr(self, 'voy_count_label'):
            total = len(self.voyages)
            shown = len(filtered)
            if total == shown:
                self.voy_count_label.config(text=f"{total} voyage(s)")
            else:
                self.voy_count_label.config(text=f"{shown}/{total} voyage(s)")

    def get_filtered_voyages(self):
        """Retourne la liste des voyages filtrés selon les critères de recherche."""
        if not hasattr(self, 'voy_search_var'):
            return list(self.voyages)

        search_text = self.voy_search_var.get().strip().lower()
        filter_type = getattr(self, 'voy_filter_type_var', tk.StringVar()).get()
        filter_actif = getattr(self, 'voy_filter_actif_var', tk.StringVar()).get()

        filtered = []
        for v in self.voyages:
            # Filtre par recherche texte
            if search_text:
                code = v.get("code", "").lower()
                country = v.get("country", "").lower()
                if search_text not in code and search_text not in country:
                    continue

            # Filtre par type
            if filter_type != "Tous" and v.get("type", "LIVRAISON") != filter_type:
                continue

            # Filtre par statut actif
            if filter_actif == "Actifs" and not v.get("actif", True):
                continue
            if filter_actif == "Inactifs" and v.get("actif", True):
                continue

            filtered.append(v)

        return filtered

    def filter_voyages_view(self):
        """Rafraîchit la vue avec les filtres appliqués."""
        self.refresh_voyages_view()

    def sort_voyages_view(self, column):
        """Trie la liste des voyages par la colonne spécifiée."""
        if not hasattr(self, 'tree_voy'):
            return

        # Toggle l'ordre si on clique sur la même colonne
        if getattr(self, 'voy_sort_column', None) == column:
            self.voy_sort_reverse = not getattr(self, 'voy_sort_reverse', False)
        else:
            self.voy_sort_column = column
            self.voy_sort_reverse = False

        # Mettre à jour les en-têtes
        cols = {"code": "Code", "type": "Type", "actif": "Actif", "country": "Pays", "duree": "Durée"}
        for col, text in cols.items():
            if col == column:
                arrow = "▼" if self.voy_sort_reverse else "▲"
                self.tree_voy.heading(col, text=f"{text} {arrow}")
            else:
                self.tree_voy.heading(col, text=text)

        self.refresh_voyages_view()

    def on_voy_delete(self):
        """Supprime le voyage sélectionné après confirmation."""
        if self.voy_form_mode != "edit" or self.voy_form_existing is None:
            messagebox.showwarning("Attention", "Veuillez d'abord sélectionner un voyage à supprimer.")
            return

        code = self.voy_form_existing.get("code", "")
        if not code:
            return

        # Confirmation
        result = messagebox.askyesno(
            "Confirmer la suppression",
            f"Êtes-vous sûr de vouloir supprimer le voyage '{code}' ?\n\n"
            "Cette action est irréversible.",
            icon="warning"
        )

        if not result:
            return

        # Supprimer le voyage
        self.voyages = [v for v in self.voyages if v.get("code") != code]

        # Logger la suppression
        activity_logger.log_action("VOYAGE_DELETE", {
            "code": code,
            "type": self.voy_form_existing.get("type"),
            "country": self.voy_form_existing.get("country"),
        })

        self.save_voyages_data()
        self.refresh_voyages_view()
        self.on_voy_new()
        messagebox.showinfo("Info", f"Le voyage '{code}' a été supprimé.")

    
    def on_voy_foreign_toggle(self):
        if getattr(self, "voy_foreign_var", None) is None:
            return
        if self.voy_foreign_var.get():
            self.voy_country_cb.configure(state="readonly")
            if self.voy_country_var.get() == "Belgique":
                self.voy_country_var.set("France")
        else:
            self.voy_country_cb.configure(state="disabled")
            self.voy_country_var.set("Belgique")
    
    def on_voy_new(self):
        self.voy_form_mode = "add"
        self.voy_form_existing = None

        # Mettre à jour le titre du formulaire
        if hasattr(self, 'voy_form_title_var'):
            self.voy_form_title_var.set("Nouveau voyage")

        # Désactiver le bouton de suppression
        if hasattr(self, 'voy_delete_btn'):
            self.voy_delete_btn.configure(state="disabled")

        if hasattr(self, 'voy_foreign_var'):
            self.voy_foreign_var.set(False)
            self.voy_country_var.set('Belgique')
            try:
                self.voy_country_cb.configure(state='disabled')
            except Exception:
                pass

        self.voy_code_var.set("")
        self.voy_type_var.set("LIVRAISON")
        self.voy_actif_var.set(True)
        if hasattr(self, 'voy_duree_var'):
            self.voy_duree_var.set("60")
        if hasattr(self, "tree_voy"):
            self.tree_voy.selection_remove(*self.tree_voy.selection())

    def on_voy_select(self, event=None):
        if not hasattr(self, "tree_voy"):
            return
        sel = self.tree_voy.selection()
        if not sel:
            return
        code = sel[0]
        v = next((x for x in self.voyages if x.get("code") == code), None)
        if not v:
            return
        self.voy_form_mode = "edit"
        self.voy_form_existing = v

        # Mettre à jour le titre du formulaire
        if hasattr(self, 'voy_form_title_var'):
            self.voy_form_title_var.set(f"Modifier : {code}")

        # Activer le bouton de suppression
        if hasattr(self, 'voy_delete_btn'):
            self.voy_delete_btn.configure(state="normal")

        self.voy_code_var.set(v.get("code", ""))
        self.voy_type_var.set(v.get("type", "LIVRAISON"))
        self.voy_actif_var.set(bool(v.get("actif", True)))
        country = v.get('country', 'Belgique')
        is_foreign = (country != 'Belgique')
        if hasattr(self, 'voy_foreign_var'):
            self.voy_foreign_var.set(is_foreign)
            self.voy_country_var.set(country if is_foreign else 'Belgique')
            try:
                self.voy_country_cb.configure(state='readonly' if is_foreign else 'disabled')
            except Exception:
                pass
        # Charger la durée
        if hasattr(self, 'voy_duree_var'):
            duree = v.get('duree', 60)
            self.voy_duree_var.set(str(duree))

    def on_voy_save(self):
        code = self.voy_code_var.get().strip()
        if not code:
            messagebox.showerror("Erreur", "Le code du voyage est obligatoire.")
            return
        type_ = self.voy_type_var.get() or "LIVRAISON"
        country = 'Belgique'
        try:
            if getattr(self, 'voy_foreign_var', None) and self.voy_foreign_var.get():
                cc = (self.voy_country_var.get() or '').strip()
                if cc:
                    country = cc
        except Exception:
            pass
        actif = bool(self.voy_actif_var.get())

        # Récupérer la durée (en minutes)
        duree = 60  # Valeur par défaut
        try:
            duree_str = self.voy_duree_var.get().strip()
            if duree_str:
                duree = int(duree_str)
                if duree <= 0:
                    duree = 60
        except ValueError:
            duree = 60

        if self.voy_form_mode == "add" or self.voy_form_existing is None:
            if any(v.get("code") == code for v in self.voyages):
                messagebox.showerror("Erreur", "Un voyage avec ce code existe déjà.")
                return
            v = {"code": code, "type": type_, "actif": actif, "country": country, "duree": duree}
            self.voyages.append(v)
            self.voy_form_existing = v
            self.voy_form_mode = "edit"
            # Logger la création
            activity_logger.log_action("VOYAGE_CREATE", {
                "code": code,
                "type": type_,
                "country": country,
                "actif": actif,
                "duree": duree,
            })
        else:
            old_code = self.voy_form_existing.get("code")
            if code != old_code and any(v.get("code") == code for v in self.voyages):
                messagebox.showerror("Erreur", "Un voyage avec ce code existe déjà.")
                return
            # Logger la modification
            before_state = self.voy_form_existing.copy()
            self.voy_form_existing["code"] = code
            self.voy_form_existing["type"] = type_
            self.voy_form_existing["actif"] = actif
            self.voy_form_existing["country"] = country
            self.voy_form_existing["duree"] = duree
            activity_logger.log_action("VOYAGE_EDIT", {
                "code": code,
                "old_code": old_code,
            }, before_state=before_state, after_state=self.voy_form_existing.copy())

        self.save_voyages_data()
        self.refresh_voyages_view()
        if hasattr(self, "tree_voy"):
            self.tree_voy.selection_set(code)
        messagebox.showinfo("Info", "Voyage enregistré.")
        self.set_user_editing(False)

    def on_voy_cancel(self):
        self.on_voy_new()
        self.set_user_editing(False)

    def build_finance_tab(self):
        self.tab_finance = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_finance, text="💰 Finance")
        
        canvas = tk.Canvas(self.tab_finance, bg="white")
        scrollbar = ttk.Scrollbar(self.tab_finance, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        main = ttk.Frame(scrollable_frame, padding=10)
        main.pack(fill="both", expand=True)
        
        ttk.Label(main, text="💰 Gestion Financière", 
                 font=("Arial", 14, "bold")).pack(pady=(0, 10))
        
        # Container principal avec 2 colonnes
        main_container = ttk.Frame(main)
        main_container.pack(fill="both", expand=True)
        
        # Colonne gauche : Tarifs et Revenus
        left_column = ttk.Frame(main_container)
        left_column.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # Colonne droite : Calcul
        right_column = ttk.Frame(main_container)
        right_column.pack(side="left", fill="both", expand=True, padx=(5, 0))
        
        sst_frame = ttk.LabelFrame(left_column, text="💶 Tarifs SST par pays et par jour", padding=10)
        sst_frame.pack(fill="x", pady=(0, 15))
        
        sst_select_frame = ttk.Frame(sst_frame)
        sst_select_frame.pack(fill="x", pady=(0, 10))
        
        # Colonne gauche: Sélection multiple de SST
        left_col = ttk.Frame(sst_select_frame)
        left_col.pack(side="left", fill="both", expand=True)
        
        ttk.Label(left_col, text="SST (sélection multiple) :").pack(anchor="w", pady=(0, 5))
        
        sst_list_frame = ttk.Frame(left_col)
        sst_list_frame.pack(fill="both", expand=True)
        
        sst_scrollbar = ttk.Scrollbar(sst_list_frame, orient="vertical")
        self.finance_sst_listbox = tk.Listbox(sst_list_frame, 
                                              selectmode="multiple",
                                              height=10,
                                              yscrollcommand=sst_scrollbar.set,
                                              exportselection=False)
        sst_scrollbar.config(command=self.finance_sst_listbox.yview)
        
        self.finance_sst_listbox.pack(side="left", fill="both", expand=True)
        sst_scrollbar.pack(side="right", fill="y")
        
        # Binding pour rafraîchir quand la sélection change
        self.finance_sst_listbox.bind("<<ListboxSelect>>", lambda e: self.refresh_sst_tarifs_display())
        
        for sst in self.sst_list:
            self.finance_sst_listbox.insert("end", sst)
        
        # Colonne droite: Pays
        right_col = ttk.Frame(sst_select_frame)
        right_col.pack(side="left", fill="both", padx=(15, 0))
        
        ttk.Label(right_col, text="Pays :").pack(anchor="w")
        self.finance_country_var = tk.StringVar(value="Belgique")
        self.finance_country_cb = ttk.Combobox(right_col, textvariable=self.finance_country_var,
                                               values=EU_COUNTRIES, width=15, state="readonly")
        self.finance_country_cb.pack(pady=5)
        
        tarif_frame = ttk.Frame(sst_frame)
        tarif_frame.pack(fill="x", pady=5)
        
        ttk.Label(tarif_frame, text="Date (JJ/MM/AAAA) :").pack(side="left", padx=(0, 5))
        self.finance_date_var = tk.StringVar(value=format_date_display(date.today()))
        ttk.Entry(tarif_frame, textvariable=self.finance_date_var, width=12).pack(side="left", padx=5)
        
        ttk.Label(tarif_frame, text="Tarif journalier/chauffeur (€) :").pack(side="left", padx=(15, 5))
        self.finance_tarif_var = tk.StringVar(value="0.00")
        ttk.Entry(tarif_frame, textvariable=self.finance_tarif_var, width=10).pack(side="left", padx=5)
        
        perms = self.rights["permissions"]
        if perms.get("manage_finance", False):
            btn_sst_frame = ttk.Frame(sst_frame)
            btn_sst_frame.pack(fill="x", pady=5)
            ttk.Button(btn_sst_frame, text="💾 Enregistrer tarif SST", 
                      command=self.save_sst_tarif).pack(side="left", padx=5)
            ttk.Button(btn_sst_frame, text="📋 Copier sur 1 mois", 
                      command=self.copy_sst_tarif_month).pack(side="left", padx=5)
            ttk.Button(btn_sst_frame, text="🗑️ Supprimer", 
                      command=self.delete_sst_tarif).pack(side="left", padx=5)
        
        self.sst_tarifs_text = tk.Text(sst_frame, height=10, width=60)
        self.sst_tarifs_text.pack(fill="x", pady=5)
        
        palette_frame = ttk.LabelFrame(left_column, text="📦 Revenus par palette, pays et par jour", padding=10)
        palette_frame.pack(fill="x", pady=(0, 15))
        
        rev_date_frame = ttk.Frame(palette_frame)
        rev_date_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(rev_date_frame, text="Date (JJ/MM/AAAA) :").pack(side="left", padx=(0, 5))
        self.rev_date_var = tk.StringVar(value=format_date_display(date.today()))
        ttk.Entry(rev_date_frame, textvariable=self.rev_date_var, width=12).pack(side="left", padx=5)
        
        ttk.Label(rev_date_frame, text="Pays :").pack(side="left", padx=(15, 5))
        self.rev_country_var = tk.StringVar(value="Belgique")
        self.rev_country_cb = ttk.Combobox(rev_date_frame, textvariable=self.rev_country_var,
                                           values=EU_COUNTRIES, width=15, state="readonly")
        self.rev_country_cb.pack(side="left", padx=5)
        self.rev_country_cb.bind("<<ComboboxSelected>>", self.on_rev_country_changed)
        
        rev_values_frame = ttk.Frame(palette_frame)
        rev_values_frame.pack(fill="x", pady=5)
        
        ttk.Label(rev_values_frame, text="Revenu/palette LIVRAISON (€) :").pack(side="left", padx=(0, 5))
        self.rev_livraison_var = tk.StringVar(value="0.00")
        ttk.Entry(rev_values_frame, textvariable=self.rev_livraison_var, width=10).pack(side="left", padx=5)
        
        ttk.Label(rev_values_frame, text="Revenu/palette RAMASSE (€) :").pack(side="left", padx=(15, 5))
        self.rev_ramasse_var = tk.StringVar(value="0.00")
        ttk.Entry(rev_values_frame, textvariable=self.rev_ramasse_var, width=10).pack(side="left", padx=5)
        
        if perms.get("manage_finance", False):
            btn_rev_frame = ttk.Frame(palette_frame)
            btn_rev_frame.pack(fill="x", pady=5)
            ttk.Button(btn_rev_frame, text="💾 Enregistrer revenus", 
                      command=self.save_revenus).pack(side="left", padx=5)
            ttk.Button(btn_rev_frame, text="📋 Copier sur 1 mois", 
                      command=self.copy_revenus_month).pack(side="left", padx=5)
            ttk.Button(btn_rev_frame, text="🗑️ Supprimer", 
                      command=self.delete_revenus).pack(side="left", padx=5)
        
        self.revenus_text = tk.Text(palette_frame, height=10, width=60)
        self.revenus_text.pack(fill="x", pady=5)
        
        calc_frame = ttk.LabelFrame(right_column, text="🧮 Calcul des coûts/revenus pour une date", padding=10)
        calc_frame.pack(fill="both", expand=True)
        
        calc_date_frame = ttk.Frame(calc_frame)
        calc_date_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(calc_date_frame, text="Date à analyser :").pack(side="left", padx=(0, 5))
        self.calc_date_var = tk.StringVar(value=format_date_display(date.today()))
        ttk.Entry(calc_date_frame, textvariable=self.calc_date_var, width=12).pack(side="left", padx=5)
        ttk.Button(calc_date_frame, text="📊 Calculer", 
                  command=self.calculate_finance).pack(side="left", padx=10)
        
        self.calc_result_text = tk.Text(calc_frame, height=30, width=60)
        self.calc_result_text.pack(fill="x")
        
        self.refresh_finance_view()



    def refresh_revenus_display(self):
        """Afficher les revenus définis"""
        if not hasattr(self, 'revenus_text'):
            return
        
        self.revenus_text.delete("1.0", "end")
        
        if not self.revenus_palettes:
            self.revenus_text.insert("1.0", "Aucun revenu défini")
            return
        
        text = "Revenus par palette définis:\n"
        text += "=" * 60 + "\n\n"
        
        for date_str in sorted(self.revenus_palettes.keys(), reverse=True)[:10]:
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d").date()
                date_display = format_date_display(d)
            except Exception:
                date_display = date_str
            
            rev = self.revenus_palettes[date_str]
            
            # Vérifier si c'est le nouveau format (par pays) ou l'ancien format
            if isinstance(rev, dict):
                # Vérifier s'il y a des sous-dictionnaires (nouveau format)
                has_countries = any(isinstance(v, dict) for v in rev.values() if v is not None)
                
                if has_countries:
                    # Nouveau format: {pays: {livraison: x, ramasse: y}}
                    text += f"Date: {date_display}\n"
                    for country, values in sorted(rev.items()):
                        if isinstance(values, dict):
                            liv = values.get("livraison", 0)
                            ram = values.get("ramasse", 0)
                            text += f"  {country}:\n"
                            text += f"    Livraison: {liv:.2f} €/palette\n"
                            text += f"    Ramasse:   {ram:.2f} €/palette\n"
                    text += "\n"
                else:
                    # Ancien format: {livraison: x, ramasse: y}
                    if "livraison" in rev or "ramasse" in rev:
                        liv = rev.get("livraison", 0)
                        ram = rev.get("ramasse", 0)
                        text += f"Date: {date_display}\n"
                        text += f"  Belgique (défaut):\n"
                        text += f"    Livraison: {liv:.2f} €/palette\n"
                        text += f"    Ramasse:   {ram:.2f} €/palette\n"
                        text += "\n"
            else:
                # Format invalide, ignorer
                continue
        
        self.revenus_text.insert("1.0", text)

    def save_revenus(self):
        try:
            d = parse_date_input(self.rev_date_var.get())
            rev_liv = float(self.rev_livraison_var.get())
            rev_ram = float(self.rev_ramasse_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date ou montant invalide")
            return
        
        date_str = format_date_internal(d)
        self.revenus_palettes[date_str] = {
            "livraison": rev_liv,
            "ramasse": rev_ram
        }
        
        save_json(self.data_dir / "revenus_palettes.json", self.revenus_palettes)
        self.refresh_revenus_display()
        messagebox.showinfo("Succès", f"Revenus enregistrés pour le {format_date_display(d)}")

    def copy_revenus_month(self):
        try:
            d = parse_date_input(self.rev_date_var.get())
            rev_liv = float(self.rev_livraison_var.get())
            rev_ram = float(self.rev_ramasse_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date ou montant invalide")
            return
        
        if not messagebox.askyesno("Confirmer", 
                                   f"Appliquer les revenus (Liv: {rev_liv:.2f} €, Ram: {rev_ram:.2f} €)\n"
                                   f"sur tout le mois de {d.strftime('%B %Y')} ?"):
            return
        
        start_date = date(d.year, d.month, 1)
        if d.month == 12:
            end_date = date(d.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(d.year, d.month + 1, 1) - timedelta(days=1)
        
        current = start_date
        while current <= end_date:
            date_str = format_date_internal(current)
            self.revenus_palettes[date_str] = {
                "livraison": rev_liv,
                "ramasse": rev_ram
            }
            current += timedelta(days=1)
        
        save_json(self.data_dir / "revenus_palettes.json", self.revenus_palettes)
        self.refresh_revenus_display()
        days = (end_date - start_date).days + 1
        messagebox.showinfo("Succès", f"Revenus copiés sur {days} jours")

    def delete_revenus(self):
        try:
            d = parse_date_input(self.rev_date_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        date_str = format_date_internal(d)
        
        if date_str in self.revenus_palettes:
            if messagebox.askyesno("Confirmer", f"Supprimer les revenus du {format_date_display(d)} ?"):
                del self.revenus_palettes[date_str]
                save_json(self.data_dir / "revenus_palettes.json", self.revenus_palettes)
                self.refresh_revenus_display()
                messagebox.showinfo("Succès", "Revenus supprimés")
        else:
            messagebox.showinfo("Info", "Aucun revenu à cette date")

    def save_sst_tarif(self):
        """Enregistrer un tarif pour plusieurs SST sélectionnés"""
        selected_indices = self.finance_sst_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Attention", "Sélectionnez au moins un SST")
            return
        
        selected_sst = [self.finance_sst_listbox.get(i) for i in selected_indices]
        country = self.finance_country_var.get()
        
        try:
            d = parse_date_input(self.finance_date_var.get())
            date_str = format_date_internal(d)
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        try:
            tarif = float(self.finance_tarif_var.get().replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erreur", "Tarif invalide")
            return
        
        # Enregistrer pour chaque SST sélectionné
        for sst in selected_sst:
            if sst not in self.tarifs_sst:
                self.tarifs_sst[sst] = {}
            if country not in self.tarifs_sst[sst]:
                self.tarifs_sst[sst][country] = {}
            
            self.tarifs_sst[sst][country][date_str] = tarif
        
        save_json(TARIFS_SST_FILE, self.tarifs_sst)
        self.refresh_sst_tarifs_display()
        
        nb_sst = len(selected_sst)
        messagebox.showinfo("Succès", 
            f"Tarif {tarif:.2f} € enregistré pour {nb_sst} SST\n"
            f"Pays: {country}\n"
            f"Date: {format_date_display(d)}")

    def copy_sst_tarif_month(self):
        """Copier un tarif SST sur tout le mois"""
        selected_indices = self.finance_sst_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Attention", "Sélectionnez au moins un SST")
            return
        
        selected_sst = [self.finance_sst_listbox.get(i) for i in selected_indices]
        country = self.finance_country_var.get()
        
        try:
            d = parse_date_input(self.finance_date_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        try:
            tarif = float(self.finance_tarif_var.get().replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erreur", "Tarif invalide")
            return
        
        # Appliquer sur tout le mois
        start_of_month = d.replace(day=1)
        if d.month == 12:
            end_of_month = d.replace(year=d.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_of_month = d.replace(month=d.month + 1, day=1) - timedelta(days=1)
        
        current = start_of_month
        count = 0
        
        for sst in selected_sst:
            if sst not in self.tarifs_sst:
                self.tarifs_sst[sst] = {}
            if country not in self.tarifs_sst[sst]:
                self.tarifs_sst[sst][country] = {}
            
            current = start_of_month
            while current <= end_of_month:
                date_str = format_date_internal(current)
                self.tarifs_sst[sst][country][date_str] = tarif
                current += timedelta(days=1)
                count += 1
        
        save_json(TARIFS_SST_FILE, self.tarifs_sst)
        self.refresh_sst_tarifs_display()
        
        nb_sst = len(selected_sst)
        nb_days = (end_of_month - start_of_month).days + 1
        messagebox.showinfo("Succès", 
            f"Tarif {tarif:.2f} € copié pour {nb_sst} SST\n"
            f"Pays: {country}\n"
            f"Période: {format_date_display(start_of_month)} au {format_date_display(end_of_month)}\n"
            f"({nb_days} jours)")

    def delete_sst_tarif(self):
        """Supprimer un tarif SST pour une date"""
        selected_indices = self.finance_sst_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Attention", "Sélectionnez au moins un SST")
            return
        
        selected_sst = [self.finance_sst_listbox.get(i) for i in selected_indices]
        country = self.finance_country_var.get()
        
        try:
            d = parse_date_input(self.finance_date_var.get())
            date_str = format_date_internal(d)
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        deleted = []
        for sst in selected_sst:
            if (sst in self.tarifs_sst and 
                country in self.tarifs_sst[sst] and 
                date_str in self.tarifs_sst[sst][country]):
                del self.tarifs_sst[sst][country][date_str]
                deleted.append(sst)
        
        if deleted:
            save_json(TARIFS_SST_FILE, self.tarifs_sst)
            self.refresh_sst_tarifs_display()
            messagebox.showinfo("Succès", 
                f"Tarif supprimé pour {len(deleted)} SST\n"
                f"Date: {format_date_display(d)}")
        else:
            messagebox.showinfo("Info", "Aucun tarif trouvé à supprimer")


    def on_rev_country_changed(self, event=None):
        """Callback quand le pays des revenus change"""
        self.refresh_revenus_display()

    def refresh_sst_tarifs_display(self):
        """Afficher uniquement les tarifs des SST sélectionnés"""
        if not hasattr(self, 'sst_tarifs_text'):
            return
        
        self.sst_tarifs_text.delete("1.0", "end")
        
        # Récupérer les SST sélectionnés
        selected_indices = self.finance_sst_listbox.curselection()
        
        if not selected_indices:
            self.sst_tarifs_text.insert("1.0", "Sélectionnez un ou plusieurs SST pour voir leurs tarifs")
            return
        
        selected_sst = [self.finance_sst_listbox.get(i) for i in selected_indices]
        
        if not self.tarifs_sst:
            self.sst_tarifs_text.insert("1.0", "Aucun tarif défini")
            return
        
        text = "Tarifs des SST sélectionnés:\n"
        text += "=" * 60 + "\n\n"
        
        # Afficher uniquement les SST sélectionnés
        for sst in selected_sst:
            if sst in self.tarifs_sst:
                text += f"SST: {sst}\n"
                text += "-" * 60 + "\n"
                
                for country in sorted(self.tarifs_sst[sst].keys()):
                    tarifs = self.tarifs_sst[sst][country]
                    if tarifs:
                        sorted_dates = sorted(tarifs.keys(), reverse=True)
                        text += f"  {country}:\n"
                        for date_str in sorted_dates[:5]:
                            tarif = tarifs[date_str]
                            try:
                                d = datetime.strptime(date_str, "%Y-%m-%d").date()
                                date_display = format_date_display(d)
                            except Exception:
                                date_display = date_str
                            text += f"    {date_display} : {tarif:.2f} €\n"
                        if len(sorted_dates) > 5:
                            text += f"    ... et {len(sorted_dates) - 5} autres dates\n"
                text += "\n"
            else:
                text += f"SST: {sst}\n"
                text += "  Aucun tarif défini\n\n"
        
        self.sst_tarifs_text.insert("1.0", text)

    def save_revenus(self):
        """Enregistrer les revenus par palette pour un pays et une date"""
        country = self.rev_country_var.get()
        
        try:
            d = parse_date_input(self.rev_date_var.get())
            date_str = format_date_internal(d)
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        try:
            rev_liv = float(self.rev_livraison_var.get().replace(',', '.'))
            rev_ram = float(self.rev_ramasse_var.get().replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erreur", "Revenus invalides")
            return
        
        # Structure: {date: {country: {livraison: x, ramasse: y}}}
        if date_str not in self.revenus_palettes:
            self.revenus_palettes[date_str] = {}
        
        self.revenus_palettes[date_str][country] = {
            "livraison": rev_liv,
            "ramasse": rev_ram
        }
        
        save_json(REVENUS_FILE, self.revenus_palettes)
        self.refresh_revenus_display()
        
        messagebox.showinfo("Succès", 
            f"Revenus enregistrés\n"
            f"Pays: {country}\n"
            f"Date: {format_date_display(d)}\n"
            f"Livraison: {rev_liv:.2f} €/palette\n"
            f"Ramasse: {rev_ram:.2f} €/palette")

    def copy_revenus_month(self):
        """Copier les revenus sur tout le mois"""
        country = self.rev_country_var.get()
        
        try:
            d = parse_date_input(self.rev_date_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        try:
            rev_liv = float(self.rev_livraison_var.get().replace(',', '.'))
            rev_ram = float(self.rev_ramasse_var.get().replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erreur", "Revenus invalides")
            return
        
        # Appliquer sur tout le mois
        start_of_month = d.replace(day=1)
        if d.month == 12:
            end_of_month = d.replace(year=d.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_of_month = d.replace(month=d.month + 1, day=1) - timedelta(days=1)
        
        current = start_of_month
        while current <= end_of_month:
            date_str = format_date_internal(current)
            
            if date_str not in self.revenus_palettes:
                self.revenus_palettes[date_str] = {}
            
            self.revenus_palettes[date_str][country] = {
                "livraison": rev_liv,
                "ramasse": rev_ram
            }
            
            current += timedelta(days=1)
        
        save_json(REVENUS_FILE, self.revenus_palettes)
        self.refresh_revenus_display()
        
        nb_days = (end_of_month - start_of_month).days + 1
        messagebox.showinfo("Succès", 
            f"Revenus copiés pour tout le mois\n"
            f"Pays: {country}\n"
            f"Période: {format_date_display(start_of_month)} au {format_date_display(end_of_month)}\n"
            f"({nb_days} jours)")

    def delete_revenus(self):
        """Supprimer les revenus pour une date et un pays"""
        country = self.rev_country_var.get()
        
        try:
            d = parse_date_input(self.rev_date_var.get())
            date_str = format_date_internal(d)
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        if (date_str in self.revenus_palettes and 
            country in self.revenus_palettes[date_str]):
            del self.revenus_palettes[date_str][country]
            
            # Supprimer la date si vide
            if not self.revenus_palettes[date_str]:
                del self.revenus_palettes[date_str]
            
            save_json(REVENUS_FILE, self.revenus_palettes)
            self.refresh_revenus_display()
            messagebox.showinfo("Succès", 
                f"Revenus supprimés\n"
                f"Pays: {country}\n"
                f"Date: {format_date_display(d)}")
        else:
            messagebox.showinfo("Info", "Aucun revenu trouvé à supprimer")

    def calculate_finance(self):
        """Calcul financier avec affichage regroupé par SST"""
        try:
            d = parse_date_input(self.calc_date_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide")
            return
        
        day_dir = get_planning_day_dir(d)
        if not day_dir or not day_dir.exists():
            self.calc_result_text.delete("1.0", "end")
            self.calc_result_text.insert("1.0", f"Aucun planning pour le {format_date_display(d)}")
            return
        
        missions = []
        for file in day_dir.glob("*.json"):
            data = load_json(file, None)
            if data:
                missions.append(data)
        
        if not missions:
            self.calc_result_text.delete("1.0", "end")
            self.calc_result_text.insert("1.0", f"Planning vide pour le {format_date_display(d)}")
            return
        
        date_str = format_date_internal(d)
        
        # Structure: {sst: {tarif_country: set(chauffeurs)}}
        sst_drivers_by_country = {}        # IMPORTANT: Compter les chauffeurs UNIQUES par SST
        # 1 chauffeur = 1 forfait par jour, peu importe le nombre de missions
        
        # DEBUG: Afficher la structure d'une mission pour comprendre le problème
        if missions and len(missions) > 0:
            sample = missions[0]
        
        for mission in missions:
            sst = mission.get("sst", "").strip()
            chauffeur_name = mission.get("chauffeur_nom", mission.get("chauffeur", "")).strip()
            
            # Ignorer les missions sans SST
            if not sst or sst == "":
                continue
            
            # Déterminer le pays de la MISSION (pas du chauffeur!)
            voyage_code = mission.get("voyage", "")
            voyage = None
            for v in self.voyages:
                if v.get("code", "") == voyage_code:
                    voyage = v
                    break
            
            if voyage:
                mission_country = voyage.get("country", "Belgique")
                # Normaliser le nom du pays
                if mission_country.lower() in ["pays-bas", "paysbas", "pays bas", "netherlands"]:
                    mission_country = "Pays-Bas"
                elif mission_country.lower() in ["belgique", "belgium", "be"]:
                    mission_country = "Belgique"
                elif mission_country.lower() in ["luxembourg", "lux", "lu"]:
                    mission_country = "Luxembourg"
                elif mission_country.lower() in ["france", "fr"]:
                    mission_country = "France"
            else:
                # Si pas de voyage, deviner depuis le code
                if voyage_code:
                    if voyage_code.upper().startswith("P-B"):
                        mission_country = "Pays-Bas"
                    elif voyage_code.upper().startswith("RES") or voyage_code.upper().startswith("ALV"):
                        mission_country = "Luxembourg"
                    else:
                        mission_country = "Belgique"
                else:
                    mission_country = "Belgique"
            
            # Initialiser les structures
            if sst not in sst_drivers_by_country:
                sst_drivers_by_country[sst] = {}
            if mission_country not in sst_drivers_by_country[sst]:
                sst_drivers_by_country[sst][mission_country] = set()
            
            # Ajouter le chauffeur UNIQUEMENT s'il a un nom
            # Un chauffeur = 1 forfait, même s'il fait plusieurs missions
            if chauffeur_name:
                sst_drivers_by_country[sst][mission_country].add(chauffeur_name)
            # Si pas de chauffeur: ne rien ajouter (les missions sans chauffeur ne coûtent rien)
        
        # Calculer les revenus par palettes et par pays
        revenue_by_country = {}
        
        # DEBUG: compter les missions par pays
        missions_by_country = {}
        
        for mission in missions:
            voyage_code = mission.get("voyage", "")
            
            # Recherche du voyage
            voyage = None
            for v in self.voyages:
                if v.get("code", "") == voyage_code:
                    voyage = v
                    break
            
            # Déterminer le pays
            if voyage:
                country = voyage.get("country", "Belgique")
                # Normaliser le nom du pays
                if country.lower() in ["pays-bas", "paysbas", "pays bas", "netherlands"]:
                    country = "Pays-Bas"
                elif country.lower() in ["belgique", "belgium", "be"]:
                    country = "Belgique"
                elif country.lower() in ["luxembourg", "lux", "lu"]:
                    country = "Luxembourg"
                elif country.lower() in ["france", "fr"]:
                    country = "France"
            else:
                # Si pas de voyage trouvé, essayer de deviner depuis le code voyage
                if voyage_code:
                    if voyage_code.upper().startswith("P-B"):
                        country = "Pays-Bas"
                    elif voyage_code.upper().startswith("RES") or voyage_code.upper().startswith("ALV"):
                        country = "Luxembourg"
                    else:
                        country = "Belgique"
                else:
                    country = "Belgique"
            
            # Initialiser le pays
            if country not in revenue_by_country:
                revenue_by_country[country] = {"pal_liv": 0, "pal_ram": 0}
            if country not in missions_by_country:
                missions_by_country[country] = 0
            
            missions_by_country[country] += 1
            
            # Compter les palettes (gérer les strings et None)
            try:
                nb_pal_str = mission.get("nb_pal", 0)
                if nb_pal_str is None or nb_pal_str == "":
                    nb_pal = 0
                else:
                    nb_pal = int(nb_pal_str)
            except (ValueError, TypeError):
                nb_pal = 0
            
            if mission.get("type") == "LIVRAISON":
                revenue_by_country[country]["pal_liv"] += nb_pal
            else:
                revenue_by_country[country]["pal_ram"] += nb_pal
        
        # Calculer les revenus totaux
        total_revenus = 0
        revenue_details = []
        
        for country, pal_data in revenue_by_country.items():
            rev_liv = 0
            rev_ram = 0
            
            if date_str in self.revenus_palettes:
                if country in self.revenus_palettes[date_str]:
                    rev_data = self.revenus_palettes[date_str][country]
                    rev_liv = rev_data.get("livraison", 0)
                    rev_ram = rev_data.get("ramasse", 0)
                elif "livraison" in self.revenus_palettes[date_str]:
                    rev_data = self.revenus_palettes[date_str]
                    rev_liv = rev_data.get("livraison", 0)
                    rev_ram = rev_data.get("ramasse", 0)
            
            if rev_liv == 0 and rev_ram == 0:
                for dt in sorted([d for d in self.revenus_palettes.keys() if d <= date_str], reverse=True):
                    if country in self.revenus_palettes[dt]:
                        rev_data = self.revenus_palettes[dt][country]
                        rev_liv = rev_data.get("livraison", 0)
                        rev_ram = rev_data.get("ramasse", 0)
                        if rev_liv > 0 or rev_ram > 0:
                            break
                    elif "livraison" in self.revenus_palettes[dt]:
                        rev_data = self.revenus_palettes[dt]
                        rev_liv = rev_data.get("livraison", 0)
                        rev_ram = rev_data.get("ramasse", 0)
                        if rev_liv > 0 or rev_ram > 0:
                            break
            
            rev_liv_total = pal_data["pal_liv"] * rev_liv
            rev_ram_total = pal_data["pal_ram"] * rev_ram
            country_revenue = rev_liv_total + rev_ram_total
            total_revenus += country_revenue
            
            revenue_details.append({
                "country": country,
                "pal_liv": pal_data["pal_liv"],
                "pal_ram": pal_data["pal_ram"],
                "rev_liv": rev_liv,
                "rev_ram": rev_ram,
                "rev_liv_total": rev_liv_total,
                "rev_ram_total": rev_ram_total,
                "total": country_revenue
            })
        
        # Afficher les résultats
        # Calculer le nombre de chauffeurs UNIQUES (pas le nombre de missions)
        unique_drivers = set()
        missions_with_sst = 0
        
        for mission in missions:
            sst = mission.get("sst", "").strip()
            if sst:
                missions_with_sst += 1
            
            chauffeur_name = mission.get("chauffeur_nom", mission.get("chauffeur", "")).strip()
            if chauffeur_name:
                unique_drivers.add(chauffeur_name)
        
        result = f"📊 ANALYSE FINANCIÈRE - {format_date_display(d)}\n"
        result += "=" * 70 + "\n"
        result += f"Missions totales: {len(missions)} | "
        result += f"Avec SST: {missions_with_sst} | "
        result += f"Chauffeurs uniques: {len(unique_drivers)}\n\n"

        for detail in revenue_details:
            result += f"\n{detail['country']}:\n"
            result += f"  Livraisons : {detail['pal_liv']} pal × {detail['rev_liv']:.2f} € = {detail['rev_liv_total']:,.2f} €\n"
            result += f"  Ramasses   : {detail['pal_ram']} pal × {detail['rev_ram']:.2f} € = {detail['rev_ram_total']:,.2f} €\n"
            result += f"  Sous-total : {detail['total']:,.2f} €\n"
        result += f"\nTOTAL REVENUS : {total_revenus:,.2f} €\n\n"
        
        result += "💶 COÛTS SST:\n"
        result += "-" * 70 + "\n"
        
        total_costs = 0
        total_sst_used = 0
        total_drivers_used = 0
        
        if sst_drivers_by_country:
            # Construire d'abord les données à afficher
            sst_display_data = {}
            
            for sst in sorted(sst_drivers_by_country.keys()):
                countries_drivers = sst_drivers_by_country[sst]
                
                # ÉTAPE 1: Pour chaque chauffeur, trouver le tarif MAX parmi tous ses pays
                driver_max_tarif = {}  # {driver_name: (country, tarif)}
                
                for country in countries_drivers.keys():
                    drivers = countries_drivers[country]
                    
                    # Récupérer le tarif pour ce pays
                    tarif_journalier = 0
                    if sst in self.tarifs_sst and country in self.tarifs_sst[sst]:
                        tarifs = self.tarifs_sst[sst][country]
                        tarif_journalier = tarifs.get(date_str, 0)
                        
                        if tarif_journalier == 0:
                            sorted_dates = sorted([dt for dt in tarifs.keys() if dt <= date_str], reverse=True)
                            if sorted_dates:
                                tarif_journalier = tarifs[sorted_dates[0]]
                    
                    # Pour chaque chauffeur de ce pays
                    for driver in drivers:
                        if driver not in driver_max_tarif or tarif_journalier > driver_max_tarif[driver][1]:
                            driver_max_tarif[driver] = (country, tarif_journalier)
                
                # ÉTAPE 2: Regrouper les chauffeurs par pays (avec le tarif max)
                country_drivers_max = {}  # {country: [drivers]}
                for driver, (country, tarif) in driver_max_tarif.items():
                    if country not in country_drivers_max:
                        country_drivers_max[country] = []
                    country_drivers_max[country].append(driver)
                
                # ÉTAPE 3: Calculer les coûts
                sst_total_cost = 0
                sst_total_drivers = len(driver_max_tarif)
                sst_lines = []
                
                for country in sorted(country_drivers_max.keys()):
                    drivers = country_drivers_max[country]
                    nb_drivers = len(drivers)
                    
                    # Récupérer le tarif pour ce pays
                    tarif_journalier = 0
                    if sst in self.tarifs_sst and country in self.tarifs_sst[sst]:
                        tarifs = self.tarifs_sst[sst][country]
                        tarif_journalier = tarifs.get(date_str, 0)
                        
                        if tarif_journalier == 0:
                            sorted_dates = sorted([dt for dt in tarifs.keys() if dt <= date_str], reverse=True)
                            if sorted_dates:
                                tarif_journalier = tarifs[sorted_dates[0]]
                    
                    # Coût pour ce pays
                    cost = nb_drivers * tarif_journalier
                    sst_total_cost += cost
                    
                    # Affichage
                    drivers_list = ", ".join(sorted(drivers))
                    sst_lines.append(f"  {country}: {nb_drivers} chauffeur(s) × {tarif_journalier:.2f} € = {cost:,.2f} €\n")
                    sst_lines.append(f"    → {drivers_list}\n")
                
                # N'ajouter que si ce SST a des chauffeurs
                if sst_total_drivers > 0:
                    sst_display_data[sst] = {
                        'lines': sst_lines,
                        'total_drivers': sst_total_drivers,
                        'total_cost': sst_total_cost
                    }
                    total_costs += sst_total_cost
                    total_drivers_used += sst_total_drivers
                    total_sst_used += 1
            
            # Afficher les SST avec des données
            if sst_display_data:
                for sst, data in sst_display_data.items():
                    result += f"\nSST: {sst}\n"
                    for line in data['lines']:
                        result += line
                    result += f"  Total SST {sst}: {data['total_drivers']} chauffeur(s) = {data['total_cost']:,.2f} €\n"
                
                result += f"\n" + "=" * 70 + "\n"
                result += f"TOTAL: {total_sst_used} SST utilisé(s), {total_drivers_used} chauffeur(s) utilisé(s)\n"
                result += f"TOTAL COÛTS : {total_costs:,.2f} €\n\n"
            else:
                result += "\n  Aucun chauffeur assigné pour cette date\n"
                result += "  Vérifiez que:\n"
                result += "  • Les missions ont un SST défini\n"
                result += "  • Les chauffeurs sont assignés aux missions\n"
                result += "  • Les tarifs SST sont définis\n\n"
        else:
            result += "\n  Aucune mission avec SST pour cette date\n\n"
        
        result += "📈 RÉSULTAT:\n"
        result += "-" * 70 + "\n"
        marge = total_revenus - total_costs
        pct = (marge / total_revenus * 100) if total_revenus > 0 else 0
        
        if marge >= 0:
            result += f"MARGE : +{marge:,.2f} € ({pct:.1f}%)\n"
        else:
            result += f"PERTE : {marge:,.2f} € ({pct:.1f}%)\n"
        
        self.calc_result_text.delete("1.0", "end")
        self.calc_result_text.insert("1.0", result)

    def refresh_finance_view(self):
        """Rafraîchir la vue finance"""
        if hasattr(self, 'finance_sst_listbox') and self.sst_list:
            self.finance_sst_listbox.delete(0, "end")
            for sst in self.sst_list:
                self.finance_sst_listbox.insert("end", sst)
            # Sélectionner le premier SST par défaut si rien n'est sélectionné
            if not self.finance_sst_listbox.curselection() and self.sst_list:
                self.finance_sst_listbox.selection_set(0)
        
        self.refresh_sst_tarifs_display()
        self.refresh_revenus_display()


    # ========== ONGLETS ANALYSE ==========
    
    def build_analyse_tab(self):
        """Construire les deux onglets d'analyse (simple et avancé)"""
        # 1. Construire l'onglet Analyse Simple (original)
        self.build_analyse_simple_tab()
        
        # 2. Construire l'onglet Analyse Avancée (nouveau module)
        self.build_analyse_advanced_tab()
    
    def build_analyse_advanced_tab(self):
        """Construire l'onglet Analyse Avancée avec le nouveau module"""
        self.analyse_module = AdvancedAnalyseModule(self)
    
    def build_analyse_simple_tab(self):
        """Construire l'onglet Analyse Simple (version originale)"""
        self.tab_analyse_simple = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_analyse_simple, text="📊 Analyse")
        
        if not MATPLOTLIB_AVAILABLE:
            msg_frame = ttk.Frame(self.tab_analyse_simple)
            msg_frame.pack(fill="both", expand=True, padx=20, pady=20)
            ttk.Label(msg_frame, text="⚠️ Module matplotlib non disponible", 
                     font=("Arial", 14, "bold")).pack(pady=10)
            ttk.Label(msg_frame, text="Pour activer les graphiques, installez matplotlib:\n\npip install matplotlib",
                     font=("Arial", 10)).pack(pady=10)
            return
        
        # Frame principale avec canvas scrollable
        main_canvas = tk.Canvas(self.tab_analyse_simple, bg="#f5f5f5")
        main_scrollbar_y = ttk.Scrollbar(self.tab_analyse_simple, orient="vertical", command=main_canvas.yview)
        main_scrollbar_x = ttk.Scrollbar(self.tab_analyse_simple, orient="horizontal", command=main_canvas.xview)
        
        self.analyse_scrollable_frame = ttk.Frame(main_canvas)
        
        self.analyse_scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=self.analyse_scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=main_scrollbar_y.set, xscrollcommand=main_scrollbar_x.set)
        
        main_scrollbar_y.pack(side="right", fill="y")
        main_scrollbar_x.pack(side="bottom", fill="x")
        main_canvas.pack(side="left", fill="both", expand=True)
        
        def _on_mousewheel(event):
            main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        self.analyse_canvas = main_canvas
        
        content_frame = ttk.Frame(self.analyse_scrollable_frame, padding=10)
        content_frame.pack(fill="both", expand=True)
        
        title_frame = ttk.Frame(content_frame)
        title_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(title_frame, text="📊 Tableau de Bord Analytique", 
                 font=("Arial", 16, "bold")).pack(side="left")
        
        controls_frame = ttk.LabelFrame(content_frame, text="🎛️ Paramètres d'analyse", padding=10)
        controls_frame.pack(fill="x", pady=(0, 15))
        
        period_frame = ttk.Frame(controls_frame)
        period_frame.pack(fill="x", pady=5)
        
        ttk.Label(period_frame, text="Période :").pack(side="left", padx=(0, 5))
        ttk.Label(period_frame, text="Du :").pack(side="left", padx=(10, 5))
        self.analyse_date_start_var = tk.StringVar(value=(date.today() - timedelta(days=30)).strftime("%d/%m/%Y"))
        ttk.Entry(period_frame, textvariable=self.analyse_date_start_var, width=12).pack(side="left", padx=5)
        
        ttk.Label(period_frame, text="Au :").pack(side="left", padx=(10, 5))
        self.analyse_date_end_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(period_frame, textvariable=self.analyse_date_end_var, width=12).pack(side="left", padx=5)
        
        ttk.Button(period_frame, text="7 jours", width=8,
                  command=lambda: self.set_analyse_period(7)).pack(side="left", padx=2)
        ttk.Button(period_frame, text="30 jours", width=8,
                  command=lambda: self.set_analyse_period(30)).pack(side="left", padx=2)
        ttk.Button(period_frame, text="90 jours", width=8,
                  command=lambda: self.set_analyse_period(90)).pack(side="left", padx=2)
        ttk.Button(period_frame, text="Cette année", width=10,
                  command=self.set_analyse_period_year).pack(side="left", padx=2)
        
        type_frame = ttk.Frame(controls_frame)
        type_frame.pack(fill="x", pady=5)
        
        ttk.Label(type_frame, text="Type de graphique :").pack(side="left", padx=(0, 5))
        self.analyse_chart_type_var = tk.StringVar(value="all")
        chart_types = [
            ("Tous les graphiques", "all"),
            ("Évolution CA", "ca_evolution"),
            ("CA par Code Voyage", "ca_voyage"),
            ("Analyse SST", "sst_analysis"),
            ("Analyse Chauffeurs", "driver_analysis"),
            ("Rentabilité", "profitability"),
        ]
        self.analyse_chart_combo = ttk.Combobox(type_frame, textvariable=self.analyse_chart_type_var,
                                                values=[t[0] for t in chart_types], width=25, state="readonly")
        self.analyse_chart_combo.pack(side="left", padx=5)
        self.analyse_chart_types_map = {t[0]: t[1] for t in chart_types}
        
        ttk.Button(type_frame, text="🔄 Générer les graphiques", 
                  command=self.generate_analyse_charts).pack(side="left", padx=15)
        ttk.Button(type_frame, text="📥 Exporter données", 
                  command=self.export_analyse_data).pack(side="left", padx=5)
        
        self.charts_container = ttk.Frame(content_frame)
        self.charts_container.pack(fill="both", expand=True, pady=10)
        
        self.analyse_placeholder = ttk.Label(self.charts_container, 
            text="📈 Cliquez sur 'Générer les graphiques' pour afficher les analyses",
            font=("Arial", 12), foreground="gray")
        self.analyse_placeholder.pack(pady=50)
        
        self.analyse_figures = []
        self.tab_analyse_simple.bind("<Configure>", self._on_analyse_resize)
    
    def refresh_analyse_view(self):
        """Rafraîchir la vue analyse (appelé par le système d'auto-refresh)"""
        pass
    
    def set_analyse_period(self, days):
        """Définir la période d'analyse"""
        end_date = date.today()
        start_date = end_date - timedelta(days=days)
        self.analyse_date_start_var.set(start_date.strftime("%d/%m/%Y"))
        self.analyse_date_end_var.set(end_date.strftime("%d/%m/%Y"))
    
    def set_analyse_period_year(self):
        """Définir la période sur l'année en cours"""
        today = date.today()
        start_date = date(today.year, 1, 1)
        self.analyse_date_start_var.set(start_date.strftime("%d/%m/%Y"))
        self.analyse_date_end_var.set(today.strftime("%d/%m/%Y"))
    
    def _on_analyse_resize(self, event=None):
        """Gérer le redimensionnement des graphiques"""
        # Ne rien faire ici pour éviter les ralentissements
        pass
    
    def get_analyse_data(self, start_date, end_date):
        """Récupérer les données d'analyse pour une période"""
        data = {
            'dates': [],
            'revenus': [],
            'couts': [],
            'marges': [],
            'missions_count': [],
            'palettes_liv': [],
            'palettes_ram': [],
            'by_voyage': {},
            'by_sst': {},
            'by_driver': {},
            'by_country': {},
        }
        
        current = start_date
        while current <= end_date:
            day_dir = get_planning_day_dir(current)
            date_str = format_date_internal(current)
            
            day_data = {
                'date': current,
                'revenus': 0,
                'couts': 0,
                'missions': 0,
                'pal_liv': 0,
                'pal_ram': 0,
            }
            
            if day_dir and day_dir.exists():
                missions = []
                for file in day_dir.glob("*.json"):
                    mission_data = load_json(file, None)
                    if mission_data:
                        missions.append(mission_data)
                
                day_data['missions'] = len(missions)
                
                # Analyser chaque mission
                drivers_by_sst = {}  # Pour éviter de compter plusieurs fois le même chauffeur
                
                for mission in missions:
                    voyage_code = mission.get("voyage", "N/A")
                    sst = mission.get("sst", "N/A")
                    driver = mission.get("chauffeur_nom", mission.get("chauffeur", "N/A"))
                    mission_type = mission.get("type", "LIVRAISON")
                    
                    # Palettes
                    try:
                        nb_pal = int(mission.get("nb_pal", 0) or 0)
                    except (ValueError, TypeError):
                        nb_pal = 0
                    
                    if mission_type == "LIVRAISON":
                        day_data['pal_liv'] += nb_pal
                    else:
                        day_data['pal_ram'] += nb_pal
                    
                    # Déterminer le pays
                    voyage = next((v for v in self.voyages if v.get("code") == voyage_code), None)
                    country = voyage.get("country", "Belgique") if voyage else "Belgique"
                    
                    # Normaliser le pays
                    country_map = {
                        "pays-bas": "Pays-Bas", "paysbas": "Pays-Bas", "netherlands": "Pays-Bas",
                        "belgique": "Belgique", "belgium": "Belgique",
                        "france": "France", "luxembourg": "Luxembourg",
                    }
                    country = country_map.get(country.lower(), country)
                    
                    # Revenus par palette
                    rev_liv = 0
                    rev_ram = 0
                    if date_str in self.revenus_palettes:
                        if country in self.revenus_palettes[date_str]:
                            rev_data = self.revenus_palettes[date_str][country]
                            rev_liv = rev_data.get("livraison", 0)
                            rev_ram = rev_data.get("ramasse", 0)
                        elif "livraison" in self.revenus_palettes[date_str]:
                            rev_data = self.revenus_palettes[date_str]
                            rev_liv = rev_data.get("livraison", 0)
                            rev_ram = rev_data.get("ramasse", 0)
                    
                    # Calculer revenus
                    if mission_type == "LIVRAISON":
                        mission_rev = nb_pal * rev_liv
                    else:
                        mission_rev = nb_pal * rev_ram
                    
                    day_data['revenus'] += mission_rev
                    
                    # Par voyage
                    if voyage_code not in data['by_voyage']:
                        data['by_voyage'][voyage_code] = {'revenus': 0, 'missions': 0, 'palettes': 0}
                    data['by_voyage'][voyage_code]['revenus'] += mission_rev
                    data['by_voyage'][voyage_code]['missions'] += 1
                    data['by_voyage'][voyage_code]['palettes'] += nb_pal
                    
                    # Par pays
                    if country not in data['by_country']:
                        data['by_country'][country] = {'revenus': 0, 'couts': 0, 'missions': 0}
                    data['by_country'][country]['revenus'] += mission_rev
                    data['by_country'][country]['missions'] += 1
                    
                    # Par SST (compter chauffeurs uniques)
                    if sst and sst != "N/A":
                        if sst not in drivers_by_sst:
                            drivers_by_sst[sst] = {}
                        if country not in drivers_by_sst[sst]:
                            drivers_by_sst[sst][country] = set()
                        if driver and driver != "N/A":
                            drivers_by_sst[sst][country].add(driver)
                    
                    # Par chauffeur
                    if driver and driver != "N/A":
                        if driver not in data['by_driver']:
                            data['by_driver'][driver] = {'revenus': 0, 'missions': 0, 'palettes': 0, 'couts': 0}
                        data['by_driver'][driver]['revenus'] += mission_rev
                        data['by_driver'][driver]['missions'] += 1
                        data['by_driver'][driver]['palettes'] += nb_pal
                
                # Calculer les coûts SST
                for sst, countries in drivers_by_sst.items():
                    if sst not in data['by_sst']:
                        data['by_sst'][sst] = {'couts': 0, 'chauffeurs': 0, 'revenus': 0}
                    
                    for country, drivers in countries.items():
                        nb_drivers = len(drivers)
                        
                        # Récupérer le tarif
                        tarif = 0
                        if sst in self.tarifs_sst and country in self.tarifs_sst[sst]:
                            tarifs = self.tarifs_sst[sst][country]
                            tarif = tarifs.get(date_str, 0)
                            if tarif == 0:
                                sorted_dates = sorted([dt for dt in tarifs.keys() if dt <= date_str], reverse=True)
                                if sorted_dates:
                                    tarif = tarifs[sorted_dates[0]]
                        
                        cost = nb_drivers * tarif
                        day_data['couts'] += cost
                        data['by_sst'][sst]['couts'] += cost
                        data['by_sst'][sst]['chauffeurs'] += nb_drivers
                        
                        # Coût par pays
                        if country in data['by_country']:
                            data['by_country'][country]['couts'] += cost
                        
                        # Coût par chauffeur (répartir équitablement)
                        if nb_drivers > 0:
                            cost_per_driver = tarif
                            for driver in drivers:
                                if driver in data['by_driver']:
                                    data['by_driver'][driver]['couts'] += cost_per_driver
            
            # Ajouter les données du jour
            data['dates'].append(current)
            data['revenus'].append(day_data['revenus'])
            data['couts'].append(day_data['couts'])
            data['marges'].append(day_data['revenus'] - day_data['couts'])
            data['missions_count'].append(day_data['missions'])
            data['palettes_liv'].append(day_data['pal_liv'])
            data['palettes_ram'].append(day_data['pal_ram'])
            
            current += timedelta(days=1)
        
        return data
    
    def generate_analyse_charts(self):
        """Générer les graphiques d'analyse"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Erreur", "matplotlib n'est pas installé")
            return
        
        # Parser les dates
        try:
            start_date = parse_date_input(self.analyse_date_start_var.get())
            end_date = parse_date_input(self.analyse_date_end_var.get())
        except ValueError as e:
            messagebox.showerror("Erreur", f"Date invalide: {e}")
            return
        
        if start_date > end_date:
            messagebox.showerror("Erreur", "La date de début doit être avant la date de fin")
            return
        
        # Récupérer les données
        data = self.get_analyse_data(start_date, end_date)
        
        # Nettoyer les anciens graphiques
        for widget in self.charts_container.winfo_children():
            widget.destroy()
        self.analyse_figures = []
        
        # Type de graphique sélectionné
        selected_type = self.analyse_chart_types_map.get(
            self.analyse_chart_combo.get(), "all")
        
        # Calculer la largeur disponible
        try:
            container_width = self.charts_container.winfo_width()
            if container_width < 100:
                container_width = 1200
        except Exception:
            container_width = 1200
        
        # Taille des figures
        fig_width = min(12, max(8, container_width / 100))
        fig_height = 5
        
        charts_to_show = []
        
        if selected_type == "all" or selected_type == "ca_evolution":
            charts_to_show.append(("ca_evolution", "📈 Évolution du CA / Coûts / Marge"))
        if selected_type == "all" or selected_type == "ca_voyage":
            charts_to_show.append(("ca_voyage", "🚛 CA par Code Voyage (Top 15)"))
        if selected_type == "all" or selected_type == "sst_analysis":
            charts_to_show.append(("sst_analysis", "💶 Analyse SST - Coûts vs Nombre de chauffeurs"))
        if selected_type == "all" or selected_type == "driver_analysis":
            charts_to_show.append(("driver_analysis", "👤 Analyse Chauffeurs - Rentabilité"))
        if selected_type == "all" or selected_type == "profitability":
            charts_to_show.append(("profitability", "📊 Rentabilité par Pays"))
        
        row = 0
        for chart_type, title in charts_to_show:
            chart_frame = ttk.LabelFrame(self.charts_container, text=title, padding=5)
            chart_frame.pack(fill="x", pady=10, padx=5)
            
            fig = Figure(figsize=(fig_width, fig_height), dpi=100)
            fig.patch.set_facecolor('#f5f5f5')
            
            if chart_type == "ca_evolution":
                self._create_ca_evolution_chart(fig, data)
            elif chart_type == "ca_voyage":
                self._create_voyage_chart(fig, data)
            elif chart_type == "sst_analysis":
                self._create_sst_chart(fig, data)
            elif chart_type == "driver_analysis":
                self._create_driver_chart(fig, data)
            elif chart_type == "profitability":
                self._create_profitability_chart(fig, data)
            
            canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            
            # Toolbar
            toolbar_frame = ttk.Frame(chart_frame)
            toolbar_frame.pack(fill="x")
            toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
            toolbar.update()
            
            self.analyse_figures.append((fig, canvas))
            row += 1
        
        # Résumé statistique
        self._create_summary_stats(data, start_date, end_date)
    
    def _create_ca_evolution_chart(self, fig, data):
        """Créer le graphique d'évolution du CA"""
        ax = fig.add_subplot(111)
        
        if not data['dates']:
            ax.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        dates = data['dates']
        revenus = data['revenus']
        couts = data['couts']
        marges = data['marges']
        
        ax.plot(dates, revenus, 'g-', linewidth=2, label='Revenus', marker='o', markersize=3)
        ax.plot(dates, couts, 'r-', linewidth=2, label='Coûts', marker='s', markersize=3)
        ax.fill_between(dates, marges, alpha=0.3, color='blue', label='Marge')
        ax.axhline(y=0, color='black', linestyle='-', linewidth=0.5)
        
        ax.set_xlabel('Date')
        ax.set_ylabel('Montant (€)')
        ax.legend(loc='upper left')
        ax.grid(True, alpha=0.3)
        
        # Format des dates
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate()
        
        # Totaux dans le coin
        total_rev = sum(revenus)
        total_cout = sum(couts)
        total_marge = sum(marges)
        ax.text(0.98, 0.98, f"Total Rev: {total_rev:,.0f}€\nTotal Coûts: {total_cout:,.0f}€\nMarge: {total_marge:,.0f}€",
                transform=ax.transAxes, fontsize=9, verticalalignment='top', horizontalalignment='right',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        fig.tight_layout()
    
    def _create_voyage_chart(self, fig, data):
        """Créer le graphique CA par voyage"""
        ax = fig.add_subplot(111)
        
        voyages_data = data['by_voyage']
        if not voyages_data:
            ax.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        # Trier par revenus et prendre le top 15
        sorted_voyages = sorted(voyages_data.items(), key=lambda x: x[1]['revenus'], reverse=True)[:15]
        
        codes = [v[0] for v in sorted_voyages]
        revenus = [v[1]['revenus'] for v in sorted_voyages]
        missions = [v[1]['missions'] for v in sorted_voyages]
        
        # Graphique à barres
        bars = ax.bar(codes, revenus, color='steelblue', alpha=0.8)
        
        # Ajouter le nombre de missions sur chaque barre
        for bar, mission_count in zip(bars, missions):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{mission_count}m',
                   ha='center', va='bottom', fontsize=8, color='darkblue')
        
        ax.set_xlabel('Code Voyage')
        ax.set_ylabel('Revenus (€)')
        ax.tick_params(axis='x', rotation=45)
        ax.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout()
    
    def _create_sst_chart(self, fig, data):
        """Créer le graphique d'analyse SST"""
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        
        sst_data = data['by_sst']
        if not sst_data:
            ax1.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            ax2.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        ssts = list(sst_data.keys())
        couts = [sst_data[s]['couts'] for s in ssts]
        chauffeurs = [sst_data[s]['chauffeurs'] for s in ssts]
        
        # Graphique 1: Coûts par SST
        colors = plt.cm.Reds([0.3 + 0.5 * i/len(ssts) for i in range(len(ssts))])
        bars1 = ax1.bar(ssts, couts, color=colors)
        ax1.set_xlabel('SST')
        ax1.set_ylabel('Coûts totaux (€)')
        ax1.set_title('Coûts par SST')
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, alpha=0.3, axis='y')
        
        # Ajouter les valeurs
        for bar, cout in zip(bars1, couts):
            if cout > 0:
                ax1.text(bar.get_x() + bar.get_width()/2., bar.get_height(),
                        f'{cout:,.0f}€', ha='center', va='bottom', fontsize=8)
        
        # Graphique 2: Nombre de chauffeurs par SST
        colors2 = plt.cm.Blues([0.3 + 0.5 * i/len(ssts) for i in range(len(ssts))])
        bars2 = ax2.bar(ssts, chauffeurs, color=colors2)
        ax2.set_xlabel('SST')
        ax2.set_ylabel('Nb jours-chauffeurs')
        ax2.set_title('Utilisation chauffeurs par SST')
        ax2.tick_params(axis='x', rotation=45)
        ax2.grid(True, alpha=0.3, axis='y')
        
        for bar, ch in zip(bars2, chauffeurs):
            if ch > 0:
                ax2.text(bar.get_x() + bar.get_width()/2., bar.get_height(),
                        str(ch), ha='center', va='bottom', fontsize=8)
        
        fig.tight_layout()
    
    def _create_driver_chart(self, fig, data):
        """Créer le graphique d'analyse des chauffeurs"""
        ax = fig.add_subplot(111)
        
        driver_data = data['by_driver']
        if not driver_data:
            ax.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        # Calculer la marge par chauffeur et trier
        drivers_with_margin = []
        for driver, vals in driver_data.items():
            marge = vals['revenus'] - vals['couts']
            drivers_with_margin.append((driver, vals['revenus'], vals['couts'], marge, vals['missions']))
        
        # Trier par marge et prendre top 15
        sorted_drivers = sorted(drivers_with_margin, key=lambda x: x[3], reverse=True)[:15]
        
        if not sorted_drivers:
            ax.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        drivers = [d[0][:15] for d in sorted_drivers]  # Tronquer les noms
        revenus = [d[1] for d in sorted_drivers]
        couts = [d[2] for d in sorted_drivers]
        marges = [d[3] for d in sorted_drivers]
        
        x = range(len(drivers))
        width = 0.35
        
        bars1 = ax.bar([i - width/2 for i in x], revenus, width, label='Revenus', color='green', alpha=0.7)
        bars2 = ax.bar([i + width/2 for i in x], couts, width, label='Coûts', color='red', alpha=0.7)
        
        # Ligne de marge
        ax2 = ax.twinx()
        ax2.plot(x, marges, 'b-', marker='D', linewidth=2, label='Marge', markersize=6)
        ax2.axhline(y=0, color='blue', linestyle='--', alpha=0.5)
        ax2.set_ylabel('Marge (€)', color='blue')
        ax2.tick_params(axis='y', labelcolor='blue')
        
        ax.set_xlabel('Chauffeur')
        ax.set_ylabel('Montant (€)')
        ax.set_xticks(x)
        ax.set_xticklabels(drivers, rotation=45, ha='right')
        ax.legend(loc='upper left')
        ax2.legend(loc='upper right')
        ax.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout()
    
    def _create_profitability_chart(self, fig, data):
        """Créer le graphique de rentabilité par pays"""
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)
        
        country_data = data['by_country']
        if not country_data:
            ax1.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            ax2.text(0.5, 0.5, "Aucune donnée disponible", ha='center', va='center', fontsize=12)
            return
        
        countries = list(country_data.keys())
        revenus = [country_data[c]['revenus'] for c in countries]
        couts = [country_data[c]['couts'] for c in countries]
        marges = [r - c for r, c in zip(revenus, couts)]
        missions = [country_data[c]['missions'] for c in countries]
        
        # Graphique 1: Camembert des revenus
        colors = ['#2ecc71', '#3498db', '#9b59b6', '#e74c3c', '#f39c12'][:len(countries)]
        if sum(revenus) > 0:
            wedges, texts, autotexts = ax1.pie(revenus, labels=countries, autopct='%1.1f%%',
                                               colors=colors, startangle=90)
            ax1.set_title('Répartition des revenus par pays')
        else:
            ax1.text(0.5, 0.5, "Aucun revenu", ha='center', va='center')
        
        # Graphique 2: Barres empilées revenus/coûts/marge
        x = range(len(countries))
        width = 0.6
        
        bars_rev = ax2.bar(x, revenus, width, label='Revenus', color='green', alpha=0.7)
        bars_cout = ax2.bar(x, [-c for c in couts], width, label='Coûts', color='red', alpha=0.7)
        
        # Ajouter la marge comme texte
        for i, (rev, cout, marge) in enumerate(zip(revenus, couts, marges)):
            color = 'green' if marge >= 0 else 'red'
            sign = '+' if marge >= 0 else ''
            ax2.text(i, rev + 50, f'{sign}{marge:,.0f}€', ha='center', va='bottom', 
                    fontsize=9, color=color, fontweight='bold')
        
        ax2.set_xlabel('Pays')
        ax2.set_ylabel('Montant (€)')
        ax2.set_xticks(x)
        ax2.set_xticklabels(countries)
        ax2.axhline(y=0, color='black', linewidth=0.5)
        ax2.legend()
        ax2.set_title('Revenus vs Coûts par pays')
        ax2.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout()
    
    def _create_summary_stats(self, data, start_date, end_date):
        """Créer le résumé statistique"""
        summary_frame = ttk.LabelFrame(self.charts_container, text="📋 Résumé statistique", padding=10)
        summary_frame.pack(fill="x", pady=10, padx=5)
        
        # Calculs
        total_revenus = sum(data['revenus'])
        total_couts = sum(data['couts'])
        total_marge = sum(data['marges'])
        total_missions = sum(data['missions_count'])
        total_pal_liv = sum(data['palettes_liv'])
        total_pal_ram = sum(data['palettes_ram'])
        
        nb_days = (end_date - start_date).days + 1
        days_with_data = len([r for r in data['revenus'] if r > 0])
        
        avg_rev_per_day = total_revenus / days_with_data if days_with_data > 0 else 0
        avg_missions_per_day = total_missions / days_with_data if days_with_data > 0 else 0
        margin_pct = (total_marge / total_revenus * 100) if total_revenus > 0 else 0
        
        # Affichage en colonnes
        cols_frame = ttk.Frame(summary_frame)
        cols_frame.pack(fill="x")
        
        # Colonne 1
        col1 = ttk.Frame(cols_frame)
        col1.pack(side="left", fill="both", expand=True, padx=10)
        ttk.Label(col1, text="💰 Finances", font=("Arial", 11, "bold")).pack(anchor="w")
        ttk.Label(col1, text=f"Revenus totaux: {total_revenus:,.2f} €").pack(anchor="w")
        ttk.Label(col1, text=f"Coûts totaux: {total_couts:,.2f} €").pack(anchor="w")
        marge_color = "green" if total_marge >= 0 else "red"
        marge_label = ttk.Label(col1, text=f"Marge: {total_marge:,.2f} € ({margin_pct:.1f}%)")
        marge_label.pack(anchor="w")
        
        # Colonne 2
        col2 = ttk.Frame(cols_frame)
        col2.pack(side="left", fill="both", expand=True, padx=10)
        ttk.Label(col2, text="📦 Activité", font=("Arial", 11, "bold")).pack(anchor="w")
        ttk.Label(col2, text=f"Missions totales: {total_missions}").pack(anchor="w")
        ttk.Label(col2, text=f"Palettes livrées: {total_pal_liv}").pack(anchor="w")
        ttk.Label(col2, text=f"Palettes ramassées: {total_pal_ram}").pack(anchor="w")
        
        # Colonne 3
        col3 = ttk.Frame(cols_frame)
        col3.pack(side="left", fill="both", expand=True, padx=10)
        ttk.Label(col3, text="📊 Moyennes", font=("Arial", 11, "bold")).pack(anchor="w")
        ttk.Label(col3, text=f"Période: {nb_days} jours ({days_with_data} avec données)").pack(anchor="w")
        ttk.Label(col3, text=f"Revenus/jour: {avg_rev_per_day:,.2f} €").pack(anchor="w")
        ttk.Label(col3, text=f"Missions/jour: {avg_missions_per_day:.1f}").pack(anchor="w")
        
        # Colonne 4 - Top performers
        col4 = ttk.Frame(cols_frame)
        col4.pack(side="left", fill="both", expand=True, padx=10)
        ttk.Label(col4, text="🏆 Top Performers", font=("Arial", 11, "bold")).pack(anchor="w")
        
        # Top voyage
        if data['by_voyage']:
            top_voyage = max(data['by_voyage'].items(), key=lambda x: x[1]['revenus'])
            ttk.Label(col4, text=f"Voyage: {top_voyage[0]} ({top_voyage[1]['revenus']:,.0f}€)").pack(anchor="w")
        
        # Top chauffeur
        if data['by_driver']:
            top_driver = max(data['by_driver'].items(), 
                           key=lambda x: x[1]['revenus'] - x[1]['couts'])
            marge_driver = top_driver[1]['revenus'] - top_driver[1]['couts']
            ttk.Label(col4, text=f"Chauffeur: {top_driver[0][:12]}... ({marge_driver:,.0f}€)").pack(anchor="w")
    
    def export_analyse_data(self):
        """Exporter les données d'analyse"""
        try:
            start_date = parse_date_input(self.analyse_date_start_var.get())
            end_date = parse_date_input(self.analyse_date_end_var.get())
        except ValueError as e:
            messagebox.showerror("Erreur", f"Date invalide: {e}")
            return
        
        data = self.get_analyse_data(start_date, end_date)
        
        if not EXCEL_AVAILABLE:
            messagebox.showwarning("Attention", "openpyxl n'est pas installé. Export Excel non disponible.")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Feuille 1: Données quotidiennes
            ws1 = wb.active
            ws1.title = "Données quotidiennes"
            headers = ["Date", "Revenus", "Coûts", "Marge", "Missions", "Pal. Livraison", "Pal. Ramasse"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row, (d, rev, cout, marge, missions, pal_liv, pal_ram) in enumerate(zip(
                data['dates'], data['revenus'], data['couts'], data['marges'],
                data['missions_count'], data['palettes_liv'], data['palettes_ram']), 2):
                ws1.cell(row=row, column=1, value=d.strftime("%d/%m/%Y"))
                ws1.cell(row=row, column=2, value=rev)
                ws1.cell(row=row, column=3, value=cout)
                ws1.cell(row=row, column=4, value=marge)
                ws1.cell(row=row, column=5, value=missions)
                ws1.cell(row=row, column=6, value=pal_liv)
                ws1.cell(row=row, column=7, value=pal_ram)
            
            # Feuille 2: Par voyage
            ws2 = wb.create_sheet("Par Voyage")
            headers2 = ["Code Voyage", "Revenus", "Missions", "Palettes"]
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row, (voyage, vals) in enumerate(sorted(data['by_voyage'].items(), 
                                                       key=lambda x: x[1]['revenus'], reverse=True), 2):
                ws2.cell(row=row, column=1, value=voyage)
                ws2.cell(row=row, column=2, value=vals['revenus'])
                ws2.cell(row=row, column=3, value=vals['missions'])
                ws2.cell(row=row, column=4, value=vals['palettes'])
            
            # Feuille 3: Par SST
            ws3 = wb.create_sheet("Par SST")
            headers3 = ["SST", "Coûts", "Jours-Chauffeurs"]
            for col, header in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row, (sst, vals) in enumerate(sorted(data['by_sst'].items(), 
                                                    key=lambda x: x[1]['couts'], reverse=True), 2):
                ws3.cell(row=row, column=1, value=sst)
                ws3.cell(row=row, column=2, value=vals['couts'])
                ws3.cell(row=row, column=3, value=vals['chauffeurs'])
            
            # Feuille 4: Par chauffeur
            ws4 = wb.create_sheet("Par Chauffeur")
            headers4 = ["Chauffeur", "Revenus", "Coûts", "Marge", "Missions", "Palettes"]
            for col, header in enumerate(headers4, 1):
                cell = ws4.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for row, (driver, vals) in enumerate(sorted(data['by_driver'].items(), 
                                                       key=lambda x: x[1]['revenus'] - x[1]['couts'], reverse=True), 2):
                ws4.cell(row=row, column=1, value=driver)
                ws4.cell(row=row, column=2, value=vals['revenus'])
                ws4.cell(row=row, column=3, value=vals['couts'])
                ws4.cell(row=row, column=4, value=vals['revenus'] - vals['couts'])
                ws4.cell(row=row, column=5, value=vals['missions'])
                ws4.cell(row=row, column=6, value=vals['palettes'])
            
            # Sauvegarder
            export_dir = ROOT_DIR / "_export"
            export_dir.mkdir(parents=True, exist_ok=True)
            filename = export_dir / f"analyse_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("Succès", f"Données exportées vers:\n{filename}")
            
            # Ouvrir le dossier
            import subprocess
            subprocess.Popen(f'explorer /select,"{filename}"')
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export: {e}")
    
    def refresh_analyse_view(self):
        """Rafraîchir la vue analyse (appelé par le système d'auto-refresh)"""
        # Ne pas rafraîchir automatiquement les graphiques car c'est coûteux
        # L'utilisateur doit cliquer sur le bouton pour générer
        pass


    def build_rights_tab(self):
        self.tab_rights = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_rights, text="Droits")

        main = ttk.Frame(self.tab_rights)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Label(main, text="Utilisateurs & rôles").pack(anchor="w")
        cols = ("user", "roles")
        self.tree_users = ttk.Treeview(main, columns=cols, show="headings", height=15, selectmode="extended")
        self.tree_users.heading("user", text="Utilisateur Windows")
        self.tree_users.heading("roles", text="Rôles")
        self.tree_users.column("user", width=150)
        self.tree_users.column("roles", width=200)
        self.tree_users.pack(fill="both", expand=True)

        btn = ttk.Frame(main)
        btn.pack(fill="x", pady=5)
        ttk.Button(btn, text="Ajouter / modifier l'utilisateur", command=self.on_add_edit_user).pack(side="left")
        ttk.Button(btn, text="Rafraîchir", command=self.refresh_users_view).pack(side="left", padx=5)

        self.refresh_users_view()

    def refresh_users_view(self):
        if not hasattr(self, "tree_users"):
            return
        for row in self.tree_users.get_children():
            self.tree_users.delete(row)
        users = self.rights.get("users_def", {})
        for user, roles in users.items():
            self.tree_users.insert("", "end", iid=user, values=(user, ", ".join(roles)))

    def on_add_edit_user(self):
        self.set_user_editing(True)
        win = tk.Toplevel(self.root)
        win.transient(self.root)
        win.grab_set()
        win.title("Ajouter / modifier utilisateur")

        user_var = tk.StringVar()

        ttk.Label(win, text="Utilisateur Windows (login) :").grid(row=0, column=0, sticky="w")
        ttk.Entry(win, textvariable=user_var, width=25).grid(row=0, column=1, sticky="w")

        roles_def = self.rights["roles_def"]
        role_names = list(roles_def.keys())
        role_vars = {}

        selected_user = None
        if hasattr(self, "tree_users"):
            sel = self.tree_users.selection()
            if sel:
                selected_user = sel[0]
                user_var.set(selected_user)

        existing_roles = []
        if selected_user:
            existing_roles = self.rights["users_def"].get(selected_user, [])

        row = 1
        ttk.Label(win, text="Rôles :").grid(row=row, column=0, sticky="nw")
        roles_frame = ttk.Frame(win)
        roles_frame.grid(row=row, column=1, sticky="w")
        for i, rname in enumerate(role_names):
            var = tk.BooleanVar(value=(rname in existing_roles))
            cb = ttk.Checkbutton(roles_frame, text=rname, variable=var)
            cb.grid(row=i, column=0, sticky="w")
            role_vars[rname] = var
        row += 1

        def on_ok():
            user = user_var.get().strip().upper()
            if not user:
                messagebox.showerror("Erreur", "Utilisateur obligatoire.")
                return
            roles = [name for name, var in role_vars.items() if var.get()]
            if not roles:
                messagebox.showerror("Erreur", "Sélectionnez au moins un rôle.")
                return

            rights_path = self.data_dir / "users_rights.json"
            rights = load_json(rights_path, {})
            if "users" not in rights:
                rights["users"] = {}
            
            # Logger la modification des droits
            old_roles = rights.get("users", {}).get(user, [])
            activity_logger.log_action("RIGHTS_CHANGE", {
                "target_user": user,
                "old_roles": old_roles,
                "new_roles": roles,
            })
            
            rights["users"][user] = roles
            save_json(rights_path, rights)

            self.rights = self.load_rights()
            self.refresh_users_view()
            self.set_user_editing(False)
            win.destroy()
            messagebox.showinfo(
                "Info",
                "Droits mis à jour. Relancez l'application pour appliquer complètement les nouveaux droits.",
            )
            
        def on_cancel():
            self.set_user_editing(False)
            win.destroy()

        ttk.Button(win, text="OK", command=on_ok).grid(row=row, column=1, pady=5, sticky="e")
        ttk.Button(win, text="Annuler", command=on_cancel).grid(row=row, column=0, pady=5, sticky="w")

    # ==========================================================================
    # ONGLET SAURON - Surveillance des activités utilisateurs
    # ==========================================================================
    
    def build_sauron_tab(self):
        """Construire l'onglet Sauron pour la surveillance des activités"""
        self.tab_sauron = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_sauron, text="👁 Sauron")
        
        # Frame principal avec PanedWindow pour être responsive
        self.sauron_paned = ttk.PanedWindow(self.tab_sauron, orient="horizontal")
        self.sauron_paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        # === PANNEAU GAUCHE : Liste des utilisateurs ===
        left_frame = ttk.Frame(self.sauron_paned)
        self.sauron_paned.add(left_frame, weight=1)
        
        # En-tête utilisateurs
        header_left = ttk.Frame(left_frame)
        header_left.pack(fill="x", padx=5, pady=5)
        ttk.Label(header_left, text="👥 Utilisateurs", font=('Arial', 11, 'bold')).pack(side="left")
        self.sauron_active_count = ttk.Label(header_left, text="(0 actif)", foreground="green")
        self.sauron_active_count.pack(side="left", padx=10)
        
        # Bouton rafraîchir
        ttk.Button(header_left, text="🔄 Actualiser", command=self.sauron_refresh_all).pack(side="right")
        
        # Filtre de recherche
        filter_frame = ttk.Frame(left_frame)
        filter_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(filter_frame, text="🔍").pack(side="left")
        self.sauron_filter_var = tk.StringVar()
        self.sauron_filter_var.trace("w", lambda *args: self.sauron_filter_users())
        filter_entry = ttk.Entry(filter_frame, textvariable=self.sauron_filter_var)
        filter_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        # Liste des utilisateurs avec scrollbar
        users_container = ttk.Frame(left_frame)
        users_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Treeview pour les utilisateurs
        columns = ("user", "status", "sessions", "temps_total", "derniere_connexion")
        self.sauron_users_tree = ttk.Treeview(users_container, columns=columns, show="headings", selectmode="browse")
        self.sauron_users_tree.heading("user", text="Utilisateur")
        self.sauron_users_tree.heading("status", text="État")
        self.sauron_users_tree.heading("sessions", text="Sessions")
        self.sauron_users_tree.heading("temps_total", text="Temps total")
        self.sauron_users_tree.heading("derniere_connexion", text="Dernière connexion")
        
        self.sauron_users_tree.column("user", width=120, minwidth=80)
        self.sauron_users_tree.column("status", width=70, minwidth=50)
        self.sauron_users_tree.column("sessions", width=60, minwidth=40)
        self.sauron_users_tree.column("temps_total", width=90, minwidth=60)
        self.sauron_users_tree.column("derniere_connexion", width=140, minwidth=100)
        
        # Scrollbar pour la liste des utilisateurs
        scrollbar_users = ttk.Scrollbar(users_container, orient="vertical", command=self.sauron_users_tree.yview)
        self.sauron_users_tree.configure(yscrollcommand=scrollbar_users.set)
        
        self.sauron_users_tree.pack(side="left", fill="both", expand=True)
        scrollbar_users.pack(side="right", fill="y")
        
        # Bind sélection utilisateur
        self.sauron_users_tree.bind("<<TreeviewSelect>>", self.sauron_on_user_select)
        
        # Tags pour les couleurs
        self.sauron_users_tree.tag_configure('active', background='#C8E6C9', foreground='#1B5E20')
        self.sauron_users_tree.tag_configure('inactive', background='#FFFFFF')
        
        # === PANNEAU DROIT : Détails et historique ===
        right_frame = ttk.Frame(self.sauron_paned)
        self.sauron_paned.add(right_frame, weight=2)
        
        # En-tête détails
        header_right = ttk.Frame(right_frame)
        header_right.pack(fill="x", padx=5, pady=5)
        self.sauron_detail_title = ttk.Label(header_right, text="📋 Sélectionnez un utilisateur", font=('Arial', 11, 'bold'))
        self.sauron_detail_title.pack(side="left")
        
        # Statistiques utilisateur
        stats_frame = ttk.LabelFrame(right_frame, text="📊 Statistiques", padding=10)
        stats_frame.pack(fill="x", padx=5, pady=5)
        
        # Grid pour les stats (responsive)
        stats_frame.columnconfigure((0, 1, 2, 3), weight=1)
        
        self.sauron_stat_sessions = ttk.Label(stats_frame, text="Sessions: -")
        self.sauron_stat_sessions.grid(row=0, column=0, sticky="w", padx=5)
        
        self.sauron_stat_time = ttk.Label(stats_frame, text="Temps total: -")
        self.sauron_stat_time.grid(row=0, column=1, sticky="w", padx=5)
        
        self.sauron_stat_actions = ttk.Label(stats_frame, text="Actions: -")
        self.sauron_stat_actions.grid(row=0, column=2, sticky="w", padx=5)
        
        self.sauron_stat_last = ttk.Label(stats_frame, text="Dernière connexion: -")
        self.sauron_stat_last.grid(row=0, column=3, sticky="w", padx=5)
        
        # Filtres pour l'historique
        filters_frame = ttk.Frame(right_frame)
        filters_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(filters_frame, text="📅 Période:").pack(side="left")
        self.sauron_date_filter = ttk.Combobox(filters_frame, values=[
            "Aujourd'hui", "Hier", "7 derniers jours", "30 derniers jours", "Tout"
        ], state="readonly", width=15)
        self.sauron_date_filter.set("Aujourd'hui")
        self.sauron_date_filter.pack(side="left", padx=5)
        self.sauron_date_filter.bind("<<ComboboxSelected>>", lambda e: self.sauron_filter_actions())
        
        ttk.Label(filters_frame, text="🏷 Type:").pack(side="left", padx=(15, 0))
        self.sauron_type_filter = ttk.Combobox(filters_frame, values=[
            "Tous", "SESSION", "MISSION", "CHAUFFEUR", "VOYAGE", "TAB_CHANGE", "Autres"
        ], state="readonly", width=12)
        self.sauron_type_filter.set("Tous")
        self.sauron_type_filter.pack(side="left", padx=5)
        self.sauron_type_filter.bind("<<ComboboxSelected>>", lambda e: self.sauron_filter_actions())
        
        # Export
        ttk.Button(filters_frame, text="📤 Exporter", command=self.sauron_export_logs).pack(side="right")
        
        # Historique des actions avec scrollbar
        history_frame = ttk.LabelFrame(right_frame, text="📜 Historique des actions", padding=5)
        history_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Treeview pour l'historique
        columns_hist = ("timestamp", "type", "details")
        self.sauron_history_tree = ttk.Treeview(history_frame, columns=columns_hist, show="headings", selectmode="browse")
        self.sauron_history_tree.heading("timestamp", text="Date/Heure")
        self.sauron_history_tree.heading("type", text="Type")
        self.sauron_history_tree.heading("details", text="Détails")
        
        self.sauron_history_tree.column("timestamp", width=150, minwidth=120)
        self.sauron_history_tree.column("type", width=120, minwidth=80)
        self.sauron_history_tree.column("details", width=400, minwidth=200)
        
        # Scrollbars
        scrollbar_hist_y = ttk.Scrollbar(history_frame, orient="vertical", command=self.sauron_history_tree.yview)
        scrollbar_hist_x = ttk.Scrollbar(history_frame, orient="horizontal", command=self.sauron_history_tree.xview)
        self.sauron_history_tree.configure(yscrollcommand=scrollbar_hist_y.set, xscrollcommand=scrollbar_hist_x.set)
        
        self.sauron_history_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_hist_y.grid(row=0, column=1, sticky="ns")
        scrollbar_hist_x.grid(row=1, column=0, sticky="ew")
        
        history_frame.columnconfigure(0, weight=1)
        history_frame.rowconfigure(0, weight=1)
        
        # Double-clic pour voir les détails
        self.sauron_history_tree.bind("<Double-1>", self.sauron_show_action_details)
        
        # Tags pour colorer les types d'actions
        self.sauron_history_tree.tag_configure('session_start', background='#E8F5E9')
        self.sauron_history_tree.tag_configure('session_end', background='#FFEBEE')
        self.sauron_history_tree.tag_configure('mission', background='#E3F2FD')
        self.sauron_history_tree.tag_configure('chauffeur', background='#FFF3E0')
        self.sauron_history_tree.tag_configure('voyage', background='#F3E5F5')
        
        # Stocker les données d'actions pour les détails
        self.sauron_actions_data = {}
        
        # Charger les données initiales
        self.sauron_refresh_all()
        
        # Auto-refresh toutes les 30 secondes
        self._sauron_refresh_job = None
        self._start_sauron_auto_refresh()
    
    def _start_sauron_auto_refresh(self):
        """Démarrer le rafraîchissement automatique de Sauron"""
        if self._sauron_refresh_job:
            self.root.after_cancel(self._sauron_refresh_job)
        
        def refresh_loop():
            if hasattr(self, 'tab_sauron') and self.tab_sauron.winfo_exists():
                # Ne rafraîchir que si l'onglet Sauron est visible
                try:
                    current_tab = self.notebook.tab(self.notebook.select(), "text")
                    if "Sauron" in current_tab:
                        self.sauron_refresh_all()
                except Exception:
                    pass
                self._sauron_refresh_job = self.root.after(30000, refresh_loop)
        
        self._sauron_refresh_job = self.root.after(30000, refresh_loop)
    
    def sauron_refresh_all(self):
        """Rafraîchir toutes les données Sauron"""
        self.sauron_refresh_users_list()
        # Rafraîchir aussi les détails si un utilisateur est sélectionné
        sel = self.sauron_users_tree.selection()
        if sel:
            self.sauron_on_user_select(None)
    
    def sauron_refresh_users_list(self):
        """Rafraîchir la liste des utilisateurs"""
        # Sauvegarder la sélection
        old_selection = self.sauron_users_tree.selection()
        old_selection_user = old_selection[0] if old_selection else None
        
        # Effacer la liste
        for item in self.sauron_users_tree.get_children():
            self.sauron_users_tree.delete(item)
        
        # Récupérer tous les logs
        all_logs = activity_logger.get_all_users_logs()
        active_users = activity_logger.get_active_sessions()
        active_usernames = {u["user"] for u in active_users}
        
        # Compter les actifs
        active_count = len(active_usernames)
        self.sauron_active_count.config(text=f"({active_count} actif{'s' if active_count > 1 else ''})")
        
        # Ajouter chaque utilisateur
        for username, logs in sorted(all_logs.items()):
            stats = activity_logger.get_user_stats(username)
            if not stats:
                continue
            
            is_active = username in active_usernames
            status = "🟢 Actif" if is_active else "⚪ Inactif"
            
            # Formater la dernière connexion
            last_login = stats.get("last_login", "-")
            if last_login and last_login != "-":
                try:
                    last_dt = datetime.fromisoformat(last_login)
                    last_login = last_dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            
            values = (
                username,
                status,
                stats.get("total_sessions", 0),
                stats.get("total_time_formatted", "-"),
                last_login
            )
            
            tag = 'active' if is_active else 'inactive'
            self.sauron_users_tree.insert("", "end", iid=username, values=values, tags=(tag,))
        
        # Restaurer la sélection
        if old_selection_user and self.sauron_users_tree.exists(old_selection_user):
            self.sauron_users_tree.selection_set(old_selection_user)
    
    def sauron_filter_users(self):
        """Filtrer la liste des utilisateurs"""
        filter_text = self.sauron_filter_var.get().upper()
        
        for item in self.sauron_users_tree.get_children():
            values = self.sauron_users_tree.item(item, "values")
            username = values[0] if values else ""
            
            if filter_text in username.upper():
                # Afficher (pas de méthode hide, on reconstruit)
                pass
        
        # Pour un vrai filtrage, il faudrait reconstruire la liste
        # Simplifié ici: juste mettre en évidence
        self.sauron_refresh_users_list()
    
    def sauron_on_user_select(self, event):
        """Quand un utilisateur est sélectionné"""
        sel = self.sauron_users_tree.selection()
        if not sel:
            return
        
        username = sel[0]
        self.sauron_selected_user = username
        self.sauron_detail_title.config(text=f"📋 Détails: {username}")
        
        # Mettre à jour les statistiques
        stats = activity_logger.get_user_stats(username)
        if stats:
            self.sauron_stat_sessions.config(text=f"Sessions: {stats.get('total_sessions', 0)}")
            self.sauron_stat_time.config(text=f"Temps total: {stats.get('total_time_formatted', '-')}")
            self.sauron_stat_actions.config(text=f"Actions: {stats.get('total_actions', 0)}")
            
            last_login = stats.get("last_login", "-")
            if last_login and last_login != "-":
                try:
                    last_dt = datetime.fromisoformat(last_login)
                    last_login = last_dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    pass
            self.sauron_stat_last.config(text=f"Dernière connexion: {last_login}")
        
        # Charger l'historique
        self.sauron_load_user_history(username)
    
    def sauron_load_user_history(self, username):
        """Charger l'historique d'un utilisateur"""
        # Effacer l'historique actuel
        for item in self.sauron_history_tree.get_children():
            self.sauron_history_tree.delete(item)
        self.sauron_actions_data = {}
        
        # Récupérer les logs
        all_logs = activity_logger.get_all_users_logs()
        if username not in all_logs:
            return
        
        actions = all_logs[username].get("actions", [])
        
        # Appliquer les filtres
        filtered_actions = self._sauron_apply_filters(actions)
        
        # Ajouter les actions (les plus récentes en premier)
        for action in reversed(filtered_actions[-2000:]):  # Afficher jusqu'à 2000 actions
            action_id = action.get("id", str(uuid.uuid4())[:12])
            timestamp = action.get("timestamp", "-")
            action_type = action.get("type", "UNKNOWN")
            details = action.get("details", {})
            
            # Formater le timestamp
            try:
                dt = datetime.fromisoformat(timestamp)
                timestamp_fmt = dt.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                timestamp_fmt = timestamp
            
            # Formater les détails pour l'affichage
            details_str = self._format_action_details(action)
            
            # Déterminer le tag de couleur
            tag = self._get_action_tag(action_type)
            
            self.sauron_history_tree.insert("", "end", iid=action_id, 
                                           values=(timestamp_fmt, action_type, details_str),
                                           tags=(tag,))
            
            # Stocker les données complètes pour les détails
            self.sauron_actions_data[action_id] = action
    
    def _sauron_apply_filters(self, actions):
        """Appliquer les filtres de date et de type"""
        # Filtre de date
        date_filter = self.sauron_date_filter.get()
        now = datetime.now()
        
        if date_filter == "Aujourd'hui":
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_filter == "Hier":
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_filter == "7 derniers jours":
            start_date = now - timedelta(days=7)
        elif date_filter == "30 derniers jours":
            start_date = now - timedelta(days=30)
        else:  # Tout
            start_date = None
        
        filtered = []
        for action in actions:
            try:
                action_time = datetime.fromisoformat(action.get("timestamp", ""))
                
                # Filtre date
                if start_date and action_time < start_date:
                    continue
                if date_filter == "Hier" and action_time >= end_date:
                    continue
                
                # Filtre type
                type_filter = self.sauron_type_filter.get()
                action_type = action.get("type", "")
                
                if type_filter != "Tous":
                    if type_filter == "SESSION" and not action_type.startswith("SESSION"):
                        continue
                    elif type_filter == "MISSION" and not action_type.startswith("MISSION"):
                        continue
                    elif type_filter == "CHAUFFEUR" and not action_type.startswith("CHAUFFEUR"):
                        continue
                    elif type_filter == "VOYAGE" and not action_type.startswith("VOYAGE"):
                        continue
                    elif type_filter == "TAB_CHANGE" and action_type != "TAB_CHANGE":
                        continue
                    elif type_filter == "Autres" and action_type.startswith(("SESSION", "MISSION", "CHAUFFEUR", "VOYAGE", "TAB_CHANGE")):
                        continue
                
                filtered.append(action)
            except Exception:
                continue
        
        return filtered
    
    def sauron_filter_actions(self):
        """Réappliquer les filtres sur l'historique"""
        if hasattr(self, 'sauron_selected_user'):
            self.sauron_load_user_history(self.sauron_selected_user)
    
    def _format_action_details(self, action):
        """Formater les détails d'une action pour l'affichage"""
        action_type = action.get("type", "")
        details = action.get("details", {})
        
        if action_type == "SESSION_START":
            return f"Début de session"
        elif action_type == "SESSION_END":
            duration = details.get("duration_formatted", "-")
            return f"Fin de session (durée: {duration})"
        elif action_type == "TAB_CHANGE":
            return f"Onglet: {details.get('tab', '-')}"
        elif action_type.startswith("MISSION"):
            voyage = details.get("voyage", "-")
            mission_id = details.get("mission_id", "-")[:8] if details.get("mission_id") else "-"
            return f"Voyage: {voyage}, ID: {mission_id}"
        elif action_type.startswith("CHAUFFEUR"):
            nom = details.get("nom", "-")
            return f"Chauffeur: {nom}"
        elif action_type.startswith("VOYAGE"):
            code = details.get("code", "-")
            return f"Code: {code}"
        else:
            # Affichage générique
            if details:
                return str(details)[:100]
            return "-"
    
    def _get_action_tag(self, action_type):
        """Obtenir le tag de couleur pour un type d'action"""
        if action_type == "SESSION_START":
            return 'session_start'
        elif action_type == "SESSION_END":
            return 'session_end'
        elif action_type.startswith("MISSION"):
            return 'mission'
        elif action_type.startswith("CHAUFFEUR"):
            return 'chauffeur'
        elif action_type.startswith("VOYAGE"):
            return 'voyage'
        return ''
    
    def sauron_show_action_details(self, event):
        """Afficher les détails complets d'une action"""
        sel = self.sauron_history_tree.selection()
        if not sel:
            return
        
        action_id = sel[0]
        action = self.sauron_actions_data.get(action_id, {})
        
        if not action:
            return
        
        # Créer une fenêtre de détails
        win = tk.Toplevel(self.root)
        win.title("Détails de l'action")
        win.geometry("600x500")
        win.transient(self.root)
        
        # Frame principal avec scrollbar
        main_frame = ttk.Frame(win, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        # Informations de base
        info_frame = ttk.LabelFrame(main_frame, text="📋 Informations", padding=10)
        info_frame.pack(fill="x", pady=5)
        
        ttk.Label(info_frame, text=f"Type: {action.get('type', '-')}", font=('Arial', 10, 'bold')).pack(anchor="w")
        
        timestamp = action.get("timestamp", "-")
        try:
            dt = datetime.fromisoformat(timestamp)
            timestamp = dt.strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            pass
        ttk.Label(info_frame, text=f"Date/Heure: {timestamp}").pack(anchor="w")
        ttk.Label(info_frame, text=f"Session: {action.get('session_id', '-')}").pack(anchor="w")
        
        # Détails
        details_frame = ttk.LabelFrame(main_frame, text="📝 Détails", padding=10)
        details_frame.pack(fill="both", expand=True, pady=5)
        
        details_text = tk.Text(details_frame, wrap="word", height=10)
        details_text.pack(fill="both", expand=True)
        
        details = action.get("details", {})
        details_text.insert("1.0", json.dumps(details, indent=2, ensure_ascii=False))
        
        # État avant/après si disponible
        if "before" in action or "after" in action:
            changes_frame = ttk.LabelFrame(main_frame, text="🔄 Modifications", padding=10)
            changes_frame.pack(fill="both", expand=True, pady=5)
            
            changes_text = tk.Text(changes_frame, wrap="word", height=8)
            changes_text.pack(fill="both", expand=True)
            
            if "before" in action:
                changes_text.insert("end", "AVANT:\n")
                changes_text.insert("end", json.dumps(action["before"], indent=2, ensure_ascii=False))
                changes_text.insert("end", "\n\n")
            
            if "after" in action:
                changes_text.insert("end", "APRÈS:\n")
                changes_text.insert("end", json.dumps(action["after"], indent=2, ensure_ascii=False))
        
        # Bouton fermer
        ttk.Button(main_frame, text="Fermer", command=win.destroy).pack(pady=10)
    
    def sauron_export_logs(self):
        """Exporter les logs de l'utilisateur sélectionné"""
        if not hasattr(self, 'sauron_selected_user'):
            messagebox.showwarning("Attention", "Sélectionnez d'abord un utilisateur.")
            return
        
        username = self.sauron_selected_user
        all_logs = activity_logger.get_all_users_logs()
        
        if username not in all_logs:
            messagebox.showerror("Erreur", "Aucun log trouvé pour cet utilisateur.")
            return
        
        # Demander le chemin de sauvegarde
        from tkinter import filedialog
        
        filename = filedialog.asksaveasfilename(
            title="Exporter les logs",
            defaultextension=".json",
            filetypes=[("Fichiers JSON", "*.json"), ("Tous les fichiers", "*.*")],
            initialfile=f"logs_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        
        if filename:
            try:
                with open(filename, "w", encoding="utf-8") as f:
                    json.dump(all_logs[username], f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Succès", f"Logs exportés vers:\n{filename}")
                activity_logger.log_action("SAURON_EXPORT", {"user_exported": username, "file": filename})
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de l'export:\n{e}")

    def reload_data_from_files(self):
        """Recharger les données depuis les fichiers JSON - seulement si modifiés.
        
        - Surveille les fichiers de référentiels (_data)
        - Surveille aussi le planning du jour courant dans le dossier _planning
        """
        try:
            import os
            from pathlib import Path
            
            # 1) Vérifier les fichiers de référentiels (_data)
            files_to_check = {
                'voyages': VOYAGES_FILE,
                'chauffeurs': CHAUFFEURS_FILE,
                'dispos': self.data_dir / "dispo_chauffeurs.json",
                'sst': self.data_dir / "sst.json",
                'tarifs_sst': TARIFS_SST_FILE,
                'revenus': REVENUS_FILE,
            }
            
            files_changed = False
            
            for key, filepath in files_to_check.items():
                try:
                    if filepath.exists():
                        current_mtime = filepath.stat().st_mtime
                        # Vérifier si le fichier a changé
                        if key not in self.file_timestamps or self.file_timestamps[key] != current_mtime:
                            files_changed = True
                            self.file_timestamps[key] = current_mtime
                except Exception:
                    # En cas de souci sur un fichier, on préfère rafraîchir quand même
                    files_changed = True
            
            # 2) Vérifier le planning du jour courant (_planning/YYYY/MM/Semaine_xx/AAAA-MM-JJ)
            missions_changed = False
            try:
                day_dir = get_planning_day_dir(self.current_date)
                if day_dir.exists():
                    latest_mtime = None
                    for f in day_dir.glob("*.json"):
                        try:
                            mtime = f.stat().st_mtime
                            if latest_mtime is None or mtime > latest_mtime:
                                latest_mtime = mtime
                        except Exception:
                            pass
                    
                    if latest_mtime is not None:
                        prev = self.file_timestamps.get("missions_current")
                        if prev is None or prev != latest_mtime:
                            missions_changed = True
                            self.file_timestamps["missions_current"] = latest_mtime
            except Exception:
                # En cas d'erreur sur le planning du jour, on ne casse pas tout
                pass
            
            # 3) Appliquer les rechargements nécessaires
            if files_changed:
                # ⚠ IMPORTANT : ne PAS toucher aux missions ici.
                # Les missions du planning sont chargées via load_planning_for_date()
                # et sauvegardées jour par jour dans le dossier _planning.
                # Ici, on ne recharge que les référentiels globaux.
                self.voyages = load_json(VOYAGES_FILE, [])
                self.chauffeurs = load_json(CHAUFFEURS_FILE, [])
                self.dispos = load_json(self.data_dir / "dispo_chauffeurs.json", [])
                self.sst_list = load_json(self.data_dir / "sst.json", [])
                self.tarifs_sst = load_json(TARIFS_SST_FILE, {})
                self.revenus_palettes = load_json(REVENUS_FILE, {})
            
            if missions_changed:
                # Recharger les missions pour la date courante,
                # en conservant au mieux l'état de l'interface
                self.load_planning_for_date(self.current_date, preserve_ui=True)
                # On retourne False pour éviter un second refresh via smart_refresh_all_views
                return False
            
            # Si seuls les référentiels ont changé → on laisse smart_refresh_all_views faire son travail
            if files_changed:
                return True  # Fichiers rechargés (référentiels)
            
            return False  # Aucun changement
                
        except Exception as e:
            print(f"Erreur lors du rechargement des données: {e}")
            return False

    def _start_auto_refresh_loop(self):
        """Démarre la boucle d'auto-rafraîchissement interne."""
        if self._auto_refresh_job is None:
            self._auto_refresh_job = self.root.after(self._auto_refresh_interval_ms, self._auto_refresh_tick)

    def _auto_refresh_tick(self):
        """Tick périodique : recharge les données et rafraîchit les vues si nécessaire."""
        self._auto_refresh_job = None
        try:
            # Ne pas rafraîchir si l'utilisateur est en train d'éditer
            if getattr(self, "user_editing", False) or getattr(self, "is_editing", False):
                return

            # Recharger les données depuis les fichiers (seulement si modifiés)
            changed = self.reload_data_from_files()
            if changed:
                # Rafraîchir intelligemment les vues (par différence, sans clignotement)
                self.smart_refresh_all_views()

            # Mettre à jour l'horodatage de dernière MAJ si quelque chose a été traité
            if changed:
                from datetime import datetime
                self.last_refresh_dt = datetime.now()
                try:
                    self.status_var.set(
                        f"Session : {self.current_user} | Dernière MAJ : {self.last_refresh_dt.strftime('%d/%m/%Y %H:%M:%S')}"
                    )
                except Exception:
                    pass
        except Exception as e:
            print(f"Erreur lors de l'auto-rafraîchissement: {e}")
        finally:
            # Reprogrammer le prochain tick
            self._start_auto_refresh_loop()

    def smart_refresh_all_views(self):
        """Refresh intelligent de toutes les vues - SANS CLIGNOTEMENT"""
        try:
            # Refresh UNIQUEMENT par différence (ZÉRO effacement)
            
            # 1. Planning principal (par pays) : mise à jour par différence
            self.diff_refresh_planning()
            
            # 2. Onglet Chauffeurs : mise à jour de la liste des chauffeurs
            if hasattr(self, 'tree_ch'):
                self.diff_refresh_chauffeurs()
            
            # 3. Onglet Voyages : mise à jour de la liste des voyages
            if hasattr(self, 'tree_voy'):
                self.diff_refresh_voyages()
            
            # 4. Vue finance (SST + revenus) : on ne touche qu'au contenu, pas à la mise en page
            if hasattr(self, 'finance_sst_listbox'):
                self.refresh_finance_view()
            
            # 5. Vue Analyse : ne pas rafraîchir automatiquement (coûteux)
            # L'utilisateur doit cliquer sur "Générer les graphiques"
            
            # 6. Vues disponibilité / planning chauffeurs
            try:
                if hasattr(self, 'drivers_available_frame'):
                    self.refresh_drivers_availability_view()
                if hasattr(self, 'drivers_used_frame'):
                    self.refresh_drivers_used_view()
                if hasattr(self, 'calendar_container'):
                    self.refresh_calendar()
            except Exception as ee:
                print(f"Erreur refresh vues chauffeurs (auto): {ee}")
                
        except Exception as e:
            print(f"Erreur lors du smart refresh: {e}")
            import traceback
            traceback.print_exc()

    def save_all_selections(self):
        """Sauvegarder toutes les sélections et positions de scroll"""
        selections = {}
        
        # Sauvegarder les sélections du planning
        if hasattr(self, 'country_trees'):
            for country, trees in self.country_trees.items():
                selections[f'planning_{country}'] = {}
                for tree_type, tree in trees.items():
                    try:
                        selections[f'planning_{country}'][tree_type] = {
                            'selection': list(tree.selection()),
                            'yview': tree.yview()[0] if tree.yview() else 0
                        }
                    except Exception:
                        pass
        
        # Sauvegarder la sélection des chauffeurs
        if hasattr(self, 'tree_ch'):
            try:
                selections['chauffeurs'] = {
                    'selection': list(self.tree_ch.selection()),
                    'yview': self.tree_ch.yview()[0] if self.tree_ch.yview() else 0
                }
            except Exception:
                pass
        
        # Sauvegarder la sélection des voyages
        if hasattr(self, 'tree_voy'):
            try:
                selections['voyages'] = {
                    'selection': list(self.tree_voy.selection()),
                    'yview': self.tree_voy.yview()[0] if self.tree_voy.yview() else 0
                }
            except Exception:
                pass
        
        return selections
    
    def restore_all_selections(self, selections):
        """Restaurer toutes les sélections et positions de scroll"""
        if not selections:
            return
        
        # Restaurer les sélections du planning
        if hasattr(self, 'country_trees'):
            for country, trees in self.country_trees.items():
                key = f'planning_{country}'
                if key in selections:
                    for tree_type, tree in trees.items():
                        if tree_type in selections[key]:
                            try:
                                saved = selections[key][tree_type]
                                # Restaurer la sélection
                                for item_id in saved.get('selection', []):
                                    if tree.exists(item_id):
                                        tree.selection_add(item_id)
                                # Restaurer le scroll
                                if 'yview' in saved:
                                    tree.yview_moveto(saved['yview'])
                            except Exception:
                                pass
        
        # Restaurer la sélection des chauffeurs
        if hasattr(self, 'tree_ch') and 'chauffeurs' in selections:
            try:
                saved = selections['chauffeurs']
                for item_id in saved.get('selection', []):
                    if self.tree_ch.exists(item_id):
                        self.tree_ch.selection_set(item_id)
                if 'yview' in saved:
                    self.tree_ch.yview_moveto(saved['yview'])
            except Exception:
                pass
        
        # Restaurer la sélection des voyages
        if hasattr(self, 'tree_voy') and 'voyages' in selections:
            try:
                saved = selections['voyages']
                for item_id in saved.get('selection', []):
                    if self.tree_voy.exists(item_id):
                        self.tree_voy.selection_set(item_id)
                if 'yview' in saved:
                    self.tree_voy.yview_moveto(saved['yview'])
            except Exception:
                pass
    
    def smart_refresh_planning(self):
        """Cette méthode n'est plus utilisée - on utilise refresh_planning_view()"""
        pass
    
    def diff_refresh_planning(self):
        """Refresh du planning par DIFFÉRENCE - ZÉRO clignotement"""
        if not hasattr(self, 'country_trees'):
            return
        
        # Pour chaque pays
        for country, trees in self.country_trees.items():
            # Pour chaque type (livraison, ramasse)
            for tree_type, tree in trees.items():
                try:
                    # Sauvegarder sélection et scroll AVANT
                    old_selection = list(tree.selection())
                    old_yview = tree.yview()[0] if tree.yview() else 0
                    
                    # Obtenir l'état ACTUEL du TreeView
                    current_items = {}
                    for item_id in tree.get_children():
                        current_items[item_id] = tree.item(item_id, 'values')
                    
                    # Obtenir les NOUVELLES données
                    new_missions = self.get_missions_for_tree(country, tree_type)
                    
                    # Créer un dict des nouvelles données
                    new_items = {}
                    for mission in new_missions:
                        new_items[mission['id']] = mission['values']
                    
                    # MISE À JOUR PAR DIFFÉRENCE
                    
                    # 1. MODIFIER les items qui ont changé
                    for item_id, new_values in new_items.items():
                        if item_id in current_items:
                            # Item existe déjà
                            if current_items[item_id] != new_values:
                                # Valeurs différentes → MODIFIER (pas supprimer+recréer)
                                tree.item(item_id, values=new_values)
                        else:
                            # Nouvel item → AJOUTER
                            tree.insert("", "end", iid=item_id, values=new_values)
                    
                    # 2. SUPPRIMER les items qui n'existent plus
                    for item_id in current_items.keys():
                        if item_id not in new_items:
                            tree.delete(item_id)
                    
                    # Restaurer sélection et scroll APRÈS
                    for item_id in old_selection:
                        if tree.exists(item_id):
                            tree.selection_add(item_id)
                    
                    tree.yview_moveto(old_yview)
                    
                    # Appliquer les couleurs alternées UNIQUEMENT si des changements
                    if len(current_items) != len(new_items) or current_items.keys() != new_items.keys():
                        for idx, item_id in enumerate(tree.get_children()):
                            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                            tree.item(item_id, tags=(tag,))
                    
                except Exception as e:
                    print(f"Erreur diff_refresh {country}/{tree_type}: {e}")
                    import traceback
                    traceback.print_exc()
    
    def get_missions_for_tree(self, country, tree_type):
        """Obtenir les données formatées pour un TreeView"""
        result = []
        
        # Filtrer les missions pour cette date
        date_str = self.current_date.strftime("%Y-%m-%d")
        missions_today = [m for m in self.missions if m.get("date") == date_str]
        
        for mission in missions_today:
            # Déterminer le pays de la mission via le voyage
            voyage_code = mission.get("voyage", "")
            voyage = next((v for v in self.voyages if v.get("code") == voyage_code), None)
            mission_country = voyage.get("country", "Belgique") if voyage else "Belgique"
            
            # Normaliser le nom du pays
            if mission_country.lower() in ["pays-bas", "paysbas", "pays bas", "netherlands"]:
                mission_country = "Pays-Bas"
            elif mission_country.lower() in ["belgique", "belgium", "be"]:
                mission_country = "Belgique"
            elif mission_country.lower() in ["luxembourg", "lux", "lu"]:
                mission_country = "Luxembourg"
            elif mission_country.lower() in ["france", "fr"]:
                mission_country = "France"
            
            # Filtrer par pays
            if mission_country != country:
                continue
            
            # Filtrer par type
            mission_type = mission.get("type", "LIVRAISON")
            if tree_type == "livraison" and mission_type != "LIVRAISON":
                continue
            if tree_type == "ramasse" and mission_type != "RAMASSE":
                continue
            
            # Formater les valeurs pour le TreeView
            mission_id = mission.get("id", "")
            heure = mission.get("heure", "")
            nb_pal = mission.get("nb_pal", "")
            numero = mission.get("numero", "")
            sst = mission.get("sst", "")
            chauffeur = mission.get("chauffeur_nom", "")
            infos = mission.get("infos", "")
            
            if tree_type == "livraison":
                values = (heure, voyage_code, nb_pal, numero, sst, chauffeur, infos)
            else:  # ramasse
                ramasse = mission.get("ramasse", "")
                values = (heure, voyage_code, nb_pal, numero, sst, chauffeur, ramasse, infos)
            
            result.append({
                'id': mission_id,
                'values': values
            })
        
        # IMPORTANT: Trier par heure pour garantir un ordre stable
        result.sort(key=lambda x: x['values'][0] if x['values'] else '')
        
        return result
    
    def smart_refresh_chauffeurs(self):
        """Cette méthode n'est plus utilisée - on utilise refresh_chauffeurs_view()"""
        pass
    
    def diff_refresh_chauffeurs(self):
        """Refresh des chauffeurs par DIFFÉRENCE - ZÉRO clignotement"""
        if not hasattr(self, 'tree_ch'):
            return
        
        try:
            tree = self.tree_ch
            
            # Sauvegarder état
            old_selection = list(tree.selection())
            old_yview = tree.yview()[0] if tree.yview() else 0
            
            # État actuel
            current_items = {}
            for item_id in tree.get_children():
                current_items[item_id] = tree.item(item_id, 'values')
            
            # Nouvelles données
            new_items = {}
            for ch in self.chauffeurs:
                ch_id = ch["id"]
                new_values = (
                    ch.get("id", ""),
                    ch.get("nom", ""),
                    ch.get("prenom", ""),
                    ch.get("sst", ""),
                    ch.get("type", ""),
                    "Oui" if ch.get("actif", True) else "Non",
                )
                new_items[ch_id] = new_values
            
            # Mise à jour par différence
            for item_id, new_values in new_items.items():
                if item_id in current_items:
                    if current_items[item_id] != new_values:
                        tree.item(item_id, values=new_values)
                else:
                    tree.insert("", "end", iid=item_id, values=new_values)
            
            for item_id in current_items.keys():
                if item_id not in new_items:
                    tree.delete(item_id)
            
            # Restaurer état
            for item_id in old_selection:
                if tree.exists(item_id):
                    tree.selection_set(item_id)
            tree.yview_moveto(old_yview)
            
        except Exception as e:
            print(f"Erreur diff_refresh_chauffeurs: {e}")
    
    def smart_refresh_voyages(self):
        """Cette méthode n'est plus utilisée - on utilise refresh_voyages_view()"""
        pass
    
    def diff_refresh_voyages(self):
        """Refresh des voyages par DIFFÉRENCE - ZÉRO clignotement"""
        if not hasattr(self, 'tree_voy'):
            return
        
        try:
            tree = self.tree_voy
            
            # Sauvegarder état
            old_selection = list(tree.selection())
            old_yview = tree.yview()[0] if tree.yview() else 0
            
            # État actuel
            current_items = {}
            for item_id in tree.get_children():
                current_items[item_id] = tree.item(item_id, 'values')
            
            # Nouvelles données
            new_items = {}
            for voy in self.voyages:
                voy_id = voy.get("code", "")
                if voy_id:
                    new_values = (
                        voy.get("code", ""),
                        voy.get("description", ""),
                        voy.get("country", ""),
                    )
                    new_items[voy_id] = new_values
            
            # Mise à jour par différence
            for item_id, new_values in new_items.items():
                if item_id in current_items:
                    if current_items[item_id] != new_values:
                        tree.item(item_id, values=new_values)
                else:
                    tree.insert("", "end", iid=item_id, values=new_values)
            
            for item_id in current_items.keys():
                if item_id not in new_items:
                    tree.delete(item_id)
            
            # Restaurer état
            for item_id in old_selection:
                if tree.exists(item_id):
                    tree.selection_set(item_id)
            tree.yview_moveto(old_yview)
            
        except Exception as e:
            print(f"Erreur diff_refresh_voyages: {e}")
    
    def stop_auto_refresh(self):
        """Arrêter le refresh automatique"""
        self.auto_refresh_enabled = False
        if self.refresh_timer_id:
            self.root.after_cancel(self.refresh_timer_id)
            self.refresh_timer_id = None


if __name__ == "__main__":
    init_default_data()
    root = tk.Tk()
    
    if not EXCEL_AVAILABLE:
        print("Note: openpyxl non installé. Export Excel non disponible.")
        print("Pour l'installer: pip install openpyxl")
    if not PDF_AVAILABLE:
        print("Note: reportlab non installé. Export PDF non disponible.")
        print("Pour l'installer: pip install reportlab")
    if not MATPLOTLIB_AVAILABLE:
        print("Note: matplotlib non installé. Graphiques d'analyse non disponibles.")
        print("Pour l'installer: pip install matplotlib")
    
    # Définir l'icône de la fenêtre (tomato_16822.ico doit se trouver dans le même dossier que ce script)
    try:
        from pathlib import Path as _Path
        icon_path = _Path(__file__).with_name("tomato_16822.ico")
        if icon_path.exists():
            root.iconbitmap(default=str(icon_path))
    except Exception as e:
        print(f"Impossible de définir l'icône de la fenêtre: {e}")
    
    app = TransportPlannerApp(root)
    root.mainloop()